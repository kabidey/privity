from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, UploadFile, File, Form, Request, WebSocket, WebSocketDisconnect, Query, Response
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import StreamingResponse, FileResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta
import bcrypt
import jwt
import io
import csv
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
import aiofiles
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import base64
import asyncio
import random
import string

# Import from modular structure
from config import (
    MONGO_URL, DB_NAME, JWT_SECRET, JWT_ALGORITHM, JWT_EXPIRATION_HOURS,
    UPLOAD_DIR, EMAIL_HOST, EMAIL_PORT, EMAIL_USERNAME, EMAIL_PASSWORD, EMAIL_FROM,
    EMERGENT_LLM_KEY, OTP_EXPIRY_MINUTES, OTP_MAX_ATTEMPTS, FRONTEND_URL,
    ROLES, ROLE_PERMISSIONS, ALLOWED_EMAIL_DOMAINS, AUDIT_ACTIONS, DEFAULT_EMAIL_TEMPLATES,
    is_pe_level, is_pe_desk_only, has_finance_access, can_manage_finance
)
from database import db, client

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# Create the main app
app = FastAPI(title="PRIVITY - Private Equity System", version="2.0.0")
api_router = APIRouter(prefix="/api")
security = HTTPBearer()

# Import WebSocket manager from services (use local for now until full migration)
from services.notification_service import ws_manager, create_notification, notify_roles
from services.audit_service import create_audit_log
from services.ocr_service import process_document_ocr
# Note: update_inventory is kept local as the service version has slightly different logic

# Models (importing from models package for new code, keeping local for backward compat)
from models import (
    AuditLog, UserCreate, UserLogin, User, TokenResponse,
    PasswordResetRequest, PasswordResetVerify, Notification,
    BankAccount, ClientDocument, ClientCreate, Client,
    StockCreate, Stock, CorporateActionCreate, CorporateAction,
    PurchaseCreate, Purchase, Inventory as InventoryModel,
    PaymentTranche, PaymentTrancheCreate, BookingCreate, Booking,
    BookingWithDetails, DPTransferRecord, DashboardStats, ClientPortfolio,
    ClientConfirmationRequest, ClientSuspensionRequest, EmailTemplateUpdate, EmailTemplatePreview
)

# Import helper functions from utils
from utils.auth import hash_password, verify_password, create_token, get_current_user, check_permission

# Import email functions from services  
from services.email_service import send_email, send_templated_email, generate_otp, send_otp_email

# Import audit service
from services.audit_service import create_audit_log

# Import modular routers
from routers import (
    email_templates_router,
    smtp_config_router,
    stocks_router,
    database_backup_router,
    users_router,
    bookings_router,
    clients_router,
    finance_router
)

# Keep local Inventory alias for backward compatibility
Inventory = InventoryModel

# Unique local helper functions
def generate_otc_ucc() -> str:
    """Generate unique OTC UCC code"""
    date_part = datetime.now(timezone.utc).strftime("%Y%m%d")
    unique_part = str(uuid.uuid4())[:8].upper()
    return f"OTC{date_part}{unique_part}"


# Lock for booking number generation to prevent race conditions
_booking_number_lock = asyncio.Lock()

async def generate_booking_number() -> str:
    """Generate unique human-readable booking number using atomic counter."""
    year = datetime.now(timezone.utc).strftime("%Y")
    
    async with _booking_number_lock:
        # Use atomic findAndModify to get next sequence
        counter = await db.counters.find_one_and_update(
            {"_id": f"booking_{year}"},
            {"$inc": {"seq": 1}},
            upsert=True,
            return_document=True
        )
        
        seq_num = counter.get("seq", 1)
        return f"BK-{year}-{seq_num:05d}"


def get_client_emails(client: dict) -> list:
    """Get all email addresses for a client (primary, secondary, tertiary)"""
    emails = []
    if client.get("email"):
        emails.append(client["email"])
    if client.get("email_secondary"):
        emails.append(client["email_secondary"])
    if client.get("email_tertiary"):
        emails.append(client["email_tertiary"])
    return emails

async def update_inventory(stock_id: str):
    """Recalculate weighted average and available quantity for a stock.
    
    Inventory Logic:
    - available_quantity: Stock available for new bookings (purchased - transferred - blocked)
    - blocked_quantity: Stock reserved for approved bookings pending transfer
    - weighted_avg_price: Calculated from total purchased value / total purchased quantity
    - Blocked stock is NOT used in weighted average calculation for new bookings
    
    Stock is blocked as soon as PE Desk approves the booking (not waiting for client confirmation)
    """
    # Get all purchases for this stock
    purchases = await db.purchases.find({"stock_id": stock_id}, {"_id": 0}).to_list(10000)
    
    # Get all bookings for this stock
    bookings = await db.bookings.find({"stock_id": stock_id}, {"_id": 0}).to_list(10000)
    
    # Calculate total purchased
    total_purchased_qty = sum(p["quantity"] for p in purchases)
    total_purchased_value = sum(p["quantity"] * p["price_per_unit"] for p in purchases)
    
    # Calculate blocked quantity (approved bookings not yet transferred)
    # Stock is blocked as soon as PE Desk approves - NOT waiting for client confirmation
    # Booking must be: approval_status=approved, NOT voided, NOT transferred
    blocked_qty = sum(
        b["quantity"] for b in bookings 
        if b.get("approval_status") == "approved" 
        and not b.get("is_voided", False)
        and not b.get("stock_transferred", False)
    )
    
    # Calculate transferred quantity (completed sales)
    transferred_qty = sum(
        b["quantity"] for b in bookings 
        if b.get("stock_transferred") == True
        and not b.get("is_voided", False)
    )
    
    # Available = purchased - transferred - blocked
    available_qty = total_purchased_qty - transferred_qty - blocked_qty
    
    # Calculate weighted average from total purchases (not affected by blocked)
    weighted_avg = total_purchased_value / total_purchased_qty if total_purchased_qty > 0 else 0
    
    # Get stock details
    stock = await db.stocks.find_one({"id": stock_id}, {"_id": 0})
    
    # Update or insert inventory
    inventory_data = {
        "stock_id": stock_id,
        "stock_symbol": stock["symbol"] if stock else "Unknown",
        "stock_name": stock["name"] if stock else "Unknown",
        "available_quantity": max(0, available_qty),  # Ensure non-negative
        "blocked_quantity": blocked_qty,
        "weighted_avg_price": weighted_avg,
        "total_value": max(0, available_qty) * weighted_avg
    }
    
    await db.inventory.update_one(
        {"stock_id": stock_id},
        {"$set": inventory_data},
        upsert=True
    )
    
    return inventory_data

# Startup event to seed admin user
@app.on_event("startup")
async def seed_admin_user():
    """Create default PE Desk super admin - ALWAYS ensures admin exists"""
    try:
        # Check if pedesk@smifs.com user exists
        pedesk_user = await db.users.find_one({"email": "pedesk@smifs.com"}, {"_id": 0})
        
        if pedesk_user:
            # User exists - update password to ensure it's correct
            await db.users.update_one(
                {"email": "pedesk@smifs.com"},
                {"$set": {
                    "password": hash_password("Kutta@123"),
                    "role": 1,
                    "name": "PE Desk Super Admin"
                }}
            )
            logging.info("PE Desk super admin password reset: pedesk@smifs.com")
        else:
            # Check if any admin exists
            admin_exists = await db.users.find_one({"role": {"$lte": 2}}, {"_id": 0})
            
            if not admin_exists:
                # Create new admin
                admin_id = str(uuid.uuid4())
                admin_doc = {
                    "id": admin_id,
                    "email": "pedesk@smifs.com",
                    "password": hash_password("Kutta@123"),
                    "name": "PE Desk Super Admin",
                    "role": 1,  # PE Desk - full access
                    "created_at": datetime.now(timezone.utc).isoformat()
                }
                await db.users.insert_one(admin_doc)
                logging.info("Default PE Desk super admin created: pedesk@smifs.com")
            else:
                # Admin exists but not pedesk - create pedesk anyway
                admin_id = str(uuid.uuid4())
                admin_doc = {
                    "id": admin_id,
                    "email": "pedesk@smifs.com",
                    "password": hash_password("Kutta@123"),
                    "name": "PE Desk Super Admin",
                    "role": 1,
                    "created_at": datetime.now(timezone.utc).isoformat()
                }
                await db.users.insert_one(admin_doc)
                logging.info("PE Desk super admin created alongside existing admin: pedesk@smifs.com")
                
    except Exception as e:
        logging.error(f"Error seeding admin user: {e}")

# Auth Routes
@api_router.post("/auth/register", response_model=TokenResponse)
async def register(user_data: UserCreate, request: Request = None):
    # Check email domain restriction
    email_domain = user_data.email.split('@')[-1].lower()
    if email_domain not in ALLOWED_EMAIL_DOMAINS:
        raise HTTPException(
            status_code=400, 
            detail=f"Registration is restricted to employees with @smifs.com email addresses"
        )
    
    existing_user = await db.users.find_one({"email": user_data.email}, {"_id": 0})
    if existing_user:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    user_id = str(uuid.uuid4())
    hashed_pw = hash_password(user_data.password)
    
    # Default role is Employee (4) for smifs.com domain
    user_role = 4  # Employee
    
    user_doc = {
        "id": user_id,
        "email": user_data.email,
        "password": hashed_pw,
        "name": user_data.name,
        "role": user_role,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.users.insert_one(user_doc)
    
    # Create audit log
    await create_audit_log(
        action="USER_REGISTER",
        entity_type="user",
        entity_id=user_id,
        user_id=user_id,
        user_name=user_data.name,
        user_role=user_role,
        entity_name=user_data.name,
        details={"email": user_data.email, "role": user_role}
    )
    
    token = create_token(user_id, user_data.email)
    user_response = User(
        id=user_id,
        email=user_data.email,
        name=user_data.name,
        role=user_role,
        role_name=ROLES.get(user_role, "Employee"),
        created_at=user_doc["created_at"]
    )
    
    return TokenResponse(token=token, user=user_response)

@api_router.post("/auth/login", response_model=TokenResponse)
async def login(login_data: UserLogin):
    user = await db.users.find_one({"email": login_data.email}, {"_id": 0})
    if not user or not verify_password(login_data.password, user["password"]):
        raise HTTPException(status_code=401, detail="Invalid email or password")
    
    # Create audit log for login
    await create_audit_log(
        action="USER_LOGIN",
        entity_type="user",
        entity_id=user["id"],
        user_id=user["id"],
        user_name=user["name"],
        user_role=user.get("role", 5),
        entity_name=user["name"]
    )
    
    token = create_token(user["id"], user["email"])
    user_response = User(
        id=user["id"],
        email=user["email"],
        name=user["name"],
        role=user.get("role", 5),
        role_name=ROLES.get(user.get("role", 5), "Viewer"),
        created_at=user["created_at"]
    )
    
    return TokenResponse(token=token, user=user_response)

@api_router.get("/auth/me", response_model=User)
async def get_me(current_user: dict = Depends(get_current_user)):
    return User(
        id=current_user["id"],
        email=current_user["email"],
        name=current_user["name"],
        role=current_user.get("role", 5),
        role_name=ROLES.get(current_user.get("role", 5), "Viewer"),
        created_at=current_user["created_at"]
    )

@api_router.post("/auth/forgot-password")
async def forgot_password(data: PasswordResetRequest):
    """Request password reset OTP"""
    user = await db.users.find_one({"email": data.email.lower()})
    if not user:
        # Don't reveal if email exists
        return {"message": "If the email exists, an OTP has been sent"}
    
    # Check rate limiting
    recent_otps = await db.password_resets.count_documents({
        "email": data.email.lower(),
        "created_at": {"$gte": datetime.now(timezone.utc) - timedelta(minutes=OTP_EXPIRY_MINUTES)}
    })
    
    if recent_otps >= OTP_MAX_ATTEMPTS:
        raise HTTPException(status_code=429, detail="Too many OTP requests. Please try again later.")
    
    # Generate OTP
    otp = generate_otp()
    
    # Store OTP
    await db.password_resets.insert_one({
        "id": str(uuid.uuid4()),
        "email": data.email.lower(),
        "otp": otp,
        "attempts": 0,
        "used": False,
        "created_at": datetime.now(timezone.utc),
        "expires_at": datetime.now(timezone.utc) + timedelta(minutes=OTP_EXPIRY_MINUTES)
    })
    
    # Send email
    await send_otp_email(data.email, otp, user["name"])
    
    return {"message": "If the email exists, an OTP has been sent"}

@api_router.post("/auth/reset-password")
async def reset_password(data: PasswordResetVerify, request: Request):
    """Reset password with OTP verification"""
    # Find valid OTP
    otp_record = await db.password_resets.find_one({
        "email": data.email.lower(),
        "otp": data.otp,
        "used": False,
        "expires_at": {"$gte": datetime.now(timezone.utc)}
    })
    
    if not otp_record:
        # Increment attempts for any matching unused OTP
        await db.password_resets.update_many(
            {"email": data.email.lower(), "used": False},
            {"$inc": {"attempts": 1}}
        )
        raise HTTPException(status_code=400, detail="Invalid or expired OTP")
    
    # Check attempts
    if otp_record.get("attempts", 0) >= OTP_MAX_ATTEMPTS:
        raise HTTPException(status_code=400, detail="Too many failed attempts. Please request a new OTP.")
    
    # Mark OTP as used
    await db.password_resets.update_one(
        {"id": otp_record["id"]},
        {"$set": {"used": True}}
    )
    
    # Update password
    user = await db.users.find_one({"email": data.email.lower()})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    hashed = bcrypt.hashpw(data.new_password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
    await db.users.update_one(
        {"email": data.email.lower()},
        {"$set": {"hashed_password": hashed}}
    )
    
    # Create audit log
    await create_audit_log(
        "USER_PASSWORD_RESET", "user", user["id"], user["id"], user["name"], user["role"],
        ip_address=request.client.host if request.client else None
    )
    
    return {"message": "Password reset successfully"}

# User Management Routes
@api_router.get("/users", response_model=List[User])
async def get_users(current_user: dict = Depends(get_current_user)):
    check_permission(current_user, "manage_users")
    users = await db.users.find({}, {"_id": 0, "password": 0}).to_list(1000)
    return [User(**{**u, "role_name": ROLES.get(u.get("role", 5), "Viewer")}) for u in users]

@api_router.put("/users/{user_id}/role")
async def update_user_role(user_id: str, role: int, current_user: dict = Depends(get_current_user)):
    check_permission(current_user, "manage_users")
    
    if role not in ROLES:
        raise HTTPException(status_code=400, detail="Invalid role")
    
    result = await db.users.update_one({"id": user_id}, {"$set": {"role": role}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    
    return {"message": "User role updated successfully"}

# Client Routes with Document Upload
@api_router.post("/clients", response_model=Client)
async def create_client(client_data: ClientCreate, current_user: dict = Depends(get_current_user)):
    user_role = current_user.get("role", 5)
    
    # Employees can only create clients (not vendors)
    if user_role == 4 and client_data.is_vendor:
        raise HTTPException(status_code=403, detail="Employees cannot create vendors")
    
    # Validate trading_ucc is required when dp_type is "smifs"
    if client_data.dp_type == "smifs" and not client_data.trading_ucc:
        raise HTTPException(status_code=400, detail="Trading UCC is required when DP is with SMIFS")
    
    # Check for duplicate client by PAN number
    existing_by_pan = await db.clients.find_one(
        {"pan_number": client_data.pan_number.upper(), "is_vendor": client_data.is_vendor},
        {"_id": 0, "name": 1, "otc_ucc": 1}
    )
    if existing_by_pan:
        entity_type = "Vendor" if client_data.is_vendor else "Client"
        raise HTTPException(
            status_code=400, 
            detail=f"{entity_type} with PAN {client_data.pan_number} already exists: {existing_by_pan['name']} ({existing_by_pan.get('otc_ucc', 'N/A')})"
        )
    
    # Check for duplicate client by DP ID
    existing_by_dp = await db.clients.find_one(
        {"dp_id": client_data.dp_id, "is_vendor": client_data.is_vendor},
        {"_id": 0, "name": 1, "otc_ucc": 1}
    )
    if existing_by_dp:
        entity_type = "Vendor" if client_data.is_vendor else "Client"
        raise HTTPException(
            status_code=400, 
            detail=f"{entity_type} with DP ID {client_data.dp_id} already exists: {existing_by_dp['name']} ({existing_by_dp.get('otc_ucc', 'N/A')})"
        )
    
    # Check permission based on whether it's a client or vendor
    if client_data.is_vendor:
        check_permission(current_user, "manage_vendors")
    else:
        if user_role == 4:
            check_permission(current_user, "create_clients")
        else:
            check_permission(current_user, "manage_clients")
    
    client_id = str(uuid.uuid4())
    # Generate unique OTC UCC code
    otc_ucc = f"OTC{datetime.now().strftime('%Y%m%d')}{client_id[:8].upper()}"
    
    # Determine approval status - employees need approval
    is_active = True
    approval_status = "approved"
    if user_role == 4 and not client_data.is_vendor:
        is_active = False
        approval_status = "pending"
    
    client_doc = {
        "id": client_id,
        "otc_ucc": otc_ucc,
        **client_data.model_dump(),
        "bank_accounts": [acc.model_dump() if hasattr(acc, 'model_dump') else acc for acc in client_data.bank_accounts] if client_data.bank_accounts else [],
        "is_active": is_active,
        "approval_status": approval_status,
        "documents": [],
        "mapped_employee_id": current_user["id"] if user_role == 4 else None,
        "mapped_employee_name": current_user["name"] if user_role == 4 else None,
        "user_id": current_user["id"],
        "created_by": current_user["id"],
        "created_by_role": user_role,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.clients.insert_one(client_doc)
    
    # Send email notification using template to all client emails
    client_emails = get_client_emails(client_doc)
    if client_emails and is_active:
        primary_email = client_emails[0]
        additional_emails = client_emails[1:] if len(client_emails) > 1 else None
        await send_templated_email(
            "welcome",
            primary_email,
            {"client_name": client_data.name},
            additional_emails=additional_emails
        )
    
    # Real-time notification for pending approval
    if approval_status == "pending":
        # Notify managers (roles 1, 2, 3) about pending client
        await notify_roles(
            [1, 2, 3],
            "client_pending",
            "New Client Pending Approval",
            f"Client '{client_data.name}' created by {current_user['name']} is pending approval",
            {"client_id": client_id, "client_name": client_data.name}
        )
    
    return Client(**{k: v for k, v in client_doc.items() if k != "user_id"})

@api_router.put("/clients/{client_id}/approve")
async def approve_client(client_id: str, approve: bool = True, current_user: dict = Depends(get_current_user)):
    """Approve or reject a client created by employee"""
    check_permission(current_user, "approve_clients")
    
    client = await db.clients.find_one({"id": client_id}, {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    
    update_data = {
        "is_active": approve,
        "approval_status": "approved" if approve else "rejected"
    }
    
    await db.clients.update_one({"id": client_id}, {"$set": update_data})
    
    # Send notification email if approved using template
    if approve and client.get("email"):
        await send_templated_email(
            "client_approved",
            client["email"],
            {"client_name": client["name"], "otc_ucc": client.get("otc_ucc", "N/A")}
        )
    
    # Real-time notification to creator
    if client.get("created_by"):
        notif_type = "client_approved" if approve else "client_rejected"
        notif_title = f"Client {'Approved' if approve else 'Rejected'}"
        notif_message = f"Client '{client['name']}' has been {'approved' if approve else 'rejected'} by {current_user['name']}"
        await create_notification(
            client["created_by"],
            notif_type,
            notif_title,
            notif_message,
            {"client_id": client_id, "client_name": client["name"]}
        )
    
    return {"message": f"Client {'approved' if approve else 'rejected'} successfully"}

@api_router.post("/clients/{client_id}/bank-account")
async def add_bank_account(client_id: str, bank_account: BankAccount, current_user: dict = Depends(get_current_user)):
    """Add another bank account to client"""
    check_permission(current_user, "manage_clients")
    
    client = await db.clients.find_one({"id": client_id}, {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    
    await db.clients.update_one(
        {"id": client_id},
        {"$push": {"bank_accounts": bank_account.model_dump()}}
    )
    
    return {"message": "Bank account added successfully"}

@api_router.post("/clients/{client_id}/documents")
async def upload_client_document(
    client_id: str,
    doc_type: str = Form(...),
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    # Allow both manage_clients (managers+) and create_clients (employees) to upload docs
    user_role = current_user.get("role", 5)
    if user_role == 4:
        check_permission(current_user, "create_clients")
        # Employees can only upload docs to their own clients
        client = await db.clients.find_one({"id": client_id}, {"_id": 0})
        if not client:
            raise HTTPException(status_code=404, detail="Client not found")
        if client.get("created_by") != current_user["id"]:
            raise HTTPException(status_code=403, detail="You can only upload documents to your own clients")
    else:
        check_permission(current_user, "manage_clients")
    
    # Verify client exists (only if not already verified for employees)
    if user_role != 4:
        client = await db.clients.find_one({"id": client_id}, {"_id": 0})
        if not client:
            raise HTTPException(status_code=404, detail="Client not found")
    
    # Create client directory
    client_dir = UPLOAD_DIR / client_id
    client_dir.mkdir(exist_ok=True)
    
    # Save file
    file_ext = Path(file.filename).suffix
    filename = f"{doc_type}_{datetime.now().strftime('%Y%m%d%H%M%S')}{file_ext}"
    file_path = client_dir / filename
    
    async with aiofiles.open(file_path, 'wb') as f:
        content = await file.read()
        await f.write(content)
    
    # Process OCR
    ocr_data = await process_document_ocr(str(file_path), doc_type)
    
    # Update client document record
    doc_record = {
        "doc_type": doc_type,
        "filename": filename,
        "file_path": str(file_path),
        "upload_date": datetime.now(timezone.utc).isoformat(),
        "ocr_data": ocr_data
    }
    
    await db.clients.update_one(
        {"id": client_id},
        {"$push": {"documents": doc_record}}
    )
    
    return {"message": "Document uploaded successfully", "document": doc_record}

@api_router.get("/clients/{client_id}/documents/{filename}")
async def download_client_document(
    client_id: str,
    filename: str,
    current_user: dict = Depends(get_current_user)
):
    file_path = UPLOAD_DIR / client_id / filename
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="Document not found")
    
    return FileResponse(file_path, filename=filename)

@api_router.post("/ocr/preview")
async def ocr_preview(
    doc_type: str = Form(...),
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """Process OCR on a document without saving - for auto-fill preview"""
    # Save temporarily
    temp_dir = UPLOAD_DIR / "temp"
    temp_dir.mkdir(exist_ok=True)
    
    file_ext = Path(file.filename).suffix
    temp_filename = f"temp_{uuid.uuid4()}{file_ext}"
    temp_path = temp_dir / temp_filename
    
    try:
        async with aiofiles.open(temp_path, 'wb') as f:
            content = await file.read()
            await f.write(content)
        
        # Process OCR
        ocr_data = await process_document_ocr(str(temp_path), doc_type)
        return ocr_data
    finally:
        # Clean up temp file
        if temp_path.exists():
            temp_path.unlink()

@api_router.get("/clients/{client_id}/documents/{filename}/ocr")
async def get_document_ocr(
    client_id: str,
    filename: str,
    current_user: dict = Depends(get_current_user)
):
    """Get OCR data for a specific document"""
    client = await db.clients.find_one({"id": client_id}, {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    
    # Find the document
    document = None
    for doc in client.get("documents", []):
        if doc["filename"] == filename:
            document = doc
            break
    
    if not document:
        raise HTTPException(status_code=404, detail="Document not found")
    
    return {
        "filename": filename,
        "doc_type": document.get("doc_type"),
        "ocr_data": document.get("ocr_data"),
        "upload_date": document.get("upload_date")
    }

@api_router.put("/clients/{client_id}/employee-mapping")
async def update_client_employee_mapping(
    client_id: str,
    employee_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Map or unmap a client to an employee (admin only)"""
    # Only PE Desk and Zonal Manager can map/unmap clients
    if current_user.get("role", 5) > 2:
        raise HTTPException(status_code=403, detail="Only admins can map/unmap clients")
    
    client = await db.clients.find_one({"id": client_id}, {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    
    update_data = {}
    
    if employee_id:
        # Map to employee
        employee = await db.users.find_one({"id": employee_id}, {"_id": 0})
        if not employee:
            raise HTTPException(status_code=404, detail="Employee not found")
        
        update_data = {
            "mapped_employee_id": employee_id,
            "mapped_employee_name": employee.get("name")
        }
    else:
        # Unmap (set to None)
        update_data = {
            "mapped_employee_id": None,
            "mapped_employee_name": None
        }
    
    await db.clients.update_one(
        {"id": client_id},
        {"$set": update_data}
    )
    
    action = "mapped" if employee_id else "unmapped"
    return {"message": f"Client successfully {action}", "client_id": client_id, **update_data}

@api_router.get("/employees", response_model=List[User])
async def get_employees(current_user: dict = Depends(get_current_user)):
    """Get list of employees for mapping"""
    users = await db.users.find({}, {"_id": 0, "password": 0}).to_list(1000)
    return [User(
        id=u["id"],
        email=u["email"],
        name=u["name"],
        role=u["role"],
        role_name=ROLES.get(u["role"], "Unknown"),
        created_at=u["created_at"]
    ) for u in users]

@api_router.get("/clients", response_model=List[Client])
async def get_clients(
    search: Optional[str] = None,
    is_vendor: Optional[bool] = None,
    pending_approval: Optional[bool] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    user_role = current_user.get("role", 5)
    user_id = current_user.get("id")
    
    # Employee and Finance role restrictions (roles 4 and 7)
    if user_role in [4, 7]:
        # Employees and Finance cannot see vendors
        if is_vendor == True:
            raise HTTPException(status_code=403, detail="You do not have access to vendors")
        query["is_vendor"] = False
        # Employees can only see their own clients (Finance role sees all clients)
        if user_role == 4:
            query["$or"] = [
                {"mapped_employee_id": user_id},
                {"created_by": user_id}
            ]
    else:
        # Vendor filter for non-employees
        if is_vendor is not None:
            query["is_vendor"] = is_vendor
    
    # Pending approval filter (for admins)
    if pending_approval and user_role <= 3:
        query["approval_status"] = "pending"
    
    # Search filter
    if search:
        search_query = [
            {"name": {"$regex": search, "$options": "i"}},
            {"email": {"$regex": search, "$options": "i"}},
            {"pan_number": {"$regex": search, "$options": "i"}},
            {"otc_ucc": {"$regex": search, "$options": "i"}}
        ]
        if "$or" in query:
            # Combine with existing $or
            existing_or = query.pop("$or")
            query["$and"] = [{"$or": existing_or}, {"$or": search_query}]
        else:
            query["$or"] = search_query
    
    clients = await db.clients.find(query, {"_id": 0, "user_id": 0}).to_list(1000)
    return clients

@api_router.get("/clients/pending-approval", response_model=List[Client])
async def get_pending_clients(current_user: dict = Depends(get_current_user)):
    """Get clients pending approval (admin only)"""
    check_permission(current_user, "approve_clients")
    
    clients = await db.clients.find(
        {"approval_status": "pending", "is_vendor": False},
        {"_id": 0, "user_id": 0}
    ).to_list(1000)
    return clients

@api_router.get("/clients/{client_id}", response_model=Client)
async def get_client(client_id: str, current_user: dict = Depends(get_current_user)):
    client = await db.clients.find_one({"id": client_id}, {"_id": 0, "user_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    return client

@api_router.put("/clients/{client_id}", response_model=Client)
async def update_client(client_id: str, client_data: ClientCreate, current_user: dict = Depends(get_current_user)):
    # PE Level (PE Desk or PE Manager) can modify clients
    if not is_pe_level(current_user.get("role", 6)):
        raise HTTPException(status_code=403, detail="Only PE Desk or PE Manager can modify clients")
    
    result = await db.clients.update_one(
        {"id": client_id},
        {"$set": client_data.model_dump()}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Client not found")
    
    updated_client = await db.clients.find_one({"id": client_id}, {"_id": 0, "user_id": 0})
    return updated_client

@api_router.put("/clients/{client_id}/employee-mapping")
async def update_client_employee_mapping(
    client_id: str,
    employee_id: str,
    employee_name: str,
    current_user: dict = Depends(get_current_user)
):
    """Map a client to an employee"""
    check_permission(current_user, "manage_clients")
    
    result = await db.clients.update_one(
        {"id": client_id},
        {"$set": {
            "mapped_employee_id": employee_id,
            "mapped_employee_name": employee_name
        }}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Client not found")
    
    return {"message": "Employee mapping updated successfully"}

@api_router.delete("/clients/{client_id}")
async def delete_client(client_id: str, current_user: dict = Depends(get_current_user)):
    # Only PE Desk (role 1) can delete clients (PE Manager cannot delete)
    if not is_pe_desk_only(current_user.get("role", 6)):
        raise HTTPException(status_code=403, detail="Only PE Desk can delete clients")
    
    result = await db.clients.delete_one({"id": client_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Client not found")
    return {"message": "Client deleted successfully"}


# Suspend/Unsuspend Client (PE Level Only)
@api_router.put("/clients/{client_id}/suspend")
async def suspend_client(
    client_id: str,
    suspension_data: ClientSuspensionRequest,
    current_user: dict = Depends(get_current_user)
):
    """Suspend a client with a reason (PE Desk or PE Manager)"""
    # PE Level can suspend clients
    if not is_pe_level(current_user.get("role", 6)):
        raise HTTPException(status_code=403, detail="Only PE Desk or PE Manager can suspend clients")
    
    client = await db.clients.find_one({"id": client_id}, {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    
    if client.get("is_suspended"):
        raise HTTPException(status_code=400, detail="Client is already suspended")
    
    # Update client suspension status
    await db.clients.update_one(
        {"id": client_id},
        {"$set": {
            "is_suspended": True,
            "suspension_reason": suspension_data.reason,
            "suspended_by": current_user["id"],
            "suspended_by_name": current_user["name"],
            "suspended_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # Log the action
    await create_audit_log(
        action="client_suspended",
        entity_type="client",
        entity_id=client_id,
        user_id=current_user["id"],
        user_name=current_user["name"],
        user_role=current_user.get("role", 5),
        entity_name=client.get("name"),
        details={"reason": suspension_data.reason}
    )
    
    return {"message": f"Client {client.get('name')} has been suspended", "reason": suspension_data.reason}


@api_router.put("/clients/{client_id}/unsuspend")
async def unsuspend_client(
    client_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Unsuspend a client (PE Desk or PE Manager)"""
    # PE Level can unsuspend clients
    if not is_pe_level(current_user.get("role", 6)):
        raise HTTPException(status_code=403, detail="Only PE Desk or PE Manager can unsuspend clients")
    
    client = await db.clients.find_one({"id": client_id}, {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    
    if not client.get("is_suspended"):
        raise HTTPException(status_code=400, detail="Client is not suspended")
    
    # Update client suspension status
    await db.clients.update_one(
        {"id": client_id},
        {"$set": {
            "is_suspended": False,
            "suspension_reason": None,
            "suspended_by": None,
            "suspended_by_name": None,
            "suspended_at": None
        }}
    )
    
    # Log the action
    await create_audit_log(
        action="client_unsuspended",
        entity_type="client",
        entity_id=client_id,
        user_id=current_user["id"],
        user_name=current_user["name"],
        user_role=current_user.get("role", 5),
        entity_name=client.get("name")
    )
    
    return {"message": f"Client {client.get('name')} has been unsuspended"}


# Bulk Upload Routes
@api_router.post("/clients/bulk-upload")
async def bulk_upload_clients(file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
    check_permission(current_user, "manage_clients")
    
    try:
        content = await file.read()
        df = pd.read_csv(io.BytesIO(content))
        
        required_columns = ["name", "pan_number", "dp_id"]
        if not all(col in df.columns for col in required_columns):
            raise HTTPException(status_code=400, detail=f"CSV must contain columns: {', '.join(required_columns)}")
        
        clients_created = 0
        for _, row in df.iterrows():
            client_id = str(uuid.uuid4())
            # Generate unique OTC UCC code
            otc_ucc = f"OTC{datetime.now().strftime('%Y%m%d')}{client_id[:8].upper()}"
            
            client_doc = {
                "id": client_id,
                "otc_ucc": otc_ucc,
                "name": str(row["name"]),
                "email": str(row.get("email", "")) if pd.notna(row.get("email")) else None,
                "phone": str(row.get("phone", "")) if pd.notna(row.get("phone")) else None,
                "pan_number": str(row["pan_number"]),
                "dp_id": str(row["dp_id"]),
                "bank_name": str(row.get("bank_name", "")) if pd.notna(row.get("bank_name")) else None,
                "account_number": str(row.get("account_number", "")) if pd.notna(row.get("account_number")) else None,
                "ifsc_code": str(row.get("ifsc_code", "")) if pd.notna(row.get("ifsc_code")) else None,
                "is_vendor": bool(row.get("is_vendor", False)) if pd.notna(row.get("is_vendor")) else False,
                "documents": [],
                "mapped_employee_id": None,
                "mapped_employee_name": None,
                "user_id": current_user["id"],
                "created_by": current_user["id"],
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            
            await db.clients.insert_one(client_doc)
            clients_created += 1
        
        return {"message": f"Successfully uploaded {clients_created} clients"}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error processing CSV: {str(e)}")

# Clone Client/Vendor Endpoint (PE Desk Only)
@api_router.post("/clients/{client_id}/clone")
async def clone_client_vendor(
    client_id: str, 
    target_type: str = Query(..., description="Target type: 'client' or 'vendor'"),
    current_user: dict = Depends(get_current_user)
):
    """Clone a client as vendor or vendor as client (PE Desk or PE Manager)"""
    # PE Level can clone
    if not is_pe_level(current_user.get("role", 6)):
        raise HTTPException(status_code=403, detail="Only PE Desk or PE Manager can clone clients/vendors")
    
    if target_type not in ["client", "vendor"]:
        raise HTTPException(status_code=400, detail="target_type must be 'client' or 'vendor'")
    
    # Get the source client/vendor
    source = await db.clients.find_one({"id": client_id}, {"_id": 0})
    if not source:
        raise HTTPException(status_code=404, detail="Source client/vendor not found")
    
    # Check if already exists as target type
    is_currently_vendor = source.get("is_vendor", False)
    if (target_type == "vendor" and is_currently_vendor) or (target_type == "client" and not is_currently_vendor):
        raise HTTPException(status_code=400, detail=f"This is already a {target_type}")
    
    # Check if a client/vendor with same PAN already exists as target type
    existing = await db.clients.find_one({
        "pan_number": source["pan_number"],
        "is_vendor": target_type == "vendor"
    }, {"_id": 0})
    
    if existing:
        raise HTTPException(
            status_code=400, 
            detail=f"A {target_type} with PAN {source['pan_number']} already exists"
        )
    
    # Create new cloned entry
    new_id = str(uuid.uuid4())
    new_otc_ucc = f"OTC{datetime.now().strftime('%Y%m%d')}{new_id[:8].upper()}"
    
    cloned_doc = {
        "id": new_id,
        "otc_ucc": new_otc_ucc,
        "name": source["name"],
        "email": source.get("email"),
        "phone": source.get("phone"),
        "mobile": source.get("mobile"),
        "pan_number": source["pan_number"],
        "dp_id": source["dp_id"],
        "dp_type": source.get("dp_type", "outside"),
        "trading_ucc": source.get("trading_ucc"),
        "address": source.get("address"),
        "pin_code": source.get("pin_code"),
        "bank_accounts": source.get("bank_accounts", []),
        "is_vendor": target_type == "vendor",
        "is_active": True,
        "approval_status": "approved",
        "documents": [],  # Documents are not cloned
        "user_id": current_user["id"],
        "created_by": current_user["id"],
        "created_by_role": 1,
        "mapped_employee_id": None,
        "mapped_employee_name": None,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.clients.insert_one(cloned_doc)
    
    # Create audit log
    source_type = "vendor" if is_currently_vendor else "client"
    await create_audit_log(
        action="CLIENT_CREATE",
        entity_type=target_type,
        entity_id=new_id,
        user_id=current_user["id"],
        user_name=current_user["name"],
        user_role=1,
        entity_name=source["name"],
        details={
            "cloned_from": client_id,
            "source_type": source_type,
            "target_type": target_type,
            "pan_number": source["pan_number"]
        }
    )
    
    return {
        "message": f"Successfully cloned {source_type} '{source['name']}' as {target_type}",
        "id": new_id,
        "otc_ucc": new_otc_ucc
    }

# Stock Routes (PE Desk Only for creation/edit)
@api_router.post("/stocks", response_model=Stock)
async def create_stock(stock_data: StockCreate, current_user: dict = Depends(get_current_user)):
    user_role = current_user.get("role", 6)
    
    # PE Level can create stocks
    if not is_pe_level(user_role):
        raise HTTPException(status_code=403, detail="Only PE Desk or PE Manager can add stocks")
    
    stock_id = str(uuid.uuid4())
    stock_doc = {
        "id": stock_id,
        **stock_data.model_dump(),
        "user_id": current_user["id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.stocks.insert_one(stock_doc)
    
    # Create audit log
    await create_audit_log(
        action="STOCK_CREATE",
        entity_type="stock",
        entity_id=stock_id,
        user_id=current_user["id"],
        user_name=current_user["name"],
        user_role=user_role,
        entity_name=stock_data.symbol,
        details={"symbol": stock_data.symbol, "name": stock_data.name, "isin": stock_data.isin_number}
    )
    
    return Stock(**{k: v for k, v in stock_doc.items() if k != "user_id"})

@api_router.get("/stocks", response_model=List[Stock])
async def get_stocks(search: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    query = {}
    
    if search:
        query["$or"] = [
            {"symbol": {"$regex": search, "$options": "i"}},
            {"name": {"$regex": search, "$options": "i"}}
        ]
    
    stocks = await db.stocks.find(query, {"_id": 0, "user_id": 0}).to_list(1000)
    return stocks

@api_router.get("/stocks/{stock_id}", response_model=Stock)
async def get_stock(stock_id: str, current_user: dict = Depends(get_current_user)):
    stock = await db.stocks.find_one({"id": stock_id}, {"_id": 0, "user_id": 0})
    if not stock:
        raise HTTPException(status_code=404, detail="Stock not found")
    return stock

@api_router.put("/stocks/{stock_id}", response_model=Stock)
async def update_stock(stock_id: str, stock_data: StockCreate, current_user: dict = Depends(get_current_user)):
    user_role = current_user.get("role", 6)
    
    # PE Level can update stocks
    if not is_pe_level(user_role):
        raise HTTPException(status_code=403, detail="Only PE Desk or PE Manager can update stocks")
    
    result = await db.stocks.update_one(
        {"id": stock_id},
        {"$set": stock_data.model_dump()}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Stock not found")
    
    updated_stock = await db.stocks.find_one({"id": stock_id}, {"_id": 0, "user_id": 0})
    return updated_stock

@api_router.delete("/stocks/{stock_id}")
async def delete_stock(stock_id: str, current_user: dict = Depends(get_current_user)):
    user_role = current_user.get("role", 6)
    
    # Only PE Desk can delete stocks (PE Manager cannot delete)
    if not is_pe_desk_only(user_role):
        raise HTTPException(status_code=403, detail="Only PE Desk can delete stocks")
    
    result = await db.stocks.delete_one({"id": stock_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Stock not found")
    return {"message": "Stock deleted successfully"}

@api_router.post("/stocks/bulk-upload")
async def bulk_upload_stocks(file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
    user_role = current_user.get("role", 6)
    
    # PE Level can bulk upload stocks
    if not is_pe_level(user_role):
        raise HTTPException(status_code=403, detail="Only PE Desk or PE Manager can bulk upload stocks")
    
    try:
        content = await file.read()
        df = pd.read_csv(io.BytesIO(content))
        
        required_columns = ["symbol", "name"]
        if not all(col in df.columns for col in required_columns):
            raise HTTPException(status_code=400, detail=f"CSV must contain columns: {', '.join(required_columns)}")
        
        stocks_created = 0
        for _, row in df.iterrows():
            stock_id = str(uuid.uuid4())
            stock_doc = {
                "id": stock_id,
                "symbol": str(row["symbol"]).upper(),
                "name": str(row["name"]),
                "exchange": str(row.get("exchange", "")) if pd.notna(row.get("exchange")) else None,
                "user_id": current_user["id"],
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            
            await db.stocks.insert_one(stock_doc)
            stocks_created += 1
        
        return {"message": f"Successfully uploaded {stocks_created} stocks"}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error processing CSV: {str(e)}")

# Corporate Actions Routes (PE Desk Only)
@api_router.post("/corporate-actions", response_model=CorporateAction)
async def create_corporate_action(
    action_data: CorporateActionCreate,
    current_user: dict = Depends(get_current_user)
):
    """Create a corporate action (Stock Split or Bonus) - PE Level only"""
    user_role = current_user.get("role", 6)
    
    if not is_pe_level(user_role):
        raise HTTPException(status_code=403, detail="Only PE Desk or PE Manager can create corporate actions")
    
    # Verify stock exists
    stock = await db.stocks.find_one({"id": action_data.stock_id}, {"_id": 0})
    if not stock:
        raise HTTPException(status_code=404, detail="Stock not found")
    
    # Validate action type
    if action_data.action_type not in ["stock_split", "bonus"]:
        raise HTTPException(status_code=400, detail="Action type must be 'stock_split' or 'bonus'")
    
    # For stock split, new_face_value is required
    if action_data.action_type == "stock_split" and not action_data.new_face_value:
        raise HTTPException(status_code=400, detail="New face value is required for stock split")
    
    action_id = str(uuid.uuid4())
    action_doc = {
        "id": action_id,
        "stock_id": action_data.stock_id,
        "action_type": action_data.action_type,
        "ratio_from": action_data.ratio_from,
        "ratio_to": action_data.ratio_to,
        "new_face_value": action_data.new_face_value,
        "record_date": action_data.record_date,
        "status": "pending",
        "applied_at": None,
        "notes": action_data.notes,
        "created_by": current_user["id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.corporate_actions.insert_one(action_doc)
    
    # Create audit log
    await create_audit_log(
        action="CORPORATE_ACTION_CREATE",
        entity_type="corporate_action",
        entity_id=action_id,
        user_id=current_user["id"],
        user_name=current_user["name"],
        user_role=user_role,
        entity_name=f"{stock['symbol']} - {action_data.action_type}",
        details={
            "stock_symbol": stock["symbol"],
            "action_type": action_data.action_type,
            "ratio": f"{action_data.ratio_from}:{action_data.ratio_to}",
            "record_date": action_data.record_date
        }
    )
    
    return CorporateAction(
        **action_doc,
        stock_symbol=stock["symbol"],
        stock_name=stock["name"]
    )

@api_router.get("/corporate-actions", response_model=List[CorporateAction])
async def get_corporate_actions(
    stock_id: Optional[str] = None,
    status: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get corporate actions - PE Level only"""
    user_role = current_user.get("role", 6)
    
    if not is_pe_level(user_role):
        raise HTTPException(status_code=403, detail="Only PE Desk or PE Manager can view corporate actions")
    
    query = {}
    if stock_id:
        query["stock_id"] = stock_id
    if status:
        query["status"] = status
    
    actions = await db.corporate_actions.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    
    # Enrich with stock details
    stock_ids = list(set(a["stock_id"] for a in actions))
    stocks = await db.stocks.find({"id": {"$in": stock_ids}}, {"_id": 0}).to_list(1000)
    stock_map = {s["id"]: s for s in stocks}
    
    enriched = []
    for a in actions:
        stock = stock_map.get(a["stock_id"], {})
        enriched.append(CorporateAction(
            **a,
            stock_symbol=stock.get("symbol", "Unknown"),
            stock_name=stock.get("name", "Unknown")
        ))
    
    return enriched

@api_router.put("/corporate-actions/{action_id}/apply")
async def apply_corporate_action(
    action_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Apply a corporate action on record date - adjusts buy prices in inventory"""
    user_role = current_user.get("role", 6)
    
    if not is_pe_level(user_role):
        raise HTTPException(status_code=403, detail="Only PE Desk or PE Manager can apply corporate actions")
    
    action = await db.corporate_actions.find_one({"id": action_id}, {"_id": 0})
    if not action:
        raise HTTPException(status_code=404, detail="Corporate action not found")
    
    if action["status"] == "applied":
        raise HTTPException(status_code=400, detail="Corporate action already applied")
    
    # Check if today is the record date
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    if action["record_date"] != today:
        raise HTTPException(
            status_code=400, 
            detail=f"Corporate action can only be applied on record date ({action['record_date']}). Today is {today}"
        )
    
    stock_id = action["stock_id"]
    ratio_from = action["ratio_from"]
    ratio_to = action["ratio_to"]
    action_type = action["action_type"]
    
    # Calculate adjustment factor
    if action_type == "stock_split":
        # Stock split: If 1:2 split, new price = old price / 2
        adjustment_factor = ratio_from / ratio_to
    else:  # bonus
        # Bonus: If 1:1 bonus, shares double, price halves
        # ratio_from:ratio_to means for every ratio_from shares, you get ratio_to bonus
        adjustment_factor = ratio_from / (ratio_from + ratio_to)
    
    # Update inventory weighted average price
    inventory = await db.inventory.find_one({"stock_id": stock_id}, {"_id": 0})
    if inventory:
        new_avg_price = inventory["weighted_avg_price"] * adjustment_factor
        
        # For bonus/split, quantity increases
        if action_type == "stock_split":
            new_quantity = int(inventory["available_quantity"] * (ratio_to / ratio_from))
        else:  # bonus
            bonus_shares = int(inventory["available_quantity"] * (ratio_to / ratio_from))
            new_quantity = inventory["available_quantity"] + bonus_shares
        
        await db.inventory.update_one(
            {"stock_id": stock_id},
            {"$set": {
                "weighted_avg_price": new_avg_price,
                "available_quantity": new_quantity
            }}
        )
    
    # Update all purchases for this stock (historical record)
    await db.purchases.update_many(
        {"stock_id": stock_id},
        {"$mul": {"price_per_unit": adjustment_factor}}
    )
    
    # Update all bookings for this stock
    await db.bookings.update_many(
        {"stock_id": stock_id},
        {"$mul": {"buying_price": adjustment_factor}}
    )
    
    # Update stock face value if stock split
    if action_type == "stock_split" and action.get("new_face_value"):
        await db.stocks.update_one(
            {"id": stock_id},
            {"$set": {"face_value": action["new_face_value"]}}
        )
    
    # Mark action as applied
    await db.corporate_actions.update_one(
        {"id": action_id},
        {"$set": {
            "status": "applied",
            "applied_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # Create audit log
    stock = await db.stocks.find_one({"id": stock_id}, {"_id": 0})
    await create_audit_log(
        action="CORPORATE_ACTION_APPLY",
        entity_type="corporate_action",
        entity_id=action_id,
        user_id=current_user["id"],
        user_name=current_user["name"],
        user_role=user_role,
        entity_name=f"{stock['symbol'] if stock else 'Unknown'} - {action_type}",
        details={
            "adjustment_factor": adjustment_factor,
            "action_type": action_type,
            "ratio": f"{ratio_from}:{ratio_to}"
        }
    )
    
    return {
        "message": f"Corporate action applied successfully. Prices adjusted by factor {adjustment_factor:.4f}",
        "adjustment_factor": adjustment_factor
    }

@api_router.delete("/corporate-actions/{action_id}")
async def delete_corporate_action(action_id: str, current_user: dict = Depends(get_current_user)):
    """Delete a pending corporate action - PE Desk only (deletion restricted)"""
    user_role = current_user.get("role", 6)
    
    if not is_pe_desk_only(user_role):
        raise HTTPException(status_code=403, detail="Only PE Desk can delete corporate actions")
    
    action = await db.corporate_actions.find_one({"id": action_id}, {"_id": 0})
    if not action:
        raise HTTPException(status_code=404, detail="Corporate action not found")
    
    if action["status"] == "applied":
        raise HTTPException(status_code=400, detail="Cannot delete applied corporate action")
    
    await db.corporate_actions.delete_one({"id": action_id})
    return {"message": "Corporate action deleted successfully"}

# Purchase Routes (from Vendors)
@api_router.post("/purchases", response_model=Purchase)
async def create_purchase(purchase_data: PurchaseCreate, current_user: dict = Depends(get_current_user)):
    user_role = current_user.get("role", 6)
    
    # PE Level can create purchases
    if not is_pe_level(user_role):
        raise HTTPException(status_code=403, detail="Only PE Desk or PE Manager can create vendor purchases")
    
    check_permission(current_user, "manage_purchases")
    
    # Verify vendor is a client with is_vendor=True
    vendor = await db.clients.find_one({"id": purchase_data.vendor_id, "is_vendor": True}, {"_id": 0})
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor not found")
    
    # Verify stock exists
    stock = await db.stocks.find_one({"id": purchase_data.stock_id}, {"_id": 0})
    if not stock:
        raise HTTPException(status_code=404, detail="Stock not found")
    
    purchase_id = str(uuid.uuid4())
    total_amount = purchase_data.quantity * purchase_data.price_per_unit
    
    purchase_doc = {
        "id": purchase_id,
        **purchase_data.model_dump(),
        "total_amount": total_amount,
        "user_id": current_user["id"],
        "created_by": current_user["id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.purchases.insert_one(purchase_doc)
    
    # Update inventory
    await update_inventory(purchase_data.stock_id)
    
    # Send email notification to vendor using template
    if vendor.get("email"):
        await send_templated_email(
            "purchase_order_created",
            vendor["email"],
            {
                "vendor_name": vendor["name"],
                "stock_symbol": stock["symbol"],
                "stock_name": stock["name"],
                "quantity": purchase_data.quantity,
                "price_per_unit": purchase_data.price_per_unit,
                "total_amount": total_amount,
                "purchase_date": purchase_data.purchase_date
            }
        )
    
    return Purchase(
        id=purchase_id,
        vendor_id=purchase_data.vendor_id,
        vendor_name=vendor["name"],
        stock_id=purchase_data.stock_id,
        stock_symbol=stock["symbol"],
        quantity=purchase_data.quantity,
        price_per_unit=purchase_data.price_per_unit,
        total_amount=total_amount,
        purchase_date=purchase_data.purchase_date,
        notes=purchase_data.notes,
        created_at=purchase_doc["created_at"],
        created_by=current_user["id"]
    )

@api_router.get("/purchases", response_model=List[Purchase])
async def get_purchases(
    vendor_id: Optional[str] = None,
    stock_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    user_role = current_user.get("role", 6)
    
    # PE Level can view purchase history
    if not is_pe_level(user_role):
        raise HTTPException(status_code=403, detail="Only PE Desk or PE Manager can access vendor purchase history")
    
    check_permission(current_user, "manage_purchases")
    
    query = {}
    if vendor_id:
        query["vendor_id"] = vendor_id
    if stock_id:
        query["stock_id"] = stock_id
    
    purchases = await db.purchases.find(query, {"_id": 0, "user_id": 0}).to_list(1000)
    
    # Enrich with vendor and stock details
    if purchases:
        vendor_ids = list(set(p["vendor_id"] for p in purchases))
        stock_ids = list(set(p["stock_id"] for p in purchases))
        
        vendors = await db.clients.find({"id": {"$in": vendor_ids}}, {"_id": 0}).to_list(1000)
        stocks = await db.stocks.find({"id": {"$in": stock_ids}}, {"_id": 0}).to_list(1000)
        
        vendor_map = {v["id"]: v for v in vendors}
        stock_map = {s["id"]: s for s in stocks}
        
        enriched_purchases = []
        for p in purchases:
            vendor = vendor_map.get(p["vendor_id"])
            stock = stock_map.get(p["stock_id"])
            enriched_purchases.append(Purchase(
                **p,
                vendor_name=vendor["name"] if vendor else "Unknown",
                stock_symbol=stock["symbol"] if stock else "Unknown"
            ))
        
        return enriched_purchases
    
    return []

# Inventory Routes
@api_router.get("/inventory", response_model=List[Inventory])
async def get_inventory(current_user: dict = Depends(get_current_user)):
    inventory_items = await db.inventory.find({}, {"_id": 0}).to_list(1000)
    return inventory_items

@api_router.get("/inventory/{stock_id}", response_model=Inventory)
async def get_stock_inventory(stock_id: str, current_user: dict = Depends(get_current_user)):
    inventory = await db.inventory.find_one({"stock_id": stock_id}, {"_id": 0})
    if not inventory:
        # Try to calculate if not exists
        inventory = await update_inventory(stock_id)
    return inventory


@api_router.delete("/inventory/{stock_id}")
async def delete_inventory(stock_id: str, current_user: dict = Depends(get_current_user)):
    """Delete inventory record for a stock (PE Desk only - deletion restricted)"""
    user_role = current_user.get("role", 6)
    
    # Only PE Desk can delete inventory
    if not is_pe_desk_only(user_role):
        raise HTTPException(status_code=403, detail="Only PE Desk can delete inventory records")
    
    # Check if inventory exists
    inventory = await db.inventory.find_one({"stock_id": stock_id}, {"_id": 0})
    if not inventory:
        raise HTTPException(status_code=404, detail="Inventory record not found")
    
    # Check for blocked quantity - don't allow deletion if stock is blocked
    if inventory.get("blocked_quantity", 0) > 0:
        raise HTTPException(
            status_code=400, 
            detail=f"Cannot delete inventory. {inventory.get('blocked_quantity')} shares are blocked for pending bookings."
        )
    
    # Delete the inventory record
    result = await db.inventory.delete_one({"stock_id": stock_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Inventory record not found")
    
    return {"message": f"Inventory for {inventory.get('stock_symbol', stock_id)} deleted successfully"}


# Booking Routes
@api_router.post("/bookings", response_model=Booking)
async def create_booking(booking_data: BookingCreate, current_user: dict = Depends(get_current_user)):
    user_role = current_user.get("role", 5)
    
    # Employees can create bookings
    if user_role == 4:
        check_permission(current_user, "create_bookings")
    else:
        check_permission(current_user, "manage_bookings")
    
    # Verify client exists and is active
    client = await db.clients.find_one({"id": booking_data.client_id}, {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    
    # Check if client is approved by PE Desk
    if client.get("approval_status") != "approved":
        raise HTTPException(
            status_code=400, 
            detail="Client must be approved by PE Desk before creating bookings. Current status: " + client.get("approval_status", "pending")
        )
    
    # Check if client is active
    if not client.get("is_active", True):
        raise HTTPException(status_code=400, detail="Client is inactive and cannot be used for bookings")
    
    # Check if client is suspended
    if client.get("is_suspended"):
        raise HTTPException(
            status_code=400, 
            detail=f"Client is suspended and cannot be used for bookings. Reason: {client.get('suspension_reason', 'No reason provided')}"
        )
    
    # Employees can only create bookings for their own clients
    if user_role == 4:
        if client.get("mapped_employee_id") != current_user["id"] and client.get("created_by") != current_user["id"]:
            raise HTTPException(status_code=403, detail="You can only create bookings for your own clients")
    
    # Verify stock exists
    stock = await db.stocks.find_one({"id": booking_data.stock_id}, {"_id": 0})
    if not stock:
        raise HTTPException(status_code=404, detail="Stock not found")
    
    # Check if stock is blocked for booking
    if stock.get("exchange") == "Blocked IPO/RTA":
        raise HTTPException(
            status_code=400, 
            detail="This stock is blocked (IPO/RTA) and not available for booking"
        )
    
    # Get inventory to check availability and weighted average
    inventory = await db.inventory.find_one({"stock_id": booking_data.stock_id}, {"_id": 0})
    
    if not inventory or inventory["available_quantity"] < booking_data.quantity:
        raise HTTPException(
            status_code=400,
            detail=f"Insufficient inventory. Available: {inventory['available_quantity'] if inventory else 0}"
        )
    
    # Validate selling_price is provided
    if booking_data.selling_price is None or booking_data.selling_price <= 0:
        raise HTTPException(
            status_code=400,
            detail="Selling price is required and must be greater than 0"
        )
    
    # Employees MUST use weighted average as buying price (cannot edit)
    if user_role == 4:
        buying_price = inventory["weighted_avg_price"]
    else:
        # For PE Desk, buying_price can be provided or default to weighted average
        buying_price = booking_data.buying_price if booking_data.buying_price else inventory["weighted_avg_price"]
    
    # Validate buying_price is valid
    if buying_price is None or buying_price <= 0:
        raise HTTPException(
            status_code=400,
            detail="Landing price is required and must be greater than 0"
        )
    
    booking_id = str(uuid.uuid4())
    booking_number = await generate_booking_number()
    
    # Check if this is a loss booking (selling price < buying price)
    is_loss_booking = False
    loss_approval_status = "not_required"
    if booking_data.selling_price is not None and booking_data.selling_price < buying_price:
        is_loss_booking = True
        loss_approval_status = "pending"  # Requires PE Desk approval for loss
    
    # Generate client confirmation token
    confirmation_token = str(uuid.uuid4())
    
    # All bookings require PE Desk approval before inventory adjustment
    booking_doc = {
        "id": booking_id,
        "booking_number": booking_number,
        "client_id": booking_data.client_id,
        "stock_id": booking_data.stock_id,
        "quantity": booking_data.quantity,
        "buying_price": buying_price,
        "selling_price": booking_data.selling_price,
        "booking_date": booking_data.booking_date,
        "status": booking_data.status,
        "approval_status": "pending",  # Requires PE Desk approval
        "approved_by": None,
        "approved_at": None,
        # Booking type (client/team/own)
        "booking_type": booking_data.booking_type,
        "insider_form_uploaded": booking_data.insider_form_uploaded,
        "insider_form_path": None,
        # Client confirmation
        "client_confirmation_status": "pending",
        "client_confirmation_token": confirmation_token,
        "client_confirmed_at": None,
        "client_denial_reason": None,
        # Loss booking
        "is_loss_booking": is_loss_booking,
        "loss_approval_status": loss_approval_status,
        "loss_approved_by": None,
        "loss_approved_at": None,
        "notes": booking_data.notes,
        "user_id": current_user["id"],
        "created_by": current_user["id"],
        "created_by_name": current_user["name"],
        "created_by_role": user_role,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.bookings.insert_one(booking_doc)
    
    # DO NOT update inventory yet - wait for PE Desk approval
    # await update_inventory(booking_data.stock_id)
    
    # Create audit log
    await create_audit_log(
        action="BOOKING_CREATE",
        entity_type="booking",
        entity_id=booking_id,
        user_id=current_user["id"],
        user_name=current_user["name"],
        user_role=user_role,
        entity_name=f"{stock['symbol']} - {client['name']} ({booking_number})",
        details={
            "client_id": booking_data.client_id,
            "client_name": client["name"],
            "stock_id": booking_data.stock_id,
            "stock_symbol": stock["symbol"],
            "quantity": booking_data.quantity,
            "buying_price": buying_price,
            "selling_price": booking_data.selling_price
        }
    )
    
    # Send detailed email notification to client with CC to user using template
    if client.get("email"):
        await send_templated_email(
            "booking_created",
            client["email"],
            {
                "client_name": client["name"],
                "booking_number": booking_number,
                "stock_symbol": stock["symbol"],
                "stock_name": stock["name"],
                "quantity": booking_data.quantity
            },
            cc_email=current_user.get("email")
        )
    
    # Real-time notification to PE Desk about pending booking
    await notify_roles(
        [1],  # PE Desk only
        "booking_pending",
        "New Booking Pending Approval",
        f"Booking {booking_number} for '{client['name']}' - {stock['symbol']} x {booking_data.quantity} awaiting PE Desk approval",
        {"booking_id": booking_id, "booking_number": booking_number, "client_name": client["name"], "stock_symbol": stock["symbol"]}
    )
    
    return Booking(**{k: v for k, v in booking_doc.items() if k not in ["user_id", "created_by_name"]})

# Client Booking Confirmation (Public endpoint - no auth required)
class ClientConfirmationRequest(BaseModel):
    reason: Optional[str] = None

@api_router.get("/booking-confirm/{booking_id}/{token}/{action}")
async def client_confirm_booking_get(booking_id: str, token: str, action: str):
    """Client confirms or denies booking via email link (GET for direct link click)"""
    return await process_client_confirmation(booking_id, token, action, None)

@api_router.post("/booking-confirm/{booking_id}/{token}/{action}")
async def client_confirm_booking_post(booking_id: str, token: str, action: str, request: ClientConfirmationRequest = None):
    """Client confirms or denies booking via email link (POST with optional reason)"""
    reason = request.reason if request else None
    return await process_client_confirmation(booking_id, token, action, reason)

async def process_client_confirmation(booking_id: str, token: str, action: str, reason: Optional[str]):
    """Process client confirmation/denial of booking"""
    if action not in ["accept", "deny"]:
        raise HTTPException(status_code=400, detail="Invalid action. Use 'accept' or 'deny'")
    
    booking = await db.bookings.find_one({"id": booking_id}, {"_id": 0})
    if not booking:
        raise HTTPException(status_code=404, detail="Booking not found")
    
    # Validate token
    if booking.get("client_confirmation_token") != token:
        raise HTTPException(status_code=403, detail="Invalid confirmation token")
    
    # Check if booking is approved by PE Desk first
    if booking.get("approval_status") != "approved":
        return {
            "message": "This booking is still pending PE Desk approval. You cannot confirm it yet.",
            "status": "pending_approval",
            "booking_number": booking.get("booking_number", booking_id[:8].upper())
        }
    
    # Check if loss booking is approved (if applicable)
    if booking.get("is_loss_booking") and booking.get("loss_approval_status") == "pending":
        return {
            "message": "This is a loss booking that requires additional approval. You cannot confirm it yet.",
            "status": "pending_loss_approval",
            "booking_number": booking.get("booking_number", booking_id[:8].upper())
        }
    
    # Check if already confirmed
    if booking.get("client_confirmation_status") != "pending":
        return {
            "message": f"Booking has already been {booking.get('client_confirmation_status')}",
            "status": booking.get("client_confirmation_status")
        }
    
    # Get client and stock info
    client = await db.clients.find_one({"id": booking["client_id"]}, {"_id": 0})
    stock = await db.stocks.find_one({"id": booking["stock_id"]}, {"_id": 0})
    booking_number = booking.get("booking_number", booking_id[:8].upper())
    
    if action == "accept":
        update_data = {
            "client_confirmation_status": "accepted",
            "client_confirmed_at": datetime.now(timezone.utc).isoformat()
        }
        message = "Booking accepted successfully! Your order is now pending PE Desk approval."
        
        # Notify PE Desk about client acceptance
        await notify_roles(
            [1],
            "client_accepted",
            "Client Accepted Booking",
            f"Client {client['name'] if client else 'Unknown'} has accepted booking {booking_number} for {stock['symbol'] if stock else 'Unknown'}",
            {"booking_id": booking_id, "booking_number": booking_number}
        )
        
        # Notify booking creator
        if booking.get("created_by"):
            await create_notification(
                booking["created_by"],
                "client_accepted",
                "Client Accepted Booking",
                f"Client {client['name'] if client else 'Unknown'} has accepted booking {booking_number}",
                {"booking_id": booking_id, "booking_number": booking_number}
            )
        
    else:  # deny
        update_data = {
            "client_confirmation_status": "denied",
            "client_confirmed_at": datetime.now(timezone.utc).isoformat(),
            "client_denial_reason": reason or "No reason provided"
        }
        message = "Booking denied. The booking creator and PE Desk have been notified."
        
        # Notify PE Desk about denial
        await notify_roles(
            [1],
            "client_denied",
            "Client Denied Booking",
            f"Client {client['name'] if client else 'Unknown'} has DENIED booking {booking_number}. Reason: {reason or 'Not specified'}",
            {"booking_id": booking_id, "booking_number": booking_number, "reason": reason}
        )
        
        # Notify booking creator
        if booking.get("created_by"):
            await create_notification(
                booking["created_by"],
                "client_denied",
                "Client Denied Booking",
                f"Client {client['name'] if client else 'Unknown'} has denied booking {booking_number}. Reason: {reason or 'Not specified'}",
                {"booking_id": booking_id, "booking_number": booking_number, "reason": reason}
            )
    
    await db.bookings.update_one({"id": booking_id}, {"$set": update_data})
    
    # Create audit log
    await create_audit_log(
        action=f"CLIENT_{'ACCEPT' if action == 'accept' else 'DENY'}",
        entity_type="booking",
        entity_id=booking_id,
        user_id=client["id"] if client else "unknown",
        user_name=client["name"] if client else "Unknown Client",
        user_role=0,  # Client
        entity_name=f"{stock['symbol'] if stock else 'Unknown'} - {booking_number}",
        details={
            "action": action,
            "reason": reason if action == "deny" else None,
            "client_email": client.get("email") if client else None
        }
    )
    
    return {
        "message": message,
        "status": "accepted" if action == "accept" else "denied",
        "booking_number": booking_number
    }

@api_router.put("/bookings/{booking_id}/approve")
async def approve_booking(
    booking_id: str,
    approve: bool = True,
    current_user: dict = Depends(get_current_user)
):
    """Approve or reject a booking for inventory adjustment (PE Level)"""
    user_role = current_user.get("role", 6)
    
    # PE Level can approve bookings
    if not is_pe_level(user_role):
        raise HTTPException(status_code=403, detail="Only PE Desk or PE Manager can approve bookings for inventory adjustment")
    
    booking = await db.bookings.find_one({"id": booking_id}, {"_id": 0})
    if not booking:
        raise HTTPException(status_code=404, detail="Booking not found")
    
    if booking.get("approval_status") != "pending":
        raise HTTPException(status_code=400, detail="Booking already processed")
    
    update_data = {
        "approval_status": "approved" if approve else "rejected",
        "approved_by": current_user["id"],
        "approved_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.bookings.update_one({"id": booking_id}, {"$set": update_data})
    
    # If approved, now update inventory
    if approve:
        await update_inventory(booking["stock_id"])
        
        # Create audit log
        await create_audit_log(
            action="BOOKING_APPROVE",
            entity_type="booking",
            entity_id=booking_id,
            user_id=current_user["id"],
            user_name=current_user["name"],
            user_role=user_role,
            details={"stock_id": booking["stock_id"], "quantity": booking["quantity"]}
        )
        
        # Get client and stock info
        client = await db.clients.find_one({"id": booking["client_id"]}, {"_id": 0})
        stock = await db.stocks.find_one({"id": booking["stock_id"]}, {"_id": 0})
        creator = await db.users.find_one({"id": booking["created_by"]}, {"_id": 0})
        booking_number = booking.get("booking_number", booking_id[:8].upper())
        confirmation_token = booking.get("client_confirmation_token")
        
        # Check if this is a loss booking that still needs loss approval
        is_loss_pending = booking.get("is_loss_booking") and booking.get("loss_approval_status") == "pending"
        
        # Get all client emails (primary, secondary, tertiary)
        client_emails = get_client_emails(client) if client else []
        primary_email = client_emails[0] if client_emails else None
        additional_emails = client_emails[1:] if len(client_emails) > 1 else None
        
        if client and primary_email:
            if is_loss_pending:
                # For loss bookings, wait for loss approval before sending client confirmation
                await send_templated_email(
                    "booking_pending_loss_review",
                    primary_email,
                    {
                        "client_name": client["name"],
                        "booking_number": booking_number,
                        "stock_symbol": stock["symbol"] if stock else "N/A"
                    },
                    cc_email=creator.get("email") if creator else None,
                    additional_emails=additional_emails
                )
            else:
                # Send client confirmation email with Accept/Deny buttons using template
                frontend_url = os.environ.get('FRONTEND_URL', 'https://booking-share-system.preview.emergentagent.com')
                await send_templated_email(
                    "booking_confirmation_request",
                    primary_email,
                    {
                        "client_name": client["name"],
                        "booking_number": booking_number,
                        "otc_ucc": client.get("otc_ucc", "N/A"),
                        "stock_symbol": stock["symbol"] if stock else "N/A",
                        "stock_name": stock["name"] if stock else "",
                        "quantity": booking["quantity"],
                        "buying_price": f"{booking.get('buying_price', 0):,.2f}",
                        "total_value": f"{(booking.get('buying_price', 0) * booking.get('quantity', 0)):,.2f}",
                        "approved_by": current_user["name"],
                        "accept_url": f"{frontend_url}/booking-confirm/{booking_id}/{confirmation_token}/accept",
                        "deny_url": f"{frontend_url}/booking-confirm/{booking_id}/{confirmation_token}/deny"
                    },
                    cc_email=creator.get("email") if creator else None,
                    additional_emails=additional_emails
                )
        
        # Real-time notification to booking creator
        if booking.get("created_by"):
            await create_notification(
                booking["created_by"],
                "booking_approved",
                "Booking Approved - Awaiting Client Confirmation",
                f"Your booking {booking_number} for '{stock['symbol'] if stock else 'N/A'}' has been approved. Client confirmation email sent.",
                {"booking_id": booking_id, "stock_symbol": stock['symbol'] if stock else None}
            )
    else:
        # Create audit log for rejection
        await create_audit_log(
            action="BOOKING_REJECT",
            entity_type="booking",
            entity_id=booking_id,
            user_id=current_user["id"],
            user_name=current_user["name"],
            user_role=user_role,
            details={"stock_id": booking["stock_id"], "quantity": booking["quantity"]}
        )
        
        # Real-time notification to booking creator for rejection
        stock = await db.stocks.find_one({"id": booking["stock_id"]}, {"_id": 0})
        if booking.get("created_by"):
            await create_notification(
                booking["created_by"],
                "booking_rejected",
                "Booking Rejected",
                f"Your booking for '{stock['symbol'] if stock else 'N/A'}' has been rejected",
                {"booking_id": booking_id, "stock_symbol": stock['symbol'] if stock else None}
            )
    
    return {"message": f"Booking {'approved' if approve else 'rejected'} successfully"}

@api_router.put("/bookings/{booking_id}/approve-loss")
async def approve_loss_booking(
    booking_id: str,
    approve: bool = True,
    current_user: dict = Depends(get_current_user)
):
    """Approve or reject a loss booking (selling price < buying price) - PE Level"""
    user_role = current_user.get("role", 6)
    
    # PE Level can approve loss bookings
    if not is_pe_level(user_role):
        raise HTTPException(status_code=403, detail="Only PE Desk or PE Manager can approve loss bookings")
    
    booking = await db.bookings.find_one({"id": booking_id}, {"_id": 0})
    if not booking:
        raise HTTPException(status_code=404, detail="Booking not found")
    
    if not booking.get("is_loss_booking", False):
        raise HTTPException(status_code=400, detail="This is not a loss booking")
    
    if booking.get("loss_approval_status") != "pending":
        raise HTTPException(status_code=400, detail="Loss booking already processed")
    
    update_data = {
        "loss_approval_status": "approved" if approve else "rejected",
        "loss_approved_by": current_user["id"],
        "loss_approved_at": datetime.now(timezone.utc).isoformat()
    }
    
    # If loss is rejected, also reject the entire booking and release inventory
    if not approve:
        update_data["status"] = "rejected"
        update_data["rejection_reason"] = "Loss booking rejected by PE"
        
        # Release blocked inventory if any was blocked
        if booking.get("approval_status") == "approved":
            inventory = await db.inventory.find_one({"stock_id": booking["stock_id"]}, {"_id": 0})
            if inventory:
                current_blocked = inventory.get("blocked_quantity", 0)
                new_blocked = max(0, current_blocked - booking["quantity"])
                await db.inventory.update_one(
                    {"stock_id": booking["stock_id"]},
                    {"$set": {"blocked_quantity": new_blocked}}
                )
    
    await db.bookings.update_one({"id": booking_id}, {"$set": update_data})
    
    # Audit logging
    action = "LOSS_BOOKING_APPROVE" if approve else "LOSS_BOOKING_REJECT"
    stock = await db.stocks.find_one({"id": booking["stock_id"]}, {"_id": 0})
    
    await create_audit_log(
        action=action,
        entity_type="booking",
        entity_id=booking_id,
        user_id=current_user["id"],
        user_name=current_user["name"],
        user_role=user_role,
        details={
            "stock_id": booking["stock_id"],
            "stock_symbol": stock['symbol'] if stock else "Unknown",
            "buying_price": booking["buying_price"],
            "selling_price": booking["selling_price"],
            "loss_amount": (booking["buying_price"] - booking["selling_price"]) * booking["quantity"]
        }
    )
    
    # Get client info
    client = await db.clients.find_one({"id": booking["client_id"]}, {"_id": 0})
    creator = await db.users.find_one({"id": booking["created_by"]}, {"_id": 0})
    booking_number = booking.get("booking_number", booking_id[:8].upper())
    confirmation_token = booking.get("client_confirmation_token")
    
    # If loss approved and booking is already PE Desk approved, send client confirmation email
    if approve and booking.get("approval_status") == "approved":
        if client and client.get("email"):
            frontend_url = os.environ.get('FRONTEND_URL', 'https://booking-share-system.preview.emergentagent.com')
            await send_templated_email(
                "loss_booking_confirmation_request",
                client["email"],
                {
                    "client_name": client["name"],
                    "booking_number": booking_number,
                    "stock_symbol": stock["symbol"] if stock else "N/A",
                    "quantity": booking["quantity"],
                    "buying_price": f"{booking.get('buying_price', 0):,.2f}",
                    "selling_price": f"{booking.get('selling_price', 0):,.2f}",
                    "accept_url": f"{frontend_url}/booking-confirm/{booking_id}/{confirmation_token}/accept",
                    "deny_url": f"{frontend_url}/booking-confirm/{booking_id}/{confirmation_token}/deny"
                },
                cc_email=creator.get("email") if creator else None
            )
    
    # Notification to booking creator
    if booking.get("created_by"):
        message = f"Loss booking {booking_number} for '{stock['symbol'] if stock else 'N/A'}' has been {'approved - client confirmation email sent' if approve else 'rejected'}"
        await create_notification(
            booking["created_by"],
            "loss_booking_approved" if approve else "loss_booking_rejected",
            "Loss Booking " + ("Approved" if approve else "Rejected"),
            message,
            {"booking_id": booking_id, "stock_symbol": stock['symbol'] if stock else None}
        )
    
    return {"message": f"Loss booking {'approved' if approve else 'rejected'} successfully"}

@api_router.get("/bookings/pending-loss-approval", response_model=List[BookingWithDetails])
async def get_pending_loss_bookings(current_user: dict = Depends(get_current_user)):
    """Get loss bookings pending approval (PE Level)"""
    user_role = current_user.get("role", 6)
    
    if not is_pe_level(user_role):
        raise HTTPException(status_code=403, detail="Only PE Desk or PE Manager can view pending loss bookings")
    
    bookings = await db.bookings.find(
        {"is_loss_booking": True, "loss_approval_status": "pending"},
        {"_id": 0, "user_id": 0}
    ).to_list(1000)
    
    if not bookings:
        return []
    
    # Enrich with client and stock details
    client_ids = list(set(b["client_id"] for b in bookings))
    stock_ids = list(set(b["stock_id"] for b in bookings))
    user_ids = list(set(b["created_by"] for b in bookings))
    
    clients = await db.clients.find({"id": {"$in": client_ids}}, {"_id": 0}).to_list(1000)
    stocks = await db.stocks.find({"id": {"$in": stock_ids}}, {"_id": 0}).to_list(1000)
    users = await db.users.find({"id": {"$in": user_ids}}, {"_id": 0}).to_list(1000)
    
    client_map = {c["id"]: c for c in clients}
    stock_map = {s["id"]: s for s in stocks}
    user_map = {u["id"]: u for u in users}
    
    enriched = []
    for b in bookings:
        client = client_map.get(b["client_id"], {})
        stock = stock_map.get(b["stock_id"], {})
        user = user_map.get(b["created_by"], {})
        
        profit_loss = None
        if b.get("selling_price"):
            profit_loss = (b["selling_price"] - b["buying_price"]) * b["quantity"]
        
        booking_data = {k: v for k, v in b.items() if k not in ["client_name", "stock_symbol", "stock_name", "created_by_name", "profit_loss"]}
        
        enriched.append(BookingWithDetails(
            **booking_data,
            client_name=client.get("name", "Unknown"),
            client_pan=client.get("pan_number"),
            client_dp_id=client.get("dp_id"),
            stock_symbol=stock.get("symbol", "Unknown"),
            stock_name=stock.get("name", "Unknown"),
            created_by_name=user.get("name", "Unknown"),
            profit_loss=profit_loss
        ))
    
    return enriched

@api_router.get("/bookings/pending-approval", response_model=List[BookingWithDetails])
async def get_pending_bookings(current_user: dict = Depends(get_current_user)):
    """Get bookings pending approval (PE Level)"""
    user_role = current_user.get("role", 6)
    
    if not is_pe_level(user_role):
        raise HTTPException(status_code=403, detail="Only PE Desk or PE Manager can view pending bookings")
    
    bookings = await db.bookings.find(
        {"approval_status": "pending"},
        {"_id": 0, "user_id": 0}
    ).to_list(1000)
    
    if not bookings:
        return []
    
    # Enrich with client and stock details
    client_ids = list(set(b["client_id"] for b in bookings))
    stock_ids = list(set(b["stock_id"] for b in bookings))
    
    clients = await db.clients.find({"id": {"$in": client_ids}}, {"_id": 0}).to_list(1000)
    stocks = await db.stocks.find({"id": {"$in": stock_ids}}, {"_id": 0}).to_list(1000)
    
    client_map = {c["id"]: c for c in clients}
    stock_map = {s["id"]: s for s in stocks}
    
    enriched = []
    for b in bookings:
        client = client_map.get(b["client_id"], {})
        stock = stock_map.get(b["stock_id"], {})
        enriched.append(BookingWithDetails(
            **b,
            client_name=client.get("name", "Unknown"),
            stock_symbol=stock.get("symbol", "Unknown"),
            stock_name=stock.get("name", "Unknown")
        ))
    
    return enriched

@api_router.get("/bookings", response_model=List[BookingWithDetails])
async def get_bookings(
    client_id: Optional[str] = None,
    stock_id: Optional[str] = None,
    status: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    
    # Role-based filtering
    if current_user.get("role", 5) >= 3:
        if "view_all" not in ROLE_PERMISSIONS.get(current_user.get("role", 5), []):
            query["created_by"] = current_user["id"]
    
    if client_id:
        query["client_id"] = client_id
    if stock_id:
        query["stock_id"] = stock_id
    if status:
        query["status"] = status
    
    bookings = await db.bookings.find(query, {"_id": 0, "user_id": 0}).to_list(1000)
    
    if not bookings:
        return []
    
    # Batch fetch related data
    client_ids = list(set(b["client_id"] for b in bookings))
    stock_ids = list(set(b["stock_id"] for b in bookings))
    user_ids = list(set(b["created_by"] for b in bookings))
    
    clients = await db.clients.find({"id": {"$in": client_ids}}, {"_id": 0}).to_list(1000)
    stocks = await db.stocks.find({"id": {"$in": stock_ids}}, {"_id": 0}).to_list(1000)
    users = await db.users.find({"id": {"$in": user_ids}}, {"_id": 0}).to_list(1000)
    
    client_map = {c["id"]: c for c in clients}
    stock_map = {s["id"]: s for s in stocks}
    user_map = {u["id"]: u for u in users}
    
    enriched_bookings = []
    for booking in bookings:
        client = client_map.get(booking["client_id"])
        stock = stock_map.get(booking["stock_id"])
        user = user_map.get(booking["created_by"])
        
        profit_loss = None
        if booking.get("selling_price") and booking["status"] == "closed":
            profit_loss = (booking["selling_price"] - booking["buying_price"]) * booking["quantity"]
        
        # Exclude fields that will be explicitly provided to avoid duplicate key errors
        booking_data = {k: v for k, v in booking.items() if k not in ["client_name", "stock_symbol", "stock_name", "created_by_name", "profit_loss"]}
        
        enriched_bookings.append(BookingWithDetails(
            **booking_data,
            client_name=client["name"] if client else "Unknown",
            stock_symbol=stock["symbol"] if stock else "Unknown",
            stock_name=stock["name"] if stock else "Unknown",
            created_by_name=user["name"] if user else "Unknown",
            profit_loss=profit_loss
        ))
    
    return enriched_bookings

@api_router.get("/bookings-export")
async def export_bookings(
    format: str = "xlsx",
    current_user: dict = Depends(get_current_user)
):
    """Export all bookings to Excel/CSV with all fields"""
    user_role = current_user.get("role", 5)
    
    # Build query based on role
    query = {}
    if user_role >= 3:
        if "view_all" not in ROLE_PERMISSIONS.get(user_role, []):
            query["created_by"] = current_user["id"]
    
    bookings = await db.bookings.find(query, {"_id": 0, "user_id": 0}).to_list(10000)
    
    if not bookings:
        raise HTTPException(status_code=404, detail="No bookings found")
    
    # Batch fetch related data
    client_ids = list(set(b["client_id"] for b in bookings))
    stock_ids = list(set(b["stock_id"] for b in bookings))
    user_ids = list(set(b.get("created_by", "") for b in bookings if b.get("created_by")))
    approved_by_ids = list(set(b.get("approved_by", "") for b in bookings if b.get("approved_by")))
    
    clients = await db.clients.find({"id": {"$in": client_ids}}, {"_id": 0}).to_list(1000)
    stocks = await db.stocks.find({"id": {"$in": stock_ids}}, {"_id": 0}).to_list(1000)
    users = await db.users.find({"id": {"$in": user_ids + approved_by_ids}}, {"_id": 0}).to_list(1000)
    
    client_map = {c["id"]: c for c in clients}
    stock_map = {s["id"]: s for s in stocks}
    user_map = {u["id"]: u for u in users}
    
    # Prepare export data
    export_data = []
    for b in bookings:
        client = client_map.get(b["client_id"], {})
        stock = stock_map.get(b["stock_id"], {})
        created_by_user = user_map.get(b.get("created_by", ""), {})
        approved_by_user = user_map.get(b.get("approved_by", ""), {})
        
        total_amount = (b.get("selling_price") or 0) * b.get("quantity", 0)
        revenue = None
        if b.get("selling_price"):
            revenue = (b["selling_price"] - b["buying_price"]) * b["quantity"]
        
        export_data.append({
            "Booking ID": b.get("booking_number", b["id"][:8].upper()),
            "Client Name": client.get("name", "Unknown"),
            "Client PAN": client.get("pan_number", ""),
            "Client DP ID": client.get("dp_id", ""),
            "Client Mobile": client.get("mobile_number", ""),
            "Stock Symbol": stock.get("symbol", "Unknown"),
            "Stock Name": stock.get("name", "Unknown"),
            "ISIN": stock.get("isin_number", ""),
            "Quantity": b.get("quantity", 0),
            "Landing Price": b.get("buying_price", 0),
            "Selling Price": b.get("selling_price", ""),
            "Total Amount": total_amount,
            "Revenue": revenue if revenue is not None else "",
            "Booking Date": b.get("booking_date", ""),
            "Status": b.get("status", "").upper(),
            "Approval Status": b.get("approval_status", "").upper(),
            "Approved By": approved_by_user.get("name", ""),
            "Approved At": b.get("approved_at", ""),
            "Is Loss Booking": "Yes" if b.get("is_loss_booking") else "No",
            "Loss Approval": b.get("loss_approval_status", "").upper() if b.get("is_loss_booking") else "N/A",
            "Payment Status": b.get("payment_status", "pending").upper(),
            "Total Paid": b.get("total_paid", 0),
            "Remaining Balance": total_amount - b.get("total_paid", 0) if total_amount else 0,
            "DP Transfer Ready": "Yes" if b.get("dp_transfer_ready") else "No",
            "Payment Completed At": b.get("payment_completed_at", ""),
            "Created By": created_by_user.get("name", "Unknown"),
            "Created At": b.get("created_at", ""),
            "Notes": b.get("notes", "")
        })
    
    df = pd.DataFrame(export_data)
    
    if format == "csv":
        output = io.StringIO()
        df.to_csv(output, index=False)
        content = output.getvalue()
        return Response(
            content=content,
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=bookings_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"}
        )
    else:
        # Excel format
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Bookings')
            
            # Auto-adjust column widths
            worksheet = writer.sheets['Bookings']
            for idx, col in enumerate(df.columns):
                max_length = max(df[col].astype(str).map(len).max(), len(col)) + 2
                worksheet.column_dimensions[chr(65 + idx) if idx < 26 else f"A{chr(65 + idx - 26)}"].width = min(max_length, 50)
        
        buffer.seek(0)
        return Response(
            content=buffer.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=bookings_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"}
        )

@api_router.get("/bookings/{booking_id}", response_model=BookingWithDetails)
async def get_booking(booking_id: str, current_user: dict = Depends(get_current_user)):
    booking = await db.bookings.find_one({"id": booking_id}, {"_id": 0, "user_id": 0})
    if not booking:
        raise HTTPException(status_code=404, detail="Booking not found")
    
    client = await db.clients.find_one({"id": booking["client_id"]}, {"_id": 0})
    stock = await db.stocks.find_one({"id": booking["stock_id"]}, {"_id": 0})
    user = await db.users.find_one({"id": booking["created_by"]}, {"_id": 0})
    
    profit_loss = None
    if booking.get("selling_price") and booking["status"] == "closed":
        profit_loss = (booking["selling_price"] - booking["buying_price"]) * booking["quantity"]
    
    # Exclude fields that will be explicitly provided
    booking_data = {k: v for k, v in booking.items() if k not in ["client_name", "stock_symbol", "stock_name", "created_by_name", "profit_loss"]}
    
    return BookingWithDetails(
        **booking_data,
        client_name=client["name"] if client else "Unknown",
        stock_symbol=stock["symbol"] if stock else "Unknown",
        stock_name=stock["name"] if stock else "Unknown",
        created_by_name=user["name"] if user else "Unknown",
        profit_loss=profit_loss
    )

@api_router.put("/bookings/{booking_id}", response_model=Booking)
async def update_booking(booking_id: str, booking_data: BookingCreate, current_user: dict = Depends(get_current_user)):
    user_role = current_user.get("role", 5)
    check_permission(current_user, "manage_bookings")
    
    # Get old booking to check status change
    old_booking = await db.bookings.find_one({"id": booking_id}, {"_id": 0})
    if not old_booking:
        raise HTTPException(status_code=404, detail="Booking not found")
    
    # Get client and stock info for alerts
    client = await db.clients.find_one({"id": booking_data.client_id}, {"_id": 0})
    stock = await db.stocks.find_one({"id": booking_data.stock_id}, {"_id": 0})
    
    result = await db.bookings.update_one(
        {"id": booking_id},
        {"$set": booking_data.model_dump()}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Booking not found")
    
    # Update inventory
    await update_inventory(booking_data.stock_id)
    
    # Audit log for update
    await create_audit_log(
        action="BOOKING_UPDATE",
        entity_type="booking",
        entity_id=booking_id,
        user_id=current_user["id"],
        user_name=current_user["name"],
        user_role=user_role,
        entity_name=f"{stock['symbol'] if stock else 'Unknown'} - {client['name'] if client else 'Unknown'}",
        details={
            "booking_number": old_booking.get("booking_number", booking_id[:8].upper()),
            "changes": {
                "quantity": {"old": old_booking.get("quantity"), "new": booking_data.quantity},
                "selling_price": {"old": old_booking.get("selling_price"), "new": booking_data.selling_price},
                "status": {"old": old_booking.get("status"), "new": booking_data.status}
            }
        }
    )
    
    # Alert: Notify relevant parties about booking update
    booking_number = old_booking.get("booking_number", booking_id[:8].upper())
    
    # Notify booking creator if different from updater
    if old_booking.get("created_by") and old_booking["created_by"] != current_user["id"]:
        await create_notification(
            old_booking["created_by"],
            "booking_updated",
            "Booking Updated",
            f"Booking {booking_number} for {stock['symbol'] if stock else 'Unknown'} has been updated by {current_user['name']}",
            {"booking_id": booking_id, "stock_symbol": stock["symbol"] if stock else None, "updated_by": current_user["name"]}
        )
    
    # Notify PE Desk about significant changes from non-PE users
    if not is_pe_level(user_role):
        await notify_roles(
            [1, 2],
            "booking_updated",
            "Booking Modified",
            f"Booking {booking_number} modified by {current_user['name']}",
            {"booking_id": booking_id, "stock_symbol": stock["symbol"] if stock else None}
        )
    
    # Send email if status changed
    if old_booking["status"] != booking_data.status:
        if client and client.get("email"):
            await send_templated_email(
                "booking_status_updated",
                client["email"],
                {
                    "client_name": client["name"],
                    "booking_number": booking_number,
                    "status": booking_data.status.upper()
                }
            )
    
    updated_booking = await db.bookings.find_one({"id": booking_id}, {"_id": 0, "user_id": 0})
    return updated_booking

@api_router.delete("/bookings/{booking_id}")
async def delete_booking(booking_id: str, current_user: dict = Depends(get_current_user)):
    user_role = current_user.get("role", 6)
    
    # Only PE Desk can delete bookings (PE Manager cannot delete)
    if not is_pe_desk_only(user_role):
        raise HTTPException(status_code=403, detail="Only PE Desk can delete bookings")
    
    # Get booking details for alert
    booking = await db.bookings.find_one({"id": booking_id}, {"_id": 0})
    if not booking:
        raise HTTPException(status_code=404, detail="Booking not found")
    
    # Prevent deletion of transferred bookings
    if booking.get("stock_transferred"):
        raise HTTPException(status_code=400, detail="Cannot delete a booking where stock has already been transferred")
    
    # Get related info for notification
    client = await db.clients.find_one({"id": booking["client_id"]}, {"_id": 0})
    stock = await db.stocks.find_one({"id": booking["stock_id"]}, {"_id": 0})
    stock_id = booking["stock_id"]
    
    result = await db.bookings.delete_one({"id": booking_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Booking not found")
    
    # Update inventory to release blocked quantity
    await update_inventory(stock_id)
    
    # Audit log
    await create_audit_log(
        action="BOOKING_DELETE",
        entity_type="booking",
        entity_id=booking_id,
        user_id=current_user["id"],
        user_name=current_user["name"],
        user_role=user_role,
        entity_name=f"{stock['symbol'] if stock else 'Unknown'} - {client['name'] if client else 'Unknown'}",
        details={
            "booking_number": booking.get("booking_number", booking_id[:8].upper()),
            "client_name": client["name"] if client else "Unknown",
            "stock_symbol": stock["symbol"] if stock else "Unknown",
            "quantity": booking.get("quantity"),
            "deleted_by": current_user["name"],
            "inventory_released": True
        }
    )
    
    # Alert: Notify booking creator about deletion
    if booking.get("created_by") and booking["created_by"] != current_user["id"]:
        await create_notification(
            booking["created_by"],
            "booking_deleted",
            "Booking Deleted",
            f"Your booking {booking.get('booking_number', booking_id[:8].upper())} for {stock['symbol'] if stock else 'Unknown'} has been deleted by PE Desk. Inventory has been released.",
            {"booking_id": booking_id, "stock_symbol": stock["symbol"] if stock else None, "deleted_by": current_user["name"]}
        )
    
    return {"message": "Booking deleted successfully. Inventory released."}


# Void Booking (PE Level) - Release inventory without deleting record
@api_router.put("/bookings/{booking_id}/void")
async def void_booking(
    booking_id: str,
    reason: str = None,
    current_user: dict = Depends(get_current_user)
):
    """Void a booking and release blocked inventory (PE Level).
    
    Use this instead of delete when you want to keep the booking record for audit purposes.
    """
    user_role = current_user.get("role", 6)
    
    if not is_pe_level(user_role):
        raise HTTPException(status_code=403, detail="Only PE Desk or PE Manager can void bookings")
    
    booking = await db.bookings.find_one({"id": booking_id}, {"_id": 0})
    if not booking:
        raise HTTPException(status_code=404, detail="Booking not found")
    
    if booking.get("is_voided"):
        raise HTTPException(status_code=400, detail="Booking is already voided")
    
    if booking.get("stock_transferred"):
        raise HTTPException(status_code=400, detail="Cannot void a booking where stock has already been transferred")
    
    # Get related info
    client = await db.clients.find_one({"id": booking["client_id"]}, {"_id": 0})
    stock = await db.stocks.find_one({"id": booking["stock_id"]}, {"_id": 0})
    
    # Update booking as voided
    update_data = {
        "is_voided": True,
        "voided_at": datetime.now(timezone.utc).isoformat(),
        "voided_by": current_user["id"],
        "voided_by_name": current_user["name"],
        "void_reason": reason or "No reason provided",
        "status": "voided"
    }
    
    await db.bookings.update_one({"id": booking_id}, {"$set": update_data})
    
    # Create refund request if there are payments
    refund_request_id = None
    total_paid = booking.get("total_paid", 0)
    if total_paid > 0:
        # Get client bank details
        bank_accounts = client.get("bank_accounts", []) if client else []
        primary_bank = bank_accounts[0] if bank_accounts else None
        
        refund_request_id = str(uuid.uuid4())
        refund_request = {
            "id": refund_request_id,
            "booking_id": booking_id,
            "booking_number": booking.get("booking_number", booking_id[:8].upper()),
            "client_id": booking["client_id"],
            "client_name": client["name"] if client else "Unknown",
            "client_email": client.get("email") if client else None,
            "client_phone": client.get("phone") if client else None,
            "stock_id": booking["stock_id"],
            "stock_symbol": stock["symbol"] if stock else "Unknown",
            "refund_amount": total_paid,
            "payment_status": booking.get("payment_status", "unknown"),
            "void_reason": reason or "No reason provided",
            "bank_details": {
                "bank_name": primary_bank.get("bank_name") if primary_bank else None,
                "account_number": primary_bank.get("account_number") if primary_bank else None,
                "ifsc_code": primary_bank.get("ifsc_code") if primary_bank else None,
                "account_holder_name": primary_bank.get("account_holder_name") if primary_bank else (client["name"] if client else None),
                "branch": primary_bank.get("branch") if primary_bank else None
            },
            "status": "pending",  # pending, processing, completed, failed
            "created_at": datetime.now(timezone.utc).isoformat(),
            "created_by": current_user["id"],
            "created_by_name": current_user["name"],
            "processed_at": None,
            "processed_by": None,
            "processed_by_name": None,
            "notes": None,
            "reference_number": None
        }
        
        await db.refund_requests.insert_one(refund_request)
    
    # Update inventory to release blocked quantity
    await update_inventory(booking["stock_id"])
    
    # Audit log
    await create_audit_log(
        action="BOOKING_VOIDED",
        entity_type="booking",
        entity_id=booking_id,
        user_id=current_user["id"],
        user_name=current_user["name"],
        user_role=user_role,
        entity_name=f"{stock['symbol'] if stock else 'Unknown'} - {client['name'] if client else 'Unknown'}",
        details={
            "booking_number": booking.get("booking_number", booking_id[:8].upper()),
            "client_name": client["name"] if client else "Unknown",
            "stock_symbol": stock["symbol"] if stock else "Unknown",
            "quantity": booking.get("quantity"),
            "void_reason": reason,
            "inventory_released": True
        }
    )
    
    # Notify booking creator
    if booking.get("created_by") and booking["created_by"] != current_user["id"]:
        refund_msg = f" Refund of ₹{total_paid:,.2f} has been initiated." if total_paid > 0 else ""
        await create_notification(
            booking["created_by"],
            "booking_voided",
            "Booking Voided",
            f"Booking {booking.get('booking_number', booking_id[:8].upper())} for {stock['symbol'] if stock else 'Unknown'} has been voided by PE Desk. Reason: {reason or 'Not specified'}.{refund_msg}",
            {"booking_id": booking_id, "stock_symbol": stock["symbol"] if stock else None, "void_reason": reason, "refund_amount": total_paid}
        )
    
    response = {
        "message": f"Booking {booking.get('booking_number', booking_id[:8].upper())} has been voided. Inventory released.",
        "quantity_released": booking.get("quantity", 0)
    }
    
    if total_paid > 0:
        response["refund_request_created"] = True
        response["refund_request_id"] = refund_request_id
        response["refund_amount"] = total_paid
        response["message"] += f" Refund request of ₹{total_paid:,.2f} has been created."
    
    return response


# ============== Payment Tracking Endpoints (PE Desk & Zonal Manager Only) ==============
@api_router.post("/bookings/{booking_id}/payments")
async def add_payment_tranche(
    booking_id: str,
    payment: PaymentTrancheCreate,
    current_user: dict = Depends(get_current_user)
):
    """Add payment tranche to approved booking (PE Desk & Zonal Manager only)"""
    user_role = current_user.get("role", 5)
    if user_role not in [1, 2]:
        raise HTTPException(status_code=403, detail="Only PE Desk and Zonal Manager can record payments")
    
    # Get booking
    booking = await db.bookings.find_one({"id": booking_id}, {"_id": 0})
    if not booking:
        raise HTTPException(status_code=404, detail="Booking not found")
    
    if booking.get("approval_status") != "approved":
        raise HTTPException(status_code=400, detail="Can only add payments to approved bookings")
    
    # Get existing payments
    payments = booking.get("payments", [])
    
    if len(payments) >= 4:
        raise HTTPException(status_code=400, detail="Maximum 4 payment tranches allowed")
    
    # Calculate total amount required
    total_amount = (booking.get("selling_price") or 0) * booking.get("quantity", 0)
    current_paid = sum(p.get("amount", 0) for p in payments)
    remaining = total_amount - current_paid
    
    if payment.amount > remaining + 0.01:  # Small tolerance for rounding
        raise HTTPException(
            status_code=400, 
            detail=f"Payment amount ({payment.amount}) exceeds remaining balance ({remaining:.2f})"
        )
    
    # Create new tranche
    new_tranche = {
        "tranche_number": len(payments) + 1,
        "amount": payment.amount,
        "payment_date": payment.payment_date,
        "recorded_by": current_user["id"],
        "recorded_at": datetime.now(timezone.utc).isoformat(),
        "notes": payment.notes,
        "proof_url": payment.proof_url
    }
    
    payments.append(new_tranche)
    new_total_paid = current_paid + payment.amount
    
    # Determine payment status
    is_complete = abs(new_total_paid - total_amount) < 0.01  # Within 1 paisa
    payment_status = "completed" if is_complete else ("partial" if new_total_paid > 0 else "pending")
    
    # Update booking
    update_data = {
        "payments": payments,
        "total_paid": new_total_paid,
        "payment_status": payment_status,
        "dp_transfer_ready": is_complete
    }
    
    # If this is the first payment, automatically set client confirmation to accepted
    # (PE Desk entering payment implies client has accepted)
    if new_tranche["tranche_number"] == 1:
        if booking.get("client_confirmation_status") != "accepted":
            update_data["client_confirmation_status"] = "accepted"
            update_data["client_confirmed_at"] = datetime.now(timezone.utc).isoformat()
    
    if is_complete:
        update_data["payment_completed_at"] = payment.payment_date
    
    await db.bookings.update_one({"id": booking_id}, {"$set": update_data})
    
    # Create audit log
    await create_audit_log(
        action="PAYMENT_RECORDED",
        entity_type="booking",
        entity_id=booking_id,
        user_id=current_user["id"],
        user_name=current_user["name"],
        user_role=user_role,
        details={
            "tranche_number": new_tranche["tranche_number"],
            "amount": payment.amount,
            "total_paid": new_total_paid,
            "payment_status": payment_status,
            "auto_client_accepted": new_tranche["tranche_number"] == 1 and booking.get("client_confirmation_status") != "accepted"
        }
    )
    
    # Get client and stock for notifications
    client = await db.clients.find_one({"id": booking["client_id"]}, {"_id": 0})
    stock = await db.stocks.find_one({"id": booking["stock_id"]}, {"_id": 0})
    booking_number = booking.get("booking_number", booking_id[:8].upper())
    
    # If first payment auto-accepted client, create notification
    if new_tranche["tranche_number"] == 1 and booking.get("client_confirmation_status") != "accepted":
        await notify_roles(
            [1],  # PE Desk
            "client_auto_accepted",
            "Client Auto-Accepted via Payment",
            f"Booking {booking_number} automatically marked as client accepted (first payment recorded)",
            {"booking_id": booking_id, "booking_number": booking_number}
        )
    
    # Notify booking creator about payment
    if booking.get("created_by"):
        await create_notification(
            booking["created_by"],
            "payment_received",
            "Payment Received",
            f"Payment of ₹{payment.amount:,.2f} recorded for {booking_number}. Total paid: ₹{new_total_paid:,.2f}",
            {"booking_id": booking_id, "booking_number": booking_number, "amount": payment.amount}
        )
    
    # If payment complete, notify all relevant parties
    if is_complete:
        # Notify PE Desk about completed payment
        await notify_roles(
            [1, 2],  # PE Desk and Zonal Manager
            "payment_complete",
            "Payment Complete - Ready for DP Transfer",
            f"Booking {booking_number} ({stock['symbol'] if stock else 'Unknown'}) is fully paid and ready for DP transfer to {client['name'] if client else 'Unknown'}",
            {"booking_id": booking_id, "booking_number": booking_number, "client_name": client["name"] if client else None}
        )
        
        # Send email to client about completion
        if client and client.get("email"):
            await send_templated_email(
                "payment_complete",
                client["email"],
                {
                    "client_name": client["name"],
                    "booking_number": booking_number,
                    "stock_symbol": stock["symbol"] if stock else "Unknown",
                    "quantity": booking.get("quantity"),
                    "total_amount": f"{total_amount:,.2f}"
                }
            )
    
    return {
        "message": f"Payment tranche {new_tranche['tranche_number']} recorded",
        "total_paid": new_total_paid,
        "remaining": total_amount - new_total_paid,
        "payment_status": payment_status,
        "dp_transfer_ready": is_complete
    }

@api_router.get("/bookings/{booking_id}/payments")
async def get_booking_payments(
    booking_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Get payment details for a booking"""
    check_permission(current_user, "manage_bookings")
    
    booking = await db.bookings.find_one({"id": booking_id}, {"_id": 0})
    if not booking:
        raise HTTPException(status_code=404, detail="Booking not found")
    
    total_amount = (booking.get("selling_price") or 0) * booking.get("quantity", 0)
    total_paid = booking.get("total_paid", 0)
    
    return {
        "payments": booking.get("payments", []),
        "total_paid": total_paid,
        "total_amount": total_amount,
        "remaining": total_amount - total_paid,
        "payment_status": booking.get("payment_status", "pending"),
        "dp_transfer_ready": booking.get("dp_transfer_ready", False)
    }


# ============== Insider Trading Form Upload ==============

@api_router.post("/bookings/{booking_id}/insider-form")
async def upload_insider_form(
    booking_id: str,
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """Upload insider trading compliance form for Own bookings"""
    booking = await db.bookings.find_one({"id": booking_id}, {"_id": 0})
    if not booking:
        raise HTTPException(status_code=404, detail="Booking not found")
    
    # Verify user owns this booking or is PE Level
    if booking.get("created_by") != current_user["id"] and not is_pe_level(current_user.get("role", 6)):
        raise HTTPException(status_code=403, detail="You can only upload forms for your own bookings")
    
    if booking.get("booking_type") != "own":
        raise HTTPException(status_code=400, detail="Insider forms are only required for 'Own' bookings")
    
    # Create directory for insider forms
    forms_dir = UPLOAD_DIR / "insider_forms" / booking_id
    forms_dir.mkdir(parents=True, exist_ok=True)
    
    # Save file
    file_ext = Path(file.filename).suffix
    filename = f"insider_form_{datetime.now().strftime('%Y%m%d%H%M%S')}{file_ext}"
    file_path = forms_dir / filename
    
    async with aiofiles.open(file_path, 'wb') as f:
        content = await file.read()
        await f.write(content)
    
    # Update booking
    await db.bookings.update_one(
        {"id": booking_id},
        {"$set": {
            "insider_form_uploaded": True,
            "insider_form_path": str(file_path),
            "insider_form_filename": filename,
            "insider_form_uploaded_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # Create audit log
    await create_audit_log(
        action="INSIDER_FORM_UPLOAD",
        entity_type="booking",
        entity_id=booking_id,
        user_id=current_user["id"],
        user_name=current_user["name"],
        user_role=current_user.get("role", 5),
        entity_name=f"Insider Form - {booking.get('booking_number', booking_id[:8])}",
        details={
            "filename": filename,
            "booking_type": "own"
        }
    )
    
    # Notify PE Desk
    await notify_roles(
        [1],
        "insider_form_uploaded",
        "Insider Trading Form Uploaded",
        f"{current_user['name']} uploaded insider trading form for booking {booking.get('booking_number', booking_id[:8])}",
        {"booking_id": booking_id, "booking_number": booking.get("booking_number")}
    )
    
    return {
        "message": "Insider trading form uploaded successfully",
        "filename": filename
    }


@api_router.get("/bookings/{booking_id}/insider-form")
async def download_insider_form(
    booking_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Download insider trading form for a booking (PE Level)"""
    if not is_pe_level(current_user.get("role", 6)):
        raise HTTPException(status_code=403, detail="Only PE Desk or PE Manager can download insider forms")
    
    booking = await db.bookings.find_one({"id": booking_id}, {"_id": 0})
    if not booking:
        raise HTTPException(status_code=404, detail="Booking not found")
    
    if not booking.get("insider_form_path"):
        raise HTTPException(status_code=404, detail="No insider form uploaded for this booking")
    
    file_path = Path(booking["insider_form_path"])
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="Form file not found")
    
    return FileResponse(
        file_path, 
        filename=booking.get("insider_form_filename", "insider_form.pdf")
    )

@api_router.delete("/bookings/{booking_id}/payments/{tranche_number}")
async def delete_payment_tranche(
    booking_id: str,
    tranche_number: int,
    current_user: dict = Depends(get_current_user)
):
    """Delete a payment tranche (PE Desk only - deletion restricted)"""
    if not is_pe_desk_only(current_user.get("role", 6)):
        raise HTTPException(status_code=403, detail="Only PE Desk can delete payment tranches")
    
    booking = await db.bookings.find_one({"id": booking_id}, {"_id": 0})
    if not booking:
        raise HTTPException(status_code=404, detail="Booking not found")
    
    payments = booking.get("payments", [])
    payments = [p for p in payments if p.get("tranche_number") != tranche_number]
    
    # Recalculate
    total_paid = sum(p.get("amount", 0) for p in payments)
    total_amount = (booking.get("selling_price") or 0) * booking.get("quantity", 0)
    is_complete = abs(total_paid - total_amount) < 0.01
    payment_status = "completed" if is_complete else ("partial" if total_paid > 0 else "pending")
    
    # Renumber tranches
    for i, p in enumerate(payments):
        p["tranche_number"] = i + 1
    
    await db.bookings.update_one(
        {"id": booking_id},
        {"$set": {
            "payments": payments,
            "total_paid": total_paid,
            "payment_status": payment_status,
            "dp_transfer_ready": is_complete,
            "payment_completed_at": None if not is_complete else booking.get("payment_completed_at")
        }}
    )
    
    return {"message": "Payment tranche deleted"}

@api_router.get("/dp-transfer-report")
async def get_dp_transfer_report(
    current_user: dict = Depends(get_current_user)
):
    """Get bookings ready for DP transfer and already transferred (PE Desk & Zonal Manager only)"""
    user_role = current_user.get("role", 5)
    if user_role not in [1, 2]:
        raise HTTPException(status_code=403, detail="Only PE Desk and Zonal Manager can access DP transfer report")
    
    # Get all bookings ready for DP transfer OR already transferred
    bookings = await db.bookings.find(
        {
            "approval_status": "approved",
            "$or": [
                {"dp_transfer_ready": True},
                {"stock_transferred": True}
            ]
        },
        {"_id": 0}
    ).to_list(10000)
    
    if not bookings:
        return []
    
    # Get client and stock details
    client_ids = list(set(b.get("client_id") for b in bookings))
    stock_ids = list(set(b.get("stock_id") for b in bookings))
    
    clients = await db.clients.find({"id": {"$in": client_ids}}, {"_id": 0}).to_list(1000)
    stocks = await db.stocks.find({"id": {"$in": stock_ids}}, {"_id": 0}).to_list(1000)
    
    client_map = {c["id"]: c for c in clients}
    stock_map = {s["id"]: s for s in stocks}
    
    result = []
    for booking in bookings:
        client = client_map.get(booking.get("client_id"), {})
        stock = stock_map.get(booking.get("stock_id"), {})
        
        total_amount = (booking.get("selling_price") or 0) * booking.get("quantity", 0)
        
        result.append({
            "booking_id": booking["id"],
            "client_name": client.get("name", "Unknown"),
            "pan_number": client.get("pan_number", "N/A"),
            "dp_id": client.get("dp_id", "N/A"),
            "stock_symbol": stock.get("symbol", "Unknown"),
            "stock_name": stock.get("name", "Unknown"),
            "isin_number": stock.get("isin_number", "N/A"),
            "quantity": booking.get("quantity", 0),
            "total_amount": total_amount,
            "total_paid": booking.get("total_paid", 0),
            "payment_completed_at": booking.get("payment_completed_at", ""),
            "booking_date": booking.get("booking_date", ""),
            "payments": booking.get("payments", []),
            "stock_transferred": booking.get("stock_transferred", False),
            "stock_transferred_at": booking.get("stock_transferred_at", "")
        })
    
    # Sort by transfer status (not transferred first), then by payment completion date
    result.sort(key=lambda x: (x.get("stock_transferred", False), x.get("payment_completed_at", "")), reverse=False)
    
    return result

@api_router.get("/dp-transfer-report/export")
async def export_dp_transfer_report(
    format: str = "csv",
    current_user: dict = Depends(get_current_user)
):
    """Export DP transfer report as CSV or Excel (PE Desk & Zonal Manager only)"""
    user_role = current_user.get("role", 5)
    if user_role not in [1, 2]:
        raise HTTPException(status_code=403, detail="Only PE Desk and Zonal Manager can export DP transfer report")
    
    # Get report data
    bookings = await db.bookings.find(
        {"dp_transfer_ready": True, "approval_status": "approved"},
        {"_id": 0}
    ).to_list(10000)
    
    if not bookings:
        raise HTTPException(status_code=404, detail="No records to export")
    
    # Get client and stock details
    client_ids = list(set(b.get("client_id") for b in bookings))
    stock_ids = list(set(b.get("stock_id") for b in bookings))
    
    clients = await db.clients.find({"id": {"$in": client_ids}}, {"_id": 0}).to_list(1000)
    stocks = await db.stocks.find({"id": {"$in": stock_ids}}, {"_id": 0}).to_list(1000)
    
    client_map = {c["id"]: c for c in clients}
    stock_map = {s["id"]: s for s in stocks}
    
    # Build export data
    rows = []
    for booking in bookings:
        client = client_map.get(booking.get("client_id"), {})
        stock = stock_map.get(booking.get("stock_id"), {})
        
        rows.append({
            "Client Name": client.get("name", "Unknown"),
            "PAN Number": client.get("pan_number", "N/A"),
            "DP ID": client.get("dp_id", "N/A"),
            "Stock Symbol": stock.get("symbol", "Unknown"),
            "Stock Name": stock.get("name", "Unknown"),
            "ISIN": stock.get("isin_number", "N/A"),
            "Quantity": booking.get("quantity", 0),
            "Total Amount": (booking.get("selling_price") or 0) * booking.get("quantity", 0),
            "Total Paid": booking.get("total_paid", 0),
            "Payment Completed Date": booking.get("payment_completed_at", ""),
            "Booking Date": booking.get("booking_date", "")
        })
    
    df = pd.DataFrame(rows)
    
    if format == "excel":
        buffer = io.BytesIO()
        df.to_excel(buffer, index=False, engine='openpyxl')
        buffer.seek(0)
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=dp_transfer_report.xlsx"}
        )
    else:
        buffer = io.StringIO()
        df.to_csv(buffer, index=False)
        buffer.seek(0)
        return StreamingResponse(
            io.BytesIO(buffer.getvalue().encode()),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=dp_transfer_report.csv"}
        )

# ============== Stock Transfer Confirmation ==============

class StockTransferConfirm(BaseModel):
    notes: Optional[str] = None

@api_router.put("/bookings/{booking_id}/confirm-transfer")
async def confirm_stock_transfer(
    booking_id: str,
    data: StockTransferConfirm = None,
    current_user: dict = Depends(get_current_user)
):
    """Mark a booking as stock transferred and notify client (PE Level)"""
    user_role = current_user.get("role", 6)
    
    if not is_pe_level(user_role):
        raise HTTPException(status_code=403, detail="Only PE Desk or PE Manager can confirm stock transfers")
    
    booking = await db.bookings.find_one({"id": booking_id}, {"_id": 0})
    if not booking:
        raise HTTPException(status_code=404, detail="Booking not found")
    
    if not booking.get("dp_transfer_ready"):
        raise HTTPException(status_code=400, detail="Booking is not ready for DP transfer. Full payment required.")
    
    if booking.get("stock_transferred"):
        raise HTTPException(status_code=400, detail="Stock has already been transferred for this booking")
    
    # Update booking with transfer confirmation
    update_data = {
        "stock_transferred": True,
        "stock_transferred_at": datetime.now(timezone.utc).isoformat(),
        "stock_transferred_by": current_user["id"],
        "stock_transfer_notes": data.notes if data else None
    }
    
    await db.bookings.update_one({"id": booking_id}, {"$set": update_data})
    
    # Update inventory - this will move quantity from blocked to transferred
    await update_inventory(booking["stock_id"])
    
    # Get client and stock details for email
    client = await db.clients.find_one({"id": booking["client_id"]}, {"_id": 0})
    stock = await db.stocks.find_one({"id": booking["stock_id"]}, {"_id": 0})
    booking_number = booking.get("booking_number", booking_id[:8].upper())
    
    # Create audit log
    await create_audit_log(
        action="STOCK_TRANSFER",
        entity_type="booking",
        entity_id=booking_id,
        user_id=current_user["id"],
        user_name=current_user["name"],
        user_role=user_role,
        entity_name=f"{stock['symbol'] if stock else 'Unknown'} - {client['name'] if client else 'Unknown'}",
        details={
            "booking_number": booking_number,
            "client_name": client["name"] if client else "Unknown",
            "client_dp_id": client.get("dp_id", "N/A") if client else "N/A",
            "stock_symbol": stock["symbol"] if stock else "Unknown",
            "quantity": booking.get("quantity"),
            "transfer_notes": data.notes if data else None
        }
    )
    
    # Send email to client using template
    if client and client.get("email"):
        await send_templated_email(
            "stock_transfer_complete",
            client["email"],
            {
                "client_name": client["name"],
                "booking_number": booking_number,
                "stock_symbol": stock["symbol"] if stock else "N/A",
                "stock_name": stock["name"] if stock else "N/A",
                "isin_number": stock.get("isin_number", "N/A") if stock else "N/A",
                "quantity": booking.get("quantity", 0),
                "dp_id": client.get("dp_id", "N/A"),
                "transfer_date": datetime.now(timezone.utc).strftime('%d %b %Y, %H:%M')
            }
        )
    
    # Notify the booking creator
    if booking.get("created_by"):
        await create_notification(
            booking["created_by"],
            "stock_transferred",
            "Stock Transfer Completed",
            f"Stock transfer completed for booking {booking_number} - {stock['symbol'] if stock else 'Unknown'} to {client['name'] if client else 'Unknown'}",
            {"booking_id": booking_id, "booking_number": booking_number, "client_name": client["name"] if client else None}
        )
    
    return {
        "message": "Stock transfer confirmed and client notified",
        "booking_id": booking_id,
        "booking_number": booking_number,
        "client_email": client.get("email") if client else None
    }


# ============== Purchase Payment Tracking ==============

class PurchasePaymentCreate(BaseModel):
    amount: float
    payment_date: str
    notes: Optional[str] = None
    proof_url: Optional[str] = None

@api_router.post("/purchases/{purchase_id}/payments")
async def add_purchase_payment(
    purchase_id: str,
    payment: PurchasePaymentCreate,
    current_user: dict = Depends(get_current_user)
):
    """Record payment for a vendor purchase and notify vendor (PE Level)"""
    user_role = current_user.get("role", 6)
    
    if not is_pe_level(user_role):
        raise HTTPException(status_code=403, detail="Only PE Desk or PE Manager can record purchase payments")
    
    purchase = await db.purchases.find_one({"id": purchase_id}, {"_id": 0})
    if not purchase:
        raise HTTPException(status_code=404, detail="Purchase not found")
    
    vendor = await db.clients.find_one({"id": purchase["vendor_id"], "is_vendor": True}, {"_id": 0})
    stock = await db.stocks.find_one({"id": purchase["stock_id"]}, {"_id": 0})
    
    # Get current payment info
    payments = purchase.get("payments", [])
    current_paid = purchase.get("total_paid", 0)
    remaining = purchase["total_amount"] - current_paid
    
    if payment.amount > remaining + 0.01:
        raise HTTPException(
            status_code=400,
            detail=f"Payment amount ({payment.amount}) exceeds remaining balance ({remaining:.2f})"
        )
    
    # Add new payment tranche
    new_payment = {
        "tranche_number": len(payments) + 1,
        "amount": payment.amount,
        "payment_date": payment.payment_date,
        "recorded_by": current_user["id"],
        "recorded_at": datetime.now(timezone.utc).isoformat(),
        "notes": payment.notes,
        "proof_url": payment.proof_url
    }
    
    payments.append(new_payment)
    new_total_paid = current_paid + payment.amount
    
    is_complete = abs(new_total_paid - purchase["total_amount"]) < 0.01
    payment_status = "completed" if is_complete else ("partial" if new_total_paid > 0 else "pending")
    
    update_data = {
        "payments": payments,
        "total_paid": new_total_paid,
        "payment_status": payment_status
    }
    
    if is_complete:
        update_data["payment_completed_at"] = payment.payment_date
    
    await db.purchases.update_one({"id": purchase_id}, {"$set": update_data})
    
    # Create audit log
    await create_audit_log(
        action="PURCHASE_PAYMENT",
        entity_type="purchase",
        entity_id=purchase_id,
        user_id=current_user["id"],
        user_name=current_user["name"],
        user_role=user_role,
        entity_name=f"{stock['symbol'] if stock else 'Unknown'} - {vendor['name'] if vendor else 'Unknown'}",
        details={
            "vendor_name": vendor["name"] if vendor else "Unknown",
            "stock_symbol": stock["symbol"] if stock else "Unknown",
            "tranche_number": new_payment["tranche_number"],
            "amount": payment.amount,
            "total_paid": new_total_paid,
            "payment_status": payment_status
        }
    )
    
    # Send email to vendor using template
    if vendor and vendor.get("email"):
        await send_templated_email(
            "vendor_payment_received",
            vendor["email"],
            {
                "vendor_name": vendor["name"],
                "stock_symbol": stock["symbol"] if stock else "N/A",
                "stock_name": stock["name"] if stock else "N/A",
                "quantity": purchase.get("quantity", 0),
                "purchase_date": purchase.get("purchase_date", "N/A"),
                "total_amount": f"{purchase['total_amount']:,.2f}",
                "payment_amount": f"{payment.amount:,.2f}",
                "total_paid": f"{new_total_paid:,.2f}",
                "remaining_balance": f"{(purchase['total_amount'] - new_total_paid):,.2f}",
                "payment_status": payment_status.upper(),
                "is_complete": is_complete
            }
        )
    
    return {
        "message": f"Payment recorded and vendor notified",
        "tranche_number": new_payment["tranche_number"],
        "amount": payment.amount,
        "total_paid": new_total_paid,
        "remaining": purchase["total_amount"] - new_total_paid,
        "payment_status": payment_status,
        "vendor_email": vendor.get("email") if vendor else None
    }


@api_router.get("/purchases/{purchase_id}/payments")
async def get_purchase_payments(
    purchase_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Get payment details for a purchase"""
    user_role = current_user.get("role", 6)
    
    if not is_pe_level(user_role):
        raise HTTPException(status_code=403, detail="Only PE Desk or PE Manager can view purchase payments")
    
    purchase = await db.purchases.find_one({"id": purchase_id}, {"_id": 0})
    if not purchase:
        raise HTTPException(status_code=404, detail="Purchase not found")
    
    return {
        "purchase_id": purchase_id,
        "total_amount": purchase.get("total_amount", 0),
        "total_paid": purchase.get("total_paid", 0),
        "remaining": purchase.get("total_amount", 0) - purchase.get("total_paid", 0),
        "payment_status": purchase.get("payment_status", "pending"),
        "payments": purchase.get("payments", []),
        "payment_completed_at": purchase.get("payment_completed_at")
    }


@api_router.delete("/purchases/{purchase_id}")
async def delete_purchase(purchase_id: str, current_user: dict = Depends(get_current_user)):
    """Delete a purchase (PE Desk only - deletion restricted)"""
    user_role = current_user.get("role", 6)
    
    if not is_pe_desk_only(user_role):
        raise HTTPException(status_code=403, detail="Only PE Desk can delete purchases")
    
    purchase = await db.purchases.find_one({"id": purchase_id}, {"_id": 0})
    if not purchase:
        raise HTTPException(status_code=404, detail="Purchase not found")
    
    # Get vendor and stock info for audit log
    vendor = await db.clients.find_one({"id": purchase["vendor_id"]}, {"_id": 0})
    stock = await db.stocks.find_one({"id": purchase["stock_id"]}, {"_id": 0})
    
    result = await db.purchases.delete_one({"id": purchase_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Purchase not found")
    
    # Update inventory after deletion
    await update_inventory(purchase["stock_id"])
    
    # Create audit log
    await create_audit_log(
        action="PURCHASE_DELETE",
        entity_type="purchase",
        entity_id=purchase_id,
        user_id=current_user["id"],
        user_name=current_user["name"],
        user_role=user_role,
        entity_name=f"{stock['symbol'] if stock else 'Unknown'} - {vendor['name'] if vendor else 'Unknown'}",
        details={
            "vendor_name": vendor["name"] if vendor else "Unknown",
            "stock_symbol": stock["symbol"] if stock else "Unknown",
            "quantity": purchase.get("quantity"),
            "total_amount": purchase.get("total_amount"),
            "deleted_by": current_user["name"]
        }
    )
    
    return {"message": "Purchase deleted successfully"}

@api_router.post("/bookings/bulk-upload")
async def bulk_upload_bookings(file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
    check_permission(current_user, "manage_bookings")
    
    try:
        content = await file.read()
        df = pd.read_csv(io.BytesIO(content))
        
        required_columns = ["client_id", "stock_id", "quantity", "booking_date"]
        if not all(col in df.columns for col in required_columns):
            raise HTTPException(status_code=400, detail=f"CSV must contain columns: {', '.join(required_columns)}")
        
        bookings_created = 0
        for _, row in df.iterrows():
            # Get inventory for weighted average
            inventory = await db.inventory.find_one({"stock_id": str(row["stock_id"])}, {"_id": 0})
            buying_price = float(row.get("buying_price", inventory["weighted_avg_price"] if inventory else 0))
            
            booking_id = str(uuid.uuid4())
            booking_doc = {
                "id": booking_id,
                "client_id": str(row["client_id"]),
                "stock_id": str(row["stock_id"]),
                "quantity": int(row["quantity"]),
                "buying_price": buying_price,
                "selling_price": float(row["selling_price"]) if pd.notna(row.get("selling_price")) else None,
                "booking_date": str(row["booking_date"]),
                "status": str(row.get("status", "open")),
                "notes": str(row.get("notes", "")) if pd.notna(row.get("notes")) else None,
                "user_id": current_user["id"],
                "created_by": current_user["id"],
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            
            await db.bookings.insert_one(booking_doc)
            await update_inventory(str(row["stock_id"]))
            bookings_created += 1
        
        return {"message": f"Successfully uploaded {bookings_created} bookings"}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error processing CSV: {str(e)}")

# Dashboard & Analytics
@api_router.get("/dashboard/stats", response_model=DashboardStats)
async def get_dashboard_stats(current_user: dict = Depends(get_current_user)):
    query_filter = {}
    if current_user.get("role", 5) >= 3:
        if "view_all" not in ROLE_PERMISSIONS.get(current_user.get("role", 5), []):
            query_filter = {"created_by": current_user["id"]}
    
    total_clients = await db.clients.count_documents({})
    total_vendors = await db.clients.count_documents({"is_vendor": True})
    total_stocks = await db.stocks.count_documents({})
    total_bookings = await db.bookings.count_documents(query_filter)
    open_bookings = await db.bookings.count_documents({**query_filter, "status": "open"})
    closed_bookings = await db.bookings.count_documents({**query_filter, "status": "closed"})
    total_purchases = await db.purchases.count_documents({})
    
    # Calculate total P&L
    bookings = await db.bookings.find({**query_filter, "status": "closed"}, {"_id": 0}).to_list(1000)
    total_profit_loss = sum(
        (b["selling_price"] - b["buying_price"]) * b["quantity"]
        for b in bookings if b.get("selling_price")
    )
    
    # Calculate total inventory value
    inventory_items = await db.inventory.find({}, {"_id": 0}).to_list(1000)
    total_inventory_value = sum(item.get("total_value", 0) for item in inventory_items)
    
    return DashboardStats(
        total_clients=total_clients,
        total_vendors=total_vendors,
        total_stocks=total_stocks,
        total_bookings=total_bookings,
        open_bookings=open_bookings,
        closed_bookings=closed_bookings,
        total_profit_loss=total_profit_loss,
        total_inventory_value=total_inventory_value,
        total_purchases=total_purchases
    )

@api_router.get("/dashboard/analytics")
async def get_dashboard_analytics(current_user: dict = Depends(get_current_user)):
    """Get analytics data for charts"""
    query_filter = {}
    if current_user.get("role", 5) >= 3:
        if "view_all" not in ROLE_PERMISSIONS.get(current_user.get("role", 5), []):
            query_filter = {"created_by": current_user["id"]}
    
    # P&L Trend over time (last 12 months)
    bookings = await db.bookings.find({**query_filter, "status": "closed"}, {"_id": 0}).to_list(10000)
    
    # Group by month
    monthly_pnl = {}
    for booking in bookings:
        if booking.get("selling_price"):
            booking_date = datetime.fromisoformat(booking["booking_date"]).strftime("%Y-%m")
            pnl = (booking["selling_price"] - booking["buying_price"]) * booking["quantity"]
            monthly_pnl[booking_date] = monthly_pnl.get(booking_date, 0) + pnl
    
    # Top performing stocks
    stock_performance = {}
    for booking in bookings:
        if booking.get("selling_price"):
            stock_id = booking["stock_id"]
            pnl = (booking["selling_price"] - booking["buying_price"]) * booking["quantity"]
            if stock_id not in stock_performance:
                stock_performance[stock_id] = {"pnl": 0, "quantity": 0}
            stock_performance[stock_id]["pnl"] += pnl
            stock_performance[stock_id]["quantity"] += booking["quantity"]
    
    # Enrich with stock names
    stock_ids = list(stock_performance.keys())
    stocks = await db.stocks.find({"id": {"$in": stock_ids}}, {"_id": 0}).to_list(1000)
    stock_map = {s["id"]: s for s in stocks}
    
    top_stocks = [
        {
            "stock_symbol": stock_map.get(stock_id, {}).get("symbol", "Unknown"),
            "pnl": data["pnl"],
            "quantity": data["quantity"]
        }
        for stock_id, data in sorted(stock_performance.items(), key=lambda x: x[1]["pnl"], reverse=True)[:10]
    ]
    
    return {
        "monthly_pnl": [{"month": k, "pnl": v} for k, v in sorted(monthly_pnl.items())],
        "top_stocks": top_stocks
    }

# Audit Logs API
@api_router.get("/audit-logs")
async def get_audit_logs(
    entity_type: Optional[str] = None,
    action: Optional[str] = None,
    user_id: Optional[str] = None,
    limit: int = 100,
    current_user: dict = Depends(get_current_user)
):
    """Get audit logs (admin only)"""
    user_role = current_user.get("role", 5)
    
    if user_role > 2:
        raise HTTPException(status_code=403, detail="Only admins can view audit logs")
    
    query = {}
    if entity_type:
        query["entity_type"] = entity_type
    if action:
        query["action"] = action
    if user_id:
        query["user_id"] = user_id
    
    logs = await db.audit_logs.find(query, {"_id": 0}).sort("timestamp", -1).to_list(limit)
    return logs

@api_router.get("/audit-logs/stats")
async def get_audit_stats(current_user: dict = Depends(get_current_user)):
    """Get audit log statistics"""
    user_role = current_user.get("role", 5)
    
    if user_role > 2:
        raise HTTPException(status_code=403, detail="Only admins can view audit stats")
    
    total_logs = await db.audit_logs.count_documents({})
    
    # Get counts by action type
    pipeline = [
        {"$group": {"_id": "$action", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}}
    ]
    action_counts = await db.audit_logs.aggregate(pipeline).to_list(100)
    
    return {
        "total_logs": total_logs,
        "by_action": {item["_id"]: item["count"] for item in action_counts}
    }

# Client Portfolio
@api_router.get("/clients/{client_id}/portfolio", response_model=ClientPortfolio)
async def get_client_portfolio(client_id: str, current_user: dict = Depends(get_current_user)):
    client = await db.clients.find_one({"id": client_id}, {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    
    bookings = await db.bookings.find({"client_id": client_id}, {"_id": 0}).to_list(1000)
    
    total_bookings = len(bookings)
    open_bookings = sum(1 for b in bookings if b["status"] == "open")
    closed_bookings = sum(1 for b in bookings if b["status"] == "closed")
    
    total_profit_loss = sum(
        (b["selling_price"] - b["buying_price"]) * b["quantity"]
        for b in bookings if b.get("selling_price") and b["status"] == "closed"
    )
    
    # Enrich bookings
    if bookings:
        stock_ids = list(set(b["stock_id"] for b in bookings))
        user_ids = list(set(b["created_by"] for b in bookings))
        
        stocks = await db.stocks.find({"id": {"$in": stock_ids}}, {"_id": 0}).to_list(1000)
        users = await db.users.find({"id": {"$in": user_ids}}, {"_id": 0}).to_list(1000)
        
        stock_map = {s["id"]: s for s in stocks}
        user_map = {u["id"]: u for u in users}
        
        enriched_bookings = []
        for booking in bookings:
            stock = stock_map.get(booking["stock_id"])
            user = user_map.get(booking["created_by"])
            
            profit_loss = None
            if booking.get("selling_price") and booking["status"] == "closed":
                profit_loss = (booking["selling_price"] - booking["buying_price"]) * booking["quantity"]
            
            # Exclude fields that will be explicitly provided
            booking_data = {k: v for k, v in booking.items() if k not in ["client_name", "stock_symbol", "stock_name", "created_by_name", "profit_loss"]}
            
            enriched_bookings.append(BookingWithDetails(
                **booking_data,
                client_name=client["name"],
                stock_symbol=stock["symbol"] if stock else "Unknown",
                stock_name=stock["name"] if stock else "Unknown",
                created_by_name=user["name"] if user else "Unknown",
                profit_loss=profit_loss
            ))
    else:
        enriched_bookings = []
    
    return ClientPortfolio(
        client_id=client_id,
        client_name=client["name"],
        total_bookings=total_bookings,
        open_bookings=open_bookings,
        closed_bookings=closed_bookings,
        total_profit_loss=total_profit_loss,
        bookings=enriched_bookings
    )

# Advanced Reports
@api_router.get("/reports/pnl")
async def get_pnl_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    client_id: Optional[str] = None,
    stock_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    check_permission(current_user, "view_reports")
    
    query = {}
    if start_date:
        query["booking_date"] = {"$gte": start_date}
    if end_date:
        if "booking_date" in query:
            query["booking_date"]["$lte"] = end_date
        else:
            query["booking_date"] = {"$lte": end_date}
    if client_id:
        query["client_id"] = client_id
    if stock_id:
        query["stock_id"] = stock_id
    
    bookings = await db.bookings.find(query, {"_id": 0, "user_id": 0}).to_list(1000)
    
    if not bookings:
        return []
    
    # Batch fetch related data
    client_ids = list(set(b["client_id"] for b in bookings))
    stock_ids = list(set(b["stock_id"] for b in bookings))
    
    clients = await db.clients.find({"id": {"$in": client_ids}}, {"_id": 0}).to_list(1000)
    stocks = await db.stocks.find({"id": {"$in": stock_ids}}, {"_id": 0}).to_list(1000)
    
    client_map = {c["id"]: c for c in clients}
    stock_map = {s["id"]: s for s in stocks}
    
    report = []
    for booking in bookings:
        client = client_map.get(booking["client_id"])
        stock = stock_map.get(booking["stock_id"])
        
        profit_loss = None
        if booking.get("selling_price") and booking["status"] == "closed":
            profit_loss = (booking["selling_price"] - booking["buying_price"]) * booking["quantity"]
        
        report.append({
            "booking_id": booking["id"],
            "client_name": client["name"] if client else "Unknown",
            "stock_symbol": stock["symbol"] if stock else "Unknown",
            "stock_name": stock["name"] if stock else "Unknown",
            "quantity": booking["quantity"],
            "buying_price": booking["buying_price"],
            "selling_price": booking.get("selling_price"),
            "booking_date": booking["booking_date"],
            "status": booking["status"],
            "profit_loss": profit_loss
        })
    
    return report

@api_router.get("/reports/export/excel")
async def export_excel(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    check_permission(current_user, "view_reports")
    
    query = {}
    if start_date:
        query["booking_date"] = {"$gte": start_date}
    if end_date:
        if "booking_date" in query:
            query["booking_date"]["$lte"] = end_date
        else:
            query["booking_date"] = {"$lte": end_date}
    
    bookings = await db.bookings.find(query, {"_id": 0, "user_id": 0}).to_list(1000)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Revenue Report"
    
    # Headers
    headers = ["Client", "Stock Symbol", "Stock Name", "Quantity", "Landing Price", "Selling Price", "Date", "Status", "Revenue"]
    ws.append(headers)
    
    # Style headers
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="064E3B", end_color="064E3B", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    if bookings:
        client_ids = list(set(b["client_id"] for b in bookings))
        stock_ids = list(set(b["stock_id"] for b in bookings))
        
        clients = await db.clients.find({"id": {"$in": client_ids}}, {"_id": 0}).to_list(1000)
        stocks = await db.stocks.find({"id": {"$in": stock_ids}}, {"_id": 0}).to_list(1000)
        
        client_map = {c["id"]: c for c in clients}
        stock_map = {s["id"]: s for s in stocks}
        
        for booking in bookings:
            client = client_map.get(booking["client_id"])
            stock = stock_map.get(booking["stock_id"])
            
            profit_loss = ""
            if booking.get("selling_price") and booking["status"] == "closed":
                profit_loss = (booking["selling_price"] - booking["buying_price"]) * booking["quantity"]
            
            ws.append([
                client["name"] if client else "Unknown",
                stock["symbol"] if stock else "Unknown",
                stock["name"] if stock else "Unknown",
                booking["quantity"],
                booking["buying_price"],
                booking.get("selling_price") or "",
                booking["booking_date"],
                booking["status"],
                profit_loss
            ])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=pnl_report.xlsx"}
    )

@api_router.get("/reports/export/pdf")
async def export_pdf(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    check_permission(current_user, "view_reports")
    
    query = {}
    if start_date:
        query["booking_date"] = {"$gte": start_date}
    if end_date:
        if "booking_date" in query:
            query["booking_date"]["$lte"] = end_date
        else:
            query["booking_date"] = {"$lte": end_date}
    
    bookings = await db.bookings.find(query, {"_id": 0, "user_id": 0}).to_list(1000)
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#064E3B'),
        alignment=TA_CENTER
    )
    elements.append(Paragraph("Profit & Loss Report", title_style))
    elements.append(Spacer(1, 0.3*inch))
    
    data = [["Client", "Stock", "Qty", "Landing Price", "Sell Price", "Status", "Revenue"]]
    
    if bookings:
        client_ids = list(set(b["client_id"] for b in bookings))
        stock_ids = list(set(b["stock_id"] for b in bookings))
        
        clients = await db.clients.find({"id": {"$in": client_ids}}, {"_id": 0}).to_list(1000)
        stocks = await db.stocks.find({"id": {"$in": stock_ids}}, {"_id": 0}).to_list(1000)
        
        client_map = {c["id"]: c for c in clients}
        stock_map = {s["id"]: s for s in stocks}
        
        for booking in bookings:
            client = client_map.get(booking["client_id"])
            stock = stock_map.get(booking["stock_id"])
            
            profit_loss = ""
            if booking.get("selling_price") and booking["status"] == "closed":
                pl = (booking["selling_price"] - booking["buying_price"]) * booking["quantity"]
                profit_loss = f"₹{pl:,.2f}"
            
            data.append([
                client["name"] if client else "Unknown",
                stock["symbol"] if stock else "Unknown",
                str(booking["quantity"]),
                f"₹{booking['buying_price']:,.2f}",
                f"₹{booking.get('selling_price', 0):,.2f}" if booking.get('selling_price') else "-",
                booking["status"].upper(),
                profit_loss
            ])
    
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#064E3B')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=pnl_report.pdf"}
    )

# Notification Routes
@api_router.get("/notifications")
async def get_notifications(
    unread_only: bool = False,
    limit: int = 50,
    current_user: dict = Depends(get_current_user)
):
    """Get user notifications"""
    query = {"user_id": current_user["id"]}
    if unread_only:
        query["read"] = False
    
    notifications = await db.notifications.find(
        query, {"_id": 0}
    ).sort("created_at", -1).limit(limit).to_list(limit)
    
    return notifications

@api_router.get("/notifications/unread-count")
async def get_unread_count(current_user: dict = Depends(get_current_user)):
    """Get count of unread notifications"""
    count = await db.notifications.count_documents({
        "user_id": current_user["id"],
        "read": False
    })
    return {"count": count}

@api_router.put("/notifications/{notification_id}/read")
async def mark_notification_read(notification_id: str, current_user: dict = Depends(get_current_user)):
    """Mark notification as read"""
    result = await db.notifications.update_one(
        {"id": notification_id, "user_id": current_user["id"]},
        {"$set": {"read": True}}
    )
    return {"message": "Notification marked as read"}

@api_router.put("/notifications/read-all")
async def mark_all_notifications_read(current_user: dict = Depends(get_current_user)):
    """Mark all notifications as read"""
    result = await db.notifications.update_many(
        {"user_id": current_user["id"], "read": False},
        {"$set": {"read": True}}
    )
    return {"message": f"Marked {result.modified_count} notifications as read"}

# ============== Finance / Payments Dashboard ==============

@api_router.get("/finance/payments")
async def get_all_payments(
    payment_type: Optional[str] = None,  # 'client', 'vendor', or None for all
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get all payments (client and vendor) for finance dashboard (PE Level or Finance role)"""
    if not has_finance_access(current_user.get("role", 6)):
        raise HTTPException(status_code=403, detail="Only PE Desk, PE Manager, or Finance can access finance data")
    
    all_payments = []
    
    # Get client payments from bookings
    if payment_type in [None, 'client']:
        booking_query = {"payments": {"$exists": True, "$ne": []}}
        if start_date:
            booking_query["booking_date"] = {"$gte": start_date}
        if end_date:
            if "booking_date" in booking_query:
                booking_query["booking_date"]["$lte"] = end_date
            else:
                booking_query["booking_date"] = {"$lte": end_date}
        
        bookings = await db.bookings.find(booking_query, {"_id": 0}).to_list(5000)
        
        # Get related data
        client_ids = list(set(b["client_id"] for b in bookings))
        stock_ids = list(set(b["stock_id"] for b in bookings))
        user_ids = list(set(p.get("recorded_by") for b in bookings for p in b.get("payments", []) if p.get("recorded_by")))
        
        clients = await db.clients.find({"id": {"$in": client_ids}}, {"_id": 0, "id": 1, "name": 1}).to_list(1000)
        stocks = await db.stocks.find({"id": {"$in": stock_ids}}, {"_id": 0, "id": 1, "symbol": 1, "name": 1}).to_list(1000)
        users = await db.users.find({"id": {"$in": user_ids}}, {"_id": 0, "id": 1, "name": 1}).to_list(100)
        
        client_map = {c["id"]: c for c in clients}
        stock_map = {s["id"]: s for s in stocks}
        user_map = {u["id"]: u["name"] for u in users}
        
        for booking in bookings:
            client = client_map.get(booking["client_id"], {})
            stock = stock_map.get(booking["stock_id"], {})
            booking_number = booking.get("booking_number", booking["id"][:8].upper())
            
            for payment in booking.get("payments", []):
                all_payments.append({
                    "id": f"client_{booking['id']}_{payment['tranche_number']}",
                    "type": "client",
                    "direction": "received",
                    "reference_id": booking["id"],
                    "reference_number": booking_number,
                    "party_name": client.get("name", "Unknown"),
                    "party_id": booking["client_id"],
                    "stock_symbol": stock.get("symbol", "Unknown"),
                    "stock_name": stock.get("name", "Unknown"),
                    "tranche_number": payment["tranche_number"],
                    "amount": payment["amount"],
                    "payment_date": payment["payment_date"],
                    "recorded_by": user_map.get(payment.get("recorded_by"), "System"),
                    "recorded_at": payment.get("recorded_at"),
                    "notes": payment.get("notes"),
                    "proof_url": payment.get("proof_url"),
                    "booking_date": booking.get("booking_date"),
                    "total_amount": (booking.get("selling_price") or 0) * booking.get("quantity", 0),
                    "total_paid": booking.get("total_paid", 0),
                    "payment_status": booking.get("payment_status", "pending")
                })
    
    # Get vendor payments from purchases
    if payment_type in [None, 'vendor']:
        purchase_query = {"payments": {"$exists": True, "$ne": []}}
        if start_date:
            purchase_query["purchase_date"] = {"$gte": start_date}
        if end_date:
            if "purchase_date" in purchase_query:
                purchase_query["purchase_date"]["$lte"] = end_date
            else:
                purchase_query["purchase_date"] = {"$lte": end_date}
        
        purchases = await db.purchases.find(purchase_query, {"_id": 0}).to_list(5000)
        
        # Get related data
        vendor_ids = list(set(p["vendor_id"] for p in purchases))
        stock_ids = list(set(p["stock_id"] for p in purchases))
        user_ids = list(set(pay.get("recorded_by") for p in purchases for pay in p.get("payments", []) if pay.get("recorded_by")))
        
        vendors = await db.clients.find({"id": {"$in": vendor_ids}, "is_vendor": True}, {"_id": 0, "id": 1, "name": 1}).to_list(1000)
        stocks = await db.stocks.find({"id": {"$in": stock_ids}}, {"_id": 0, "id": 1, "symbol": 1, "name": 1}).to_list(1000)
        users = await db.users.find({"id": {"$in": user_ids}}, {"_id": 0, "id": 1, "name": 1}).to_list(100)
        
        vendor_map = {v["id"]: v for v in vendors}
        stock_map = {s["id"]: s for s in stocks}
        user_map = {u["id"]: u["name"] for u in users}
        
        for purchase in purchases:
            vendor = vendor_map.get(purchase["vendor_id"], {})
            stock = stock_map.get(purchase["stock_id"], {})
            
            for payment in purchase.get("payments", []):
                all_payments.append({
                    "id": f"vendor_{purchase['id']}_{payment['tranche_number']}",
                    "type": "vendor",
                    "direction": "sent",
                    "reference_id": purchase["id"],
                    "reference_number": purchase["id"][:8].upper(),
                    "party_name": vendor.get("name", "Unknown"),
                    "party_id": purchase["vendor_id"],
                    "stock_symbol": stock.get("symbol", "Unknown"),
                    "stock_name": stock.get("name", "Unknown"),
                    "tranche_number": payment["tranche_number"],
                    "amount": payment["amount"],
                    "payment_date": payment["payment_date"],
                    "recorded_by": user_map.get(payment.get("recorded_by"), "System"),
                    "recorded_at": payment.get("recorded_at"),
                    "notes": payment.get("notes"),
                    "proof_url": payment.get("proof_url"),
                    "purchase_date": purchase.get("purchase_date"),
                    "total_amount": purchase.get("total_amount", 0),
                    "total_paid": purchase.get("total_paid", 0),
                    "payment_status": purchase.get("payment_status", "pending")
                })
    
    # Sort by payment date (newest first)
    all_payments.sort(key=lambda x: x.get("payment_date", ""), reverse=True)
    
    return all_payments


@api_router.get("/finance/summary")
async def get_finance_summary(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get finance summary statistics (PE Level or Finance role)"""
    if not has_finance_access(current_user.get("role", 6)):
        raise HTTPException(status_code=403, detail="Only PE Desk, PE Manager, or Finance can access finance data")
    
    # Get all payments first
    payments = await get_all_payments(None, start_date, end_date, current_user)
    
    client_payments = [p for p in payments if p["type"] == "client"]
    vendor_payments = [p for p in payments if p["type"] == "vendor"]
    
    # Get refund requests stats
    refund_requests = await db.refund_requests.find({}, {"_id": 0}).to_list(1000)
    pending_refunds = [r for r in refund_requests if r.get("status") == "pending"]
    completed_refunds = [r for r in refund_requests if r.get("status") == "completed"]
    
    return {
        "total_received": sum(p["amount"] for p in client_payments),
        "total_sent": sum(p["amount"] for p in vendor_payments),
        "client_payments_count": len(client_payments),
        "vendor_payments_count": len(vendor_payments),
        "net_flow": sum(p["amount"] for p in client_payments) - sum(p["amount"] for p in vendor_payments),
        "pending_refunds_count": len(pending_refunds),
        "pending_refunds_amount": sum(r.get("refund_amount", 0) for r in pending_refunds),
        "completed_refunds_count": len(completed_refunds),
        "completed_refunds_amount": sum(r.get("refund_amount", 0) for r in completed_refunds)
    }


# ============== Refund Requests Endpoints ==============
@api_router.get("/finance/refund-requests")
async def get_refund_requests(
    status: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get all refund requests (PE Level or Finance role)"""
    if not has_finance_access(current_user.get("role", 6)):
        raise HTTPException(status_code=403, detail="Only PE Desk, PE Manager, or Finance can access refund requests")
    
    query = {}
    if status:
        query["status"] = status
    
    refund_requests = await db.refund_requests.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return refund_requests


@api_router.get("/finance/refund-requests/{request_id}")
async def get_refund_request(request_id: str, current_user: dict = Depends(get_current_user)):
    """Get a specific refund request"""
    if not has_finance_access(current_user.get("role", 6)):
        raise HTTPException(status_code=403, detail="Only PE Desk, PE Manager, or Finance can access refund requests")
    
    refund = await db.refund_requests.find_one({"id": request_id}, {"_id": 0})
    if not refund:
        raise HTTPException(status_code=404, detail="Refund request not found")
    
    return refund


class RefundStatusUpdate(BaseModel):
    status: str  # processing, completed, failed
    notes: Optional[str] = None
    reference_number: Optional[str] = None


@api_router.put("/finance/refund-requests/{request_id}")
async def update_refund_request(
    request_id: str,
    update_data: RefundStatusUpdate,
    current_user: dict = Depends(get_current_user)
):
    """Update refund request status (PE Level or Finance role)"""
    if not can_manage_finance(current_user.get("role", 6)):
        raise HTTPException(status_code=403, detail="Only PE Desk, PE Manager, or Finance can update refund requests")
    
    if update_data.status not in ["processing", "completed", "failed"]:
        raise HTTPException(status_code=400, detail="Invalid status. Use: processing, completed, or failed")
    
    refund = await db.refund_requests.find_one({"id": request_id}, {"_id": 0})
    if not refund:
        raise HTTPException(status_code=404, detail="Refund request not found")
    
    update_fields = {
        "status": update_data.status,
        "processed_at": datetime.now(timezone.utc).isoformat(),
        "processed_by": current_user["id"],
        "processed_by_name": current_user["name"]
    }
    
    if update_data.notes:
        update_fields["notes"] = update_data.notes
    if update_data.reference_number:
        update_fields["reference_number"] = update_data.reference_number
    
    await db.refund_requests.update_one({"id": request_id}, {"$set": update_fields})
    
    # Notify about refund completion
    if update_data.status == "completed":
        # Get client info
        client = await db.clients.find_one({"id": refund["client_id"]}, {"_id": 0})
        if client and client.get("email"):
            await send_templated_email(
                "refund_completed",
                client["email"],
                {
                    "client_name": client["name"],
                    "booking_number": refund["booking_number"],
                    "refund_amount": f"{refund['refund_amount']:,.2f}",
                    "reference_number": update_data.reference_number or "N/A",
                    "stock_symbol": refund.get("stock_symbol", "N/A")
                }
            )
    
    # Audit log
    await create_audit_log(
        action=f"REFUND_{update_data.status.upper()}",
        entity_type="refund_request",
        entity_id=request_id,
        user_id=current_user["id"],
        user_name=current_user["name"],
        user_role=current_user.get("role", 6),
        details={
            "booking_number": refund["booking_number"],
            "client_name": refund["client_name"],
            "refund_amount": refund["refund_amount"],
            "status": update_data.status,
            "reference_number": update_data.reference_number
        }
    )
    
    return {"message": f"Refund request updated to {update_data.status}"}


@api_router.put("/finance/refund-requests/{request_id}/bank-details")
async def update_refund_bank_details(
    request_id: str,
    bank_name: str,
    account_number: str,
    ifsc_code: str,
    account_holder_name: str,
    branch: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Update bank details for a refund request (PE Level or Finance role)"""
    if not can_manage_finance(current_user.get("role", 6)):
        raise HTTPException(status_code=403, detail="Only PE Desk, PE Manager, or Finance can update refund requests")
    
    refund = await db.refund_requests.find_one({"id": request_id}, {"_id": 0})
    if not refund:
        raise HTTPException(status_code=404, detail="Refund request not found")
    
    bank_details = {
        "bank_name": bank_name,
        "account_number": account_number,
        "ifsc_code": ifsc_code,
        "account_holder_name": account_holder_name,
        "branch": branch
    }
    
    await db.refund_requests.update_one({"id": request_id}, {"$set": {"bank_details": bank_details}})
    
    return {"message": "Bank details updated successfully"}


@api_router.get("/finance/export/excel")
async def export_finance_excel(
    payment_type: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Export finance data to Excel (PE Level or Finance role)"""
    if not has_finance_access(current_user.get("role", 6)):
        raise HTTPException(status_code=403, detail="Only PE Desk, PE Manager, or Finance can export finance data")
    
    payments = await get_all_payments(payment_type, start_date, end_date, current_user)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Payments Report"
    
    # Headers
    headers = [
        "Payment Date", "Type", "Direction", "Party Name", "Stock Symbol", 
        "Stock Name", "Tranche #", "Amount", "Total Amount", "Total Paid",
        "Status", "Reference #", "Recorded By", "Notes"
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
        ws.cell(row=1, column=col).font = Font(bold=True)
    
    # Data
    for row_idx, payment in enumerate(payments, 2):
        ws.cell(row=row_idx, column=1, value=payment.get("payment_date", ""))
        ws.cell(row=row_idx, column=2, value=payment.get("type", "").upper())
        ws.cell(row=row_idx, column=3, value=payment.get("direction", "").upper())
        ws.cell(row=row_idx, column=4, value=payment.get("party_name", ""))
        ws.cell(row=row_idx, column=5, value=payment.get("stock_symbol", ""))
        ws.cell(row=row_idx, column=6, value=payment.get("stock_name", ""))
        ws.cell(row=row_idx, column=7, value=payment.get("tranche_number", 0))
        ws.cell(row=row_idx, column=8, value=payment.get("amount", 0))
        ws.cell(row=row_idx, column=9, value=payment.get("total_amount", 0))
        ws.cell(row=row_idx, column=10, value=payment.get("total_paid", 0))
        ws.cell(row=row_idx, column=11, value=payment.get("payment_status", "").upper())
        ws.cell(row=row_idx, column=12, value=payment.get("reference_number", ""))
        ws.cell(row=row_idx, column=13, value=payment.get("recorded_by", ""))
        ws.cell(row=row_idx, column=14, value=payment.get("notes", ""))
    
    # Summary section
    summary_row = len(payments) + 4
    ws.cell(row=summary_row, column=1, value="SUMMARY")
    ws.cell(row=summary_row, column=1).font = Font(bold=True)
    
    client_total = sum(p["amount"] for p in payments if p["type"] == "client")
    vendor_total = sum(p["amount"] for p in payments if p["type"] == "vendor")
    
    ws.cell(row=summary_row + 1, column=1, value="Total Received (Clients):")
    ws.cell(row=summary_row + 1, column=2, value=client_total)
    ws.cell(row=summary_row + 2, column=1, value="Total Sent (Vendors):")
    ws.cell(row=summary_row + 2, column=2, value=vendor_total)
    ws.cell(row=summary_row + 3, column=1, value="Net Flow:")
    ws.cell(row=summary_row + 3, column=2, value=client_total - vendor_total)
    ws.cell(row=summary_row + 3, column=1).font = Font(bold=True)
    ws.cell(row=summary_row + 3, column=2).font = Font(bold=True)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"finance_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@api_router.post("/payments/upload-proof")
async def upload_payment_proof(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """Upload payment proof document"""
    if not is_pe_level(current_user.get("role", 6)):
        raise HTTPException(status_code=403, detail="Only PE Desk or PE Manager can upload payment proofs")
    
    # Validate file type
    allowed_extensions = ['.pdf', '.jpg', '.jpeg', '.png', '.doc', '.docx']
    file_ext = os.path.splitext(file.filename)[1].lower()
    if file_ext not in allowed_extensions:
        raise HTTPException(status_code=400, detail=f"File type not allowed. Allowed: {', '.join(allowed_extensions)}")
    
    # Create upload directory
    proof_dir = os.path.join(UPLOAD_DIR, "payment_proofs")
    os.makedirs(proof_dir, exist_ok=True)
    
    # Generate unique filename
    unique_filename = f"payment_proof_{uuid.uuid4().hex}{file_ext}"
    file_path = os.path.join(proof_dir, unique_filename)
    
    # Save file
    with open(file_path, "wb") as f:
        content = await file.read()
        f.write(content)
    
    return {
        "filename": unique_filename,
        "original_name": file.filename,
        "url": f"/uploads/payment_proofs/{unique_filename}"
    }

# ============== Email Templates Routes (PE Level) ==============
@api_router.get("/email-templates")
async def get_email_templates(current_user: dict = Depends(get_current_user)):
    """Get all email templates (PE Level)"""
    if not is_pe_level(current_user.get("role", 6)):
        raise HTTPException(status_code=403, detail="Only PE Desk or PE Manager can manage email templates")
    
    templates = await db.email_templates.find({}, {"_id": 0}).to_list(100)
    
    # Add default templates if not in database
    from config import DEFAULT_EMAIL_TEMPLATES
    existing_keys = {t["key"] for t in templates}
    
    for key, default_data in DEFAULT_EMAIL_TEMPLATES.items():
        if key not in existing_keys:
            templates.append({
                "id": key,
                "key": key,
                **default_data,
                "is_active": True,
                "updated_at": None,
                "updated_by": None
            })
    
    return templates

@api_router.get("/email-templates/{template_key}")
async def get_email_template(template_key: str, current_user: dict = Depends(get_current_user)):
    """Get specific email template"""
    if not is_pe_level(current_user.get("role", 6)):
        raise HTTPException(status_code=403, detail="Only PE Desk or PE Manager can manage email templates")
    
    template = await db.email_templates.find_one({"key": template_key}, {"_id": 0})
    
    if not template:
        from config import DEFAULT_EMAIL_TEMPLATES
        default = DEFAULT_EMAIL_TEMPLATES.get(template_key)
        if default:
            return {
                "id": template_key,
                "key": template_key,
                **default,
                "is_active": True,
                "updated_at": None,
                "updated_by": None
            }
        raise HTTPException(status_code=404, detail="Template not found")
    
    return template

@api_router.put("/email-templates/{template_key}")
async def update_email_template(
    template_key: str,
    subject: Optional[str] = None,
    body: Optional[str] = None,
    is_active: Optional[bool] = None,
    current_user: dict = Depends(get_current_user)
):
    """Update email template (PE Level)"""
    if not is_pe_level(current_user.get("role", 6)):
        raise HTTPException(status_code=403, detail="Only PE Desk or PE Manager can manage email templates")
    
    # Check if template exists in database
    existing = await db.email_templates.find_one({"key": template_key})
    
    update_data = {
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "updated_by": current_user["id"]
    }
    
    if subject is not None:
        update_data["subject"] = subject
    if body is not None:
        update_data["body"] = body
    if is_active is not None:
        update_data["is_active"] = is_active
    
    if existing:
        await db.email_templates.update_one(
            {"key": template_key},
            {"$set": update_data}
        )
    else:
        # Create from default
        from config import DEFAULT_EMAIL_TEMPLATES
        default = DEFAULT_EMAIL_TEMPLATES.get(template_key)
        if not default:
            raise HTTPException(status_code=404, detail="Template not found")
        
        new_template = {
            "id": str(uuid.uuid4()),
            "key": template_key,
            **default,
            **update_data
        }
        await db.email_templates.insert_one(new_template)
    
    # Create audit log
    await create_audit_log(
        action="EMAIL_TEMPLATE_UPDATE",
        entity_type="email_template",
        entity_id=template_key,
        user_id=current_user["id"],
        user_name=current_user["name"],
        user_role=current_user.get("role", 5),
        details={"template_key": template_key, "updates": list(update_data.keys())}
    )
    
    return {"message": "Template updated successfully"}

@api_router.post("/email-templates/{template_key}/reset")
async def reset_email_template(template_key: str, current_user: dict = Depends(get_current_user)):
    """Reset email template to default (PE Level)"""
    if not is_pe_level(current_user.get("role", 6)):
        raise HTTPException(status_code=403, detail="Only PE Desk or PE Manager can manage email templates")
    
    from config import DEFAULT_EMAIL_TEMPLATES
    default = DEFAULT_EMAIL_TEMPLATES.get(template_key)
    if not default:
        raise HTTPException(status_code=404, detail="Default template not found")
    
    await db.email_templates.delete_one({"key": template_key})
    
    return {"message": "Template reset to default"}

@api_router.post("/email-templates/{template_key}/preview")
async def preview_email_template(
    template_key: str,
    variables: Dict[str, str] = {},
    current_user: dict = Depends(get_current_user)
):
    """Preview email template with sample variables"""
    if not is_pe_level(current_user.get("role", 6)):
        raise HTTPException(status_code=403, detail="Only PE Desk or PE Manager can manage email templates")
    
    template = await db.email_templates.find_one({"key": template_key}, {"_id": 0})
    
    if not template:
        from config import DEFAULT_EMAIL_TEMPLATES
        default = DEFAULT_EMAIL_TEMPLATES.get(template_key)
        if not default:
            raise HTTPException(status_code=404, detail="Template not found")
        template = default
    
    # Replace variables
    subject = template["subject"]
    body = template["body"]
    
    for key, value in variables.items():
        subject = subject.replace(f"{{{{{key}}}}}", str(value))
        body = body.replace(f"{{{{{key}}}}}", str(value))
    
    return {
        "subject": subject,
        "body": body,
        "variables": template.get("variables", [])
    }

# ============== Advanced Analytics Routes (PE Desk Only) ==============
@api_router.get("/analytics/summary")
async def get_analytics_summary(
    days: int = 30,
    current_user: dict = Depends(get_current_user)
):
    """Get comprehensive analytics summary (PE Level)"""
    if not is_pe_level(current_user.get("role", 6)):
        raise HTTPException(status_code=403, detail="Only PE Desk or PE Manager can access advanced analytics")
    
    start_date = (datetime.now(timezone.utc) - timedelta(days=days)).isoformat()
    
    # Get all closed bookings
    bookings = await db.bookings.find(
        {"status": "closed", "approval_status": "approved"},
        {"_id": 0}
    ).to_list(10000)
    
    # Calculate totals
    total_revenue = sum(b.get("selling_price", 0) * b.get("quantity", 0) for b in bookings)
    total_cost = sum(b.get("buying_price", 0) * b.get("quantity", 0) for b in bookings)
    total_profit = total_revenue - total_cost
    
    # Get clients count
    clients_count = await db.clients.count_documents({"is_vendor": False, "is_active": True})
    
    avg_booking_value = total_revenue / len(bookings) if bookings else 0
    profit_margin = (total_profit / total_revenue * 100) if total_revenue > 0 else 0
    
    return {
        "total_revenue": round(total_revenue, 2),
        "total_profit": round(total_profit, 2),
        "total_bookings": len(bookings),
        "total_clients": clients_count,
        "avg_booking_value": round(avg_booking_value, 2),
        "profit_margin": round(profit_margin, 2)
    }

@api_router.get("/analytics/stock-performance")
async def get_stock_performance(
    days: int = 30,
    limit: int = 10,
    current_user: dict = Depends(get_current_user)
):
    """Get stock performance analytics (PE Level)"""
    if not is_pe_level(current_user.get("role", 6)):
        raise HTTPException(status_code=403, detail="Only PE Desk or PE Manager can access advanced analytics")
    
    start_date = (datetime.now(timezone.utc) - timedelta(days=days)).isoformat()
    
    bookings = await db.bookings.find(
        {"status": "closed", "approval_status": "approved"},
        {"_id": 0}
    ).to_list(10000)
    
    stock_stats = {}
    for booking in bookings:
        stock_id = booking.get("stock_id")
        if not stock_id:
            continue
        
        if stock_id not in stock_stats:
            stock_stats[stock_id] = {"total_qty": 0, "total_revenue": 0, "total_cost": 0}
        
        qty = booking.get("quantity", 0)
        stock_stats[stock_id]["total_qty"] += qty
        stock_stats[stock_id]["total_revenue"] += booking.get("selling_price", 0) * qty
        stock_stats[stock_id]["total_cost"] += booking.get("buying_price", 0) * qty
    
    # Get stock details
    stock_ids = list(stock_stats.keys())
    stocks = await db.stocks.find({"id": {"$in": stock_ids}}, {"_id": 0}).to_list(1000)
    stock_map = {s["id"]: s for s in stocks}
    
    result = []
    for stock_id, stats in stock_stats.items():
        stock = stock_map.get(stock_id, {})
        profit = stats["total_revenue"] - stats["total_cost"]
        margin = (profit / stats["total_revenue"] * 100) if stats["total_revenue"] > 0 else 0
        
        result.append({
            "stock_id": stock_id,
            "stock_symbol": stock.get("symbol", "Unknown"),
            "stock_name": stock.get("name", "Unknown"),
            "sector": stock.get("sector", "Unknown"),
            "total_quantity_sold": stats["total_qty"],
            "total_revenue": round(stats["total_revenue"], 2),
            "total_cost": round(stats["total_cost"], 2),
            "profit_loss": round(profit, 2),
            "profit_margin": round(margin, 2)
        })
    
    result.sort(key=lambda x: x["profit_loss"], reverse=True)
    return result[:limit]

@api_router.get("/analytics/employee-performance")
async def get_employee_performance(
    days: int = 30,
    limit: int = 10,
    current_user: dict = Depends(get_current_user)
):
    """Get employee performance analytics (PE Level)"""
    if not is_pe_level(current_user.get("role", 6)):
        raise HTTPException(status_code=403, detail="Only PE Desk or PE Manager can access advanced analytics")
    
    bookings = await db.bookings.find(
        {"status": "closed", "approval_status": "approved"},
        {"_id": 0}
    ).to_list(10000)
    
    emp_stats = {}
    for booking in bookings:
        user_id = booking.get("created_by")
        if not user_id:
            continue
        
        if user_id not in emp_stats:
            emp_stats[user_id] = {
                "total_bookings": 0,
                "total_value": 0,
                "total_profit": 0,
                "client_ids": set()
            }
        
        qty = booking.get("quantity", 0)
        revenue = booking.get("selling_price", 0) * qty
        cost = booking.get("buying_price", 0) * qty
        
        emp_stats[user_id]["total_bookings"] += 1
        emp_stats[user_id]["total_value"] += revenue
        emp_stats[user_id]["total_profit"] += revenue - cost
        emp_stats[user_id]["client_ids"].add(booking.get("client_id"))
    
    # Get user details
    user_ids = list(emp_stats.keys())
    users = await db.users.find({"id": {"$in": user_ids}}, {"_id": 0, "hashed_password": 0}).to_list(1000)
    user_map = {u["id"]: u for u in users}
    
    result = []
    for user_id, stats in emp_stats.items():
        user = user_map.get(user_id, {})
        result.append({
            "user_id": user_id,
            "user_name": user.get("name", "Unknown"),
            "role": user.get("role", 5),
            "role_name": ROLES.get(user.get("role", 5), "Unknown"),
            "total_bookings": stats["total_bookings"],
            "total_value": round(stats["total_value"], 2),
            "total_profit": round(stats["total_profit"], 2),
            "clients_count": len(stats["client_ids"])
        })
    
    result.sort(key=lambda x: x["total_profit"], reverse=True)
    return result[:limit]

@api_router.get("/analytics/daily-trend")
async def get_daily_trend(
    days: int = 30,
    current_user: dict = Depends(get_current_user)
):
    """Get daily booking trend (PE Level)"""
    if not is_pe_level(current_user.get("role", 6)):
        raise HTTPException(status_code=403, detail="Only PE Desk or PE Manager can access advanced analytics")
    
    result = []
    
    for i in range(days, -1, -1):
        date = (datetime.now(timezone.utc) - timedelta(days=i)).strftime("%Y-%m-%d")
        date_start = f"{date}T00:00:00"
        date_end = f"{date}T23:59:59"
        
        # Get bookings for this day
        bookings = await db.bookings.find({
            "created_at": {"$gte": date_start, "$lte": date_end},
            "approval_status": "approved"
        }, {"_id": 0}).to_list(1000)
        
        # Get new clients for this day
        new_clients = await db.clients.count_documents({
            "created_at": {"$gte": date_start, "$lte": date_end},
            "is_vendor": False
        })
        
        bookings_value = sum(b.get("selling_price", 0) * b.get("quantity", 0) for b in bookings)
        profit_loss = sum(
            (b.get("selling_price", 0) - b.get("buying_price", 0)) * b.get("quantity", 0)
            for b in bookings if b.get("status") == "closed"
        )
        
        result.append({
            "date": date,
            "bookings_count": len(bookings),
            "bookings_value": round(bookings_value, 2),
            "profit_loss": round(profit_loss, 2),
            "new_clients": new_clients
        })
    
    return result

@api_router.get("/analytics/sector-distribution")
async def get_sector_distribution(current_user: dict = Depends(get_current_user)):
    """Get booking distribution by sector (PE Level)"""
    if not is_pe_level(current_user.get("role", 6)):
        raise HTTPException(status_code=403, detail="Only PE Desk or PE Manager can access advanced analytics")
    
    bookings = await db.bookings.find(
        {"status": "closed", "approval_status": "approved"},
        {"_id": 0}
    ).to_list(10000)
    
    # Get all stocks with sectors
    stock_ids = list(set(b.get("stock_id") for b in bookings if b.get("stock_id")))
    stocks = await db.stocks.find({"id": {"$in": stock_ids}}, {"_id": 0}).to_list(1000)
    stock_map = {s["id"]: s for s in stocks}
    
    sector_stats = {}
    for booking in bookings:
        stock = stock_map.get(booking.get("stock_id"), {})
        sector = stock.get("sector") or "Unknown"
        
        if sector not in sector_stats:
            sector_stats[sector] = {"bookings_count": 0, "total_value": 0, "total_profit": 0}
        
        qty = booking.get("quantity", 0)
        revenue = booking.get("selling_price", 0) * qty
        cost = booking.get("buying_price", 0) * qty
        
        sector_stats[sector]["bookings_count"] += 1
        sector_stats[sector]["total_value"] += revenue
        sector_stats[sector]["total_profit"] += revenue - cost
    
    result = [
        {
            "sector": sector,
            **stats,
            "total_value": round(stats["total_value"], 2),
            "total_profit": round(stats["total_profit"], 2)
        }
        for sector, stats in sector_stats.items()
    ]
    
    result.sort(key=lambda x: x["total_value"], reverse=True)
    return result


# ============== Email Server Configuration (Microsoft 365 Compatible) ==============

class SMTPConfigUpdate(BaseModel):
    smtp_host: str
    smtp_port: int
    smtp_username: str
    smtp_password: Optional[str] = None  # Don't require password on every update
    smtp_from_email: str
    smtp_from_name: Optional[str] = "SMIFS Private Equity System"
    use_tls: bool = True
    use_ssl: bool = False
    timeout: int = 30
    is_enabled: bool = True

class SMTPTestRequest(BaseModel):
    test_email: str

@api_router.get("/email-config")
async def get_email_config(current_user: dict = Depends(get_current_user)):
    """Get current SMTP configuration (PE Level)"""
    if not is_pe_level(current_user.get("role", 6)):
        raise HTTPException(status_code=403, detail="Only PE Desk or PE Manager can access email configuration")
    
    config = await db.email_config.find_one({"_id": "smtp_config"})
    
    if not config:
        # Return default configuration from environment
        default_config = {
            "_id": "smtp_config",
            "smtp_host": os.environ.get('EMAIL_HOST', 'smtp.office365.com'),
            "smtp_port": int(os.environ.get('EMAIL_PORT', '587')),
            "smtp_username": os.environ.get('EMAIL_USERNAME', ''),
            "smtp_password": "",  # Don't expose password
            "smtp_from_email": os.environ.get('EMAIL_FROM', os.environ.get('EMAIL_USERNAME', '')),
            "smtp_from_name": "SMIFS Private Equity System",
            "use_tls": True,
            "use_ssl": False,
            "timeout": 30,
            "is_enabled": bool(os.environ.get('EMAIL_USERNAME')),
            "last_updated": None,
            "updated_by": None,
            "connection_status": "not_configured" if not os.environ.get('EMAIL_USERNAME') else "configured"
        }
        return {k: v for k, v in default_config.items() if k != "_id"}
    
    # Don't expose password
    config_dict = {k: v for k, v in config.items() if k not in ["_id", "smtp_password"]}
    config_dict["smtp_password"] = "********" if config.get("smtp_password") else ""
    return config_dict


@api_router.put("/email-config")
async def update_email_config(
    config: SMTPConfigUpdate,
    current_user: dict = Depends(get_current_user)
):
    """Update SMTP configuration (PE Level)"""
    if not is_pe_level(current_user.get("role", 6)):
        raise HTTPException(status_code=403, detail="Only PE Desk or PE Manager can update email configuration")
    
    # Get existing config to preserve password if not provided
    existing = await db.email_config.find_one({"_id": "smtp_config"})
    
    update_data = {
        "_id": "smtp_config",
        "smtp_host": config.smtp_host,
        "smtp_port": config.smtp_port,
        "smtp_username": config.smtp_username,
        "smtp_from_email": config.smtp_from_email,
        "smtp_from_name": config.smtp_from_name or "SMIFS Private Equity System",
        "use_tls": config.use_tls,
        "use_ssl": config.use_ssl,
        "timeout": config.timeout,
        "is_enabled": config.is_enabled,
        "last_updated": datetime.now(timezone.utc).isoformat(),
        "updated_by": current_user["id"],
        "updated_by_name": current_user["name"]
    }
    
    # Only update password if provided and not placeholder
    if config.smtp_password and config.smtp_password != "********":
        update_data["smtp_password"] = config.smtp_password
    elif existing and existing.get("smtp_password"):
        update_data["smtp_password"] = existing["smtp_password"]
    else:
        update_data["smtp_password"] = ""
    
    await db.email_config.replace_one(
        {"_id": "smtp_config"},
        update_data,
        upsert=True
    )
    
    # Create audit log
    await create_audit_log(
        action="EMAIL_CONFIG_UPDATE",
        entity_type="system_config",
        entity_id="smtp_config",
        user_id=current_user["id"],
        user_name=current_user["name"],
        user_role=current_user["role"],
        entity_name="SMTP Configuration",
        details={
            "smtp_host": config.smtp_host,
            "smtp_port": config.smtp_port,
            "smtp_username": config.smtp_username,
            "use_tls": config.use_tls,
            "is_enabled": config.is_enabled
        }
    )
    
    return {"message": "Email configuration updated successfully"}


@api_router.post("/email-config/test")
async def test_email_config(
    test_request: SMTPTestRequest,
    current_user: dict = Depends(get_current_user)
):
    """Test SMTP configuration by sending a test email (PE Level)"""
    if not is_pe_level(current_user.get("role", 6)):
        raise HTTPException(status_code=403, detail="Only PE Desk or PE Manager can test email configuration")
    
    config = await db.email_config.find_one({"_id": "smtp_config"})
    
    if not config:
        raise HTTPException(status_code=400, detail="Email configuration not found. Please save configuration first.")
    
    if not config.get("smtp_password"):
        raise HTTPException(status_code=400, detail="SMTP password not configured")
    
    try:
        # Create test email
        msg = MIMEMultipart()
        from_name = config.get("smtp_from_name", "SMIFS Private Equity System")
        msg['From'] = f"{from_name} <{config['smtp_from_email']}>"
        msg['To'] = test_request.test_email
        msg['Subject'] = "Test Email - SMIFS Private Equity System"
        
        body = f"""
        <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto;">
            <h2 style="color: #064E3B;">✓ Email Configuration Test Successful</h2>
            <p>This is a test email from SMIFS Private Equity System.</p>
            
            <table style="width: 100%; border-collapse: collapse; margin: 20px 0;">
                <tr style="background-color: #f3f4f6;">
                    <td style="padding: 10px; border: 1px solid #e5e7eb;"><strong>SMTP Host</strong></td>
                    <td style="padding: 10px; border: 1px solid #e5e7eb;">{config['smtp_host']}</td>
                </tr>
                <tr>
                    <td style="padding: 10px; border: 1px solid #e5e7eb;"><strong>SMTP Port</strong></td>
                    <td style="padding: 10px; border: 1px solid #e5e7eb;">{config['smtp_port']}</td>
                </tr>
                <tr style="background-color: #f3f4f6;">
                    <td style="padding: 10px; border: 1px solid #e5e7eb;"><strong>From Email</strong></td>
                    <td style="padding: 10px; border: 1px solid #e5e7eb;">{config['smtp_from_email']}</td>
                </tr>
                <tr>
                    <td style="padding: 10px; border: 1px solid #e5e7eb;"><strong>TLS Enabled</strong></td>
                    <td style="padding: 10px; border: 1px solid #e5e7eb;">{'Yes' if config.get('use_tls') else 'No'}</td>
                </tr>
                <tr style="background-color: #f3f4f6;">
                    <td style="padding: 10px; border: 1px solid #e5e7eb;"><strong>Tested By</strong></td>
                    <td style="padding: 10px; border: 1px solid #e5e7eb;">{current_user['name']}</td>
                </tr>
                <tr>
                    <td style="padding: 10px; border: 1px solid #e5e7eb;"><strong>Test Time</strong></td>
                    <td style="padding: 10px; border: 1px solid #e5e7eb;">{datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')}</td>
                </tr>
            </table>
            
            <div style="background-color: #d1fae5; border-left: 4px solid #10b981; padding: 12px; margin: 20px 0;">
                <p style="margin: 0; color: #065f46;"><strong>Your email server is configured correctly!</strong></p>
            </div>
            
            <p>Best regards,<br><strong>SMIFS Private Equity System</strong></p>
        </div>
        """
        
        msg.attach(MIMEText(body, 'html'))
        
        # Connect and send
        if config.get('use_ssl'):
            server = smtplib.SMTP_SSL(config['smtp_host'], config['smtp_port'], timeout=config.get('timeout', 30))
        else:
            server = smtplib.SMTP(config['smtp_host'], config['smtp_port'], timeout=config.get('timeout', 30))
            if config.get('use_tls'):
                server.starttls()
        
        server.login(config['smtp_username'], config['smtp_password'])
        server.sendmail(config['smtp_from_email'], test_request.test_email, msg.as_string())
        server.quit()
        
        # Update connection status
        await db.email_config.update_one(
            {"_id": "smtp_config"},
            {"$set": {
                "connection_status": "connected",
                "last_test": datetime.now(timezone.utc).isoformat(),
                "last_test_result": "success",
                "last_test_email": test_request.test_email
            }}
        )
        
        return {
            "message": f"Test email sent successfully to {test_request.test_email}",
            "status": "success"
        }
        
    except smtplib.SMTPAuthenticationError as e:
        await db.email_config.update_one(
            {"_id": "smtp_config"},
            {"$set": {
                "connection_status": "auth_failed",
                "last_test": datetime.now(timezone.utc).isoformat(),
                "last_test_result": "auth_failed",
                "last_test_error": str(e)
            }}
        )
        raise HTTPException(
            status_code=400, 
            detail="Authentication failed. Please check your username and password. For Microsoft 365, ensure you're using an App Password if MFA is enabled."
        )
    except smtplib.SMTPConnectError as e:
        await db.email_config.update_one(
            {"_id": "smtp_config"},
            {"$set": {
                "connection_status": "connection_failed",
                "last_test": datetime.now(timezone.utc).isoformat(),
                "last_test_result": "connection_failed",
                "last_test_error": str(e)
            }}
        )
        raise HTTPException(status_code=400, detail=f"Connection failed: {str(e)}")
    except Exception as e:
        await db.email_config.update_one(
            {"_id": "smtp_config"},
            {"$set": {
                "connection_status": "error",
                "last_test": datetime.now(timezone.utc).isoformat(),
                "last_test_result": "error",
                "last_test_error": str(e)
            }}
        )
        raise HTTPException(status_code=400, detail=f"Failed to send test email: {str(e)}")


# Microsoft 365 preset configuration
@api_router.get("/email-config/presets")
async def get_email_presets(current_user: dict = Depends(get_current_user)):
    """Get SMTP preset configurations for common providers"""
    if not is_pe_level(current_user.get("role", 6)):
        raise HTTPException(status_code=403, detail="Only PE Desk or PE Manager can access email presets")
    
    return {
        "presets": [
            {
                "name": "Microsoft 365 / Office 365",
                "smtp_host": "smtp.office365.com",
                "smtp_port": 587,
                "use_tls": True,
                "use_ssl": False,
                "notes": "Use your Microsoft 365 email as username. If MFA is enabled, create an App Password in your Microsoft account security settings."
            },
            {
                "name": "Microsoft Exchange (On-Premises)",
                "smtp_host": "mail.yourdomain.com",
                "smtp_port": 587,
                "use_tls": True,
                "use_ssl": False,
                "notes": "Replace mail.yourdomain.com with your Exchange server address."
            },
            {
                "name": "Gmail / Google Workspace",
                "smtp_host": "smtp.gmail.com",
                "smtp_port": 587,
                "use_tls": True,
                "use_ssl": False,
                "notes": "Enable 'Less secure app access' or use an App Password if 2FA is enabled."
            },
            {
                "name": "Amazon SES",
                "smtp_host": "email-smtp.us-east-1.amazonaws.com",
                "smtp_port": 587,
                "use_tls": True,
                "use_ssl": False,
                "notes": "Use SMTP credentials from AWS SES console (not IAM credentials). Replace region as needed."
            },
            {
                "name": "SendGrid",
                "smtp_host": "smtp.sendgrid.net",
                "smtp_port": 587,
                "use_tls": True,
                "use_ssl": False,
                "notes": "Use 'apikey' as username and your SendGrid API key as password."
            },
            {
                "name": "Mailgun",
                "smtp_host": "smtp.mailgun.org",
                "smtp_port": 587,
                "use_tls": True,
                "use_ssl": False,
                "notes": "Use your Mailgun SMTP credentials from the domain settings."
            }
        ]
    }


# WebSocket endpoint for real-time notifications
@app.websocket("/api/ws/notifications")
async def websocket_notifications(websocket: WebSocket, token: str = Query(...)):
    """WebSocket endpoint for real-time notifications"""
    try:
        # Verify token
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user_id = payload["user_id"]
        
        await ws_manager.connect(websocket, user_id)
        
        try:
            while True:
                data = await websocket.receive_text()
                if data == "ping":
                    await websocket.send_text("pong")
        except WebSocketDisconnect:
            ws_manager.disconnect(websocket, user_id)
    except jwt.ExpiredSignatureError:
        await websocket.close(code=4001, reason="Token expired")
    except jwt.InvalidTokenError:
        await websocket.close(code=4002, reason="Invalid token")
    except Exception as e:
        logging.error(f"WebSocket error: {e}")
        await websocket.close(code=1011)

# Include modular routers FIRST (with /api prefix) - these take precedence
app.include_router(bookings_router, prefix="/api")
app.include_router(clients_router, prefix="/api")
app.include_router(finance_router, prefix="/api")
app.include_router(email_templates_router, prefix="/api")
app.include_router(smtp_config_router, prefix="/api")
app.include_router(stocks_router, prefix="/api")
app.include_router(database_backup_router, prefix="/api")
app.include_router(users_router, prefix="/api")

# Include the legacy api_router (endpoints here will be overridden by modular routers)
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
