"""
Clients Router

Handles all client and vendor management operations.
"""
from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Form, Query
from fastapi.responses import FileResponse
from typing import List, Optional
from datetime import datetime, timezone
import uuid
import os
import aiofiles
from pathlib import Path

from database import db
from config import ROLES, UPLOAD_DIR
from models import ClientCreate, Client, BankAccount, ClientSuspensionRequest
from utils.auth import get_current_user
from services.permission_service import (
    check_permission,
    check_permission as check_dynamic_permission,
    has_permission,
    require_permission,
    is_pe_level_dynamic
)
from services.notification_service import notify_roles, create_notification
from services.audit_service import create_audit_log
from services.email_service import send_templated_email, send_email, get_email_template
from services.ocr_service import process_document_ocr
from services.file_storage import upload_file_to_gridfs, get_file_url
from utils.demo_isolation import is_demo_user, add_demo_filter, mark_as_demo, require_demo_access

router = APIRouter(tags=["Clients"])


# Helper functions for backward compatibility
def is_pe_level(role: int) -> bool:
    """Check if role is PE level (PE Desk or PE Manager)."""
    return role in [1, 2]


def is_pe_desk_only(role: int) -> bool:
    """Check if role is PE Desk only."""
    return role == 1


async def generate_otc_ucc() -> str:
    """Generate unique OTC UCC code"""
    counter = await db.counters.find_one_and_update(
        {"_id": "otc_ucc"},
        {"$inc": {"seq": 1}},
        upsert=True,
        return_document=True
    )
    seq = counter.get("seq", 1)
    return f"OTC{str(seq).zfill(6)}"


@router.post("/clients", response_model=Client)
async def create_client(
    client_data: ClientCreate,
    current_user: dict = Depends(get_current_user),
    _: None = Depends(require_permission("clients.create", "create clients"))
):
    """Create a new client or vendor."""
    user_role = current_user.get("role", 5)
    
    # Employees and Finance cannot create vendors
    if user_role in [4, 7] and client_data.is_vendor:
        raise HTTPException(status_code=403, detail="You do not have permission to create vendors")
    
    # STRICT RULE: Clients MUST be created with documents using /clients-with-documents endpoint
    # This endpoint is only allowed for VENDORS (who don't require documents)
    if not client_data.is_vendor:
        raise HTTPException(
            status_code=400,
            detail="Clients must be created with mandatory documents (PAN Card, CML Copy). Please use the document upload form to create clients. Documents must be uploaded and stored before client creation."
        )
    
    # Check for duplicate PAN
    existing = await db.clients.find_one(
        {"pan_number": client_data.pan_number.upper()},
        {"_id": 0}
    )
    if existing:
        raise HTTPException(
            status_code=400,
            detail=f"{'Vendor' if client_data.is_vendor else 'Client'} with PAN {client_data.pan_number} already exists"
        )
    
    # STRICT RULE: Client cannot be an RP - Check by PAN
    existing_rp_pan = await db.referral_partners.find_one(
        {"pan_number": client_data.pan_number.upper()},
        {"_id": 0, "name": 1, "pan_number": 1, "rp_code": 1}
    )
    if existing_rp_pan:
        raise HTTPException(
            status_code=400,
            detail=f"Cannot create {'Vendor' if client_data.is_vendor else 'Client'}: This PAN ({client_data.pan_number.upper()}) belongs to an existing Referral Partner ({existing_rp_pan.get('name', 'Unknown')} - {existing_rp_pan.get('rp_code', '')}). An RP cannot be a Client."
        )
    
    # STRICT RULE: Client cannot be an RP - Check by Email (if email provided)
    if client_data.email:
        existing_rp_email = await db.referral_partners.find_one(
            {"email": client_data.email.lower()},
            {"_id": 0, "name": 1, "email": 1, "rp_code": 1}
        )
        if existing_rp_email:
            raise HTTPException(
                status_code=400,
                detail=f"Cannot create {'Vendor' if client_data.is_vendor else 'Client'}: This email ({client_data.email}) belongs to an existing Referral Partner ({existing_rp_email.get('name', 'Unknown')} - {existing_rp_email.get('rp_code', '')}). An RP cannot be a Client."
            )
    
    # STRICT RULE: System users (those with accounts) cannot be created as clients - Check by PAN
    existing_user_pan = await db.users.find_one(
        {"pan_number": client_data.pan_number.upper()},
        {"_id": 0, "name": 1, "email": 1, "pan_number": 1}
    )
    if existing_user_pan:
        raise HTTPException(
            status_code=400,
            detail=f"Cannot create {'Vendor' if client_data.is_vendor else 'Client'}: This PAN ({client_data.pan_number.upper()}) belongs to an existing system user ({existing_user_pan.get('name', 'Unknown')} - {existing_user_pan.get('email', '')}). System users cannot be mapped as Clients."
        )
    
    # STRICT RULE: System users cannot be created as clients - Check by Email
    if client_data.email:
        existing_user_email = await db.users.find_one(
            {"email": client_data.email.lower()},
            {"_id": 0, "name": 1, "email": 1}
        )
        if existing_user_email:
            raise HTTPException(
                status_code=400,
                detail=f"Cannot create {'Vendor' if client_data.is_vendor else 'Client'}: This email ({client_data.email}) belongs to an existing system user ({existing_user_email.get('name', 'Unknown')}). System users cannot be mapped as Clients."
            )
    
    client_id = str(uuid.uuid4())
    
    # PE Desk/Manager created clients are auto-approved
    approval_status = "approved" if is_pe_level(user_role) else "pending"
    
    # Generate OTC UCC
    otc_ucc = await generate_otc_ucc()
    
    # Auto-map client to creator (all roles)
    client_doc = {
        "id": client_id,
        "otc_ucc": otc_ucc,
        "name": client_data.name,
        "email": client_data.email,
        "email_secondary": client_data.email_secondary,
        "email_tertiary": client_data.email_tertiary,
        "phone": client_data.phone,
        "mobile": client_data.mobile,
        "pan_number": client_data.pan_number.upper(),
        "dp_id": client_data.dp_id,
        "dp_type": client_data.dp_type,
        "trading_ucc": client_data.trading_ucc,
        "address": client_data.address,
        "pin_code": client_data.pin_code,
        "is_vendor": client_data.is_vendor,
        "is_proprietor": client_data.is_proprietor,
        "has_name_mismatch": client_data.has_name_mismatch,
        "bank_accounts": [acc.model_dump() for acc in client_data.bank_accounts] if client_data.bank_accounts else [],
        "documents": [],
        "approval_status": approval_status,
        "approved_by": current_user["id"] if is_pe_level(user_role) else None,
        "approved_at": datetime.now(timezone.utc).isoformat() if is_pe_level(user_role) else None,
        "is_active": True,
        "is_suspended": False,
        "suspension_reason": None,
        "suspended_at": None,
        "suspended_by": None,
        # Auto-map client to creator for ALL roles
        "mapped_employee_id": current_user["id"],
        "mapped_employee_name": current_user["name"],
        "mapped_employee_email": current_user.get("email"),
        "is_cloned": False,
        "cloned_from_id": None,
        "cloned_from_type": None,
        "created_by": current_user["id"],
        "created_by_role": user_role,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    # Mark as demo data if created by demo user
    client_doc = mark_as_demo(client_doc, current_user)
    
    await db.clients.insert_one(client_doc)
    
    # Create audit log
    await create_audit_log(
        action="CLIENT_CREATE" if not client_data.is_vendor else "VENDOR_CREATE",
        entity_type="client" if not client_data.is_vendor else "vendor",
        entity_id=client_id,
        user_id=current_user["id"],
        user_name=current_user["name"],
        user_role=user_role,
        entity_name=client_data.name,
        details={
            "pan_number": client_data.pan_number, 
            "is_vendor": client_data.is_vendor,
            "is_proprietor": client_data.is_proprietor,
            "has_name_mismatch": client_data.has_name_mismatch
        }
    )
    
    # Notify PE Desk if client needs approval
    if approval_status == "pending":
        await notify_roles(
            [1, 2],
            "client_pending",
            f"New {'Vendor' if client_data.is_vendor else 'Client'} Pending Approval",
            f"'{client_data.name}' ({client_data.pan_number}) requires approval",
            {"client_id": client_id, "client_name": client_data.name}
        )
    
    # Send welcome email if auto-approved (PE Level created)
    if approval_status == "approved" and client_data.email:
        entity_type = "vendor" if client_data.is_vendor else "client"
        subject = f"Welcome - Your {entity_type.title()} Account is Active"
        body = f"""Dear {client_data.name},

Welcome to SMIFS Private Equity!

Your {entity_type} account has been successfully created and is now active.

Your Account Details:
- Name: {client_data.name}
- OTC UCC: {otc_ucc}
- PAN: {client_data.pan_number}
- DP ID: {client_data.dp_id}

{"You can now receive purchase orders and stock transfer requests from us." if client_data.is_vendor else "You can now place orders and manage your portfolio with us."}

If you have any questions, please contact our PE Desk.

Best Regards,
SMIFS Private Equity Team"""
        
        await send_email(
            to_email=client_data.email,
            subject=subject,
            body=body,
            template_key="welcome",
            variables={"client_name": client_data.name, "otc_ucc": otc_ucc},
            related_entity_type=entity_type,
            related_entity_id=client_id
        )
    
    return Client(**{k: v for k, v in client_doc.items() if k != "_id"})


@router.get("/clients", response_model=List[Client])
async def get_clients(
    search: Optional[str] = None,
    is_vendor: Optional[bool] = None,
    pending_approval: Optional[bool] = None,
    include_unmapped: Optional[bool] = False,
    current_user: dict = Depends(get_current_user),
    _: None = Depends(require_permission("clients.view", "view clients"))
):
    """Get all clients with optional filters based on hierarchy."""
    from services.hierarchy_service import get_team_user_ids
    
    query = {}
    user_role = current_user.get("role", 7)
    user_id = current_user.get("id")
    
    # Role-based access:
    # 1 (PE Desk), 2 (PE Manager) - See all clients/vendors
    # 4 (Viewer) - See all clients/vendors (read-only)
    # Others - See clients based on hierarchy (self + subordinates)
    
    if is_pe_level(user_role):
        # PE Level sees everything
        if is_vendor is not None:
            query["is_vendor"] = is_vendor
    elif user_role == 4:
        # Viewer sees all (read-only)
        if is_vendor is not None:
            query["is_vendor"] = is_vendor
    else:
        # All other roles - use hierarchy to show team's clients
        if is_vendor == True:
            raise HTTPException(status_code=403, detail="You do not have access to vendors")
        query["is_vendor"] = False
        
        # Get team user IDs based on hierarchy (self + all subordinates)
        team_ids = await get_team_user_ids(user_id, include_self=True)
        
        # Show clients that are:
        # 1. Mapped to anyone in the team, OR
        # 2. Created by anyone in the team AND approved
        query["$or"] = [
            {"mapped_employee_id": {"$in": team_ids}},
            {"$and": [
                {"created_by": {"$in": team_ids}}, 
                {"approval_status": "approved"}
            ]}
        ]
    
    # Pending approval filter (for PE Level only)
    if pending_approval and is_pe_level(user_role):
        query["approval_status"] = "pending"
    
    # Unmapped clients - only PE Level can see
    if include_unmapped and is_pe_level(user_role):
        # Remove the team filter and add unmapped filter
        if "$or" in query:
            existing_or = query.pop("$or")
            query["$or"] = existing_or + [
                {"mapped_employee_id": None},
                {"mapped_employee_id": {"$exists": False}},
                {"mapped_employee_id": ""}
            ]
    
    # Search filter
    if search:
        search_regex = {"$regex": search, "$options": "i"}
        search_conditions = [
            {"name": search_regex},
            {"pan_number": search_regex},
            {"email": search_regex},
            {"phone": search_regex}
        ]
        if "$or" in query:
            # Combine with existing $or using $and
            query = {"$and": [{"$or": query["$or"]}, {"$or": search_conditions}]}
        else:
            query["$or"] = search_conditions
    
    # CRITICAL: Add demo data isolation filter
    # Demo users only see demo data, live users don't see demo data
    query = add_demo_filter(query, current_user)
    
    clients = await db.clients.find(query, {"_id": 0}).sort("created_at", -1).to_list(10000)
    
    # Add can_book field to each client
    # PE Level can book for any client, others can only book for clients mapped to them OR created by them
    result = []
    for c in clients:
        client_data = dict(c)
        if is_pe_level(user_role):
            client_data["can_book"] = True
        else:
            # User can book if:
            # 1. Client is mapped to them
            # 2. OR they created the client (and it's approved)
            mapped_to_user = c.get("mapped_employee_id") == user_id
            created_by_user = c.get("created_by") == user_id
            is_approved = c.get("approval_status") == "approved"
            client_data["can_book"] = mapped_to_user or (created_by_user and is_approved)
        result.append(Client(**client_data))
    
    return result


@router.get("/clients/pending-approval", response_model=List[Client])
async def get_pending_clients(
    current_user: dict = Depends(get_current_user),
    _: None = Depends(require_permission("client_approval.view", "view pending clients"))
):
    """Get clients pending approval (requires client_approval.view permission)."""
    query = {"approval_status": "pending"}
    # Add demo isolation filter
    query = add_demo_filter(query, current_user)
    clients = await db.clients.find(query, {"_id": 0}).to_list(1000)
    return [Client(**c) for c in clients]


@router.get("/clients/{client_id}", response_model=Client)
async def get_client(
    client_id: str,
    current_user: dict = Depends(get_current_user),
    _: None = Depends(require_permission("clients.view", "view client details"))
):
    """Get a specific client by ID."""
    client = await db.clients.find_one({"id": client_id}, {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    
    # Verify demo data access
    require_demo_access(client, current_user)
    
    return Client(**client)


@router.put("/clients/{client_id}", response_model=Client)
async def update_client(
    client_id: str,
    client_data: ClientCreate,
    current_user: dict = Depends(get_current_user),
    _: None = Depends(require_permission("clients.edit", "edit clients"))
):
    """Update an existing client."""
    existing = await db.clients.find_one({"id": client_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Client not found")
    
    update_data = {
        "name": client_data.name,
        "email": client_data.email,
        "email_secondary": client_data.email_secondary,
        "email_tertiary": client_data.email_tertiary,
        "phone": client_data.phone,
        "mobile": client_data.mobile,
        "pan_number": client_data.pan_number.upper(),
        "dp_id": client_data.dp_id,
        "dp_type": client_data.dp_type,
        "trading_ucc": client_data.trading_ucc,
        "address": client_data.address,
        "pin_code": client_data.pin_code,
        "is_vendor": client_data.is_vendor,
        "bank_accounts": [acc.model_dump() for acc in client_data.bank_accounts] if client_data.bank_accounts else existing.get("bank_accounts", []),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.clients.update_one({"id": client_id}, {"$set": update_data})
    
    updated = await db.clients.find_one({"id": client_id}, {"_id": 0})
    return Client(**updated)


@router.put("/clients/{client_id}/approve")
async def approve_client(
    client_id: str,
    approve: bool = True,
    current_user: dict = Depends(get_current_user),
    _: None = Depends(require_permission("client_approval.approve", "approve clients"))
):
    """Approve or reject a client (requires client_approval.approve permission)."""
    user_role = current_user.get("role", 6)
    
    client = await db.clients.find_one({"id": client_id}, {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    
    # Check if user has permission to skip cancelled cheque
    can_skip_cancelled_cheque = await has_permission(current_user, "clients.skip_cancelled_cheque")
    
    # MANDATORY DOCUMENT CHECK: Require all 3 documents before approval
    if approve:
        documents = client.get("documents", [])
        doc_types = [d.get("doc_type") for d in documents]
        
        missing_docs = []
        if "pan_card" not in doc_types:
            missing_docs.append("PAN Card")
        if "cml_copy" not in doc_types:
            missing_docs.append("CML Copy")
        # Only require cancelled cheque if user doesn't have skip permission
        if "cancelled_cheque" not in doc_types and not can_skip_cancelled_cheque:
            missing_docs.append("Cancelled Cheque")
        
        if missing_docs:
            raise HTTPException(
                status_code=400,
                detail=f"Cannot approve: Missing mandatory documents - {', '.join(missing_docs)}. All documents must be uploaded before approval."
            )
        
        # Verify documents are actually stored in GridFS (have file_id)
        docs_without_files = []
        for doc in documents:
            file_id = doc.get("file_id")
            if not file_id or file_id == "None" or file_id == "null":
                docs_without_files.append(doc.get("doc_type", "Unknown"))
        
        if docs_without_files:
            raise HTTPException(
                status_code=400,
                detail=f"Cannot approve: Documents not properly uploaded to storage - {', '.join(docs_without_files)}. Please re-upload these documents."
            )
    
    # PE Manager cannot approve proprietor clients with name mismatch unless bank proof is uploaded
    # PE Desk (role 1) can bypass this check
    if approve and user_role == 2:  # PE Manager
        is_proprietor = client.get("is_proprietor", False)
        has_name_mismatch = client.get("has_name_mismatch", False)
        bank_proof_url = client.get("bank_proof_url")
        
        if is_proprietor and has_name_mismatch and not bank_proof_url:
            raise HTTPException(
                status_code=400, 
                detail="Cannot approve: This is a proprietorship client with name mismatch. Bank Proof must be uploaded before approval."
            )
    
    update_data = {
        "approval_status": "approved" if approve else "rejected",
        "approved_by": current_user["id"],
        "approved_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.clients.update_one({"id": client_id}, {"$set": update_data})
    
    # Create audit log
    await create_audit_log(
        action="CLIENT_APPROVE" if approve else "CLIENT_REJECT",
        entity_type="client",
        entity_id=client_id,
        user_id=current_user["id"],
        user_name=current_user["name"],
        user_role=current_user.get("role", 6),
        entity_name=client["name"],
        details={"approved": approve}
    )
    
    # Notify creator
    if client.get("created_by"):
        await create_notification(
            client["created_by"],
            "client_approved" if approve else "client_rejected",
            f"Client {'Approved' if approve else 'Rejected'}",
            f"Client '{client['name']}' has been {'approved' if approve else 'rejected'}",
            {"client_id": client_id}
        )
    
    # Send email to client
    client_email = client.get("email")
    if client_email:
        template_key = "client_approved" if approve else "client_rejected"
        template = await get_email_template(template_key)
        
        if template:
            subject = template["subject"].replace("{{client_name}}", client["name"])
            body = template["body"].replace("{{client_name}}", client["name"])
            body = body.replace("{{otc_ucc}}", client.get("otc_ucc", "N/A"))
            
            await send_email(
                to_email=client_email,
                subject=subject,
                body=body,
                template_key=template_key,
                variables={"client_name": client["name"], "otc_ucc": client.get("otc_ucc")},
                related_entity_type="client",
                related_entity_id=client_id
            )
    
    return {"message": f"Client {'approved' if approve else 'rejected'} successfully"}


@router.delete("/clients/{client_id}")
async def delete_client(
    client_id: str,
    current_user: dict = Depends(get_current_user),
    _: None = Depends(require_permission("clients.delete", "delete clients"))
):
    """Delete a client (requires clients.delete permission)."""
    
    client = await db.clients.find_one({"id": client_id}, {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    
    # Check for existing bookings
    booking_count = await db.bookings.count_documents({"client_id": client_id})
    if booking_count > 0:
        raise HTTPException(
            status_code=400,
            detail=f"Cannot delete client with existing bookings ({booking_count} bookings)"
        )
    
    await db.clients.delete_one({"id": client_id})
    
    # Create audit log
    await create_audit_log(
        action="CLIENT_DELETE",
        entity_type="client",
        entity_id=client_id,
        user_id=current_user["id"],
        user_name=current_user["name"],
        user_role=current_user.get("role", 6),
        entity_name=client["name"]
    )
    
    return {"message": f"Client '{client['name']}' deleted successfully"}


@router.put("/clients/{client_id}/suspend")
async def suspend_client(
    client_id: str,
    suspension: ClientSuspensionRequest,
    current_user: dict = Depends(get_current_user),
    _: None = Depends(require_permission("clients.suspend", "suspend clients"))
):
    """Suspend a client (requires clients.suspend permission)."""
    
    client = await db.clients.find_one({"id": client_id}, {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    
    if client.get("is_suspended"):
        raise HTTPException(status_code=400, detail="Client is already suspended")
    
    update_data = {
        "is_suspended": True,
        "suspension_reason": suspension.reason,
        "suspended_at": datetime.now(timezone.utc).isoformat(),
        "suspended_by": current_user["id"]
    }
    
    await db.clients.update_one({"id": client_id}, {"$set": update_data})
    
    # Create audit log
    await create_audit_log(
        action="CLIENT_SUSPEND",
        entity_type="client",
        entity_id=client_id,
        user_id=current_user["id"],
        user_name=current_user["name"],
        user_role=current_user.get("role", 6),
        entity_name=client["name"],
        details={"reason": suspension.reason}
    )
    
    return {"message": f"Client '{client['name']}' suspended successfully"}


@router.put("/clients/{client_id}/unsuspend")
async def unsuspend_client(
    client_id: str,
    current_user: dict = Depends(get_current_user),
    _: None = Depends(require_permission("clients.suspend", "unsuspend clients"))
):
    """Unsuspend a client (requires clients.suspend permission)."""
    
    client = await db.clients.find_one({"id": client_id}, {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    
    if not client.get("is_suspended"):
        raise HTTPException(status_code=400, detail="Client is not suspended")
    
    update_data = {
        "is_suspended": False,
        "suspension_reason": None,
        "suspended_at": None,
        "suspended_by": None,
        "unsuspended_at": datetime.now(timezone.utc).isoformat(),
        "unsuspended_by": current_user["id"]
    }
    
    await db.clients.update_one({"id": client_id}, {"$set": update_data})
    
    # Create audit log
    await create_audit_log(
        action="CLIENT_UNSUSPEND",
        entity_type="client",
        entity_id=client_id,
        user_id=current_user["id"],
        user_name=current_user["name"],
        user_role=current_user.get("role", 6),
        entity_name=client["name"]
    )
    
    return {"message": f"Client '{client['name']}' unsuspended successfully"}


@router.post("/clients/{client_id}/bank-account")
async def add_bank_account(
    client_id: str,
    bank_account: BankAccount,
    current_user: dict = Depends(get_current_user),
    _: None = Depends(require_permission("clients.edit", "add bank account"))
):
    """Add a bank account to a client."""
    client = await db.clients.find_one({"id": client_id}, {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    
    bank_accounts = client.get("bank_accounts", [])
    bank_accounts.append(bank_account.model_dump())
    
    await db.clients.update_one({"id": client_id}, {"$set": {"bank_accounts": bank_accounts}})
    
    return {"message": "Bank account added successfully"}


@router.put("/clients/{client_id}/employee-mapping")
async def update_client_employee_mapping(
    client_id: str,
    employee_id: str,
    current_user: dict = Depends(get_current_user),
    _: None = Depends(require_permission("clients.map", "update employee mapping"))
):
    """Map a client to an employee (requires clients.map permission)."""
    
    client = await db.clients.find_one({"id": client_id}, {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    
    # Verify employee exists
    employee = await db.users.find_one({"id": employee_id}, {"_id": 0})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    await db.clients.update_one(
        {"id": client_id},
        {"$set": {
            "mapped_employee_id": employee_id,
            "mapped_employee_name": employee.get("name"),
            "mapped_employee_email": employee.get("email")
        }}
    )
    
    return {"message": f"Client mapped to {employee['name']} successfully"}


# Create client with documents atomically
@router.post("/clients-with-documents")
async def create_client_with_documents(
    name: str = Form(...),
    email: str = Form(None),
    email_secondary: str = Form(None),
    email_tertiary: str = Form(None),
    phone: str = Form(None),
    mobile: str = Form(None),
    pan_number: str = Form(...),
    dp_id: str = Form(None),
    dp_type: str = Form("outside"),
    trading_ucc: str = Form(None),
    address: str = Form(None),
    pin_code: str = Form(None),
    is_vendor: bool = Form(False),
    is_proprietor: bool = Form(False),
    has_name_mismatch: bool = Form(False),
    bank_accounts: str = Form("[]"),  # JSON string
    pan_card: UploadFile = File(...),
    cml_copy: UploadFile = File(...),
    cancelled_cheque: UploadFile = File(None),
    current_user: dict = Depends(get_current_user),
    _: None = Depends(require_permission("clients.create", "create client"))
):
    """
    Create a client with mandatory documents in a single atomic operation.
    Documents are uploaded to GridFS FIRST, then the client is created.
    If any step fails, no client is created.
    """
    import json
    
    user_role = current_user.get("role", 5)
    
    # Check for duplicate PAN
    existing = await db.clients.find_one(
        {"pan_number": pan_number.upper(), "is_active": True},
        {"_id": 0, "name": 1, "otc_ucc": 1}
    )
    if existing:
        raise HTTPException(
            status_code=400,
            detail=f"A client with PAN {pan_number.upper()} already exists: {existing.get('name')} ({existing.get('otc_ucc')})"
        )
    
    # Check if PAN belongs to a system user
    existing_user_pan = await db.users.find_one(
        {"pan_number": pan_number.upper()},
        {"_id": 0, "name": 1, "email": 1}
    )
    if existing_user_pan:
        raise HTTPException(
            status_code=400,
            detail=f"Cannot create client: This PAN belongs to system user {existing_user_pan.get('name')}"
        )
    
    # Generate IDs
    client_id = str(uuid.uuid4())
    otc_ucc = await generate_otc_ucc()
    
    # STEP 1: Upload documents to GridFS FIRST
    documents = []
    upload_errors = []
    
    async def upload_doc_to_gridfs(file: UploadFile, doc_type: str) -> dict:
        """Upload a single document to GridFS and return the document record."""
        try:
            content = await file.read()
            file_ext = Path(file.filename).suffix
            filename = f"{doc_type}_{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}{file_ext}"
            
            # Upload to GridFS
            file_id = await upload_file_to_gridfs(
                content,
                filename,
                file.content_type or "application/octet-stream",
                {
                    "category": "client_documents",
                    "entity_id": client_id,
                    "doc_type": doc_type,
                    "uploaded_by": current_user.get("id"),
                    "uploaded_by_name": current_user.get("name")
                }
            )
            
            if not file_id:
                raise Exception(f"Failed to store {doc_type} in GridFS")
            
            # Save locally for OCR processing
            client_dir = UPLOAD_DIR / client_id
            client_dir.mkdir(exist_ok=True)
            file_path = client_dir / filename
            
            async with aiofiles.open(file_path, 'wb') as f:
                await f.write(content)
            
            # Process OCR
            ocr_data = await process_document_ocr(str(file_path), doc_type)
            
            return {
                "doc_type": doc_type,
                "filename": filename,
                "original_filename": file.filename,
                "file_id": file_id,
                "file_url": get_file_url(file_id),
                "file_path": str(file_path),
                "upload_date": datetime.now(timezone.utc).isoformat(),
                "ocr_data": ocr_data,
                "stored_in_gridfs": True
            }
        except Exception as e:
            upload_errors.append(f"{doc_type}: {str(e)}")
            return None
    
    # Upload mandatory documents
    pan_doc = await upload_doc_to_gridfs(pan_card, "pan_card")
    if pan_doc:
        documents.append(pan_doc)
    
    cml_doc = await upload_doc_to_gridfs(cml_copy, "cml_copy")
    if cml_doc:
        documents.append(cml_doc)
    
    # Upload optional cancelled cheque
    if cancelled_cheque and cancelled_cheque.filename:
        cheque_doc = await upload_doc_to_gridfs(cancelled_cheque, "cancelled_cheque")
        if cheque_doc:
            documents.append(cheque_doc)
    
    # STEP 2: Verify all mandatory documents were uploaded
    doc_types_uploaded = [d["doc_type"] for d in documents]
    if "pan_card" not in doc_types_uploaded or "cml_copy" not in doc_types_uploaded:
        # Clean up any uploaded files from GridFS
        for doc in documents:
            try:
                from services.file_storage import delete_file_from_gridfs
                await delete_file_from_gridfs(doc["file_id"])
            except:
                pass
        
        error_msg = "STRICT REQUIREMENT: Both PAN Card and CML Copy documents must be uploaded. "
        if upload_errors:
            error_msg += f"Upload errors: {', '.join(upload_errors)}"
        raise HTTPException(status_code=400, detail=error_msg)
    
    # STEP 2.5: STRICT VERIFICATION - Confirm documents exist in GridFS before proceeding
    from bson import ObjectId
    verified_documents = []
    for doc in documents:
        try:
            # Verify the file actually exists in GridFS
            gridfs_file = await db["fs.files"].find_one({"_id": ObjectId(doc["file_id"])})
            if not gridfs_file:
                raise Exception(f"Document {doc['doc_type']} not found in GridFS after upload")
            
            # Mark as verified
            doc["gridfs_verified"] = True
            doc["gridfs_upload_date"] = gridfs_file.get("uploadDate").isoformat() if gridfs_file.get("uploadDate") else None
            verified_documents.append(doc)
        except Exception as e:
            # If verification fails, clean up and abort
            for d in documents:
                try:
                    from services.file_storage import delete_file_from_gridfs
                    await delete_file_from_gridfs(d["file_id"])
                except:
                    pass
            raise HTTPException(
                status_code=500, 
                detail=f"STRICT VERIFICATION FAILED: Could not verify document {doc.get('doc_type', 'unknown')} in GridFS. Error: {str(e)}. Client NOT created."
            )
    
    # Replace documents with verified ones
    documents = verified_documents
    
    # STEP 3: Create the client with document references
    approval_status = "approved" if is_pe_level(user_role) else "pending"
    
    # Parse bank accounts
    try:
        bank_accounts_list = json.loads(bank_accounts) if bank_accounts else []
    except:
        bank_accounts_list = []
    
    client_doc = {
        "id": client_id,
        "otc_ucc": otc_ucc,
        "name": name,
        "email": email,
        "email_secondary": email_secondary,
        "email_tertiary": email_tertiary,
        "phone": phone,
        "mobile": mobile,
        "pan_number": pan_number.upper(),
        "dp_id": dp_id,
        "dp_type": dp_type,
        "trading_ucc": trading_ucc,
        "address": address,
        "pin_code": pin_code,
        "is_vendor": is_vendor,
        "is_proprietor": is_proprietor,
        "has_name_mismatch": has_name_mismatch,
        "bank_accounts": bank_accounts_list,
        "documents": documents,  # Documents are already stored in GridFS
        "approval_status": approval_status,
        "approved_by": current_user["id"] if is_pe_level(user_role) else None,
        "approved_at": datetime.now(timezone.utc).isoformat() if is_pe_level(user_role) else None,
        "is_active": True,
        "is_suspended": False,
        "suspension_reason": None,
        "suspended_at": None,
        "suspended_by": None,
        "mapped_employee_id": current_user["id"],
        "mapped_employee_name": current_user["name"],
        "mapped_employee_email": current_user.get("email"),
        "is_cloned": False,
        "cloned_from_id": None,
        "cloned_from_type": None,
        "created_by": current_user["id"],
        "created_by_name": current_user["name"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.clients.insert_one(client_doc)
    
    # Create audit log
    await create_audit_log(
        action="create_client_with_docs",
        entity_type="client",
        entity_id=client_id,
        user_id=current_user["id"],
        user_name=current_user["name"],
        user_role=current_user.get("role", 7),
        entity_name=name,
        details={
            "pan_number": pan_number.upper(),
            "documents_count": len(documents),
            "documents_stored_in_gridfs": True
        }
    )
    
    # Notify PE Desk if client needs approval
    if approval_status == "pending":
        await notify_roles(
            [1, 2],  # PE Desk and PE Manager
            f"New client '{name}' pending approval",
            "client_pending",
            {"client_id": client_id, "client_name": name}
        )
    
    return {
        "id": client_id,
        "otc_ucc": otc_ucc,
        "message": f"Client created with {len(documents)} documents stored in GridFS",
        "documents_count": len(documents),
        "approval_status": approval_status
    }


# Document Management Endpoints
@router.post("/clients/{client_id}/documents")
async def upload_client_document(
    client_id: str,
    doc_type: str = Form(...),
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user),
    _: None = Depends(require_permission("clients.upload_docs", "upload client documents"))
):
    """Upload a document for a client with OCR processing. Stored in GridFS for persistence."""
    client = await db.clients.find_one({"id": client_id}, {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    
    # Permission check
    await check_permission(current_user, "clients.upload_docs", "upload client documents")
    
    # Read file content
    content = await file.read()
    
    # Generate filename
    file_ext = Path(file.filename).suffix
    filename = f"{doc_type}_{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}{file_ext}"
    
    # Upload to GridFS for persistent storage
    file_id = await upload_file_to_gridfs(
        content,
        filename,
        file.content_type or "application/octet-stream",
        {
            "category": "client_documents",
            "entity_id": client_id,
            "doc_type": doc_type,
            "uploaded_by": current_user.get("id"),
            "uploaded_by_name": current_user.get("name")
        }
    )
    
    # Also save locally for OCR processing (temporary)
    client_dir = UPLOAD_DIR / client_id
    client_dir.mkdir(exist_ok=True)
    file_path = client_dir / filename
    
    async with aiofiles.open(file_path, 'wb') as f:
        await f.write(content)
    
    # Process OCR
    ocr_data = await process_document_ocr(str(file_path), doc_type)
    
    # Update client document record with GridFS file_id
    doc_record = {
        "doc_type": doc_type,
        "filename": filename,
        "file_id": file_id,  # GridFS file ID for persistent access
        "file_url": get_file_url(file_id),  # URL to access file
        "file_path": str(file_path),  # Local path (may not persist)
        "upload_date": datetime.now(timezone.utc).isoformat(),
        "ocr_data": ocr_data
    }
    
    await db.clients.update_one(
        {"id": client_id},
        {"$push": {"documents": doc_record}}
    )
    
    return {"message": "Document uploaded successfully", "document": doc_record}


@router.post("/clients/{client_id}/bank-proof")
async def upload_bank_proof(
    client_id: str,
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user),
    _: None = Depends(require_permission("clients.upload_docs", "upload bank proof"))
):
    """Upload bank proof document for proprietor clients with name mismatch (PE Level only). Stored in GridFS."""
    user_role = current_user.get("role", 6)
    client = await db.clients.find_one({"id": client_id}, {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    
    # Validate this is a proprietor client with name mismatch
    if not client.get("is_proprietor") or not client.get("has_name_mismatch"):
        raise HTTPException(status_code=400, detail="Bank proof upload is only required for proprietor clients with name mismatch")
    
    # Read file content
    content = await file.read()
    
    # Generate filename
    file_ext = Path(file.filename).suffix
    filename = f"bank_proof_{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}{file_ext}"
    
    # Upload to GridFS for persistent storage
    file_id = await upload_file_to_gridfs(
        content,
        filename,
        file.content_type or "application/octet-stream",
        {
            "category": "client_bank_proof",
            "entity_id": client_id,
            "doc_type": "bank_proof",
            "uploaded_by": current_user.get("id"),
            "uploaded_by_name": current_user.get("name")
        }
    )
    
    # Also save locally (for backward compatibility)
    client_dir = UPLOAD_DIR / client_id
    client_dir.mkdir(exist_ok=True)
    file_path = client_dir / filename
    
    async with aiofiles.open(file_path, 'wb') as f:
        await f.write(content)
    
    # Update client with bank proof URL (use GridFS URL)
    bank_proof_url = get_file_url(file_id)
    update_data = {
        "bank_proof_url": bank_proof_url,
        "bank_proof_file_id": file_id,
        "bank_proof_uploaded_by": current_user["id"],
        "bank_proof_uploaded_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.clients.update_one({"id": client_id}, {"$set": update_data})
    
    # Also add to documents array for consistency
    doc_record = {
        "doc_type": "bank_proof",
        "filename": filename,
        "file_id": file_id,
        "file_url": bank_proof_url,
        "file_path": str(file_path),
        "upload_date": datetime.now(timezone.utc).isoformat(),
        "ocr_data": None
    }
    
    await db.clients.update_one(
        {"id": client_id},
        {"$push": {"documents": doc_record}}
    )
    
    # Create audit log
    await create_audit_log(
        action="BANK_PROOF_UPLOAD",
        entity_type="client",
        entity_id=client_id,
        user_id=current_user["id"],
        user_name=current_user["name"],
        user_role=user_role,
        entity_name=client["name"],
        details={"filename": filename, "file_id": file_id, "is_proprietor": True, "has_name_mismatch": True}
    )
    
    return {
        "message": "Bank proof uploaded successfully", 
        "bank_proof_url": bank_proof_url,
        "file_id": file_id,
        "uploaded_by": current_user["name"],
        "uploaded_at": update_data["bank_proof_uploaded_at"]
    }


@router.get("/clients/{client_id}/documents/{filename}")
async def download_client_document(
    client_id: str,
    filename: str,
    current_user: dict = Depends(get_current_user),
    _: None = Depends(require_permission("clients.view_docs", "download client documents"))
):
    """Download a client document from GridFS or local storage."""
    from fastapi.responses import Response
    from services.file_storage import download_file_from_gridfs
    
    # Get client to find document
    client = await db.clients.find_one({"id": client_id}, {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    
    # Find the document by filename or original_filename
    document = None
    for doc in client.get("documents", []):
        if doc.get("filename") == filename or doc.get("original_filename") == filename:
            document = doc
            break
    
    if not document:
        raise HTTPException(status_code=404, detail="Document not found")
    
    # Try GridFS first - check both file_id (new) and gridfs_id (legacy) fields
    gridfs_id = document.get("file_id") or document.get("gridfs_id")
    if gridfs_id and gridfs_id not in ["None", "null", ""]:
        try:
            content, metadata = await download_file_from_gridfs(gridfs_id)
            content_type = metadata.get("content_type", "application/octet-stream")
            original_filename = document.get("original_filename") or document.get("filename", filename)
            
            return Response(
                content=content,
                media_type=content_type,
                headers={
                    "Content-Disposition": f'attachment; filename="{original_filename}"',
                }
            )
        except Exception as e:
            # GridFS failed, try local file
            print(f"GridFS download failed for {gridfs_id}: {e}")
            pass
    
    # Fallback to local file storage
    file_path = UPLOAD_DIR / client_id / filename
    if file_path.exists():
        return FileResponse(file_path, filename=filename)
    
    # Also try with original_filename
    original_filename = document.get("original_filename")
    if original_filename:
        file_path = UPLOAD_DIR / client_id / original_filename
        if file_path.exists():
            return FileResponse(file_path, filename=original_filename)
    
    raise HTTPException(status_code=404, detail="Document file not found in storage")


@router.get("/clients/{client_id}/documents/{filename}/ocr")
async def get_document_ocr(
    client_id: str,
    filename: str,
    current_user: dict = Depends(get_current_user),
    _: None = Depends(require_permission("clients.view_docs", "view document OCR"))
):
    """Get OCR data for a specific document."""
    client = await db.clients.find_one({"id": client_id}, {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    
    document = None
    for doc in client.get("documents", []):
        if doc["filename"] == filename:
            document = doc
            break
    
    if not document:
        raise HTTPException(status_code=404, detail="Document not found")
    
    return {
        "filename": filename,
        "doc_type": document.get("doc_type"),
        "ocr_data": document.get("ocr_data"),
        "upload_date": document.get("upload_date")
    }


@router.post("/ocr/preview")
async def ocr_preview(
    doc_type: str = Form(...),
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """Process OCR on a document without saving - for auto-fill preview."""
    import uuid as uuid_module
    
    temp_dir = UPLOAD_DIR / "temp"
    temp_dir.mkdir(exist_ok=True)
    
    file_ext = Path(file.filename).suffix
    temp_filename = f"temp_{uuid_module.uuid4()}{file_ext}"
    temp_path = temp_dir / temp_filename
    
    try:
        async with aiofiles.open(temp_path, 'wb') as f:
            content = await file.read()
            await f.write(content)
        
        ocr_data = await process_document_ocr(str(temp_path), doc_type)
        return ocr_data
    finally:
        if temp_path.exists():
            temp_path.unlink()


@router.post("/clients/{client_id}/rerun-ocr")
async def rerun_client_ocr(
    client_id: str,
    doc_types: Optional[List[str]] = Query(None, description="Specific document types to re-run OCR for. If not provided, runs on all documents."),
    update_client: bool = Query(False, description="If true, update client data with new OCR results"),
    current_user: dict = Depends(get_current_user),
    _: None = Depends(require_permission("clients.rerun_ocr", "re-run OCR on client documents"))
):
    """
    Re-run OCR on client documents.
    
    This endpoint allows admins to re-trigger OCR processing on uploaded documents,
    useful when the initial OCR extraction was incorrect or incomplete.
    
    - PE Level users can re-run OCR and optionally update client data
    - Returns comparison of old vs new OCR results
    """
    from services.ocr_service import rerun_ocr_for_client, compare_ocr_with_client_data
    from services.file_storage import get_file_from_gridfs
    
    user_role = current_user.get("role", 6)
    
    # Only PE Level can re-run OCR
    if not is_pe_level(user_role):
        raise HTTPException(
            status_code=403,
            detail="Only PE Desk or PE Manager can re-run OCR on documents"
        )
    
    client = await db.clients.find_one({"id": client_id}, {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    
    documents = client.get("documents", [])
    if not documents:
        raise HTTPException(status_code=400, detail="No documents found for this client")
    
    # Filter documents by type if specified
    if doc_types:
        target_docs = [d for d in documents if d.get("doc_type") in doc_types]
        if not target_docs:
            raise HTTPException(
                status_code=400, 
                detail=f"No documents found with types: {', '.join(doc_types)}"
            )
    else:
        target_docs = documents
    
    results = {
        "client_id": client_id,
        "client_name": client.get("name"),
        "rerun_at": datetime.now(timezone.utc).isoformat(),
        "documents_processed": [],
        "old_vs_new": {},
        "update_applied": False,
        "errors": []
    }
    
    import tempfile
    
    for doc in target_docs:
        doc_type = doc.get("doc_type")
        file_id = doc.get("file_id")
        old_ocr = doc.get("ocr_data", {})
        
        if not file_id:
            results["errors"].append(f"No file_id found for {doc_type}")
            continue
        
        try:
            # Retrieve file from GridFS
            file_result = await get_file_from_gridfs(file_id)
            if not file_result:
                results["errors"].append(f"Could not retrieve file for {doc_type} from storage")
                continue
            
            # Extract actual content bytes from the result dict
            file_content = file_result.get("content") if isinstance(file_result, dict) else file_result
            if not file_content:
                results["errors"].append(f"No content found in file for {doc_type}")
                continue
            
            # Save to temp file for OCR processing
            file_ext = doc.get("filename", "file.jpg").split(".")[-1] or "jpg"
            with tempfile.NamedTemporaryFile(delete=False, suffix=f".{file_ext}") as tmp:
                tmp.write(file_content)
                tmp_path = tmp.name
            
            try:
                # Run OCR
                new_ocr = await process_document_ocr(tmp_path, doc_type)
                
                # Store comparison
                results["old_vs_new"][doc_type] = {
                    "old_data": old_ocr.get("extracted_data", {}) if isinstance(old_ocr, dict) else {},
                    "new_data": new_ocr.get("extracted_data", {}),
                    "old_confidence": old_ocr.get("confidence", 0) if isinstance(old_ocr, dict) else 0,
                    "new_confidence": new_ocr.get("confidence", 0)
                }
                
                results["documents_processed"].append({
                    "doc_type": doc_type,
                    "old_confidence": old_ocr.get("confidence", 0) if isinstance(old_ocr, dict) else 0,
                    "new_confidence": new_ocr.get("confidence", 0),
                    "status": new_ocr.get("status", "unknown")
                })
                
                # Update document OCR data in database
                await db.clients.update_one(
                    {"id": client_id, "documents.doc_type": doc_type},
                    {
                        "$set": {
                            "documents.$.ocr_data": new_ocr,
                            "documents.$.ocr_rerun_at": datetime.now(timezone.utc).isoformat(),
                            "documents.$.ocr_rerun_by": current_user.get("id")
                        }
                    }
                )
                
            finally:
                # Clean up temp file
                if os.path.exists(tmp_path):
                    os.unlink(tmp_path)
                    
        except Exception as e:
            results["errors"].append(f"Error processing {doc_type}: {str(e)}")
    
    # Optionally update client data with new OCR results
    if update_client and results["documents_processed"]:
        updates = {}
        
        # Extract data from CML
        cml_data = results["old_vs_new"].get("cml_copy", {}).get("new_data", {})
        if cml_data:
            if cml_data.get("client_name"):
                updates["name"] = cml_data["client_name"]
            if cml_data.get("pan_number"):
                updates["pan_number"] = cml_data["pan_number"]
            if cml_data.get("email"):
                updates["email"] = cml_data["email"]
            if cml_data.get("mobile"):
                updates["mobile"] = cml_data["mobile"]
            if cml_data.get("address"):
                updates["address"] = cml_data["address"]
            if cml_data.get("pin_code"):
                updates["pin_code"] = cml_data["pin_code"]
            
            # Construct full DP ID
            dp_id = cml_data.get("dp_id", "")
            client_id_from_cml = cml_data.get("client_id", "")
            if dp_id and client_id_from_cml:
                updates["dp_id"] = f"{dp_id}{client_id_from_cml}"
            elif cml_data.get("full_dp_client_id"):
                updates["dp_id"] = cml_data["full_dp_client_id"].replace("-", "")
        
        # Extract data from PAN card
        pan_data = results["old_vs_new"].get("pan_card", {}).get("new_data", {})
        if pan_data:
            if pan_data.get("pan_number") and not updates.get("pan_number"):
                updates["pan_number"] = pan_data["pan_number"]
            if pan_data.get("name") and not updates.get("name"):
                updates["name"] = pan_data["name"]
        
        # Extract bank data from cancelled cheque
        cheque_data = results["old_vs_new"].get("cancelled_cheque", {}).get("new_data", {})
        if cheque_data:
            # Add or update bank account
            if cheque_data.get("account_number") and cheque_data.get("ifsc_code"):
                new_bank = {
                    "bank_name": cheque_data.get("bank_name", ""),
                    "account_number": cheque_data["account_number"],
                    "ifsc_code": cheque_data["ifsc_code"],
                    "branch_name": cheque_data.get("branch_name", ""),
                    "account_holder_name": cheque_data.get("account_holder_name", ""),
                    "source": "cancelled_cheque_rerun"
                }
                
                # Check if this bank account already exists
                existing_banks = client.get("bank_accounts", [])
                bank_exists = any(b.get("account_number") == new_bank["account_number"] for b in existing_banks)
                
                if not bank_exists:
                    await db.clients.update_one(
                        {"id": client_id},
                        {"$push": {"bank_accounts": new_bank}}
                    )
        
        if updates:
            updates["ocr_updated_at"] = datetime.now(timezone.utc).isoformat()
            updates["ocr_updated_by"] = current_user.get("id")
            
            await db.clients.update_one(
                {"id": client_id},
                {"$set": updates}
            )
            results["update_applied"] = True
            results["fields_updated"] = list(updates.keys())
    
    # Create audit log
    await create_audit_log(
        action="OCR_RERUN",
        entity_type="client",
        entity_id=client_id,
        user_id=current_user.get("id"),
        user_name=current_user.get("name"),
        user_role=user_role,
        entity_name=client.get("name"),
        details={
            "documents_processed": len(results["documents_processed"]),
            "update_applied": results.get("update_applied", False)
        }
    )
    
    return results


@router.get("/clients/{client_id}/document-status")
async def get_client_document_status(
    client_id: str,
    current_user: dict = Depends(get_current_user),
    _: None = Depends(require_permission("clients.view", "view document status"))
):
    """
    Get document verification status for a client.
    Returns which mandatory documents are uploaded and properly stored.
    """
    client = await db.clients.find_one({"id": client_id}, {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    
    documents = client.get("documents", [])
    
    # Check each mandatory document type
    mandatory_docs = ["pan_card", "cml_copy", "cancelled_cheque"]
    doc_status = {}
    
    for doc_type in mandatory_docs:
        doc = next((d for d in documents if d.get("doc_type") == doc_type), None)
        
        if doc:
            file_id = doc.get("file_id")
            is_stored = file_id and file_id != "None" and file_id != "null"
            
            doc_status[doc_type] = {
                "uploaded": True,
                "stored_in_gridfs": is_stored,
                "file_id": file_id if is_stored else None,
                "filename": doc.get("filename"),
                "upload_date": doc.get("upload_date"),
                "has_ocr_data": bool(doc.get("ocr_data", {}).get("extracted_data"))
            }
        else:
            doc_status[doc_type] = {
                "uploaded": False,
                "stored_in_gridfs": False,
                "file_id": None,
                "filename": None,
                "upload_date": None,
                "has_ocr_data": False
            }
    
    # Check optional bank_proof for proprietor clients
    bank_proof = next((d for d in documents if d.get("doc_type") == "bank_proof"), None)
    if bank_proof:
        file_id = bank_proof.get("file_id")
        is_stored = file_id and file_id != "None" and file_id != "null"
        doc_status["bank_proof"] = {
            "uploaded": True,
            "stored_in_gridfs": is_stored,
            "file_id": file_id if is_stored else None,
            "filename": bank_proof.get("filename"),
            "upload_date": bank_proof.get("upload_date")
        }
    
    # Calculate overall status
    all_mandatory_uploaded = all(doc_status[dt]["uploaded"] for dt in mandatory_docs)
    all_mandatory_stored = all(doc_status[dt]["stored_in_gridfs"] for dt in mandatory_docs)
    
    return {
        "client_id": client_id,
        "client_name": client.get("name"),
        "is_vendor": client.get("is_vendor", False),
        "is_proprietor": client.get("is_proprietor", False),
        "has_name_mismatch": client.get("has_name_mismatch", False),
        "approval_status": client.get("approval_status"),
        "documents": doc_status,
        "summary": {
            "all_mandatory_uploaded": all_mandatory_uploaded,
            "all_mandatory_stored": all_mandatory_stored,
            "can_be_approved": all_mandatory_stored,
            "missing_documents": [dt for dt in mandatory_docs if not doc_status[dt]["uploaded"]],
            "documents_needing_reupload": [dt for dt in mandatory_docs if doc_status[dt]["uploaded"] and not doc_status[dt]["stored_in_gridfs"]]
        }
    }


@router.post("/clients/{client_id}/reveal-documents")
async def reveal_documents_from_gridfs(
    client_id: str,
    current_user: dict = Depends(get_current_user),
    _: None = Depends(require_permission("clients.view", "reveal documents"))
):
    """
    Deep search GridFS to find and link any orphaned documents for a client.
    This searches by:
    1. entity_id in metadata (exact match)
    2. PAN number in filename/metadata
    3. Client name patterns
    
    If documents are found in GridFS but not linked to the client, this will link them.
    """
    from bson import ObjectId
    
    client = await db.clients.find_one({"id": client_id}, {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    
    client_name = client.get("name", "")
    pan_number = client.get("pan_number", "")
    current_docs = client.get("documents", [])
    current_file_ids = set(d.get("file_id") for d in current_docs if d.get("file_id"))
    
    found_documents = []
    linked_documents = []
    
    # Search 1: By entity_id (exact match)
    entity_search = await db["fs.files"].find({
        "metadata.entity_id": client_id
    }).to_list(100)
    
    for f in entity_search:
        file_id = str(f["_id"])
        if file_id not in current_file_ids:
            found_documents.append({
                "file_id": file_id,
                "filename": f.get("filename"),
                "doc_type": f.get("metadata", {}).get("doc_type", "unknown"),
                "upload_date": f.get("uploadDate").isoformat() if f.get("uploadDate") else None,
                "search_method": "entity_id",
                "metadata": f.get("metadata", {})
            })
    
    # Search 2: By PAN number in filename or metadata
    if pan_number:
        pan_search = await db["fs.files"].find({
            "$or": [
                {"filename": {"$regex": pan_number, "$options": "i"}},
                {"metadata.pan_number": pan_number.upper()}
            ]
        }).to_list(100)
        
        for f in pan_search:
            file_id = str(f["_id"])
            if file_id not in current_file_ids and file_id not in [d["file_id"] for d in found_documents]:
                found_documents.append({
                    "file_id": file_id,
                    "filename": f.get("filename"),
                    "doc_type": f.get("metadata", {}).get("doc_type", "unknown"),
                    "upload_date": f.get("uploadDate").isoformat() if f.get("uploadDate") else None,
                    "search_method": "pan_number",
                    "metadata": f.get("metadata", {})
                })
    
    # Search 3: By client_documents category with loose entity matching
    category_search = await db["fs.files"].find({
        "metadata.category": "client_documents"
    }).to_list(500)
    
    for f in category_search:
        file_id = str(f["_id"])
        entity_id = f.get("metadata", {}).get("entity_id", "")
        
        # Check if this file's entity_id might be a partial match or old format
        if entity_id and entity_id != client_id:
            # Check if entity_id contains part of our client_id or vice versa
            if client_id.startswith(entity_id[:8]) or entity_id.startswith(client_id[:8]):
                if file_id not in current_file_ids and file_id not in [d["file_id"] for d in found_documents]:
                    found_documents.append({
                        "file_id": file_id,
                        "filename": f.get("filename"),
                        "doc_type": f.get("metadata", {}).get("doc_type", "unknown"),
                        "upload_date": f.get("uploadDate").isoformat() if f.get("uploadDate") else None,
                        "search_method": "partial_entity_match",
                        "metadata": f.get("metadata", {})
                    })
    
    # If documents found, offer to link them
    if found_documents:
        # Auto-link documents found by entity_id (most reliable)
        for doc in found_documents:
            if doc["search_method"] == "entity_id":
                doc_type = doc["doc_type"]
                
                # Check if this doc_type already exists
                existing = next((d for d in current_docs if d.get("doc_type") == doc_type), None)
                
                if not existing:
                    new_doc = {
                        "doc_type": doc_type,
                        "filename": doc["filename"],
                        "file_id": doc["file_id"],
                        "file_url": f"/api/files/{doc['file_id']}",
                        "upload_date": doc["upload_date"],
                        "stored_in_gridfs": True,
                        "revealed_at": datetime.now(timezone.utc).isoformat(),
                        "revealed_by": current_user.get("name")
                    }
                    
                    await db.clients.update_one(
                        {"id": client_id},
                        {"$push": {"documents": new_doc}}
                    )
                    linked_documents.append(new_doc)
    
    # Get updated client document list
    updated_client = await db.clients.find_one({"id": client_id}, {"_id": 0, "documents": 1})
    
    # Create audit log
    await create_audit_log(
        action="REVEAL_DOCUMENTS",
        entity_type="client",
        entity_id=client_id,
        user_id=current_user.get("id"),
        user_name=current_user.get("name"),
        user_role=current_user.get("role", 7),
        entity_name=client_name,
        details={
            "found_in_gridfs": len(found_documents),
            "auto_linked": len(linked_documents)
        }
    )
    
    return {
        "client_id": client_id,
        "client_name": client_name,
        "search_completed": True,
        "found_in_gridfs": found_documents,
        "auto_linked": linked_documents,
        "current_documents": updated_client.get("documents", []),
        "summary": {
            "total_found": len(found_documents),
            "auto_linked": len(linked_documents),
            "needs_manual_review": len([d for d in found_documents if d["search_method"] != "entity_id"])
        }
    }


@router.post("/clients/{client_id}/clone")
async def clone_client_vendor(
    client_id: str, 
    target_type: str = Query(..., description="Target type: 'client' or 'vendor'"),
    current_user: dict = Depends(get_current_user),
    _: None = Depends(require_permission("clients.create", "clone clients/vendors"))
):
    """Clone a client as vendor or vendor as client (requires clients.create permission)."""
    
    if target_type not in ["client", "vendor"]:
        raise HTTPException(status_code=400, detail="target_type must be 'client' or 'vendor'")
    
    source = await db.clients.find_one({"id": client_id}, {"_id": 0})
    if not source:
        raise HTTPException(status_code=404, detail="Source client/vendor not found")
    
    is_currently_vendor = source.get("is_vendor", False)
    source_type = "vendor" if is_currently_vendor else "client"
    
    # Check if this is already the target type
    if (target_type == "vendor" and is_currently_vendor) or (target_type == "client" and not is_currently_vendor):
        raise HTTPException(status_code=400, detail=f"This is already a {target_type}")
    
    # Check if a clone already exists (by PAN and target type)
    existing = await db.clients.find_one({
        "pan_number": source["pan_number"],
        "is_vendor": target_type == "vendor"
    }, {"_id": 0})
    
    if existing:
        raise HTTPException(
            status_code=400, 
            detail=f"A {target_type} with PAN {source['pan_number']} already exists (OTC UCC: {existing.get('otc_ucc', 'N/A')}). Cannot create duplicate."
        )
    
    # Check if this entity was already cloned to prevent re-cloning
    already_cloned = await db.clients.find_one({
        "cloned_from_id": client_id,
        "is_vendor": target_type == "vendor"
    }, {"_id": 0})
    
    if already_cloned:
        raise HTTPException(
            status_code=400, 
            detail=f"This {source_type} has already been cloned as a {target_type} (OTC UCC: {already_cloned.get('otc_ucc', 'N/A')}). Cannot clone again."
        )
    
    new_id = str(uuid.uuid4())
    otc_ucc = await generate_otc_ucc()
    
    # Deep copy documents with all fields
    cloned_documents = []
    for doc in source.get("documents", []):
        cloned_doc_entry = {
            "doc_type": doc.get("doc_type"),
            "file_path": doc.get("file_path"),
            "file_url": doc.get("file_url"),
            "gridfs_id": doc.get("gridfs_id"),
            "original_filename": doc.get("original_filename"),
            "uploaded_at": doc.get("uploaded_at"),
            "uploaded_by": doc.get("uploaded_by"),
            "upload_date": doc.get("upload_date"),
            "ocr_data": doc.get("ocr_data"),
            "ocr_status": doc.get("ocr_status"),
            "ocr_processed_at": doc.get("ocr_processed_at"),
        }
        cloned_documents.append(cloned_doc_entry)
    
    # Deep copy bank accounts
    cloned_bank_accounts = []
    for acc in source.get("bank_accounts", []):
        cloned_acc = {
            "account_name": acc.get("account_name"),
            "bank_name": acc.get("bank_name"),
            "account_number": acc.get("account_number"),
            "ifsc_code": acc.get("ifsc_code"),
            "branch": acc.get("branch"),
            "is_primary": acc.get("is_primary", False),
        }
        cloned_bank_accounts.append(cloned_acc)
    
    # Create cloned document with ALL fields
    cloned_doc = {
        "id": new_id,
        "otc_ucc": otc_ucc,
        # Basic info
        "name": source["name"],
        "email": source.get("email"),
        "email_secondary": source.get("email_secondary"),
        "email_tertiary": source.get("email_tertiary"),
        "phone": source.get("phone"),
        "mobile": source.get("mobile"),
        "pan_number": source["pan_number"],
        # DP info
        "dp_id": source.get("dp_id"),
        "dp_type": source.get("dp_type", "outside"),
        "trading_ucc": source.get("trading_ucc"),
        # Address
        "address": source.get("address"),
        "pin_code": source.get("pin_code"),
        # Banking
        "bank_accounts": cloned_bank_accounts,
        # Type flags
        "is_vendor": target_type == "vendor",
        "is_active": True,
        "is_suspended": False,
        "suspension_reason": None,
        "suspended_at": None,
        "suspended_by": None,
        "suspended_by_name": None,
        "approval_status": "approved",
        "approved_by": current_user["id"],
        "approved_at": datetime.now(timezone.utc).isoformat(),
        # Documents - deep copied
        "documents": cloned_documents,
        # Proprietor flags
        "is_proprietor": source.get("is_proprietor", False),
        "has_name_mismatch": source.get("has_name_mismatch", False),
        "bank_proof_url": source.get("bank_proof_url"),
        "bank_proof_uploaded_by": source.get("bank_proof_uploaded_by"),
        "bank_proof_uploaded_at": source.get("bank_proof_uploaded_at"),
        # Creator info
        "user_id": current_user["id"],
        "created_by": current_user["id"],
        "created_by_role": current_user.get("role", 1),
        # Auto-map to creator
        "mapped_employee_id": current_user["id"],
        "mapped_employee_name": current_user["name"],
        "mapped_employee_email": current_user.get("email"),
        # Clone tracking
        "is_cloned": True,
        "cloned_from_id": client_id,
        "cloned_from_type": source_type,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.clients.insert_one(cloned_doc)
    await create_audit_log(
        action="CLIENT_CREATE",
        entity_type=target_type,
        entity_id=new_id,
        user_id=current_user["id"],
        user_name=current_user["name"],
        user_role=current_user.get("role", 1),
        entity_name=source["name"],
        details={
            "cloned_from": client_id,
            "source_type": source_type,
            "target_type": target_type,
            "pan_number": source["pan_number"]
        }
    )
    
    # Send email notification for conversion
    entity_email = source.get("email")
    if entity_email:
        if target_type == "vendor":
            # Client converted to Vendor
            subject = f"Welcome as Vendor - {source['name']}"
            body = f"""Dear {source['name']},

We are pleased to inform you that your account has been registered as a Vendor in our system.

Your Vendor Details:
- Name: {source['name']}
- OTC UCC: {otc_ucc}
- PAN: {source['pan_number']}
- DP ID: {source['dp_id']}

You can now receive purchase orders and stock transfer requests from us.

If you have any questions, please contact our PE Desk.

Best Regards,
SMIFS Private Equity Team"""
        else:
            # Vendor converted to Client
            subject = f"Welcome as Client - {source['name']}"
            body = f"""Dear {source['name']},

We are pleased to inform you that your account has been registered as a Client in our system.

Your Client Details:
- Name: {source['name']}
- OTC UCC: {otc_ucc}
- PAN: {source['pan_number']}
- DP ID: {source['dp_id']}

You can now place orders and manage your portfolio with us.

If you have any questions, please contact our PE Desk.

Best Regards,
SMIFS Private Equity Team"""
        
        await send_email(
            to_email=entity_email,
            subject=subject,
            body=body,
            template_key=f"{source_type}_to_{target_type}_conversion",
            variables={"name": source["name"], "otc_ucc": otc_ucc, "pan_number": source["pan_number"]},
            related_entity_type=target_type,
            related_entity_id=new_id
        )
    
    return {
        "message": f"Successfully cloned {source_type} '{source['name']}' as {target_type}",
        "id": new_id,
        "otc_ucc": otc_ucc
    }


@router.get("/clients/{client_id}/portfolio")
async def get_client_portfolio(
    client_id: str,
    current_user: dict = Depends(get_current_user),
    _: None = Depends(require_permission("clients.view", "view client portfolio"))
):
    """Get a client's portfolio showing all their holdings."""
    client = await db.clients.find_one({"id": client_id}, {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    
    # Get all completed bookings for this client
    bookings = await db.bookings.find({
        "client_id": client_id,
        "stock_transferred": True,
        "is_voided": {"$ne": True}
    }, {"_id": 0}).to_list(10000)
    
    # Group by stock and calculate holdings
    holdings = {}
    for booking in bookings:
        stock_id = booking.get("stock_id")
        if stock_id not in holdings:
            holdings[stock_id] = {
                "stock_id": stock_id,
                "stock_symbol": booking.get("stock_symbol", ""),
                "stock_name": booking.get("stock_name", ""),
                "quantity": 0,
                "total_value": 0,
                "avg_price": 0,
                "bookings": []
            }
        
        qty = booking.get("quantity", 0)
        price = booking.get("selling_price", 0)
        holdings[stock_id]["quantity"] += qty
        holdings[stock_id]["total_value"] += qty * price
        holdings[stock_id]["bookings"].append({
            "booking_number": booking.get("booking_number"),
            "quantity": qty,
            "price": price,
            "date": booking.get("created_at")
        })
    
    # Calculate average price
    for stock_id, holding in holdings.items():
        if holding["quantity"] > 0:
            holding["avg_price"] = holding["total_value"] / holding["quantity"]
    
    portfolio_list = list(holdings.values())
    total_portfolio_value = sum(h["total_value"] for h in portfolio_list)
    
    return {
        "client_id": client_id,
        "client_name": client.get("name"),
        "otc_ucc": client.get("otc_ucc"),
        "holdings": portfolio_list,
        "total_value": total_portfolio_value,
        "total_stocks": len(portfolio_list)
    }



# ============== EXPORT ENDPOINTS ==============

@router.get("/clients-export")
async def export_clients(
    format: str = Query("xlsx", enum=["xlsx", "csv"]),
    is_vendor: bool = Query(False),
    approval_status: Optional[str] = Query(None),
    current_user: dict = Depends(get_current_user),
    _: None = Depends(require_permission("clients.view", "export clients"))
):
    """Export clients to Excel or CSV"""
    import io
    from fastapi.responses import StreamingResponse
    
    # Build query
    query = {"is_vendor": is_vendor}
    if approval_status:
        query["approval_status"] = approval_status
    
    # Get clients
    clients_data = await db.clients.find(query, {"_id": 0}).sort("name", 1).to_list(10000)
    
    # Get mapped employees for lookup
    users = {u["id"]: u for u in await db.users.find({}, {"_id": 0, "id": 1, "name": 1, "email": 1}).to_list(10000)}
    
    if format == "csv":
        import csv
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Headers
        headers = [
            "OTC UCC", "Name", "PAN Number", "DP ID", "DP Type",
            "Email", "Phone", "Address", "PIN Code",
            "Approval Status", "Is Active", "Mapped To",
            "Bank Account", "IFSC Code", "Bank Name",
            "Created At", "Notes"
        ]
        writer.writerow(headers)
        
        # Data rows
        for client in clients_data:
            mapped_employee = users.get(client.get("mapped_employee_id"), {})
            bank_accounts = client.get("bank_accounts", [])
            primary_bank = bank_accounts[0] if bank_accounts else {}
            
            row = [
                client.get("otc_ucc", ""),
                client.get("name", ""),
                client.get("pan_number", ""),
                client.get("dp_id", ""),
                client.get("dp_type", ""),
                client.get("email", ""),
                client.get("phone", client.get("mobile", "")),
                client.get("address", ""),
                client.get("pin_code", ""),
                client.get("approval_status", ""),
                "Yes" if client.get("is_active") else "No",
                mapped_employee.get("name", ""),
                primary_bank.get("account_number", ""),
                primary_bank.get("ifsc_code", ""),
                primary_bank.get("bank_name", ""),
                client.get("created_at", ""),
                client.get("notes", "")
            ]
            writer.writerow(row)
        
        output.seek(0)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=clients_export.csv"}
        )
    
    else:  # xlsx
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Vendors" if is_vendor else "Clients"
        
        # Headers
        headers = [
            "OTC UCC", "Name", "PAN Number", "DP ID", "DP Type",
            "Email", "Phone", "Address", "PIN Code",
            "Approval Status", "Is Active", "Mapped To",
            "Bank Account", "IFSC Code", "Bank Name",
            "Created At", "Notes"
        ]
        
        # Style headers
        header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Data rows
        for row_num, client in enumerate(clients_data, 2):
            mapped_employee = users.get(client.get("mapped_employee_id"), {})
            bank_accounts = client.get("bank_accounts", [])
            primary_bank = bank_accounts[0] if bank_accounts else {}
            
            data = [
                client.get("otc_ucc", ""),
                client.get("name", ""),
                client.get("pan_number", ""),
                client.get("dp_id", ""),
                client.get("dp_type", ""),
                client.get("email", ""),
                client.get("phone", client.get("mobile", "")),
                client.get("address", ""),
                client.get("pin_code", ""),
                client.get("approval_status", ""),
                "Yes" if client.get("is_active") else "No",
                mapped_employee.get("name", ""),
                primary_bank.get("account_number", ""),
                primary_bank.get("ifsc_code", ""),
                primary_bank.get("bank_name", ""),
                client.get("created_at", ""),
                client.get("notes", "")
            ]
            
            for col, value in enumerate(data, 1):
                ws.cell(row=row_num, column=col, value=value)
        
        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=clients_export.xlsx"}
        )
