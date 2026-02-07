"""
Inventory Router
Handles inventory management endpoints
"""
from typing import Optional
from datetime import datetime, timezone
from fastapi import APIRouter, HTTPException, Depends, Query
from pydantic import BaseModel
import logging

from database import db
from models import Inventory as InventoryModel
from utils.auth import get_current_user
from services.permission_service import (
    require_permission
)
from utils.demo_isolation import add_demo_filter

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/inventory", tags=["Inventory"])


# Helper function for backward compatibility
def is_pe_level(role: int) -> bool:
    """Check if role is PE level (PE Desk or PE Manager)."""
    return role in [1, 2]


class UpdateLandingPriceRequest(BaseModel):
    landing_price: float


async def calculate_weighted_avg_for_stock(stock_id: str) -> dict:
    """
    Calculate weighted average price from ALL purchases for a stock.
    This ensures accurate pricing based on actual purchase history.
    """
    # Get ALL purchases for this stock (to calculate true weighted average)
    purchases = await db.purchases.find(
        {"stock_id": stock_id},
        {"_id": 0, "quantity": 1, "total_amount": 1}
    ).to_list(10000)
    
    if not purchases:
        # Fallback: Use stored inventory values if no purchases
        inventory = await db.inventory.find_one({"stock_id": stock_id}, {"_id": 0})
        if inventory:
            return {
                "weighted_avg_price": inventory.get("weighted_avg_price", 0),
                "total_value": inventory.get("total_value", 0)
            }
        return {"weighted_avg_price": 0, "total_value": 0}
    
    # Calculate weighted average from all purchases
    total_quantity = sum(p.get("quantity", 0) for p in purchases)
    total_purchase_value = sum(p.get("total_amount", 0) for p in purchases)
    
    weighted_avg = total_purchase_value / total_quantity if total_quantity > 0 else 0
    
    # Get current inventory quantity for total_value calculation
    inventory = await db.inventory.find_one({"stock_id": stock_id}, {"_id": 0, "available_quantity": 1})
    current_qty = inventory.get("available_quantity", 0) if inventory else 0
    
    # Total value = current inventory quantity * weighted average price
    total_value = current_qty * weighted_avg
    
    return {
        "weighted_avg_price": round(weighted_avg, 2),
        "total_value": round(total_value, 2)
    }


@router.get("")
async def get_inventory(
    search: Optional[str] = None,
    skip: int = Query(0, ge=0),
    limit: int = Query(1000, ge=1, le=10000),
    current_user: dict = Depends(get_current_user),
    _: None = Depends(require_permission("inventory.view", "view inventory"))
):
    """
    Get all inventory items with dynamically calculated weighted average pricing.
    
    Args:
        search: Search query to filter by stock symbol, name, or ISIN
        skip: Number of records to skip (for pagination)
        limit: Maximum number of records to return
    
    - PE Desk/Manager: See both WAP (Weighted Avg Price) and LP (Landing Price)
    - Other users: Only see LP (Landing Price) - WAP is hidden
    """
    user_role = current_user.get("role", 6)
    is_pe = is_pe_level(user_role)
    
    # Build query with demo isolation filter
    query = {}
    query = add_demo_filter(query, current_user)
    
    # If search is provided, first find matching stock IDs
    stock_filter = {}
    if search:
        search_regex = {"$regex": search, "$options": "i"}
        stock_filter = {"$or": [
            {"symbol": search_regex},
            {"name": search_regex},
            {"isin": search_regex}
        ]}
    
    # Get matching stock IDs
    if stock_filter:
        matching_stocks = await db.stocks.find(stock_filter, {"_id": 0, "id": 1}).to_list(10000)
        matching_stock_ids = [s["id"] for s in matching_stocks]
        if matching_stock_ids:
            query["stock_id"] = {"$in": matching_stock_ids}
        else:
            # No matching stocks, return empty
            return []
    
    inventory = await db.inventory.find(query, {"_id": 0}).skip(skip).limit(limit).to_list(limit)
    
    # Enrich with stock details
    stock_ids = list(set(item.get("stock_id") for item in inventory if item.get("stock_id")))
    stocks = await db.stocks.find({"id": {"$in": stock_ids}}, {"_id": 0}).to_list(1000)
    stock_map = {s["id"]: s for s in stocks}
    
    result = []
    for item in inventory:
        stock_id = item.get("stock_id")
        stock = stock_map.get(stock_id, {})
        
        # Skip orphaned inventory records (stock doesn't exist)
        if not stock:
            continue
        
        item["stock_symbol"] = stock.get("symbol", "Unknown")
        item["stock_name"] = stock.get("name", "Unknown")
        
        # ALWAYS calculate weighted average dynamically from purchases
        calc = await calculate_weighted_avg_for_stock(stock_id)
        
        # Get landing price (or default to WAP if not set)
        landing_price = item.get("landing_price")
        if landing_price is None or landing_price <= 0:
            landing_price = calc["weighted_avg_price"]
        
        item["landing_price"] = round(landing_price, 2) if landing_price else 0.0
        
        # Calculate total value based on landing price for non-PE users
        if is_pe:
            # PE Desk/Manager see both WAP and LP
            item["weighted_avg_price"] = calc["weighted_avg_price"]
            item["total_value"] = calc["total_value"]  # WAP-based value
            item["lp_total_value"] = round(item["available_quantity"] * landing_price, 2)  # LP-based value
        else:
            # Other users only see LP (WAP is hidden)
            item["weighted_avg_price"] = landing_price  # Show LP as the "price"
            item["total_value"] = round(item["available_quantity"] * landing_price, 2)
            # Remove actual WAP from response
            item.pop("weighted_avg_price_actual", None)
        
        result.append(item)
    
    return result


@router.get("/export")
async def export_inventory(
    format: str = Query("xlsx", enum=["xlsx", "csv"]),
    current_user: dict = Depends(get_current_user),
    _: None = Depends(require_permission("inventory.view", "export inventory"))
):
    """Export inventory to Excel or CSV"""
    import io
    from fastapi.responses import StreamingResponse
    
    # Get inventory with stock details
    inventory_data = await db.inventory.find({}, {"_id": 0}).to_list(10000)
    
    # Get stocks for lookup
    stocks = {s["id"]: s for s in await db.stocks.find({}, {"_id": 0, "id": 1, "name": 1, "symbol": 1, "face_value": 1}).to_list(10000)}
    
    if format == "csv":
        import csv
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Headers
        headers = [
            "Stock Symbol", "Stock Name", "Available Qty", "Reserved Qty",
            "Total Qty", "Weighted Avg Price", "Total Value", "Face Value",
            "Last Updated"
        ]
        writer.writerow(headers)
        
        # Data rows
        for inv in inventory_data:
            stock = stocks.get(inv.get("stock_id"), {})
            available = inv.get("available_quantity", 0)
            reserved = inv.get("reserved_quantity", 0)
            total = available + reserved
            avg_price = inv.get("weighted_average_price", 0)
            total_value = total * avg_price
            
            row = [
                stock.get("symbol", ""),
                stock.get("name", ""),
                available,
                reserved,
                total,
                round(avg_price, 2),
                round(total_value, 2),
                stock.get("face_value", ""),
                inv.get("updated_at", "")
            ]
            writer.writerow(row)
        
        output.seek(0)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=inventory_export.csv"}
        )
    
    else:  # xlsx
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventory"
        
        # Headers
        headers = [
            "Stock Symbol", "Stock Name", "Available Qty", "Reserved Qty",
            "Total Qty", "Weighted Avg Price", "Total Value", "Face Value",
            "Last Updated"
        ]
        
        # Style headers
        header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Data rows
        for row_num, inv in enumerate(inventory_data, 2):
            stock = stocks.get(inv.get("stock_id"), {})
            available = inv.get("available_quantity", 0)
            reserved = inv.get("reserved_quantity", 0)
            total = available + reserved
            avg_price = inv.get("weighted_average_price", 0)
            total_value = total * avg_price
            
            data = [
                stock.get("symbol", ""),
                stock.get("name", ""),
                available,
                reserved,
                total,
                round(avg_price, 2),
                round(total_value, 2),
                stock.get("face_value", ""),
                inv.get("updated_at", "")
            ]
            
            for col, value in enumerate(data, 1):
                ws.cell(row=row_num, column=col, value=value)
        
        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=inventory_export.xlsx"}
        )


@router.get("/{stock_id}", response_model=InventoryModel)
async def get_inventory_item(
    stock_id: str,
    current_user: dict = Depends(get_current_user),
    _: None = Depends(require_permission("inventory.view", "view inventory details"))
):
    """Get inventory for a specific stock"""
    user_role = current_user.get("role", 6)
    is_pe = is_pe_level(user_role)
    
    inventory = await db.inventory.find_one({"stock_id": stock_id}, {"_id": 0})
    if not inventory:
        raise HTTPException(status_code=404, detail="Inventory not found")
    
    # Enrich with stock details
    stock = await db.stocks.find_one({"id": stock_id}, {"_id": 0})
    if stock:
        inventory["stock_symbol"] = stock.get("symbol", "Unknown")
        inventory["stock_name"] = stock.get("name", "Unknown")
    
    # Calculate WAP
    calc = await calculate_weighted_avg_for_stock(stock_id)
    
    # Get landing price (or default to WAP)
    landing_price = inventory.get("landing_price")
    if landing_price is None or landing_price <= 0:
        landing_price = calc["weighted_avg_price"]
    
    inventory["landing_price"] = round(landing_price, 2)
    
    if is_pe:
        inventory["weighted_avg_price"] = calc["weighted_avg_price"]
        inventory["total_value"] = calc["total_value"]
    else:
        # Non-PE users see LP as the price
        inventory["weighted_avg_price"] = landing_price
        inventory["total_value"] = round(inventory.get("available_quantity", 0) * landing_price, 2)
    
    return inventory


@router.get("/{stock_id}/landing-price")
async def get_landing_price(
    stock_id: str,
    current_user: dict = Depends(get_current_user),
    _: None = Depends(require_permission("inventory.view", "view landing price"))
):
    """Get the landing price for a stock (used by booking to get buying price)"""
    inventory = await db.inventory.find_one({"stock_id": stock_id}, {"_id": 0})
    if not inventory:
        raise HTTPException(status_code=404, detail="Inventory not found")
    
    # Calculate WAP
    calc = await calculate_weighted_avg_for_stock(stock_id)
    
    # Get landing price (or default to WAP)
    landing_price = inventory.get("landing_price")
    if landing_price is None or landing_price <= 0:
        landing_price = calc["weighted_avg_price"]
    
    return {
        "stock_id": stock_id,
        "landing_price": round(landing_price, 2),
        "available_quantity": inventory.get("available_quantity", 0)
    }


@router.put("/{stock_id}/landing-price")
async def update_landing_price(
    stock_id: str,
    data: UpdateLandingPriceRequest,
    current_user: dict = Depends(get_current_user),
    _: None = Depends(require_permission("inventory.edit_landing_price", "update landing price"))
):
    """
    Update landing price for a stock (requires inventory.edit_landing_price permission).
    Landing Price (LP) is the price shown to non-PE users and used for booking revenue calculation.
    """
    user_role = current_user.get("role", 6)
    
    if data.landing_price <= 0:
        raise HTTPException(status_code=400, detail="Landing price must be greater than 0")
    
    # Check inventory exists
    inventory = await db.inventory.find_one({"stock_id": stock_id})
    if not inventory:
        raise HTTPException(status_code=404, detail="Inventory not found")
    
    # Get stock info for audit
    stock = await db.stocks.find_one({"id": stock_id}, {"_id": 0, "symbol": 1, "name": 1})
    
    old_lp = inventory.get("landing_price", 0)
    new_lp = round(data.landing_price, 2)
    
    # Store LP history entry
    lp_history_entry = {
        "stock_id": stock_id,
        "stock_symbol": stock.get("symbol") if stock else None,
        "old_price": old_lp,
        "new_price": new_lp,
        "change": round(new_lp - old_lp, 2) if old_lp else 0,
        "change_percent": round(((new_lp - old_lp) / old_lp) * 100, 2) if old_lp and old_lp > 0 else 0,
        "updated_by": current_user["id"],
        "updated_by_name": current_user["name"],
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    await db.lp_history.insert_one(lp_history_entry)
    
    # Update landing price with previous LP reference
    result = await db.inventory.update_one(
        {"stock_id": stock_id},
        {"$set": {
            "landing_price": new_lp,
            "previous_landing_price": old_lp,
            "landing_price_updated_at": datetime.now(timezone.utc).isoformat(),
            "landing_price_updated_by": current_user["id"],
            "landing_price_updated_by_name": current_user["name"]
        }}
    )
    
    # Create audit log
    from services.audit_service import create_audit_log
    await create_audit_log(
        action="LANDING_PRICE_UPDATE",
        entity_type="inventory",
        entity_id=stock_id,
        user_id=current_user["id"],
        user_name=current_user["name"],
        user_role=user_role,
        entity_name=stock.get("symbol", stock_id) if stock else stock_id,
        details={
            "stock_symbol": stock.get("symbol") if stock else None,
            "stock_name": stock.get("name") if stock else None,
            "new_landing_price": new_lp,
            "old_landing_price": old_lp
        }
    )
    
    return {
        "message": "Landing price updated successfully",
        "stock_id": stock_id,
        "landing_price": new_lp,
        "previous_landing_price": old_lp,
        "change": round(new_lp - old_lp, 2) if old_lp else 0
    }


@router.get("/{stock_id}/lp-history")
async def get_lp_history(
    stock_id: str,
    current_user: dict = Depends(get_current_user),
    _: None = Depends(require_permission("inventory.view_lp_history", "view LP history"))
):
    """Get landing price history for a stock"""
    # Get LP history from dedicated collection
    history = await db.lp_history.find(
        {"stock_id": stock_id},
        {"_id": 0}
    ).sort("updated_at", -1).to_list(100)
    
    # Also get current inventory data
    inventory = await db.inventory.find_one({"stock_id": stock_id}, {"_id": 0})
    
    return {
        "stock_id": stock_id,
        "stock_symbol": inventory.get("stock_symbol") if inventory else None,
        "current_lp": inventory.get("landing_price") if inventory else None,
        "previous_lp": inventory.get("previous_landing_price") if inventory else None,
        "history": history
    }


@router.delete("/{stock_id}")
async def delete_inventory(
    stock_id: str,
    current_user: dict = Depends(get_current_user),
    _: None = Depends(require_permission("inventory.delete", "delete inventory"))
):
    """Delete inventory for a stock (requires inventory.delete permission)"""
    user_role = current_user.get("role", 6)
    
    result = await db.inventory.delete_one({"stock_id": stock_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Inventory not found")
    
    return {"message": "Inventory deleted successfully"}


# ========== PE REPORT ENDPOINT ==========

@router.post("/send-pe-report")
async def send_consolidated_pe_report(
    current_user: dict = Depends(get_current_user),
    _: None = Depends(require_permission("inventory.send_report", "send PE report"))
):
    """
    Send Consolidated PE Inventory Report to all users via Email.
    
    A single report containing ALL stocks in tabular format:
    - Stock Code
    - Stock Name  
    - Landing Price (LP)
    
    Includes company logo, timestamp, footer with company info and disclaimer.
    CC: pe@smifs.com
    """
    from services.email_service import send_email
    from datetime import datetime, timezone
    
    # Get ALL inventory items (regardless of quantity)
    inventory_items = await db.inventory.find(
        {},  # No filter - include all stocks
        {"_id": 0}
    ).sort("stock_symbol", 1).to_list(1000)
    
    if not inventory_items:
        raise HTTPException(status_code=400, detail="No inventory items found")
    
    # Get stock master data for LP fallback
    stocks_master = {}
    all_stocks = await db.stocks.find({}, {"_id": 0, "id": 1, "symbol": 1, "landing_price": 1}).to_list(1000)
    for stock in all_stocks:
        if stock.get("id"):
            stocks_master[stock["id"]] = stock
        if stock.get("symbol"):
            stocks_master[stock["symbol"]] = stock
    
    # Filter out items without proper stock symbol (data quality) and enrich with LP
    valid_items = []
    for item in inventory_items:
        symbol = item.get("stock_symbol")
        if not symbol or symbol == "Unknown":
            continue
        
        # Try to get LP from inventory first, then from stocks master
        lp = item.get("landing_price")
        if not lp or lp <= 0:
            stock_id = item.get("stock_id")
            master = stocks_master.get(stock_id) or stocks_master.get(symbol)
            if master:
                lp = master.get("landing_price")
        
        item["_resolved_lp"] = lp  # Store resolved LP
        valid_items.append(item)
    
    items_without_symbol = len(inventory_items) - len(valid_items)
    
    # Count items with and without LP
    items_with_lp = [item for item in valid_items if item.get("_resolved_lp") and item.get("_resolved_lp") > 0]
    items_without_lp = len(valid_items) - len(items_with_lp)
    
    # Count items with available stock vs out of stock
    items_in_stock = [item for item in valid_items if item.get("available_quantity", 0) > 0]
    items_out_of_stock = len(valid_items) - len(items_in_stock)
    
    if not valid_items:
        raise HTTPException(status_code=400, detail="No valid inventory items found (all items have missing stock symbols)")
    
    # Get company master details
    company = await db.company_master.find_one({"_id": "company_settings"}, {"_id": 0})
    if not company:
        company = {
            "company_name": "SMIFS Private Equity",
            "address": "12, India Exchange Place, Kolkata - 700001",
            "phone": "+91-33-4012-1234",
            "email": "pe@smifs.com",
            "website": "www.smifs.com",
            "logo_url": ""
        }
    
    # Get all active users with email
    users = await db.users.find(
        {"is_active": {"$ne": False}, "email": {"$exists": True, "$ne": ""}},
        {"_id": 0, "id": 1, "name": 1, "email": 1}
    ).to_list(10000)
    
    if not users:
        raise HTTPException(status_code=400, detail="No users found to send report")
    
    # Generate timestamp
    report_time = datetime.now(timezone.utc)
    report_time_ist = report_time.strftime("%d %B %Y, %I:%M %p IST")
    
    # Build inventory table rows - compact design for many stocks
    table_rows = ""
    for idx, item in enumerate(valid_items):
        bg_color = "#fff" if idx % 2 == 0 else "#f8f9fa"
        lp = item.get("_resolved_lp", 0)
        lp_formatted = f"₹{lp:,.2f}" if lp and lp > 0 else '<span style="color:#999;">-</span>'
        
        # Show availability status - compact
        avail_qty = item.get("available_quantity", 0)
        if avail_qty > 0:
            availability = '<span style="color:#16a34a;">●</span>'
        else:
            availability = '<span style="color:#dc2626;">○</span>'
        
        table_rows += f"""<tr style="background:{bg_color};">
            <td style="padding:6px 8px;border-bottom:1px solid #e5e7eb;font-family:monospace;font-size:11px;font-weight:600;color:#064E3B;">{item.get('stock_symbol', '-')}</td>
            <td style="padding:6px 8px;border-bottom:1px solid #e5e7eb;font-size:11px;color:#374151;">{item.get('stock_name', '-')}</td>
            <td style="padding:6px 8px;border-bottom:1px solid #e5e7eb;text-align:right;font-weight:600;color:#059669;font-size:11px;">{lp_formatted}</td>
            <td style="padding:6px 8px;border-bottom:1px solid #e5e7eb;text-align:center;font-size:11px;">{availability}</td>
        </tr>"""
    
    # Build the compact HTML email
    logo_html = ""
    if company.get("logo_url"):
        logo_html = f'<img src="{company["logo_url"]}" alt="{company["company_name"]}" style="max-height:40px;max-width:150px;">'
    else:
        logo_html = f'<div style="font-size:18px;font-weight:bold;color:#fff;">{company.get("company_name", "SMIFS PE")}</div>'
    
    # Compact warning note if some items don't have LP
    lp_warning = ""
    if items_without_lp > 0:
        lp_warning = f'<p style="background:#fef3c7;padding:8px 12px;border-radius:4px;margin:0 0 15px 0;font-size:11px;color:#92400e;"><b>Note:</b> {items_without_lp} stock(s) without LP. Contact PE Desk.</p>'
    
    html_content = f"""<!DOCTYPE html>
<html>
<head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0"></head>
<body style="margin:0;padding:0;font-family:Arial,sans-serif;background:#f3f4f6;font-size:12px;">
<div style="max-width:600px;margin:0 auto;background:#fff;">
    
    <!-- Header -->
    <div style="background:#064E3B;padding:15px 20px;text-align:center;">
        {logo_html}
        <h1 style="color:#fff;margin:8px 0 0 0;font-size:16px;font-weight:600;">PE Inventory Report</h1>
    </div>
    
    <!-- Timestamp -->
    <div style="background:#ecfdf5;padding:8px 20px;border-bottom:1px solid #10b981;">
        <p style="margin:0;color:#065f46;font-size:11px;"><b>Generated:</b> {report_time_ist}</p>
    </div>
    
    <!-- Content -->
    <div style="padding:15px 20px;">
        <p style="color:#374151;font-size:12px;line-height:1.5;margin:0 0 12px 0;">
            Current inventory of unlisted and pre-IPO shares. <span style="color:#16a34a;">●</span> Available <span style="color:#dc2626;">○</span> Out of Stock
        </p>
        
        {lp_warning}
        
        <!-- Table -->
        <table style="width:100%;border-collapse:collapse;border:1px solid #e5e7eb;font-size:11px;">
            <thead>
                <tr style="background:#064E3B;">
                    <th style="padding:8px 6px;text-align:left;color:#fff;font-size:10px;font-weight:600;">CODE</th>
                    <th style="padding:8px 6px;text-align:left;color:#fff;font-size:10px;font-weight:600;">STOCK NAME</th>
                    <th style="padding:8px 6px;text-align:right;color:#fff;font-size:10px;font-weight:600;">LP</th>
                    <th style="padding:8px 6px;text-align:center;color:#fff;font-size:10px;font-weight:600;width:30px;"></th>
                </tr>
            </thead>
            <tbody>
                {table_rows}
            </tbody>
        </table>
        
        <p style="color:#6b7280;font-size:10px;margin:10px 0;">
            Total: <b>{len(valid_items)}</b> ({len(items_in_stock)} available) {f'• {len(items_with_lp)} priced' if items_without_lp > 0 else ''}
        </p>
        
        <!-- CTA -->
        <div style="text-align:center;margin:15px 0;">
            <a href="{company.get('custom_domain', 'https://pesmifs.com')}/bookings" 
               style="display:inline-block;background:#064E3B;color:#fff;padding:10px 25px;text-decoration:none;border-radius:5px;font-size:12px;font-weight:600;">
                Book Now
            </a>
        </div>
    </div>
    
    <!-- Disclaimer -->
    <div style="background:#fef3c7;padding:10px 20px;border-top:1px solid #fbbf24;">
        <p style="margin:0;color:#92400e;font-size:9px;line-height:1.4;">
            <b>DISCLAIMER:</b> Indicative report for information only. Prices subject to change. Not investment advice. Consult financial advisor. SMIFS Ltd is SEBI registered.
        </p>
    </div>
    
    <!-- Footer -->
    <div style="background:#1f2937;padding:15px 20px;text-align:center;">
        <p style="color:#fff;font-size:11px;margin:0 0 5px 0;font-weight:600;">{company.get('company_name', 'SMIFS PE')}</p>
        <p style="color:#9ca3af;font-size:9px;margin:0;line-height:1.4;">
            {company.get('address', '')}
        </p>
        <p style="color:#9ca3af;font-size:9px;margin:8px 0 0 0;">
            {company.get('phone', '')} | {company.get('email', 'pe@smifs.com')} | {company.get('website', 'www.smifs.com')}
        </p>
        <p style="color:#6b7280;font-size:8px;margin:10px 0 0 0;">
            © {datetime.now().year} {company.get('company_name', 'SMIFS')}
        </p>
    </div>
</div>
</body>
</html>"""
    
    # Send to all users with better error tracking
    results = {
        "total_users": len(users),
        "emails_sent": 0,
        "emails_failed": 0,
        "errors": []
    }
    
    subject = f"📊 PE Inventory Report - {report_time.strftime('%d %b %Y')}"
    
    for user in users:
        user_email = user.get("email")
        if user_email:
            try:
                await send_email(
                    to_email=user_email,
                    subject=subject,
                    body=html_content,
                    cc_email="pe@smifs.com"
                )
                results["emails_sent"] += 1
            except Exception as e:
                results["emails_failed"] += 1
                logger.error(f"Failed to send PE Report to {user_email}: {str(e)}")
                results["errors"].append({"email": user_email, "error": str(e)})
    
    # Create audit log
    from services.audit_service import create_audit_log
    await create_audit_log(
        action="PE_INVENTORY_REPORT_SENT",
        entity_type="inventory",
        entity_id="consolidated_report",
        user_id=current_user["id"],
        user_name=current_user["name"],
        user_role=current_user.get("role", 6),
        entity_name="PE Inventory Report",
        details={
            "total_stocks": len(valid_items),
            "stocks_in_stock": len(items_in_stock),
            "stocks_out_of_stock": items_out_of_stock,
            "stocks_with_lp": len(items_with_lp),
            "stocks_without_lp": items_without_lp,
            "skipped_invalid": items_without_symbol,
            "total_users": results["total_users"],
            "emails_sent": results["emails_sent"],
            "emails_failed": results["emails_failed"],
            "errors_count": len(results["errors"]),
            "report_time": report_time_ist
        }
    )
    
    logger.info(f"PE Inventory Report: {results['emails_sent']} sent, {results['emails_failed']} failed by {current_user['name']}")
    
    # Build detailed response message
    if results["emails_failed"] == 0:
        message = f"✅ PE Report sent successfully to {results['emails_sent']} users"
    elif results["emails_sent"] > 0:
        message = f"⚠️ PE Report sent to {results['emails_sent']} users, {results['emails_failed']} failed"
    else:
        message = f"❌ Failed to send PE Report. {results['emails_failed']} emails failed."
    
    return {
        "success": results["emails_sent"] > 0,
        "message": message,
        "report_time": report_time_ist,
        "stocks_included": len(valid_items),
        "stocks_in_stock": len(items_in_stock),
        "stocks_out_of_stock": items_out_of_stock,
        "stocks_with_lp": len(items_with_lp),
        "stocks_without_lp": items_without_lp,
        "results": results
    }


@router.post("/recalculate")
async def recalculate_all_inventory(current_user: dict = Depends(get_current_user)):
    """
    Recalculate inventory for ALL stocks (PE Desk only).
    
    This is a manual fail-safe to ensure inventory data is accurate.
    It recalculates available_quantity, blocked_quantity, and weighted_avg_price
    for every stock based on actual purchases and bookings.
    """
    from services.permission_service import check_permission
    
    # Check permission dynamically
    await check_permission(current_user, "inventory.recalculate", "recalculate inventory")
    
    from services.inventory_service import update_inventory
    
    # Get all unique stock IDs from purchases, inventory, AND bookings
    purchase_stock_ids = await db.purchases.distinct("stock_id")
    inventory_stock_ids = await db.inventory.distinct("stock_id")
    booking_stock_ids = await db.bookings.distinct("stock_id")
    all_stock_ids = list(set(purchase_stock_ids + inventory_stock_ids + booking_stock_ids))
    
    results = {
        "total_stocks": len(all_stock_ids),
        "recalculated": 0,
        "errors": []
    }
    
    for stock_id in all_stock_ids:
        try:
            await update_inventory(stock_id)
            results["recalculated"] += 1
        except Exception as e:
            logger.error(f"Error recalculating inventory for stock {stock_id}: {str(e)}")
            results["errors"].append({"stock_id": stock_id, "error": str(e)})
    
    # Create audit log
    from services.audit_service import create_audit_log
    await create_audit_log(
        action="INVENTORY_RECALCULATE_ALL",
        entity_type="inventory",
        entity_id="all",
        user_id=current_user["id"],
        user_name=current_user["name"],
        user_role=current_user.get("role", 6),
        entity_name="All Stocks",
        details={
            "total_stocks": results["total_stocks"],
            "recalculated": results["recalculated"],
            "errors_count": len(results["errors"])
        }
    )
    
    logger.info(f"Inventory recalculated for {results['recalculated']}/{results['total_stocks']} stocks by {current_user['name']}")
    
    return {
        "message": f"Inventory recalculated for {results['recalculated']} of {results['total_stocks']} stocks",
        "details": results
    }



# ========== DATA QUALITY ENDPOINTS ==========

@router.get("/data-quality/report")
async def get_data_quality_report(
    current_user: dict = Depends(get_current_user),
    _: None = Depends(require_permission("inventory.view", "view data quality report"))
):
    """
    Get a data quality report for inventory and stocks collections.
    Reports issues like orphaned records, missing landing prices, etc.
    """
    from scripts.data_cleanup import generate_data_quality_report
    from database import db as database
    
    report = await generate_data_quality_report(database)
    return report


@router.post("/data-quality/cleanup")
async def run_data_cleanup(
    dry_run: bool = True,
    current_user: dict = Depends(get_current_user),
    _: None = Depends(require_permission("inventory.edit_landing_price", "run data cleanup"))
):
    """
    Run data cleanup operations on inventory and stocks collections.
    
    Args:
        dry_run: If True (default), only report what would be changed without making changes.
                 Set to False to actually apply fixes.
    
    Operations performed:
    1. Remove or fix orphaned inventory records (stock_symbol = "Unknown")
    2. Set missing landing_price to weighted_avg_price or stock master LP
    3. Validate and update stock references in inventory
    
    PE Desk/Manager only.
    """
    from scripts.data_cleanup import generate_data_quality_report
    from database import db as database
    from services.audit_service import create_audit_log
    
    # Run cleanup
    results = {
        "dry_run": dry_run,
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "report_before": await generate_data_quality_report(database)
    }
    
    # Cleanup orphaned inventory
    from scripts.data_cleanup import (
        cleanup_orphaned_inventory, 
        fix_missing_landing_prices, 
        validate_inventory_stock_references
    )
    
    results["orphaned_cleanup"] = await cleanup_orphaned_inventory(database, dry_run)
    results["landing_price_fix"] = await fix_missing_landing_prices(database, dry_run)
    results["reference_validation"] = await validate_inventory_stock_references(database, dry_run)
    
    if not dry_run:
        results["report_after"] = await generate_data_quality_report(database)
        
        # Create audit log
        await create_audit_log(
            action="DATA_CLEANUP_EXECUTED",
            entity_type="inventory",
            entity_id="all",
            user_id=current_user["id"],
            user_name=current_user["name"],
            user_role=current_user.get("role", 6),
            entity_name="Data Quality Cleanup",
            details={
                "orphaned_deleted": results["orphaned_cleanup"]["orphaned_deleted"],
                "lp_fixed": results["landing_price_fix"]["lp_fixed"],
                "metadata_updated": results["reference_validation"]["updated_metadata"],
                "health_before": results["report_before"]["summary"]["health_score"],
                "health_after": results["report_after"]["summary"]["health_score"]
            }
        )
        
        logger.info(f"Data cleanup executed by {current_user['name']}: "
                   f"{results['orphaned_cleanup']['orphaned_deleted']} orphaned deleted, "
                   f"{results['landing_price_fix']['lp_fixed']} LP fixed")
    
    return {
        "success": True,
        "message": "Dry run completed" if dry_run else "Data cleanup completed",
        "dry_run": dry_run,
        "results": results
    }

