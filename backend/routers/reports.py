"""
Reports Router
Handles P&L reports, exports (Excel, PDF), and financial reporting
"""
from typing import List, Optional
from datetime import datetime, timezone
from fastapi import APIRouter, HTTPException, Depends, Query
from fastapi.responses import StreamingResponse
import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT

from database import db
from config import ROLES
from utils.auth import get_current_user
from services.permission_service import (
    has_permission,
    check_permission as check_dynamic_permission,
    require_permission
)
from utils.demo_isolation import is_demo_user, add_demo_filter, mark_as_demo, require_demo_access

router = APIRouter(prefix="/reports", tags=["Reports"])


# Helper function for backward compatibility
def is_pe_level(role: int) -> bool:
    """Check if role is PE level (PE Desk or PE Manager)."""
    return role in [1, 2]


@router.get("/pnl", dependencies=[Depends(require_permission("reports.view", "view P&L reports"))])
async def get_pnl_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    stock_id: Optional[str] = None,
    client_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get P&L report with optional filters"""
    query = {"status": {"$ne": "cancelled"}, "is_voided": {"$ne": True}}
    
    if start_date:
        query["booking_date"] = {"$gte": start_date}
    if end_date:
        if "booking_date" in query:
            query["booking_date"]["$lte"] = end_date
        else:
            query["booking_date"] = {"$lte": end_date}
    if stock_id:
        query["stock_id"] = stock_id
    if client_id:
        query["client_id"] = client_id
    
    # CRITICAL: Add demo data isolation filter
    query = add_demo_filter(query, current_user)
    
    bookings = await db.bookings.find(query, {"_id": 0}).to_list(10000)
    
    # Calculate P&L
    total_revenue = 0
    total_cost = 0
    pnl_items = []
    
    for booking in bookings:
        qty = booking.get("quantity", 0)
        selling = booking.get("selling_price", 0)
        buying = booking.get("buying_price", 0)
        
        revenue = qty * selling
        cost = qty * buying
        profit = revenue - cost
        
        total_revenue += revenue
        total_cost += cost
        
        # Get related entities
        client = await db.clients.find_one({"id": booking.get("client_id")}, {"_id": 0, "name": 1})
        stock = await db.stocks.find_one({"id": booking.get("stock_id")}, {"_id": 0, "symbol": 1, "name": 1})
        
        pnl_items.append({
            "booking_id": booking.get("id"),
            "booking_number": booking.get("booking_number"),
            "booking_date": booking.get("booking_date"),
            "client_name": client.get("name") if client else "Unknown",
            "stock_symbol": stock.get("symbol") if stock else "Unknown",
            "stock_name": stock.get("name") if stock else "Unknown",
            "quantity": qty,
            "buying_price": buying,
            "selling_price": selling,
            "total_cost": cost,
            "total_revenue": revenue,
            "profit_loss": profit,
            "profit_margin": (profit / revenue * 100) if revenue > 0 else 0
        })
    
    return {
        "items": pnl_items,
        "summary": {
            "total_revenue": total_revenue,
            "total_cost": total_cost,
            "gross_profit": total_revenue - total_cost,
            "profit_margin": ((total_revenue - total_cost) / total_revenue * 100) if total_revenue > 0 else 0,
            "total_bookings": len(pnl_items)
        }
    }


@router.get("/export/excel", dependencies=[Depends(require_permission("reports.export", "export P&L reports"))])
async def export_pnl_excel(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    stock_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Export P&L report to Excel"""
    # Get P&L data
    query = {"status": {"$ne": "cancelled"}, "is_voided": {"$ne": True}}
    
    if start_date:
        query["booking_date"] = {"$gte": start_date}
    if end_date:
        if "booking_date" in query:
            query["booking_date"]["$lte"] = end_date
        else:
            query["booking_date"] = {"$lte": end_date}
    if stock_id:
        query["stock_id"] = stock_id
    
    bookings = await db.bookings.find(query, {"_id": 0}).to_list(10000)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "P&L Report"
    
    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="064E3B", end_color="064E3B", fill_type="solid")
    
    # Headers
    headers = ["Booking #", "Date", "Client", "Stock", "Qty", "Buy Price", "Sell Price", "Cost", "Revenue", "Profit/Loss"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Data rows
    total_revenue = 0
    total_cost = 0
    
    for row_idx, booking in enumerate(bookings, 2):
        client = await db.clients.find_one({"id": booking.get("client_id")}, {"_id": 0, "name": 1})
        stock = await db.stocks.find_one({"id": booking.get("stock_id")}, {"_id": 0, "symbol": 1})
        
        qty = booking.get("quantity", 0)
        buying = booking.get("buying_price", 0)
        selling = booking.get("selling_price", 0)
        cost = qty * buying
        revenue = qty * selling
        profit = revenue - cost
        
        total_revenue += revenue
        total_cost += cost
        
        ws.cell(row=row_idx, column=1, value=booking.get("booking_number", ""))
        ws.cell(row=row_idx, column=2, value=booking.get("booking_date", ""))
        ws.cell(row=row_idx, column=3, value=client.get("name") if client else "Unknown")
        ws.cell(row=row_idx, column=4, value=stock.get("symbol") if stock else "Unknown")
        ws.cell(row=row_idx, column=5, value=qty)
        ws.cell(row=row_idx, column=6, value=buying)
        ws.cell(row=row_idx, column=7, value=selling)
        ws.cell(row=row_idx, column=8, value=cost)
        ws.cell(row=row_idx, column=9, value=revenue)
        ws.cell(row=row_idx, column=10, value=profit)
    
    # Summary row
    summary_row = len(bookings) + 3
    ws.cell(row=summary_row, column=1, value="TOTALS").font = Font(bold=True)
    ws.cell(row=summary_row, column=8, value=total_cost).font = Font(bold=True)
    ws.cell(row=summary_row, column=9, value=total_revenue).font = Font(bold=True)
    ws.cell(row=summary_row, column=10, value=total_revenue - total_cost).font = Font(bold=True)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 30)
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"pnl_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/export/pdf", dependencies=[Depends(require_permission("reports.export", "export P&L reports"))])
async def export_pnl_pdf(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    stock_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Export P&L report to PDF"""
    # Get P&L data
    query = {"status": {"$ne": "cancelled"}, "is_voided": {"$ne": True}}
    
    if start_date:
        query["booking_date"] = {"$gte": start_date}
    if end_date:
        if "booking_date" in query:
            query["booking_date"]["$lte"] = end_date
        else:
            query["booking_date"] = {"$lte": end_date}
    if stock_id:
        query["stock_id"] = stock_id
    
    bookings = await db.bookings.find(query, {"_id": 0}).to_list(10000)
    
    # Create PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'Title',
        parent=styles['Heading1'],
        fontSize=18,
        alignment=TA_CENTER,
        spaceAfter=30
    )
    
    # Title
    elements.append(Paragraph("Profit & Loss Report", title_style))
    elements.append(Spacer(1, 12))
    
    # Date range
    date_range = f"Period: {start_date or 'All'} to {end_date or 'Present'}"
    elements.append(Paragraph(date_range, styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Table data
    table_data = [["Booking #", "Date", "Client", "Stock", "Qty", "P/L"]]
    
    total_profit = 0
    for booking in bookings:
        client = await db.clients.find_one({"id": booking.get("client_id")}, {"_id": 0, "name": 1})
        stock = await db.stocks.find_one({"id": booking.get("stock_id")}, {"_id": 0, "symbol": 1})
        
        qty = booking.get("quantity", 0)
        profit = (booking.get("selling_price", 0) - booking.get("buying_price", 0)) * qty
        total_profit += profit
        
        table_data.append([
            booking.get("booking_number", "")[:12],
            booking.get("booking_date", "")[:10],
            (client.get("name") if client else "Unknown")[:15],
            (stock.get("symbol") if stock else "Unknown")[:10],
            str(qty),
            f"₹{profit:,.0f}"
        ])
    
    # Summary row
    table_data.append(["", "", "", "", "TOTAL:", f"₹{total_profit:,.0f}"])
    
    # Create table
    table = Table(table_data, colWidths=[1.2*inch, 0.9*inch, 1.2*inch, 0.8*inch, 0.5*inch, 1*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.02, 0.3, 0.23)),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, -1), (-1, -1), colors.Color(0.9, 0.9, 0.9)),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
    ]))
    
    elements.append(table)
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    filename = f"pnl_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/client-portfolio/{client_id}", dependencies=[Depends(require_permission("reports.view", "view client portfolio"))])
async def get_client_portfolio(
    client_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Get complete portfolio for a client"""
    client = await db.clients.find_one({"id": client_id}, {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    
    # Get all bookings for this client
    bookings = await db.bookings.find(
        {"client_id": client_id, "is_voided": {"$ne": True}},
        {"_id": 0}
    ).to_list(10000)
    
    # Group by stock
    stock_holdings = {}
    for booking in bookings:
        stock_id = booking.get("stock_id")
        if stock_id not in stock_holdings:
            stock = await db.stocks.find_one({"id": stock_id}, {"_id": 0})
            stock_holdings[stock_id] = {
                "stock_id": stock_id,
                "stock_symbol": stock.get("symbol") if stock else "Unknown",
                "stock_name": stock.get("name") if stock else "Unknown",
                "total_quantity": 0,
                "total_investment": 0,
                "bookings": []
            }
        
        qty = booking.get("quantity", 0)
        buying_price = booking.get("buying_price", 0)
        
        # Only count approved and transferred bookings
        if booking.get("stock_transferred"):
            stock_holdings[stock_id]["total_quantity"] += qty
            stock_holdings[stock_id]["total_investment"] += qty * buying_price
        
        stock_holdings[stock_id]["bookings"].append({
            "booking_number": booking.get("booking_number"),
            "booking_date": booking.get("booking_date"),
            "quantity": qty,
            "buying_price": buying_price,
            "selling_price": booking.get("selling_price"),
            "status": booking.get("status"),
            "approval_status": booking.get("approval_status"),
            "stock_transferred": booking.get("stock_transferred", False)
        })
    
    # Calculate average price
    for holding in stock_holdings.values():
        if holding["total_quantity"] > 0:
            holding["average_price"] = holding["total_investment"] / holding["total_quantity"]
        else:
            holding["average_price"] = 0
    
    return {
        "client": {
            "id": client.get("id"),
            "name": client.get("name"),
            "otc_ucc": client.get("otc_ucc"),
            "pan_number": client.get("pan_number")
        },
        "holdings": list(stock_holdings.values()),
        "summary": {
            "total_stocks": len(stock_holdings),
            "total_investment": sum(h["total_investment"] for h in stock_holdings.values()),
            "total_bookings": len(bookings)
        }
    }


# ============== PE Desk HIT Report ==============

@router.get("/pe-desk-hit", dependencies=[Depends(require_permission("reports.pe_hit", "view PE HIT report"))])
async def get_pe_desk_hit_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    stock_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """
    PE Desk HIT Report - Shows (LP - WAP) × Quantity for completed DP transfers.
    
    This report calculates the difference between:
    - LP (Landing Price): What the booking was made at
    - WAP (Weighted Average Price): The actual cost price
    
    The HIT is the margin captured by PE Desk before showing prices to users.
    Only accessible to PE Desk/Manager.
    """
    user_role = current_user.get("role", 6)
    
    if not is_pe_level(user_role):
        raise HTTPException(status_code=403, detail="Only PE Desk or PE Manager can access HIT report")
    
    # Query for completed DP transfers
    query = {
        "status": "completed",  # Only completed bookings
        "is_voided": {"$ne": True}
    }
    
    if start_date:
        query["dp_transfer_date"] = {"$gte": start_date}
    if end_date:
        if "dp_transfer_date" in query:
            query["dp_transfer_date"]["$lte"] = end_date
        else:
            query["dp_transfer_date"] = {"$lte": end_date}
    if stock_id:
        query["stock_id"] = stock_id
    
    bookings = await db.bookings.find(query, {"_id": 0}).to_list(10000)
    
    # Get stock and client details
    stock_ids = list(set(b.get("stock_id") for b in bookings if b.get("stock_id")))
    client_ids = list(set(b.get("client_id") for b in bookings if b.get("client_id")))
    
    stocks = await db.stocks.find({"id": {"$in": stock_ids}}, {"_id": 0}).to_list(1000)
    clients = await db.clients.find({"id": {"$in": client_ids}}, {"_id": 0}).to_list(1000)
    
    stock_map = {s["id"]: s for s in stocks}
    client_map = {c["id"]: c for c in clients}
    
    # Calculate HIT for each booking
    report_items = []
    total_hit = 0
    total_quantity = 0
    
    for booking in bookings:
        stock = stock_map.get(booking.get("stock_id"), {})
        client = client_map.get(booking.get("client_id"), {})
        
        quantity = booking.get("quantity", 0)
        wap = booking.get("weighted_avg_price") or booking.get("buying_price", 0)
        lp = booking.get("landing_price") or booking.get("buying_price", 0)
        
        # HIT = (LP - WAP) × Quantity
        hit = (lp - wap) * quantity
        
        report_items.append({
            "booking_id": booking.get("id"),
            "booking_number": booking.get("booking_number"),
            "booking_date": booking.get("booking_date"),
            "dp_transfer_date": booking.get("dp_transfer_date"),
            "stock_id": booking.get("stock_id"),
            "stock_symbol": stock.get("symbol", "Unknown"),
            "stock_name": stock.get("name", "Unknown"),
            "client_id": booking.get("client_id"),
            "client_name": client.get("name", "Unknown"),
            "client_otc_ucc": client.get("otc_ucc"),
            "quantity": quantity,
            "wap": round(wap, 2),  # Weighted Average Price
            "lp": round(lp, 2),  # Landing Price
            "lp_wap_diff": round(lp - wap, 2),  # LP - WAP per share
            "hit": round(hit, 2),  # Total HIT for this booking
            "selling_price": booking.get("selling_price", 0),
            "revenue": round((booking.get("selling_price", 0) - lp) * quantity, 2)
        })
        
        total_hit += hit
        total_quantity += quantity
    
    # Group by stock for summary
    stock_summary = {}
    for item in report_items:
        sid = item["stock_id"]
        if sid not in stock_summary:
            stock_summary[sid] = {
                "stock_symbol": item["stock_symbol"],
                "stock_name": item["stock_name"],
                "total_quantity": 0,
                "total_hit": 0,
                "avg_lp_wap_diff": 0,
                "booking_count": 0
            }
        stock_summary[sid]["total_quantity"] += item["quantity"]
        stock_summary[sid]["total_hit"] += item["hit"]
        stock_summary[sid]["booking_count"] += 1
    
    # Calculate average diff
    for sid in stock_summary:
        if stock_summary[sid]["total_quantity"] > 0:
            stock_summary[sid]["avg_lp_wap_diff"] = round(
                stock_summary[sid]["total_hit"] / stock_summary[sid]["total_quantity"], 2
            )
    
    return {
        "report_type": "PE Desk HIT Report",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "generated_by": current_user["name"],
        "filters": {
            "start_date": start_date,
            "end_date": end_date,
            "stock_id": stock_id
        },
        "summary": {
            "total_bookings": len(report_items),
            "total_quantity": total_quantity,
            "total_hit": round(total_hit, 2),
            "avg_hit_per_share": round(total_hit / total_quantity, 2) if total_quantity > 0 else 0
        },
        "by_stock": list(stock_summary.values()),
        "details": report_items
    }


@router.get("/pe-desk-hit/export", dependencies=[Depends(require_permission("reports.pe_hit", "export PE HIT report"))])
async def export_pe_desk_hit_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    stock_id: Optional[str] = None,
    format: str = Query("xlsx", enum=["xlsx", "pdf"]),
    current_user: dict = Depends(get_current_user)
):
    """Export PE Desk HIT Report to Excel or PDF"""
    user_role = current_user.get("role", 6)
    
    if not is_pe_level(user_role):
        raise HTTPException(status_code=403, detail="Only PE Desk or PE Manager can export HIT report")
    
    # Get report data
    report = await get_pe_desk_hit_report(start_date, end_date, stock_id, current_user)
    
    if format == "xlsx":
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "PE Desk HIT Report"
        
        # Header styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        
        # Title
        ws.merge_cells('A1:L1')
        ws['A1'] = f"PE Desk HIT Report - {start_date or 'All'} to {end_date or 'All'}"
        ws['A1'].font = Font(bold=True, size=14)
        
        # Summary
        ws['A3'] = "Summary"
        ws['A3'].font = Font(bold=True, size=12)
        ws['A4'] = f"Total Bookings: {report['summary']['total_bookings']}"
        ws['A5'] = f"Total Quantity: {report['summary']['total_quantity']:,}"
        ws['A6'] = f"Total HIT: ₹{report['summary']['total_hit']:,.2f}"
        ws['A7'] = f"Avg HIT/Share: ₹{report['summary']['avg_hit_per_share']:,.2f}"
        
        # Details header
        row = 9
        headers = ["Booking #", "Date", "DP Transfer", "Stock", "Client", "Qty", "WAP", "LP", "Diff", "HIT", "SP", "Revenue"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Details data
        for item in report['details']:
            row += 1
            ws.cell(row=row, column=1, value=item['booking_number'])
            ws.cell(row=row, column=2, value=item['booking_date'])
            ws.cell(row=row, column=3, value=item['dp_transfer_date'])
            ws.cell(row=row, column=4, value=item['stock_symbol'])
            ws.cell(row=row, column=5, value=item['client_name'])
            ws.cell(row=row, column=6, value=item['quantity'])
            ws.cell(row=row, column=7, value=item['wap'])
            ws.cell(row=row, column=8, value=item['lp'])
            ws.cell(row=row, column=9, value=item['lp_wap_diff'])
            ws.cell(row=row, column=10, value=item['hit'])
            ws.cell(row=row, column=11, value=item['selling_price'])
            ws.cell(row=row, column=12, value=item['revenue'])
        
        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="PE_Desk_HIT_Report_{datetime.now().strftime("%Y%m%d")}.xlsx"'}
        )
    
    else:  # PDF
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle('Title', parent=styles['Heading1'], alignment=TA_CENTER)
        elements.append(Paragraph("PE Desk HIT Report", title_style))
        elements.append(Spacer(1, 20))
        
        # Summary
        summary_data = [
            ["Total Bookings", str(report['summary']['total_bookings'])],
            ["Total Quantity", f"{report['summary']['total_quantity']:,}"],
            ["Total HIT", f"₹{report['summary']['total_hit']:,.2f}"],
            ["Avg HIT/Share", f"₹{report['summary']['avg_hit_per_share']:,.2f}"]
        ]
        summary_table = Table(summary_data, colWidths=[150, 150])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 20))
        
        # Details
        elements.append(Paragraph("Details", styles['Heading2']))
        elements.append(Spacer(1, 10))
        
        detail_data = [["#", "Stock", "Client", "Qty", "WAP", "LP", "HIT"]]
        for item in report['details'][:50]:  # Limit to 50 for PDF
            detail_data.append([
                item['booking_number'],
                item['stock_symbol'],
                item['client_name'][:15],
                str(item['quantity']),
                f"₹{item['wap']:.2f}",
                f"₹{item['lp']:.2f}",
                f"₹{item['hit']:.2f}"
            ])
        
        detail_table = Table(detail_data, colWidths=[70, 50, 80, 40, 55, 55, 60])
        detail_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.green),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (3, 0), (-1, -1), 'RIGHT'),
        ]))
        elements.append(detail_table)
        
        doc.build(elements)
        buffer.seek(0)
        
        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="PE_Desk_HIT_Report_{datetime.now().strftime("%Y%m%d")}.pdf"'}
        )

