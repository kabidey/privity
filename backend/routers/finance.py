"""
Finance Router

Handles all finance-related operations including payments tracking, refund requests,
and financial reports/exports.
"""
from fastapi import APIRouter, HTTPException, Depends, Query
from fastapi.responses import StreamingResponse
from typing import List, Optional
from datetime import datetime, timezone
from pydantic import BaseModel
import uuid
import io

from database import db
from config import is_pe_level, has_finance_access, can_manage_finance
from utils.auth import get_current_user
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

router = APIRouter(tags=["Finance"])


class RefundStatusUpdate(BaseModel):
    status: str  # processing, completed, failed
    notes: Optional[str] = None
    reference_number: Optional[str] = None


@router.get("/finance/payments")
async def get_all_payments(
    payment_type: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get all payments (client and vendor) for finance dashboard."""
    if not has_finance_access(current_user.get("role", 6)):
        raise HTTPException(status_code=403, detail="Only PE Desk, PE Manager, or Finance can access finance data")
    
    all_payments = []
    
    # Get client payments from bookings
    if payment_type in [None, "client"]:
        bookings = await db.bookings.find(
            {"payments": {"$exists": True, "$ne": []}},
            {"_id": 0}
        ).to_list(10000)
        
        for booking in bookings:
            client = await db.clients.find_one({"id": booking["client_id"]}, {"_id": 0, "name": 1})
            stock = await db.stocks.find_one({"id": booking["stock_id"]}, {"_id": 0, "symbol": 1})
            
            for payment in booking.get("payments", []):
                payment_date = payment.get("payment_date", "")
                
                # Date filter
                if start_date and payment_date < start_date:
                    continue
                if end_date and payment_date > end_date:
                    continue
                
                all_payments.append({
                    "id": payment.get("id", str(uuid.uuid4())),
                    "type": "client",
                    "direction": "received",
                    "entity_id": booking["client_id"],
                    "entity_name": client["name"] if client else "Unknown",
                    "reference_id": booking["id"],
                    "reference_number": booking.get("booking_number", ""),
                    "stock_symbol": stock["symbol"] if stock else "Unknown",
                    "amount": payment.get("amount", 0),
                    "payment_date": payment_date,
                    "notes": payment.get("notes", ""),
                    "proof_url": payment.get("proof_url"),
                    "recorded_by": payment.get("recorded_by_name", "")
                })
    
    # Get vendor payments from purchases
    if payment_type in [None, "vendor"]:
        purchases = await db.purchases.find(
            {"payments": {"$exists": True, "$ne": []}},
            {"_id": 0}
        ).to_list(10000)
        
        for purchase in purchases:
            vendor = await db.clients.find_one({"id": purchase["vendor_id"]}, {"_id": 0, "name": 1})
            stock = await db.stocks.find_one({"id": purchase["stock_id"]}, {"_id": 0, "symbol": 1})
            
            for payment in purchase.get("payments", []):
                payment_date = payment.get("payment_date", "")
                
                # Date filter
                if start_date and payment_date < start_date:
                    continue
                if end_date and payment_date > end_date:
                    continue
                
                all_payments.append({
                    "id": payment.get("id", str(uuid.uuid4())),
                    "type": "vendor",
                    "direction": "sent",
                    "entity_id": purchase["vendor_id"],
                    "entity_name": vendor["name"] if vendor else "Unknown",
                    "reference_id": purchase["id"],
                    "reference_number": purchase.get("purchase_number", purchase["id"][:8].upper()),
                    "stock_symbol": stock["symbol"] if stock else "Unknown",
                    "amount": payment.get("amount", 0),
                    "payment_date": payment_date,
                    "notes": payment.get("notes", ""),
                    "proof_url": payment.get("proof_url"),
                    "recorded_by": payment.get("recorded_by_name", "")
                })
    
    # Sort by date descending
    all_payments.sort(key=lambda x: x.get("payment_date", ""), reverse=True)
    
    return all_payments


@router.get("/finance/summary")
async def get_finance_summary(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get finance summary statistics."""
    if not has_finance_access(current_user.get("role", 6)):
        raise HTTPException(status_code=403, detail="Only PE Desk, PE Manager, or Finance can access finance data")
    
    # Get all payments
    payments = await get_all_payments(None, start_date, end_date, current_user)
    
    # Calculate summaries
    total_received = sum(p["amount"] for p in payments if p["direction"] == "received")
    total_sent = sum(p["amount"] for p in payments if p["direction"] == "sent")
    
    client_payments_count = len([p for p in payments if p["type"] == "client"])
    vendor_payments_count = len([p for p in payments if p["type"] == "vendor"])
    
    # Get refund request stats
    pending_refunds = await db.refund_requests.find(
        {"status": {"$in": ["pending", "processing"]}},
        {"_id": 0}
    ).to_list(1000)
    
    completed_refunds = await db.refund_requests.find(
        {"status": "completed"},
        {"_id": 0}
    ).to_list(1000)
    
    # Get RP payment stats
    pending_rp_payments = await db.rp_payments.find(
        {"status": {"$in": ["pending", "processing"]}},
        {"_id": 0}
    ).to_list(1000)
    
    paid_rp_payments = await db.rp_payments.find(
        {"status": "paid"},
        {"_id": 0}
    ).to_list(1000)
    
    return {
        "total_received": total_received,
        "total_sent": total_sent,
        "net_flow": total_received - total_sent,
        "client_payments_count": client_payments_count,
        "vendor_payments_count": vendor_payments_count,
        "pending_refunds_count": len(pending_refunds),
        "pending_refunds_amount": sum(r.get("refund_amount", 0) for r in pending_refunds),
        "completed_refunds_count": len(completed_refunds),
        "completed_refunds_amount": sum(r.get("refund_amount", 0) for r in completed_refunds),
        "pending_rp_payments_count": len(pending_rp_payments),
        "pending_rp_payments_amount": sum(p.get("payment_amount", 0) for p in pending_rp_payments),
        "paid_rp_payments_count": len(paid_rp_payments),
        "paid_rp_payments_amount": sum(p.get("payment_amount", 0) for p in paid_rp_payments)
    }


@router.get("/finance/refund-requests")
async def get_refund_requests(
    status: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get all refund requests."""
    if not has_finance_access(current_user.get("role", 6)):
        raise HTTPException(status_code=403, detail="Only PE Desk, PE Manager, or Finance can access refund requests")
    
    query = {}
    if status:
        query["status"] = status
    
    refunds = await db.refund_requests.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return refunds


@router.get("/finance/refund-requests/{request_id}")
async def get_refund_request(request_id: str, current_user: dict = Depends(get_current_user)):
    """Get a specific refund request."""
    if not has_finance_access(current_user.get("role", 6)):
        raise HTTPException(status_code=403, detail="Only PE Desk, PE Manager, or Finance can access refund requests")
    
    refund = await db.refund_requests.find_one({"id": request_id}, {"_id": 0})
    if not refund:
        raise HTTPException(status_code=404, detail="Refund request not found")
    
    return refund


@router.put("/finance/refund-requests/{request_id}")
async def update_refund_request(
    request_id: str,
    update_data: RefundStatusUpdate,
    current_user: dict = Depends(get_current_user)
):
    """Update refund request status."""
    if not can_manage_finance(current_user.get("role", 6)):
        raise HTTPException(status_code=403, detail="Only PE Desk, PE Manager, or Finance can update refund requests")
    
    refund = await db.refund_requests.find_one({"id": request_id}, {"_id": 0})
    if not refund:
        raise HTTPException(status_code=404, detail="Refund request not found")
    
    update_fields = {
        "status": update_data.status,
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "updated_by": current_user["id"]
    }
    
    if update_data.notes:
        update_fields["notes"] = update_data.notes
    if update_data.reference_number:
        update_fields["reference_number"] = update_data.reference_number
    
    await db.refund_requests.update_one({"id": request_id}, {"$set": update_fields})
    
    return {"message": f"Refund request updated to {update_data.status}"}


@router.get("/finance/tcs-payments")
async def get_tcs_payments(
    financial_year: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get all vendor payments with TCS deducted."""
    if not has_finance_access(current_user.get("role", 6)):
        raise HTTPException(status_code=403, detail="Only PE Desk, PE Manager, or Finance can access TCS data")
    
    # Build query for payments with TCS
    query = {"tcs_applicable": True}
    
    if financial_year:
        query["financial_year"] = financial_year
    
    payments = await db.purchase_payments.find(query, {"_id": 0}).sort("payment_date", -1).to_list(1000)
    
    # Enrich with vendor details
    enriched_payments = []
    for payment in payments:
        vendor_id = payment.get("vendor_id")
        purchase_id = payment.get("purchase_id")
        
        vendor = await db.clients.find_one({"id": vendor_id}, {"_id": 0, "name": 1, "pan_number": 1})
        purchase = await db.purchases.find_one({"id": purchase_id}, {"_id": 0, "stock_symbol": 1, "stock_name": 1})
        
        enriched_payments.append({
            **payment,
            "vendor_name": vendor.get("name") if vendor else "Unknown",
            "vendor_pan": vendor.get("pan_number") if vendor else "Unknown",
            "stock_symbol": purchase.get("stock_symbol") if purchase else "Unknown",
            "stock_name": purchase.get("stock_name") if purchase else "Unknown"
        })
    
    return enriched_payments


@router.get("/finance/tcs-summary")
async def get_tcs_summary(
    financial_year: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get TCS summary grouped by vendor."""
    if not has_finance_access(current_user.get("role", 6)):
        raise HTTPException(status_code=403, detail="Only PE Desk, PE Manager, or Finance can access TCS data")
    
    # Get current FY if not specified
    if not financial_year:
        now = datetime.now()
        if now.month >= 4:
            financial_year = f"{now.year}-{now.year + 1}"
        else:
            financial_year = f"{now.year - 1}-{now.year}"
    
    # Aggregate TCS by vendor
    pipeline = [
        {"$match": {"tcs_applicable": True, "financial_year": financial_year}},
        {"$group": {
            "_id": "$vendor_id",
            "total_tcs": {"$sum": "$tcs_amount"},
            "total_payments": {"$sum": "$amount"},
            "payment_count": {"$sum": 1}
        }}
    ]
    
    results = await db.purchase_payments.aggregate(pipeline).to_list(1000)
    
    # Enrich with vendor details
    summary = []
    for result in results:
        vendor = await db.clients.find_one({"id": result["_id"]}, {"_id": 0, "name": 1, "pan_number": 1})
        summary.append({
            "vendor_id": result["_id"],
            "vendor_name": vendor.get("name") if vendor else "Unknown",
            "vendor_pan": vendor.get("pan_number") if vendor else "Unknown",
            "total_tcs": result["total_tcs"],
            "total_payments": result["total_payments"],
            "payment_count": result["payment_count"],
            "financial_year": financial_year
        })
    
    return {
        "financial_year": financial_year,
        "total_tcs_collected": sum(s["total_tcs"] for s in summary),
        "total_payments": sum(s["total_payments"] for s in summary),
        "vendor_count": len(summary),
        "vendors": sorted(summary, key=lambda x: x["total_tcs"], reverse=True)
    }


@router.put("/finance/refund-requests/{request_id}/bank-details")
async def update_refund_bank_details(
    request_id: str,
    bank_name: str,
    account_number: str,
    ifsc_code: str,
    account_holder_name: str,
    branch: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Update bank details for a refund request."""
    if not can_manage_finance(current_user.get("role", 6)):
        raise HTTPException(status_code=403, detail="Only PE Desk, PE Manager, or Finance can update refund requests")
    
    refund = await db.refund_requests.find_one({"id": request_id}, {"_id": 0})
    if not refund:
        raise HTTPException(status_code=404, detail="Refund request not found")
    
    bank_details = {
        "bank_name": bank_name,
        "account_number": account_number,
        "ifsc_code": ifsc_code,
        "account_holder_name": account_holder_name,
        "branch": branch
    }
    
    await db.refund_requests.update_one({"id": request_id}, {"$set": {"bank_details": bank_details}})
    
    return {"message": "Bank details updated successfully"}


# ============== RP Payments Endpoints ==============
class RPPaymentUpdate(BaseModel):
    status: str  # processing, paid
    payment_reference: Optional[str] = None
    notes: Optional[str] = None
    payment_date: Optional[str] = None


@router.get("/finance/rp-payments")
async def get_rp_payments(
    status: Optional[str] = None,
    rp_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get all RP payments for finance dashboard."""
    if not has_finance_access(current_user.get("role", 6)):
        raise HTTPException(status_code=403, detail="Only PE Desk, PE Manager, or Finance can access RP payments")
    
    query = {}
    if status:
        query["status"] = status
    if rp_id:
        query["referral_partner_id"] = rp_id
    
    payments = await db.rp_payments.find(query, {"_id": 0}).sort("created_at", -1).to_list(10000)
    return payments


@router.get("/finance/rp-payments/summary")
async def get_rp_payments_summary(
    current_user: dict = Depends(get_current_user)
):
    """Get RP payments summary statistics."""
    if not has_finance_access(current_user.get("role", 6)):
        raise HTTPException(status_code=403, detail="Only PE Desk, PE Manager, or Finance can access RP payments")
    
    all_payments = await db.rp_payments.find({}, {"_id": 0}).to_list(10000)
    
    pending = [p for p in all_payments if p.get("status") == "pending"]
    processing = [p for p in all_payments if p.get("status") == "processing"]
    paid = [p for p in all_payments if p.get("status") == "paid"]
    
    return {
        "pending_count": len(pending),
        "pending_amount": sum(p.get("payment_amount", 0) for p in pending),
        "processing_count": len(processing),
        "processing_amount": sum(p.get("payment_amount", 0) for p in processing),
        "paid_count": len(paid),
        "paid_amount": sum(p.get("payment_amount", 0) for p in paid),
        "total_count": len(all_payments),
        "total_amount": sum(p.get("payment_amount", 0) for p in all_payments)
    }


@router.put("/finance/rp-payments/{payment_id}")
async def update_rp_payment(
    payment_id: str,
    update_data: RPPaymentUpdate,
    current_user: dict = Depends(get_current_user)
):
    """Update RP payment status (mark as paid)."""
    if not can_manage_finance(current_user.get("role", 6)):
        raise HTTPException(status_code=403, detail="Only PE Desk, PE Manager, or Finance can update RP payments")
    
    payment = await db.rp_payments.find_one({"id": payment_id}, {"_id": 0})
    if not payment:
        raise HTTPException(status_code=404, detail="RP payment not found")
    
    update_fields = {
        "status": update_data.status,
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "updated_by": current_user["id"]
    }
    
    if update_data.payment_reference:
        update_fields["payment_reference"] = update_data.payment_reference
    if update_data.notes:
        update_fields["notes"] = update_data.notes
    if update_data.payment_date:
        update_fields["payment_date"] = update_data.payment_date
    elif update_data.status == "paid":
        update_fields["payment_date"] = datetime.now(timezone.utc).isoformat()
    
    await db.rp_payments.update_one({"id": payment_id}, {"$set": update_fields})
    
    # Also update the booking's RP payment status
    if payment.get("booking_id"):
        await db.bookings.update_one(
            {"id": payment["booking_id"]},
            {"$set": {"rp_payment_status": update_data.status}}
        )
    
    return {"message": f"RP payment updated to {update_data.status}"}


@router.get("/finance/employee-commissions")
async def get_employee_commissions(
    employee_id: Optional[str] = None,
    status: Optional[str] = None,  # pending, calculated, paid
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get employee commissions from confirmed bookings"""
    if not has_finance_access(current_user.get("role", 6)):
        raise HTTPException(status_code=403, detail="Only PE Desk, PE Manager, or Finance can access commission data")
    
    query = {
        "stock_transferred": True,
        "employee_commission_amount": {"$exists": True, "$gt": 0}
    }
    
    if employee_id:
        query["created_by"] = employee_id
    
    if status:
        query["employee_commission_status"] = status
    
    if start_date:
        query["created_at"] = {"$gte": start_date}
    if end_date:
        if "created_at" in query:
            query["created_at"]["$lte"] = end_date
        else:
            query["created_at"] = {"$lte": end_date}
    
    bookings = await db.bookings.find(query, {"_id": 0}).sort("created_at", -1).to_list(5000)
    
    # Get employee and stock details
    employee_ids = list(set(b.get("created_by") for b in bookings if b.get("created_by")))
    stock_ids = list(set(b.get("stock_id") for b in bookings if b.get("stock_id")))
    client_ids = list(set(b.get("client_id") for b in bookings if b.get("client_id")))
    
    employees = await db.users.find({"id": {"$in": employee_ids}}, {"_id": 0, "hashed_password": 0}).to_list(1000)
    stocks = await db.stocks.find({"id": {"$in": stock_ids}}, {"_id": 0}).to_list(1000)
    clients = await db.clients.find({"id": {"$in": client_ids}}, {"_id": 0}).to_list(1000)
    
    emp_map = {e["id"]: e for e in employees}
    stock_map = {s["id"]: s for s in stocks}
    client_map = {c["id"]: c for c in clients}
    
    result = []
    for booking in bookings:
        emp = emp_map.get(booking.get("created_by"), {})
        stock = stock_map.get(booking.get("stock_id"), {})
        client = client_map.get(booking.get("client_id"), {})
        
        result.append({
            "booking_id": booking.get("id"),
            "booking_number": booking.get("booking_number"),
            "employee_id": booking.get("created_by"),
            "employee_name": emp.get("name", "Unknown"),
            "client_name": client.get("name", "Unknown"),
            "stock_symbol": stock.get("symbol", "Unknown"),
            "quantity": booking.get("quantity", 0),
            "profit": round((booking.get("selling_price", 0) - booking.get("buying_price", 0)) * booking.get("quantity", 0), 2),
            "rp_share_percent": booking.get("rp_revenue_share_percent", 0) or 0,
            "employee_share_percent": booking.get("employee_revenue_share_percent", 100),
            "employee_commission_amount": booking.get("employee_commission_amount", 0),
            "rp_payment_amount": booking.get("rp_payment_amount", 0) or 0,
            "status": booking.get("employee_commission_status", "pending"),
            "transfer_date": booking.get("stock_transferred_at"),
            "created_at": booking.get("created_at")
        })
    
    return result


@router.get("/finance/employee-commissions/summary")
async def get_employee_commissions_summary(
    current_user: dict = Depends(get_current_user)
):
    """Get summary of employee commissions by employee"""
    if not has_finance_access(current_user.get("role", 6)):
        raise HTTPException(status_code=403, detail="Only PE Desk, PE Manager, or Finance can access commission data")
    
    bookings = await db.bookings.find(
        {
            "stock_transferred": True,
            "employee_commission_amount": {"$exists": True, "$gt": 0}
        },
        {"_id": 0}
    ).to_list(10000)
    
    # Aggregate by employee
    emp_stats = {}
    for booking in bookings:
        emp_id = booking.get("created_by")
        if not emp_id:
            continue
        
        if emp_id not in emp_stats:
            emp_stats[emp_id] = {
                "total_bookings": 0,
                "total_profit": 0,
                "total_commission": 0,
                "total_rp_share": 0,
                "pending_commission": 0,
                "calculated_commission": 0,
                "paid_commission": 0
            }
        
        profit = (booking.get("selling_price", 0) - booking.get("buying_price", 0)) * booking.get("quantity", 0)
        commission = booking.get("employee_commission_amount", 0)
        rp_amount = booking.get("rp_payment_amount", 0) or 0
        status = booking.get("employee_commission_status", "pending")
        
        emp_stats[emp_id]["total_bookings"] += 1
        emp_stats[emp_id]["total_profit"] += profit
        emp_stats[emp_id]["total_commission"] += commission
        emp_stats[emp_id]["total_rp_share"] += rp_amount
        
        if status == "pending":
            emp_stats[emp_id]["pending_commission"] += commission
        elif status == "calculated":
            emp_stats[emp_id]["calculated_commission"] += commission
        elif status == "paid":
            emp_stats[emp_id]["paid_commission"] += commission
    
    # Get employee names
    emp_ids = list(emp_stats.keys())
    employees = await db.users.find({"id": {"$in": emp_ids}}, {"_id": 0, "id": 1, "name": 1, "role": 1}).to_list(1000)
    emp_map = {e["id"]: e for e in employees}
    
    result = []
    for emp_id, stats in emp_stats.items():
        emp = emp_map.get(emp_id, {})
        result.append({
            "employee_id": emp_id,
            "employee_name": emp.get("name", "Unknown"),
            "role": emp.get("role", 5),
            **{k: round(v, 2) if isinstance(v, float) else v for k, v in stats.items()}
        })
    
    result.sort(key=lambda x: x["total_commission"], reverse=True)
    return result


@router.put("/finance/employee-commissions/{booking_id}/mark-paid")
async def mark_commission_paid(
    booking_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Mark an employee commission as paid"""
    if not can_manage_finance(current_user.get("role", 6)):
        raise HTTPException(status_code=403, detail="Only PE Desk can mark commissions as paid")
    
    booking = await db.bookings.find_one({"id": booking_id}, {"_id": 0})
    if not booking:
        raise HTTPException(status_code=404, detail="Booking not found")
    
    if not booking.get("employee_commission_amount"):
        raise HTTPException(status_code=400, detail="No commission calculated for this booking")
    
    await db.bookings.update_one(
        {"id": booking_id},
        {"$set": {
            "employee_commission_status": "paid",
            "employee_commission_paid_at": datetime.now(timezone.utc).isoformat(),
            "employee_commission_paid_by": current_user["id"]
        }}
    )
    
    return {"message": "Commission marked as paid"}


# ============== BP Payments Endpoints ==============
class BPPaymentUpdate(BaseModel):
    status: str  # processing, paid
    payment_reference: Optional[str] = None
    notes: Optional[str] = None
    payment_date: Optional[str] = None


@router.get("/finance/bp-payments")
async def get_bp_payments(
    status: Optional[str] = None,
    bp_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get all Business Partner payments for finance dashboard."""
    if not has_finance_access(current_user.get("role", 6)):
        raise HTTPException(status_code=403, detail="Only PE Desk, PE Manager, or Finance can access BP payments")
    
    # Get all bookings with BP revenue share
    query = {
        "is_bp_booking": True,
        "status": "completed"
    }
    if bp_id:
        query["business_partner_id"] = bp_id
    
    bookings = await db.bookings.find(query, {"_id": 0}).sort("created_at", -1).to_list(10000)
    
    # Get BP payments from dedicated collection
    bp_payments_query = {}
    if status:
        bp_payments_query["status"] = status
    if bp_id:
        bp_payments_query["business_partner_id"] = bp_id
    
    existing_payments = await db.bp_payments.find(bp_payments_query, {"_id": 0}).to_list(10000)
    existing_booking_ids = {p.get("booking_id") for p in existing_payments}
    
    # Get BP names
    bp_ids = list(set(b.get("business_partner_id") for b in bookings if b.get("business_partner_id")))
    bps = await db.business_partners.find({"id": {"$in": bp_ids}}, {"_id": 0, "id": 1, "name": 1, "email": 1, "pan_number": 1, "bank_details": 1}).to_list(1000)
    bp_map = {bp["id"]: bp for bp in bps}
    
    # Get client and stock names
    client_ids = list(set(b.get("client_id") for b in bookings if b.get("client_id")))
    stock_ids = list(set(b.get("stock_id") for b in bookings if b.get("stock_id")))
    
    clients = await db.clients.find({"id": {"$in": client_ids}}, {"_id": 0, "id": 1, "name": 1}).to_list(1000)
    stocks = await db.stocks.find({"id": {"$in": stock_ids}}, {"_id": 0, "id": 1, "symbol": 1}).to_list(1000)
    
    client_map = {c["id"]: c for c in clients}
    stock_map = {s["id"]: s for s in stocks}
    
    # Create payment records for new bookings
    payments_to_create = []
    for booking in bookings:
        if booking.get("id") not in existing_booking_ids:
            profit = (booking.get("selling_price", 0) - booking.get("buying_price", 0)) * booking.get("quantity", 0)
            bp_share = booking.get("bp_revenue_share_percent", 0)
            payment_amount = profit * (bp_share / 100)
            
            if payment_amount > 0:
                bp_info = bp_map.get(booking.get("business_partner_id"), {})
                payment_record = {
                    "id": str(uuid.uuid4()),
                    "booking_id": booking.get("id"),
                    "booking_number": booking.get("booking_number"),
                    "business_partner_id": booking.get("business_partner_id"),
                    "bp_name": bp_info.get("name", booking.get("bp_name", "Unknown")),
                    "bp_email": bp_info.get("email"),
                    "bp_pan": bp_info.get("pan_number"),
                    "client_name": client_map.get(booking.get("client_id"), {}).get("name", "Unknown"),
                    "stock_symbol": stock_map.get(booking.get("stock_id"), {}).get("symbol", "Unknown"),
                    "quantity": booking.get("quantity", 0),
                    "profit": round(profit, 2),
                    "bp_share_percent": bp_share,
                    "payment_amount": round(payment_amount, 2),
                    "status": "pending",
                    "payment_reference": None,
                    "payment_date": None,
                    "notes": None,
                    "bank_details": bp_info.get("bank_details"),
                    "created_at": datetime.now(timezone.utc).isoformat()
                }
                payments_to_create.append(payment_record)
    
    # Insert new payment records
    if payments_to_create:
        await db.bp_payments.insert_many(payments_to_create)
    
    # Fetch all BP payments again with status filter
    final_query = {}
    if status:
        final_query["status"] = status
    if bp_id:
        final_query["business_partner_id"] = bp_id
    
    all_payments = await db.bp_payments.find(final_query, {"_id": 0}).sort("created_at", -1).to_list(10000)
    return all_payments


@router.get("/finance/bp-payments/summary")
async def get_bp_payments_summary(
    current_user: dict = Depends(get_current_user)
):
    """Get BP payments summary statistics."""
    if not has_finance_access(current_user.get("role", 6)):
        raise HTTPException(status_code=403, detail="Only PE Desk, PE Manager, or Finance can access BP payments")
    
    all_payments = await db.bp_payments.find({}, {"_id": 0}).to_list(10000)
    
    pending = [p for p in all_payments if p.get("status") == "pending"]
    processing = [p for p in all_payments if p.get("status") == "processing"]
    paid = [p for p in all_payments if p.get("status") == "paid"]
    
    return {
        "pending_count": len(pending),
        "pending_amount": round(sum(p.get("payment_amount", 0) for p in pending), 2),
        "processing_count": len(processing),
        "processing_amount": round(sum(p.get("payment_amount", 0) for p in processing), 2),
        "paid_count": len(paid),
        "paid_amount": round(sum(p.get("payment_amount", 0) for p in paid), 2),
        "total_count": len(all_payments),
        "total_amount": round(sum(p.get("payment_amount", 0) for p in all_payments), 2)
    }


@router.put("/finance/bp-payments/{payment_id}")
async def update_bp_payment(
    payment_id: str,
    update_data: BPPaymentUpdate,
    current_user: dict = Depends(get_current_user)
):
    """Update a BP payment status."""
    if not can_manage_finance(current_user.get("role", 6)):
        raise HTTPException(status_code=403, detail="Only PE Desk or Finance can update BP payments")
    
    payment = await db.bp_payments.find_one({"id": payment_id}, {"_id": 0})
    if not payment:
        raise HTTPException(status_code=404, detail="BP payment not found")
    
    update_fields = {
        "status": update_data.status,
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "updated_by": current_user["id"]
    }
    
    if update_data.payment_reference:
        update_fields["payment_reference"] = update_data.payment_reference
    if update_data.notes:
        update_fields["notes"] = update_data.notes
    if update_data.payment_date:
        update_fields["payment_date"] = update_data.payment_date
    elif update_data.status == "paid":
        update_fields["payment_date"] = datetime.now(timezone.utc).isoformat()
    
    await db.bp_payments.update_one({"id": payment_id}, {"$set": update_fields})
    
    # Also update the booking's BP payment status
    if payment.get("booking_id"):
        await db.bookings.update_one(
            {"id": payment["booking_id"]},
            {"$set": {"bp_payment_status": update_data.status}}
        )
    
    return {"message": f"BP payment updated to {update_data.status}"}


@router.get("/finance/export/excel")
async def export_finance_excel(
    payment_type: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Export finance data to Excel."""
    if not has_finance_access(current_user.get("role", 6)):
        raise HTTPException(status_code=403, detail="Only PE Desk, PE Manager, or Finance can export finance data")
    
    payments = await get_all_payments(payment_type, start_date, end_date, current_user)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Payments Report"
    
    # Headers
    headers = ["Date", "Type", "Direction", "Entity", "Reference", "Stock", "Amount", "Notes", "Recorded By"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Data rows
    for row, payment in enumerate(payments, 2):
        ws.cell(row=row, column=1, value=payment.get("payment_date", ""))
        ws.cell(row=row, column=2, value=payment.get("type", "").title())
        ws.cell(row=row, column=3, value=payment.get("direction", "").title())
        ws.cell(row=row, column=4, value=payment.get("entity_name", ""))
        ws.cell(row=row, column=5, value=payment.get("reference_number", ""))
        ws.cell(row=row, column=6, value=payment.get("stock_symbol", ""))
        ws.cell(row=row, column=7, value=payment.get("amount", 0))
        ws.cell(row=row, column=8, value=payment.get("notes", ""))
        ws.cell(row=row, column=9, value=payment.get("recorded_by", ""))
    
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except (TypeError, AttributeError):
                pass
        ws.column_dimensions[column].width = min(max_length + 2, 50)
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"finance_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/finance/tcs-export")
async def export_tcs_report(
    financial_year: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """
    Export TCS (Tax Collected at Source) report to Excel.
    Report includes vendor-wise TCS collected for compliance filing.
    
    Format follows standard TCS filing requirements:
    - Vendor Name, PAN, Total Payments, TCS Amount, Payment Details
    """
    if not has_finance_access(current_user.get("role", 6)):
        raise HTTPException(status_code=403, detail="Only PE Desk, PE Manager, or Finance can export TCS data")
    
    # Get current FY if not specified
    if not financial_year:
        now = datetime.now()
        if now.month >= 4:
            financial_year = f"{now.year}-{now.year + 1}"
        else:
            financial_year = f"{now.year - 1}-{now.year}"
    
    # Get all TCS payments for the financial year
    query = {"tcs_applicable": True}
    if financial_year:
        query["financial_year"] = financial_year
    
    payments = await db.purchase_payments.find(query, {"_id": 0}).sort("payment_date", 1).to_list(10000)
    
    # Create workbook
    wb = Workbook()
    
    # ===== Sheet 1: TCS Summary by Vendor =====
    ws_summary = wb.active
    ws_summary.title = "TCS Summary"
    
    # Header styling
    header_fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")  # Amber
    header_font = Font(bold=True, color="000000")
    subheader_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")  # Light amber
    
    # Title row
    ws_summary.merge_cells('A1:H1')
    title_cell = ws_summary['A1']
    title_cell.value = f"TCS COMPLIANCE REPORT - FY {financial_year}"
    title_cell.font = Font(bold=True, size=14, color="000000")
    title_cell.alignment = Alignment(horizontal="center")
    title_cell.fill = PatternFill(start_color="FBBF24", end_color="FBBF24", fill_type="solid")
    
    # Generated on
    ws_summary.merge_cells('A2:H2')
    ws_summary['A2'].value = f"Generated on: {datetime.now().strftime('%d-%b-%Y %H:%M:%S')}"
    ws_summary['A2'].alignment = Alignment(horizontal="center")
    ws_summary['A2'].font = Font(italic=True)
    
    # TCS Rate info
    ws_summary.merge_cells('A3:H3')
    ws_summary['A3'].value = "TCS Rate: 0.1% under Section 194Q (applicable on payments exceeding ₹50 lakhs)"
    ws_summary['A3'].alignment = Alignment(horizontal="center")
    ws_summary['A3'].font = Font(italic=True, size=10)
    
    # Summary Headers (row 5)
    summary_headers = ["S.No.", "Vendor Name", "PAN Number", "Total Payments (₹)", "TCS Collected (₹)", "No. of Transactions", "First Payment Date", "Last Payment Date"]
    for col, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=5, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Aggregate by vendor
    vendor_data = {}
    for payment in payments:
        vendor_id = payment.get("vendor_id")
        if vendor_id not in vendor_data:
            vendor_data[vendor_id] = {
                "vendor_id": vendor_id,
                "total_payments": 0,
                "total_tcs": 0,
                "payment_count": 0,
                "first_date": payment.get("payment_date"),
                "last_date": payment.get("payment_date")
            }
        
        vendor_data[vendor_id]["total_payments"] += payment.get("amount", 0)
        vendor_data[vendor_id]["total_tcs"] += payment.get("tcs_amount", 0)
        vendor_data[vendor_id]["payment_count"] += 1
        
        # Track first and last dates
        pdate = payment.get("payment_date", "")
        if pdate < vendor_data[vendor_id]["first_date"]:
            vendor_data[vendor_id]["first_date"] = pdate
        if pdate > vendor_data[vendor_id]["last_date"]:
            vendor_data[vendor_id]["last_date"] = pdate
    
    # Fetch vendor details and write rows
    row = 6
    serial = 1
    total_payments_sum = 0
    total_tcs_sum = 0
    
    for vendor_id, data in sorted(vendor_data.items(), key=lambda x: x[1]["total_tcs"], reverse=True):
        vendor = await db.clients.find_one({"id": vendor_id}, {"_id": 0, "name": 1, "pan_number": 1})
        
        ws_summary.cell(row=row, column=1, value=serial)
        ws_summary.cell(row=row, column=2, value=vendor.get("name") if vendor else "Unknown")
        ws_summary.cell(row=row, column=3, value=vendor.get("pan_number") if vendor else "N/A")
        ws_summary.cell(row=row, column=4, value=round(data["total_payments"], 2))
        ws_summary.cell(row=row, column=5, value=round(data["total_tcs"], 2))
        ws_summary.cell(row=row, column=6, value=data["payment_count"])
        ws_summary.cell(row=row, column=7, value=data["first_date"][:10] if data["first_date"] else "")
        ws_summary.cell(row=row, column=8, value=data["last_date"][:10] if data["last_date"] else "")
        
        # Format numbers
        ws_summary.cell(row=row, column=4).number_format = '#,##0.00'
        ws_summary.cell(row=row, column=5).number_format = '#,##0.00'
        
        total_payments_sum += data["total_payments"]
        total_tcs_sum += data["total_tcs"]
        
        row += 1
        serial += 1
    
    # Total row
    total_row = row
    ws_summary.cell(row=total_row, column=1, value="")
    ws_summary.cell(row=total_row, column=2, value="TOTAL")
    ws_summary.cell(row=total_row, column=3, value="")
    ws_summary.cell(row=total_row, column=4, value=round(total_payments_sum, 2))
    ws_summary.cell(row=total_row, column=5, value=round(total_tcs_sum, 2))
    ws_summary.cell(row=total_row, column=6, value=len(payments))
    
    for col in range(1, 9):
        ws_summary.cell(row=total_row, column=col).fill = subheader_fill
        ws_summary.cell(row=total_row, column=col).font = Font(bold=True)
    ws_summary.cell(row=total_row, column=4).number_format = '#,##0.00'
    ws_summary.cell(row=total_row, column=5).number_format = '#,##0.00'
    
    # Column widths for summary
    summary_widths = [8, 35, 15, 22, 20, 18, 18, 18]
    for i, width in enumerate(summary_widths, 1):
        ws_summary.column_dimensions[chr(64 + i)].width = width
    
    # ===== Sheet 2: Detailed TCS Transactions =====
    ws_detail = wb.create_sheet(title="TCS Transactions")
    
    detail_headers = ["S.No.", "Payment Date", "Vendor Name", "Vendor PAN", "Stock", "Purchase Ref", 
                      "Payment Amount (₹)", "TCS Amount (₹)", "Net Payment (₹)", "FY Cumulative Before", "FY Cumulative After", "Reference #"]
    
    for col, header in enumerate(detail_headers, 1):
        cell = ws_detail.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Detail rows
    row = 2
    for serial, payment in enumerate(payments, 1):
        vendor_id = payment.get("vendor_id")
        purchase_id = payment.get("purchase_id")
        
        vendor = await db.clients.find_one({"id": vendor_id}, {"_id": 0, "name": 1, "pan_number": 1})
        purchase = await db.purchases.find_one({"id": purchase_id}, {"_id": 0, "purchase_number": 1, "stock_symbol": 1})
        
        ws_detail.cell(row=row, column=1, value=serial)
        ws_detail.cell(row=row, column=2, value=payment.get("payment_date", "")[:10] if payment.get("payment_date") else "")
        ws_detail.cell(row=row, column=3, value=vendor.get("name") if vendor else "Unknown")
        ws_detail.cell(row=row, column=4, value=vendor.get("pan_number") if vendor else "N/A")
        ws_detail.cell(row=row, column=5, value=purchase.get("stock_symbol") if purchase else "N/A")
        ws_detail.cell(row=row, column=6, value=purchase.get("purchase_number") if purchase else "N/A")
        ws_detail.cell(row=row, column=7, value=round(payment.get("amount", 0), 2))
        ws_detail.cell(row=row, column=8, value=round(payment.get("tcs_amount", 0), 2))
        ws_detail.cell(row=row, column=9, value=round(payment.get("net_payment", 0), 2))
        ws_detail.cell(row=row, column=10, value=round(payment.get("vendor_fy_cumulative_before", 0), 2))
        ws_detail.cell(row=row, column=11, value=round(payment.get("vendor_fy_cumulative_after", 0), 2))
        ws_detail.cell(row=row, column=12, value=payment.get("reference_number", ""))
        
        # Format numbers
        for c in [7, 8, 9, 10, 11]:
            ws_detail.cell(row=row, column=c).number_format = '#,##0.00'
        
        row += 1
    
    # Column widths for detail
    detail_widths = [8, 14, 35, 15, 12, 20, 18, 16, 16, 20, 20, 20]
    for i, width in enumerate(detail_widths, 1):
        ws_detail.column_dimensions[chr(64 + i) if i <= 26 else 'A' + chr(64 + i - 26)].width = width
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Parse FY for filename
    fy_short = financial_year.replace("-", "_") if financial_year else "current"
    filename = f"TCS_Report_FY{fy_short}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ==================== VENDOR TO PAY CALCULATIONS ====================

# Constants for calculations
TO_PAY_TCS_RATE = 0.001  # 0.1% TCS
TO_PAY_TCS_THRESHOLD = 5000000  # 50 lakhs threshold for TCS
TO_PAY_STAMP_DUTY_RATE = 0.00015  # 0.015% Stamp Duty


def get_current_fy_range() -> tuple:
    """Get Indian Financial Year (April to March) start and end dates"""
    today = datetime.now()
    year = today.year
    month = today.month
    
    if month >= 4:
        fy_start = f"{year}-04-01"
        fy_end = f"{year + 1}-03-31"
        fy_label = f"FY {year}-{str(year + 1)[2:]}"
    else:
        fy_start = f"{year - 1}-04-01"
        fy_end = f"{year}-03-31"
        fy_label = f"FY {year - 1}-{str(year)[2:]}"
    
    return fy_start, fy_end, fy_label


async def get_vendor_fy_cumulative(vendor_id: str, exclude_purchase_id: str = None) -> float:
    """Get total amount already paid to vendor in current FY (for TCS calculation)"""
    fy_start, fy_end, _ = get_current_fy_range()
    
    purchases = await db.purchases.find({"vendor_id": vendor_id}, {"_id": 0, "id": 1}).to_list(1000)
    purchase_ids = [p["id"] for p in purchases]
    
    if not purchase_ids:
        return 0.0
    
    if exclude_purchase_id and exclude_purchase_id in purchase_ids:
        purchase_ids = [pid for pid in purchase_ids if pid != exclude_purchase_id]
    
    if not purchase_ids:
        return 0.0
    
    payments = await db.purchase_payments.find({
        "purchase_id": {"$in": purchase_ids},
        "payment_date": {"$gte": fy_start, "$lte": fy_end},
        "status": {"$in": ["completed", "paid"]}
    }, {"_id": 0, "net_payment": 1, "amount": 1}).to_list(10000)
    
    total = sum(p.get("net_payment", p.get("amount", 0)) for p in payments)
    return total


@router.get("/finance/to-pay")
async def get_vendor_to_pay_list(
    status: Optional[str] = Query(None, description="Filter by status"),
    current_user: dict = Depends(get_current_user),
    _: None = Depends(require_permission("finance.view", "view vendor payments"))
):
    """
    Get list of purchases pending payment with full calculation breakdown.
    
    A = Gross Consideration = Purchase Price x Number of Shares
    B = TCS = 0.1% if FY cumulative payments exceed 50 lakhs
    C = Stamp Duty = 0.015% of Gross Consideration
    D = Net Payable = A - B - C (Amount to pay to vendor)
    """
    query = {}
    if status:
        query["status"] = status
    else:
        query["status"] = {"$in": ["pending", "received", "partial"]}
    
    purchases = await db.purchases.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    
    result = []
    fy_start, fy_end, fy_label = get_current_fy_range()
    
    for purchase in purchases:
        vendor_id = purchase.get("vendor_id")
        stock_id = purchase.get("stock_id")
        
        vendor = await db.clients.find_one({"id": vendor_id}, {"_id": 0, "name": 1, "pan_number": 1, "dp_id": 1})
        if not vendor:
            continue
        
        stock = await db.stocks.find_one({"id": stock_id}, {"_id": 0, "symbol": 1, "name": 1})
        
        price_per_unit = float(purchase.get("price_per_unit", 0))
        quantity = int(purchase.get("quantity", 0))
        gross_consideration = price_per_unit * quantity
        
        vendor_fy_paid = await get_vendor_fy_cumulative(vendor_id, purchase.get("id"))
        cumulative_after_this = vendor_fy_paid + gross_consideration
        tcs_applicable = cumulative_after_this > TO_PAY_TCS_THRESHOLD
        
        if tcs_applicable:
            if vendor_fy_paid >= TO_PAY_TCS_THRESHOLD:
                tcs_amount = gross_consideration * TO_PAY_TCS_RATE
            else:
                amount_exceeding = cumulative_after_this - TO_PAY_TCS_THRESHOLD
                tcs_amount = amount_exceeding * TO_PAY_TCS_RATE
        else:
            tcs_amount = 0.0
        
        stamp_duty = gross_consideration * TO_PAY_STAMP_DUTY_RATE
        net_payable = gross_consideration - tcs_amount - stamp_duty
        
        existing_payments = await db.purchase_payments.find(
            {"purchase_id": purchase.get("id"), "status": {"$in": ["completed", "paid"]}},
            {"_id": 0, "net_payment": 1, "amount": 1}
        ).to_list(100)
        already_paid = sum(p.get("net_payment", p.get("amount", 0)) for p in existing_payments)
        remaining_to_pay = max(0, net_payable - already_paid)
        
        result.append({
            "purchase_id": purchase.get("id"),
            "purchase_number": purchase.get("purchase_number"),
            "purchase_date": purchase.get("purchase_date") or purchase.get("created_at"),
            "status": purchase.get("status"),
            "vendor_id": vendor_id,
            "vendor_name": vendor.get("name"),
            "vendor_pan": vendor.get("pan_number"),
            "vendor_dp_id": vendor.get("dp_id"),
            "stock_id": stock_id,
            "stock_symbol": stock.get("symbol") if stock else purchase.get("stock_symbol", "N/A"),
            "stock_name": stock.get("name") if stock else "N/A",
            "quantity": quantity,
            "price_per_unit": round(price_per_unit, 2),
            "gross_consideration": round(gross_consideration, 2),
            "tcs_applicable": tcs_applicable,
            "tcs_rate": TO_PAY_TCS_RATE * 100 if tcs_applicable else 0,
            "tcs_amount": round(tcs_amount, 2),
            "stamp_duty_rate": TO_PAY_STAMP_DUTY_RATE * 100,
            "stamp_duty_amount": round(stamp_duty, 2),
            "net_payable": round(net_payable, 2),
            "already_paid": round(already_paid, 2),
            "remaining_to_pay": round(remaining_to_pay, 2),
            "financial_year": fy_label,
            "vendor_fy_cumulative_before": round(vendor_fy_paid, 2),
            "vendor_fy_cumulative_after": round(cumulative_after_this, 2),
            "tcs_threshold": TO_PAY_TCS_THRESHOLD,
            "tcs_threshold_exceeded": cumulative_after_this > TO_PAY_TCS_THRESHOLD
        })
    
    return {
        "financial_year": fy_label,
        "tcs_threshold": TO_PAY_TCS_THRESHOLD,
        "tcs_rate_percent": TO_PAY_TCS_RATE * 100,
        "stamp_duty_rate_percent": TO_PAY_STAMP_DUTY_RATE * 100,
        "to_pay_list": result,
        "summary": {
            "total_purchases": len(result),
            "total_gross_consideration": round(sum(r["gross_consideration"] for r in result), 2),
            "total_tcs": round(sum(r["tcs_amount"] for r in result), 2),
            "total_stamp_duty": round(sum(r["stamp_duty_amount"] for r in result), 2),
            "total_net_payable": round(sum(r["net_payable"] for r in result), 2),
            "total_remaining": round(sum(r["remaining_to_pay"] for r in result), 2)
        }
    }
