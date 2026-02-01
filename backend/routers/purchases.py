"""
Purchases Router
Handles purchase orders and vendor payments
"""
from typing import List, Optional
from datetime import datetime, timezone
from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Body
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
import uuid
import os
import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from database import db
from config import is_pe_level, UPLOAD_DIR
from models import Purchase, PurchaseCreate
from utils.auth import get_current_user
from services.audit_service import create_audit_log
from services.email_service import send_stock_transfer_request_email, send_email, get_email_template
from services.contract_note_service import create_and_save_vendor_contract_note

router = APIRouter(prefix="/purchases", tags=["Purchases"])

# TCS Constants
TCS_RATE = 0.001  # 0.1%
TCS_THRESHOLD = 5000000  # 50 lakhs


def get_indian_financial_year(date_str: str = None) -> tuple:
    """Get Indian Financial Year (April to March) start and end dates"""
    from datetime import datetime
    
    if date_str:
        date = datetime.fromisoformat(date_str.replace('Z', '+00:00')) if 'T' in date_str else datetime.strptime(date_str, '%Y-%m-%d')
    else:
        date = datetime.now()
    
    year = date.year
    month = date.month
    
    # Indian FY starts in April
    if month >= 4:  # April to December - FY is current year to next year
        fy_start = f"{year}-04-01"
        fy_end = f"{year + 1}-03-31"
    else:  # January to March - FY is previous year to current year
        fy_start = f"{year - 1}-04-01"
        fy_end = f"{year}-03-31"
    
    return fy_start, fy_end


async def get_vendor_fy_payments(vendor_id: str, payment_date: str = None) -> float:
    """Get total payments made to a vendor in the current Indian Financial Year"""
    fy_start, fy_end = get_indian_financial_year(payment_date)
    
    # Get all purchases for this vendor
    purchases = await db.purchases.find(
        {"vendor_id": vendor_id},
        {"_id": 0, "id": 1}
    ).to_list(1000)
    
    purchase_ids = [p["id"] for p in purchases]
    
    if not purchase_ids:
        return 0.0
    
    # Get all payments for these purchases within the financial year
    pipeline = [
        {
            "$match": {
                "purchase_id": {"$in": purchase_ids},
                "payment_date": {"$gte": fy_start, "$lte": fy_end}
            }
        },
        {
            "$group": {
                "_id": None,
                "total": {"$sum": "$amount"}
            }
        }
    ]
    
    result = await db.purchase_payments.aggregate(pipeline).to_list(1)
    
    return result[0]["total"] if result else 0.0


def calculate_tcs(payment_amount: float, cumulative_payments: float) -> dict:
    """Calculate TCS based on payment amount and cumulative payments"""
    # TCS applies only if cumulative payments exceed threshold
    amount_above_threshold = (cumulative_payments + payment_amount) - TCS_THRESHOLD
    
    if amount_above_threshold <= 0:
        # Total payments still under threshold
        return {
            "tcs_applicable": False,
            "tcs_amount": 0.0,
            "tcs_rate": TCS_RATE,
            "threshold": TCS_THRESHOLD,
            "cumulative_before": cumulative_payments,
            "cumulative_after": cumulative_payments + payment_amount,
            "amount_above_threshold": 0.0
        }
    
    # Calculate how much of this payment is above threshold
    if cumulative_payments >= TCS_THRESHOLD:
        # Already above threshold, TCS on entire payment
        tcs_applicable_amount = payment_amount
    else:
        # Partially above threshold
        tcs_applicable_amount = amount_above_threshold
    
    tcs_amount = round(tcs_applicable_amount * TCS_RATE, 2)
    
    return {
        "tcs_applicable": True,
        "tcs_amount": tcs_amount,
        "tcs_rate": TCS_RATE,
        "threshold": TCS_THRESHOLD,
        "cumulative_before": cumulative_payments,
        "cumulative_after": cumulative_payments + payment_amount,
        "amount_above_threshold": amount_above_threshold,
        "tcs_applicable_amount": tcs_applicable_amount
    }


# Pydantic model for payment request
class PaymentRequest(BaseModel):
    amount: float
    payment_date: str
    payment_mode: Optional[str] = "bank_transfer"
    reference_number: Optional[str] = None
    notes: Optional[str] = None
    proof_url: Optional[str] = None
    tcs_amount: Optional[float] = None  # Manual TCS override
    tcs_applicable: Optional[bool] = None  # Manual TCS flag


@router.post("", response_model=Purchase)
async def create_purchase(
    purchase_data: PurchaseCreate,
    current_user: dict = Depends(get_current_user)
):
    """Create a new purchase order"""
    user_role = current_user.get("role", 6)
    
    if not is_pe_level(user_role):
        raise HTTPException(status_code=403, detail="Only PE level can create purchases")
    
    # Validate vendor
    vendor = await db.clients.find_one(
        {"id": purchase_data.vendor_id, "is_vendor": True},
        {"_id": 0}
    )
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor not found")
    
    # Validate stock
    stock = await db.stocks.find_one({"id": purchase_data.stock_id}, {"_id": 0})
    if not stock:
        raise HTTPException(status_code=404, detail="Stock not found")
    
    purchase_id = str(uuid.uuid4())
    purchase_number = f"PO-{datetime.now(timezone.utc).strftime('%Y%m%d')}-{purchase_id[:8].upper()}"
    
    purchase_doc = {
        "id": purchase_id,
        "purchase_number": purchase_number,
        "vendor_id": purchase_data.vendor_id,
        "vendor_name": vendor.get("name"),
        "stock_id": purchase_data.stock_id,
        "stock_symbol": stock.get("symbol"),
        "quantity": purchase_data.quantity,
        "price_per_share": purchase_data.price_per_unit,
        "total_amount": purchase_data.quantity * purchase_data.price_per_unit,
        "status": "pending",
        "payment_status": "pending",
        "notes": purchase_data.notes,
        "created_by": current_user["id"],
        "created_by_name": current_user["name"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.purchases.insert_one(purchase_doc)
    
    await create_audit_log(
        action="PURCHASE_CREATE",
        entity_type="purchase",
        entity_id=purchase_id,
        user_id=current_user["id"],
        user_name=current_user["name"],
        user_role=user_role,
        entity_name=purchase_number
    )
    
    # Send email notification to vendor
    vendor_email = vendor.get("email")
    if vendor_email:
        template = await get_email_template("purchase_order_created")
        
        if template:
            # Format amounts with Indian locale
            formatted_price = f"₹{purchase_data.price_per_unit:,.2f}"
            formatted_total = f"₹{purchase_doc['total_amount']:,.2f}"
            formatted_quantity = f"{purchase_data.quantity:,}"
            
            subject = template["subject"]
            subject = subject.replace("{{stock_symbol}}", stock.get("symbol", ""))
            
            body = template["body"]
            body = body.replace("{{vendor_name}}", vendor.get("name", ""))
            body = body.replace("{{stock_symbol}}", stock.get("symbol", ""))
            body = body.replace("{{quantity}}", formatted_quantity)
            body = body.replace("{{price_per_unit}}", formatted_price)
            body = body.replace("{{total_amount}}", formatted_total)
            
            await send_email(
                to_email=vendor_email,
                subject=subject,
                body=body,
                template_key="purchase_order_created",
                variables={
                    "vendor_name": vendor.get("name"),
                    "stock_symbol": stock.get("symbol"),
                    "quantity": formatted_quantity,
                    "price_per_unit": formatted_price,
                    "total_amount": formatted_total
                },
                related_entity_type="purchase",
                related_entity_id=purchase_id
            )
    
    return purchase_doc


@router.get("", response_model=List[Purchase])
async def get_purchases(
    status: Optional[str] = None,
    vendor_id: Optional[str] = None,
    stock_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get all purchases with vendor and stock details"""
    query = {}
    if status:
        query["status"] = status
    if vendor_id:
        query["vendor_id"] = vendor_id
    if stock_id:
        query["stock_id"] = stock_id
    
    purchases = await db.purchases.find(query, {"_id": 0}).sort("created_at", -1).to_list(10000)
    
    if not purchases:
        return []
    
    # Enrich with vendor and stock details
    vendor_ids = list(set(p["vendor_id"] for p in purchases))
    stock_ids = list(set(p["stock_id"] for p in purchases))
    purchase_ids = [p["id"] for p in purchases]
    
    vendors = await db.clients.find({"id": {"$in": vendor_ids}}, {"_id": 0}).to_list(1000)
    stocks = await db.stocks.find({"id": {"$in": stock_ids}}, {"_id": 0}).to_list(1000)
    
    # Get all payments for these purchases
    all_payments = await db.purchase_payments.find(
        {"purchase_id": {"$in": purchase_ids}},
        {"_id": 0, "purchase_id": 1, "amount": 1}
    ).to_list(10000)
    
    # Calculate total paid per purchase
    payments_by_purchase = {}
    for payment in all_payments:
        pid = payment["purchase_id"]
        if pid not in payments_by_purchase:
            payments_by_purchase[pid] = 0
        payments_by_purchase[pid] += payment.get("amount", 0)
    
    vendor_map = {v["id"]: v for v in vendors}
    stock_map = {s["id"]: s for s in stocks}
    
    enriched_purchases = []
    for p in purchases:
        vendor = vendor_map.get(p["vendor_id"])
        stock = stock_map.get(p["stock_id"])
        
        # Add vendor_name and stock_symbol if not present
        p["vendor_name"] = p.get("vendor_name") or (vendor["name"] if vendor else "Unknown")
        p["stock_symbol"] = p.get("stock_symbol") or (stock["symbol"] if stock else "Unknown")
        
        # Calculate payment status
        total_paid = payments_by_purchase.get(p["id"], 0)
        total_amount = p.get("total_amount", 0)
        
        p["total_paid"] = total_paid
        
        if total_paid >= total_amount:
            p["payment_status"] = "completed"
        elif total_paid > 0:
            p["payment_status"] = "partial"
        else:
            p["payment_status"] = "pending"
        
        enriched_purchases.append(Purchase(**p))
    
    return enriched_purchases


@router.get("/{purchase_id}/payments")
async def get_purchase_payments(
    purchase_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Get payment tranches for a purchase"""
    purchase = await db.purchases.find_one({"id": purchase_id}, {"_id": 0})
    if not purchase:
        raise HTTPException(status_code=404, detail="Purchase not found")
    
    payments = await db.purchase_payments.find(
        {"purchase_id": purchase_id},
        {"_id": 0}
    ).sort("tranche_number", 1).to_list(100)
    
    return payments


@router.get("/{purchase_id}/tcs-preview")
async def get_tcs_preview(
    purchase_id: str,
    amount: float,
    payment_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get TCS calculation preview for a payment amount"""
    purchase = await db.purchases.find_one({"id": purchase_id}, {"_id": 0})
    if not purchase:
        raise HTTPException(status_code=404, detail="Purchase not found")
    
    vendor_id = purchase.get("vendor_id")
    vendor = await db.clients.find_one({"id": vendor_id, "is_vendor": True}, {"_id": 0, "name": 1, "pan_number": 1})
    
    # Get cumulative payments for this vendor in FY
    cumulative_payments = await get_vendor_fy_payments(vendor_id, payment_date)
    
    # Calculate TCS
    tcs_info = calculate_tcs(amount, cumulative_payments)
    
    # Get FY info
    fy_start, fy_end = get_indian_financial_year(payment_date)
    
    return {
        "vendor_name": vendor.get("name") if vendor else "Unknown",
        "vendor_pan": vendor.get("pan_number") if vendor else "Unknown",
        "payment_amount": amount,
        "financial_year": f"{fy_start[:4]}-{fy_end[:4]}",
        "fy_start": fy_start,
        "fy_end": fy_end,
        "tcs_rate_percent": TCS_RATE * 100,
        "tcs_threshold": TCS_THRESHOLD,
        "tcs_threshold_formatted": "₹50,00,000",
        **tcs_info,
        "net_payment": round(amount - tcs_info["tcs_amount"], 2) if tcs_info["tcs_applicable"] else amount
    }


@router.post("/{purchase_id}/payments")
async def add_purchase_payment(
    purchase_id: str,
    payment_data: PaymentRequest,
    current_user: dict = Depends(get_current_user)
):
    """Add a payment tranche to a purchase with TCS calculation"""
    user_role = current_user.get("role", 6)
    
    # Extract payment data from Pydantic model
    amount = payment_data.amount
    payment_date = payment_data.payment_date
    payment_mode = payment_data.payment_mode or "bank_transfer"
    reference_number = payment_data.reference_number
    notes = payment_data.notes
    proof_url = payment_data.proof_url
    manual_tcs_amount = payment_data.tcs_amount
    manual_tcs_applicable = payment_data.tcs_applicable
    
    if not is_pe_level(user_role):
        raise HTTPException(status_code=403, detail="Only PE level can add payments")
    
    purchase = await db.purchases.find_one({"id": purchase_id}, {"_id": 0})
    if not purchase:
        raise HTTPException(status_code=404, detail="Purchase not found")
    
    vendor_id = purchase.get("vendor_id")
    
    # Get existing payments for this purchase
    existing_payments = await db.purchase_payments.find(
        {"purchase_id": purchase_id},
        {"_id": 0}
    ).to_list(100)
    
    total_paid = sum(p.get("amount", 0) for p in existing_payments)
    remaining = purchase.get("total_amount", 0) - total_paid
    
    if amount > remaining:
        raise HTTPException(
            status_code=400,
            detail=f"Payment amount ({amount}) exceeds remaining balance ({remaining})"
        )
    
    # Calculate TCS
    cumulative_vendor_payments = await get_vendor_fy_payments(vendor_id, payment_date)
    tcs_info = calculate_tcs(amount, cumulative_vendor_payments)
    
    # Use manual TCS if provided, otherwise use calculated
    if manual_tcs_applicable is not None:
        tcs_applicable = manual_tcs_applicable
        tcs_amount = manual_tcs_amount if manual_tcs_amount is not None else tcs_info["tcs_amount"]
    else:
        tcs_applicable = tcs_info["tcs_applicable"]
        tcs_amount = tcs_info["tcs_amount"]
    
    tranche_number = len(existing_payments) + 1
    payment_id = str(uuid.uuid4())
    
    # Get Indian FY info
    fy_start, fy_end = get_indian_financial_year(payment_date)
    
    payment_doc = {
        "id": payment_id,
        "purchase_id": purchase_id,
        "vendor_id": vendor_id,
        "tranche_number": tranche_number,
        "amount": amount,
        "payment_date": payment_date,
        "payment_mode": payment_mode,
        "reference_number": reference_number,
        "notes": notes,
        "proof_url": proof_url,
        # TCS fields
        "tcs_applicable": tcs_applicable,
        "tcs_amount": tcs_amount if tcs_applicable else 0.0,
        "tcs_rate": TCS_RATE,
        "tcs_threshold": TCS_THRESHOLD,
        "vendor_fy_cumulative_before": cumulative_vendor_payments,
        "vendor_fy_cumulative_after": cumulative_vendor_payments + amount,
        "financial_year": f"{fy_start[:4]}-{fy_end[:4]}",
        # Net payment (amount - TCS)
        "net_payment": round(amount - (tcs_amount if tcs_applicable else 0), 2),
        "status": "completed",
        "created_by": current_user["id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.purchase_payments.insert_one(payment_doc)
    
    # Remove MongoDB _id from response
    payment_doc.pop("_id", None)
    
    # Update purchase payment status
    new_total_paid = total_paid + amount
    if new_total_paid >= purchase.get("total_amount", 0):
        await db.purchases.update_one(
            {"id": purchase_id},
            {"$set": {
                "payment_status": "completed",
                "dp_status": "receivable",
                "dp_receivable_at": datetime.now(timezone.utc).isoformat()
            }}
        )
        
        # Payment completed - send stock transfer request email to vendor
        vendor = await db.clients.find_one(
            {"id": purchase.get("vendor_id"), "is_vendor": True},
            {"_id": 0}
        )
        stock = await db.stocks.find_one({"id": purchase.get("stock_id")}, {"_id": 0})
        company_master = await db.company_master.find_one({"_id": "company_settings"}, {"_id": 0})
        
        if vendor and company_master:
            await send_stock_transfer_request_email(
                purchase_id=purchase_id,
                vendor=vendor,
                purchase=purchase,
                stock=stock,
                total_paid=new_total_paid,
                payment_date=payment_doc["created_at"],
                company_master=company_master,
                cc_email=current_user.get("email")
            )
    else:
        await db.purchases.update_one(
            {"id": purchase_id},
            {"$set": {"payment_status": "partial"}}
        )
    
    return payment_doc


@router.delete("/{purchase_id}")
async def delete_purchase(
    purchase_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Delete a purchase order (PE Desk only)"""
    user_role = current_user.get("role", 6)
    
    if user_role != 1:
        raise HTTPException(status_code=403, detail="Only PE Desk can delete purchases")
    
    purchase = await db.purchases.find_one({"id": purchase_id}, {"_id": 0})
    if not purchase:
        raise HTTPException(status_code=404, detail="Purchase not found")
    
    # Check if payments exist
    payments_count = await db.purchase_payments.count_documents({"purchase_id": purchase_id})
    if payments_count > 0:
        raise HTTPException(
            status_code=400,
            detail="Cannot delete purchase with existing payments. Delete payments first."
        )
    
    await db.purchases.delete_one({"id": purchase_id})
    
    await create_audit_log(
        action="PURCHASE_DELETE",
        entity_type="purchase",
        entity_id=purchase_id,
        user_id=current_user["id"],
        user_name=current_user["name"],
        user_role=user_role,
        entity_name=purchase.get("purchase_number", purchase_id)
    )
    
    return {"message": "Purchase deleted successfully"}


# ============== DP Receivable Endpoints ==============

@router.get("/dp-receivables")
async def get_dp_receivables(current_user: dict = Depends(get_current_user)):
    """Get all purchases with DP receivable status (PE Level only)"""
    user_role = current_user.get("role", 6)
    
    if not is_pe_level(user_role):
        raise HTTPException(status_code=403, detail="Only PE level can view DP receivables")
    
    # Get purchases with dp_status = "receivable"
    purchases = await db.purchases.find(
        {"dp_status": "receivable"},
        {"_id": 0}
    ).sort("dp_receivable_at", -1).to_list(1000)
    
    # Enrich with vendor and stock details
    for purchase in purchases:
        vendor = await db.clients.find_one(
            {"id": purchase.get("vendor_id")},
            {"_id": 0, "name": 1, "email": 1}
        )
        if vendor:
            purchase["vendor_name"] = vendor.get("name")
            purchase["vendor_email"] = vendor.get("email")
        
        stock = await db.stocks.find_one(
            {"id": purchase.get("stock_id")},
            {"_id": 0, "name": 1, "symbol": 1}
        )
        if stock:
            purchase["stock_name"] = stock.get("name")
            purchase["stock_symbol"] = stock.get("symbol")
    
    return purchases


@router.get("/dp-received")
async def get_dp_received(current_user: dict = Depends(get_current_user)):
    """Get all purchases with DP received status (PE Level only)"""
    user_role = current_user.get("role", 6)
    
    if not is_pe_level(user_role):
        raise HTTPException(status_code=403, detail="Only PE level can view DP received records")
    
    # Get purchases with dp_status = "received"
    purchases = await db.purchases.find(
        {"dp_status": "received"},
        {"_id": 0}
    ).sort("dp_received_at", -1).to_list(1000)
    
    # Enrich with vendor and stock details
    for purchase in purchases:
        vendor = await db.clients.find_one(
            {"id": purchase.get("vendor_id")},
            {"_id": 0, "name": 1, "email": 1}
        )
        if vendor:
            purchase["vendor_name"] = vendor.get("name")
            purchase["vendor_email"] = vendor.get("email")
        
        stock = await db.stocks.find_one(
            {"id": purchase.get("stock_id")},
            {"_id": 0, "name": 1, "symbol": 1}
        )
        if stock:
            purchase["stock_name"] = stock.get("name")
            purchase["stock_symbol"] = stock.get("symbol")
    
    return purchases


@router.put("/{purchase_id}/mark-dp-received")
async def mark_dp_received(
    purchase_id: str,
    dp_type: str,  # "NSDL" or "CDSL"
    current_user: dict = Depends(get_current_user)
):
    """Mark a purchase as DP received (PE Level only)
    
    Args:
        purchase_id: Purchase ID
        dp_type: Either "NSDL" or "CDSL"
    """
    user_role = current_user.get("role", 6)
    
    if not is_pe_level(user_role):
        raise HTTPException(status_code=403, detail="Only PE level can mark DP as received")
    
    if dp_type not in ["NSDL", "CDSL"]:
        raise HTTPException(status_code=400, detail="dp_type must be 'NSDL' or 'CDSL'")
    
    purchase = await db.purchases.find_one({"id": purchase_id}, {"_id": 0})
    if not purchase:
        raise HTTPException(status_code=404, detail="Purchase not found")
    
    if purchase.get("dp_status") != "receivable":
        raise HTTPException(status_code=400, detail="Purchase is not in receivable status")
    
    received_at = datetime.now(timezone.utc).isoformat()
    
    # Update purchase with received status
    await db.purchases.update_one(
        {"id": purchase_id},
        {"$set": {
            "dp_status": "received",
            "dp_type": dp_type,
            "dp_received_at": received_at,
            "dp_received_by": current_user["id"],
            "dp_received_by_name": current_user["name"]
        }}
    )
    
    # Update inventory - add received quantity and calculate weighted average price
    stock = await db.stocks.find_one({"id": purchase.get("stock_id")}, {"_id": 0})
    if stock:
        new_quantity = purchase.get("quantity", 0)
        new_total_value = purchase.get("total_amount", 0)
        new_price_per_share = new_total_value / new_quantity if new_quantity > 0 else 0
        
        # Update inventory with received quantity and weighted average
        inventory = await db.inventory.find_one({"stock_id": purchase.get("stock_id")}, {"_id": 0})
        if inventory:
            # Calculate new weighted average price
            old_quantity = inventory.get("available_quantity", 0)
            old_total_value = inventory.get("total_value", 0)
            
            # New totals
            combined_quantity = old_quantity + new_quantity
            combined_total_value = old_total_value + new_total_value
            
            # Weighted average price
            weighted_avg_price = combined_total_value / combined_quantity if combined_quantity > 0 else 0
            
            await db.inventory.update_one(
                {"stock_id": purchase.get("stock_id")},
                {"$set": {
                    "available_quantity": combined_quantity,
                    "total_value": combined_total_value,
                    "weighted_avg_price": round(weighted_avg_price, 2),
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }}
            )
        else:
            # Create new inventory record
            weighted_avg_price = new_price_per_share
            await db.inventory.insert_one({
                "id": str(uuid.uuid4()),
                "stock_id": purchase.get("stock_id"),
                "available_quantity": new_quantity,
                "blocked_quantity": 0,
                "total_value": new_total_value,
                "weighted_avg_price": round(weighted_avg_price, 2),
                "created_at": datetime.now(timezone.utc).isoformat()
            })
    
    await create_audit_log(
        action="DP_RECEIVED",
        entity_type="purchase",
        entity_id=purchase_id,
        user_id=current_user["id"],
        user_name=current_user["name"],
        user_role=user_role,
        details={
            "dp_type": dp_type,
            "quantity": purchase.get("quantity"),
            "stock_id": purchase.get("stock_id")
        },
        entity_name=purchase.get("purchase_number", purchase_id)
    )
    
    # Send email notification to vendor
    vendor = await db.clients.find_one(
        {"id": purchase.get("vendor_id"), "is_vendor": True},
        {"_id": 0}
    )
    
    if vendor and vendor.get("email"):
        template = await get_email_template("vendor_stock_received")
        
        if template:
            # Format the received date
            received_date = datetime.fromisoformat(received_at.replace('Z', '+00:00')).strftime('%d-%b-%Y %H:%M')
            
            # Format total amount with Indian locale
            total_amount = f"{purchase.get('total_amount', 0):,.2f}"
            
            subject = template["subject"]
            subject = subject.replace("{{stock_symbol}}", stock.get("symbol", "") if stock else "")
            subject = subject.replace("{{purchase_number}}", purchase.get("purchase_number", ""))
            
            body = template["body"]
            body = body.replace("{{vendor_name}}", vendor.get("name", ""))
            body = body.replace("{{purchase_number}}", purchase.get("purchase_number", ""))
            body = body.replace("{{stock_symbol}}", stock.get("symbol", "") if stock else "")
            body = body.replace("{{stock_name}}", stock.get("name", "") if stock else "")
            body = body.replace("{{isin_number}}", stock.get("isin", "") if stock else "")
            body = body.replace("{{quantity}}", f"{purchase.get('quantity', 0):,}")
            body = body.replace("{{dp_type}}", dp_type)
            body = body.replace("{{received_date}}", received_date)
            body = body.replace("{{total_amount}}", total_amount)
            
            await send_email(
                to_email=vendor.get("email"),
                subject=subject,
                body=body,
                template_key="vendor_stock_received",
                variables={
                    "vendor_name": vendor.get("name"),
                    "purchase_number": purchase.get("purchase_number"),
                    "stock_symbol": stock.get("symbol") if stock else "",
                    "stock_name": stock.get("name") if stock else "",
                    "isin_number": stock.get("isin") if stock else "",
                    "quantity": purchase.get("quantity"),
                    "dp_type": dp_type,
                    "received_date": received_date,
                    "total_amount": total_amount
                },
                related_entity_type="purchase",
                related_entity_id=purchase_id
            )
    
    # Generate Vendor Purchase Contract Note
    contract_note_doc = None
    try:
        contract_note_doc = await create_and_save_vendor_contract_note(
            purchase_id=purchase_id,
            user_id=current_user["id"],
            user_name=current_user["name"]
        )
        
        # Send contract note email to vendor
        if vendor and vendor.get("email") and contract_note_doc:
            cn_template = await get_email_template("contract_note")
            
            if cn_template:
                cn_subject = f"Purchase Contract Note - {contract_note_doc['contract_note_number']} | {stock.get('symbol', '') if stock else ''}"
                
                cn_body = f"""
                <div style="font-family: 'Segoe UI', Arial, sans-serif; max-width: 650px; margin: 0 auto; background: linear-gradient(135deg, #f0fdf4 0%, #dcfce7 100%); padding: 30px; border-radius: 16px;">
                    <div style="background: white; padding: 30px; border-radius: 12px; box-shadow: 0 4px 15px rgba(0,0,0,0.1);">
                        <div style="text-align: center; margin-bottom: 25px;">
                            <div style="background: linear-gradient(135deg, #10b981 0%, #059669 100%); color: white; padding: 15px; border-radius: 12px; display: inline-block;">
                                <h2 style="margin: 0; font-size: 20px;">📄 Purchase Contract Note</h2>
                            </div>
                        </div>
                        
                        <p style="font-size: 16px; color: #374151;">Dear <strong>{vendor.get('name', 'Vendor')}</strong>,</p>
                        
                        <p style="color: #4b5563; line-height: 1.7;">
                            Please find attached the Purchase Contract Note for the stock transfer completed today.
                        </p>
                        
                        <div style="background: #f8fafc; border-radius: 12px; padding: 20px; margin: 25px 0; border: 1px solid #e2e8f0;">
                            <table style="width: 100%; border-collapse: collapse;">
                                <tr>
                                    <td style="padding: 10px 0; color: #6b7280; width: 40%;">Contract Note No:</td>
                                    <td style="padding: 10px 0; color: #111827; font-weight: 600;">{contract_note_doc['contract_note_number']}</td>
                                </tr>
                                <tr style="background: #f1f5f9;">
                                    <td style="padding: 10px; color: #6b7280;">Purchase Order:</td>
                                    <td style="padding: 10px; color: #111827; font-weight: 600;">{purchase.get('purchase_number', '')}</td>
                                </tr>
                                <tr>
                                    <td style="padding: 10px 0; color: #6b7280;">Stock:</td>
                                    <td style="padding: 10px 0; color: #111827;">{stock.get('symbol', '')} - {stock.get('name', '')}</td>
                                </tr>
                                <tr style="background: #f1f5f9;">
                                    <td style="padding: 10px; color: #6b7280;">Quantity:</td>
                                    <td style="padding: 10px; color: #111827; font-weight: 700;">{purchase.get('quantity', 0):,} shares</td>
                                </tr>
                                <tr>
                                    <td style="padding: 10px 0; color: #6b7280;">Total Amount:</td>
                                    <td style="padding: 10px 0; color: #059669; font-weight: 700;">₹{purchase.get('total_amount', 0):,.2f}</td>
                                </tr>
                            </table>
                        </div>
                        
                        <div style="background: #d1fae5; border-radius: 12px; padding: 15px; margin: 25px 0; border: 1px solid #6ee7b7;">
                            <p style="color: #065f46; margin: 0; font-size: 14px;">
                                <strong>📎 Attachment:</strong> The Purchase Contract Note PDF is attached to this email.
                            </p>
                        </div>
                        
                        <p style="color: #4b5563; line-height: 1.7;">
                            Please download and retain the attached contract note for your records.
                        </p>
                        
                        <p>Best regards,<br><strong>SMIFS Private Equity Team</strong></p>
                    </div>
                </div>
                """
                
                # Read PDF file for attachment
                attachment_content = None
                pdf_path = f"/app{contract_note_doc['pdf_url']}"
                try:
                    with open(pdf_path, 'rb') as f:
                        attachment_content = f.read()
                except Exception as e:
                    print(f"Warning: Could not read PDF file for attachment: {e}")
                
                attachments = []
                if attachment_content:
                    attachments = [{
                        'filename': f"Purchase_Contract_Note_{contract_note_doc['contract_note_number'].replace('/', '_')}.pdf",
                        'content': attachment_content,
                        'content_type': 'application/pdf'
                    }]
                
                await send_email(
                    to_email=vendor.get("email"),
                    subject=cn_subject,
                    body=cn_body,
                    template_key="vendor_contract_note",
                    variables={
                        "vendor_name": vendor.get("name"),
                        "contract_note_number": contract_note_doc['contract_note_number'],
                        "purchase_number": purchase.get("purchase_number"),
                        "stock_symbol": stock.get("symbol") if stock else "",
                        "quantity": purchase.get("quantity"),
                        "total_amount": purchase.get("total_amount")
                    },
                    related_entity_type="vendor_contract_note",
                    related_entity_id=contract_note_doc["id"],
                    attachments=attachments if attachments else None
                )
                
                # Update contract note email sent status
                await db.vendor_contract_notes.update_one(
                    {"id": contract_note_doc["id"]},
                    {"$set": {"email_sent": True, "email_sent_at": datetime.now(timezone.utc).isoformat()}}
                )
    except Exception as e:
        # Log error but don't fail the DP received operation
        print(f"Warning: Failed to generate vendor contract note: {str(e)}")
    
    return {
        "message": f"Stock received via {dp_type} and inventory updated",
        "dp_type": dp_type,
        "quantity": purchase.get("quantity"),
        "received_at": received_at,
        "email_sent": bool(vendor and vendor.get("email")),
        "contract_note_generated": bool(contract_note_doc),
        "contract_note_number": contract_note_doc.get("contract_note_number") if contract_note_doc else None,
        "contract_note_pdf_url": contract_note_doc.get("pdf_url") if contract_note_doc else None
    }



@router.get("/dp-receivables/export")
async def export_dp_receivables_excel(
    status: str = "all",  # "receivable", "received", or "all"
    current_user: dict = Depends(get_current_user)
):
    """Export DP receivables data to Excel"""
    user_role = current_user.get("role", 6)
    
    if not is_pe_level(user_role):
        raise HTTPException(status_code=403, detail="Only PE level can export DP data")
    
    # Build query based on status
    query = {}
    if status == "receivable":
        query["dp_status"] = "receivable"
    elif status == "received":
        query["dp_status"] = "received"
    else:
        query["dp_status"] = {"$in": ["receivable", "received"]}
    
    # Get purchases
    purchases = await db.purchases.find(query, {"_id": 0}).to_list(10000)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "DP Receivables"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        "Purchase #", "Vendor Name", "Vendor DP ID", "Vendor PAN",
        "Stock Symbol", "Stock Name", "ISIN", "Quantity",
        "Total Amount", "Status", "DP Type", "Received Date"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Data rows
    row = 2
    for purchase in purchases:
        # Get vendor details
        vendor = await db.clients.find_one(
            {"id": purchase.get("vendor_id")},
            {"_id": 0, "name": 1, "dp_id": 1, "pan": 1}
        )
        
        # Get stock details
        stock = await db.stocks.find_one(
            {"id": purchase.get("stock_id")},
            {"_id": 0, "name": 1, "symbol": 1, "isin": 1}
        )
        
        data = [
            purchase.get("purchase_number", ""),
            vendor.get("name", "") if vendor else "",
            vendor.get("dp_id", "") if vendor else "",
            vendor.get("pan", "") if vendor else "",
            stock.get("symbol", "") if stock else "",
            stock.get("name", "") if stock else "",
            stock.get("isin", "") if stock else "",
            purchase.get("quantity", 0),
            purchase.get("total_amount", 0),
            purchase.get("dp_status", "").upper(),
            purchase.get("dp_type", ""),
            purchase.get("dp_received_at", "")[:10] if purchase.get("dp_received_at") else ""
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin_border
            if col in [8, 9]:  # Quantity and Amount columns
                cell.alignment = Alignment(horizontal="right")
        
        row += 1
    
    # Adjust column widths
    column_widths = [15, 25, 20, 15, 15, 30, 20, 12, 15, 12, 10, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"dp_receivables_{status}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

