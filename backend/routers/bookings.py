"""
Bookings Router with High-Concurrency Support

This router handles all booking operations with proper locking and atomic updates
to prevent race conditions during simultaneous booking requests.
"""
from fastapi import APIRouter, HTTPException, Depends, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import List, Optional
from datetime import datetime, timezone, timedelta
import uuid
import os
import io
import logging

logger = logging.getLogger(__name__)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from database import db
from config import is_viewer, check_viewer_restriction, ROLES
from models import BookingCreate, Booking, BookingWithDetails
from utils.auth import get_current_user
from services.permission_service import (
    check_permission,
    check_permission as check_dynamic_permission,
    has_permission,
    require_permission,
    is_pe_level_dynamic
)
from services.notification_service import notify_roles, create_notification
from services.audit_service import create_audit_log
from services.email_service import send_templated_email, send_payment_request_email
from services.inventory_service import (
    update_inventory,
    check_and_reserve_inventory,
    release_inventory_reservation,
    check_stock_availability
)
from utils.demo_isolation import is_demo_user, add_demo_filter, mark_as_demo, require_demo_access

router = APIRouter(tags=["Bookings"])

# Lock for booking number generation
import asyncio
_booking_number_lock = asyncio.Lock()


# Helper function for backward compatibility with is_pe_level
def is_pe_level(role: int) -> bool:
    """Check if role is PE level (PE Desk or PE Manager). For sync contexts."""
    return role in [1, 2]


def is_pe_desk_only(role: int) -> bool:
    """Check if role is PE Desk only. For sync contexts."""
    return role == 1


async def generate_booking_number() -> str:
    """Generate a unique booking number in format BK-YYYY-NNNNN using atomic counter."""
    year = datetime.now().year
    
    async with _booking_number_lock:
        # Use atomic findAndModify to get next sequence
        counter = await db.counters.find_one_and_update(
            {"_id": f"booking_{year}"},
            {"$inc": {"seq": 1}},
            upsert=True,
            return_document=True
        )
        
        seq_num = counter.get("seq", 1)
        return f"BK-{year}-{seq_num:05d}"


def get_client_emails(client: dict) -> list:
    """Get all email addresses for a client"""
    emails = []
    if client.get("email"):
        emails.append(client["email"])
    if client.get("secondary_email"):
        emails.append(client["secondary_email"])
    if client.get("tertiary_email"):
        emails.append(client["tertiary_email"])
    return emails


@router.get("/bookings/check-client-rp-conflict/{client_id}")
async def check_client_rp_conflict(
    client_id: str, 
    current_user: dict = Depends(get_current_user),
    _: None = Depends(require_permission("bookings.create", "check client RP conflict"))
):
    """
    Check if a client is also registered as a Referral Partner.
    Returns conflict info if found, null otherwise.
    STRICT RULE: A Client cannot be an RP and vice versa.
    """
    client = await db.clients.find_one({"id": client_id}, {"_id": 0, "pan_number": 1, "name": 1, "otc_ucc": 1})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    
    # Check if client's PAN matches any RP
    client_pan = client.get("pan_number", "").upper()
    matching_rp = await db.referral_partners.find_one(
        {"pan_number": client_pan},
        {"_id": 0, "rp_code": 1, "name": 1, "id": 1}
    )
    
    if matching_rp:
        return {
            "has_conflict": True,
            "client_name": client.get("name"),
            "client_ucc": client.get("otc_ucc"),
            "rp_code": matching_rp.get("rp_code"),
            "rp_name": matching_rp.get("name"),
            "rp_id": matching_rp.get("id"),
            "message": f"This client ({client.get('name')}) is also registered as RP {matching_rp.get('rp_code')}. RP revenue share will be automatically set to 0%."
        }
    
    return {"has_conflict": False}


@router.post("/bookings", response_model=Booking)
async def create_booking(
    booking_data: BookingCreate, 
    current_user: dict = Depends(get_current_user),
    _: None = Depends(require_permission("bookings.create", "create bookings"))
):
    """
    Create a new booking with high-concurrency safety.
    
    This endpoint uses atomic inventory operations to prevent race conditions
    when multiple users try to book the same stock simultaneously.
    
    Booking permissions:
    - PE Desk/PE Manager: Can book for any client
    - All other roles: Can only book for clients mapped to them
    - Viewers: Cannot create bookings (read-only access)
    """
    user_role = current_user.get("role", 7)
    user_id = current_user.get("id")
    
    # Viewer restriction - Viewers cannot create bookings
    check_viewer_restriction(user_role, "create bookings")
    
    is_business_partner = user_role == 6 or current_user.get("is_bp", False)
    bp_info = None
    
    # Permission check - Business Partners can create bookings
    if is_business_partner:
        await check_permission(current_user, "bookings.create", "create bookings")
    elif not is_pe_level(user_role):
        await check_permission(current_user, "bookings.create", "create bookings")
    else:
        await check_permission(current_user, "bookings.create", "create bookings")
    
    # If Business Partner, get their BP profile for revenue share
    bp_info = None
    bp_revenue_override = None
    bp_override_requires_approval = False
    
    if is_business_partner:
        bp_id = current_user.get("user_id") or current_user.get("id")
        bp_info = await db.business_partners.find_one({"id": bp_id}, {"_id": 0})
        if not bp_info:
            raise HTTPException(status_code=404, detail="Business Partner profile not found")
        if not bp_info.get("is_active", True):
            raise HTTPException(status_code=403, detail="Your Business Partner account is inactive")
        # BPs cannot select RPs - clear any RP data
        booking_data.referral_partner_id = None
        booking_data.rp_revenue_share_percent = None
        
        # Handle BP revenue share override
        if booking_data.bp_revenue_share_override is not None:
            default_bp_share = bp_info.get("revenue_share_percent", 0)
            override_share = booking_data.bp_revenue_share_override
            
            # Override can only be LOWER than the default (giving away less revenue)
            if override_share > default_bp_share:
                raise HTTPException(
                    status_code=400, 
                    detail=f"Revenue share override ({override_share}%) cannot be higher than your default ({default_bp_share}%)"
                )
            if override_share < 0:
                raise HTTPException(status_code=400, detail="Revenue share cannot be negative")
            if override_share > 100:
                raise HTTPException(status_code=400, detail="Revenue share cannot exceed 100%")
            
            bp_revenue_override = override_share
            bp_override_requires_approval = True
    
    # Partners Desk creating booking for a BP - check for linked BP
    elif user_role == 5:  # Partners Desk
        # Check if booking has a BP linked via the client
        # Partners Desk can also apply overrides for their linked BPs
        linked_employee = await db.users.find_one({"id": current_user["id"]}, {"_id": 0, "linked_bp_id": 1})
        if linked_employee and linked_employee.get("linked_bp_id"):
            bp_info = await db.business_partners.find_one({"id": linked_employee["linked_bp_id"]}, {"_id": 0})
            if bp_info and bp_info.get("is_active", True):
                # Handle override from Partners Desk
                if booking_data.bp_revenue_share_override is not None:
                    default_bp_share = bp_info.get("revenue_share_percent", 0)
                    override_share = booking_data.bp_revenue_share_override
                    
                    if override_share > default_bp_share:
                        raise HTTPException(
                            status_code=400, 
                            detail=f"Revenue share override ({override_share}%) cannot be higher than BP default ({default_bp_share}%)"
                        )
                    if override_share < 0 or override_share > 100:
                        raise HTTPException(status_code=400, detail="Revenue share must be between 0 and 100")
                    
                    bp_revenue_override = override_share
                    bp_override_requires_approval = True
    
    # Verify client exists and is active
    client = await db.clients.find_one({"id": booking_data.client_id}, {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    
    # Auto-assign creator's RP if no RP is specified and user is not a Business Partner
    if not is_business_partner and not booking_data.referral_partner_id:
        # Find RP created by the current user (default to creator's RP)
        creator_rp = await db.referral_partners.find_one(
            {"created_by": current_user["id"], "is_active": True, "approval_status": "approved"},
            {"_id": 0, "id": 1, "rp_code": 1, "name": 1}
        )
        if creator_rp:
            booking_data.referral_partner_id = creator_rp["id"]
            # Set default RP revenue share if not specified
            if not booking_data.rp_revenue_share_percent:
                booking_data.rp_revenue_share_percent = 30.0  # Default 30%
    
    # Check client approval status
    if client.get("approval_status") != "approved":
        raise HTTPException(
            status_code=400, 
            detail=f"Client must be approved by PE Desk before creating bookings. Current status: {client.get('approval_status', 'pending')}"
        )
    
    # Check if client is active
    if not client.get("is_active", True):
        raise HTTPException(status_code=400, detail="Client is inactive and cannot be used for bookings")
    
    # Check if client is suspended
    if client.get("is_suspended"):
        raise HTTPException(
            status_code=400, 
            detail=f"Client is suspended and cannot be used for bookings. Reason: {client.get('suspension_reason', 'No reason provided')}"
        )
    
    # BOOKING RESTRICTION: Non-PE users can only book for clients mapped to them OR clients they created
    if not is_pe_level(user_role):
        mapped_employee_id = client.get("mapped_employee_id")
        created_by = client.get("created_by")
        
        # User can book if they're mapped to the client OR they created the client
        if mapped_employee_id != user_id and created_by != user_id:
            raise HTTPException(
                status_code=403, 
                detail=f"You can only create bookings for clients mapped to you or clients you created. This client is mapped to someone else."
            )
    
    # Verify stock exists
    stock = await db.stocks.find_one({"id": booking_data.stock_id}, {"_id": 0})
    if not stock:
        raise HTTPException(status_code=404, detail="Stock not found")
    
    # Check if stock is blocked
    if stock.get("exchange") == "Blocked IPO/RTA":
        raise HTTPException(
            status_code=400, 
            detail="This stock is blocked (IPO/RTA) and not available for booking"
        )
    
    # ============== DUPLICATE BOOKING PREVENTION ==============
    # Check for duplicate bookings with same client, stock, quantity, and booking date
    # Also check for recent bookings (within 30 seconds) to prevent accidental double-clicks
    
    # Check 1: Exact duplicate (same client, stock, quantity, date) that is not voided
    exact_duplicate = await db.bookings.find_one({
        "client_id": booking_data.client_id,
        "stock_id": booking_data.stock_id,
        "quantity": booking_data.quantity,
        "booking_date": booking_data.booking_date,
        "is_voided": {"$ne": True}
    }, {"_id": 0, "booking_number": 1, "created_at": 1})
    
    if exact_duplicate:
        raise HTTPException(
            status_code=400,
            detail=f"Duplicate booking detected! A booking with the same client, stock, quantity, and date already exists: {exact_duplicate.get('booking_number', 'Unknown')}"
        )
    
    # Check 2: Recent duplicate (same client, stock within last 30 seconds) - prevents double-clicks
    thirty_seconds_ago = (datetime.now(timezone.utc) - timedelta(seconds=30)).isoformat()
    recent_duplicate = await db.bookings.find_one({
        "client_id": booking_data.client_id,
        "stock_id": booking_data.stock_id,
        "created_by": current_user["id"],
        "created_at": {"$gte": thirty_seconds_ago},
        "is_voided": {"$ne": True}
    }, {"_id": 0, "booking_number": 1})
    
    if recent_duplicate:
        raise HTTPException(
            status_code=400,
            detail=f"A booking for this client and stock was just created: {recent_duplicate.get('booking_number', 'Unknown')}. Please wait before creating another booking."
        )
    
    # Check 3: Same client + stock with pending approval (prevents multiple pending bookings for same stock)
    pending_same_stock = await db.bookings.find_one({
        "client_id": booking_data.client_id,
        "stock_id": booking_data.stock_id,
        "approval_status": "pending",
        "is_voided": {"$ne": True}
    }, {"_id": 0, "booking_number": 1})
    
    if pending_same_stock:
        raise HTTPException(
            status_code=400,
            detail=f"There is already a pending booking for this client and stock: {pending_same_stock.get('booking_number', 'Unknown')}. Please approve or reject it first."
        )
    # ============== END DUPLICATE BOOKING PREVENTION ==============
    
    # HIGH-CONCURRENCY: Atomic inventory check
    is_available, available_qty, weighted_avg = await check_stock_availability(
        booking_data.stock_id, 
        booking_data.quantity
    )
    
    if not is_available:
        raise HTTPException(
            status_code=400,
            detail=f"Insufficient inventory. Available: {available_qty}, Requested: {booking_data.quantity}"
        )
    
    # Validate selling price
    if booking_data.selling_price is None or booking_data.selling_price <= 0:
        raise HTTPException(
            status_code=400,
            detail="Selling price is required and must be greater than 0"
        )
    
    # Get Landing Price (LP) from inventory - this is what non-PE users see and use
    inventory = await db.inventory.find_one({"stock_id": booking_data.stock_id}, {"_id": 0})
    landing_price = inventory.get("landing_price") if inventory else None
    if landing_price is None or landing_price <= 0:
        landing_price = weighted_avg  # Default to WAP if LP not set
    
    # Determine buying price - Use Landing Price (LP) instead of WAP
    if user_role in [5, 4]:  # Employee/Manager must use landing price
        buying_price = landing_price
    else:
        # PE Desk/PE Manager can override but default to LP
        buying_price = booking_data.buying_price if booking_data.buying_price else landing_price
    
    if buying_price is None or buying_price <= 0:
        raise HTTPException(
            status_code=400,
            detail="Landing price is required and must be greater than 0"
        )
    
    # Validate RP revenue share percentage cap at 30%
    if booking_data.rp_revenue_share_percent is not None:
        if booking_data.rp_revenue_share_percent > 30:
            raise HTTPException(
                status_code=400,
                detail="Referral Partner revenue share cannot exceed 30%"
            )
        if booking_data.rp_revenue_share_percent < 0:
            raise HTTPException(
                status_code=400,
                detail="Referral Partner revenue share cannot be negative"
            )
    
    # STRICT RULE: A Client cannot be an RP - Auto-zero RP share if client's PAN matches any RP
    rp_share_auto_zeroed = False
    if booking_data.referral_partner_id and booking_data.rp_revenue_share_percent:
        # Check if client's PAN matches any RP's PAN
        client_pan = client.get("pan_number", "").upper()
        matching_rp = await db.referral_partners.find_one(
            {"pan_number": client_pan},
            {"_id": 0, "rp_code": 1, "name": 1}
        )
        if matching_rp:
            # Client is also an RP - auto-zero the RP revenue share
            booking_data.rp_revenue_share_percent = 0
            booking_data.referral_partner_id = None  # Remove RP assignment
            rp_share_auto_zeroed = True
    
    # Get RP details if specified
    rp_code = None
    rp_name = None
    if booking_data.referral_partner_id:
        rp = await db.referral_partners.find_one({"id": booking_data.referral_partner_id}, {"_id": 0})
        if rp:
            # Additional check: Verify the selected RP is not the same person as the client
            rp_pan = rp.get("pan_number", "").upper()
            client_pan = client.get("pan_number", "").upper()
            if rp_pan == client_pan:
                # Same person - zero out RP share
                booking_data.rp_revenue_share_percent = 0
                booking_data.referral_partner_id = None
                rp_share_auto_zeroed = True
            else:
                rp_code = rp.get("rp_code")
                rp_name = rp.get("name")
        else:
            raise HTTPException(status_code=404, detail="Referral Partner not found")
    
    # Generate booking identifiers
    booking_id = str(uuid.uuid4())
    booking_number = await generate_booking_number()
    confirmation_token = str(uuid.uuid4())
    
    # Check if loss booking
    is_loss_booking = booking_data.selling_price < buying_price
    loss_approval_status = "pending" if is_loss_booking else "not_required"
    
    # Determine if BP booking - BP revenue share takes precedence
    is_bp_booking = is_business_partner and bp_info is not None
    bp_revenue_share = bp_info.get("revenue_share_percent", 0) if bp_info else 0
    
    # Use override if provided, otherwise use default BP share
    effective_bp_share = bp_revenue_override if bp_revenue_override is not None else bp_revenue_share
    
    # Create booking document
    booking_doc = {
        "id": booking_id,
        "booking_number": booking_number,
        "client_id": booking_data.client_id,
        "stock_id": booking_data.stock_id,
        "quantity": booking_data.quantity,
        "buying_price": buying_price,  # This is now Landing Price (LP)
        "weighted_avg_price": weighted_avg,  # Store actual WAP for PE Desk HIT report
        "landing_price": landing_price,  # Store LP explicitly
        "selling_price": booking_data.selling_price,
        "booking_date": booking_data.booking_date,
        "status": booking_data.status,
        "approval_status": "pending",  # Requires PE approval
        "approved_by": None,
        "approved_at": None,
        "booking_type": booking_data.booking_type,
        "insider_form_uploaded": booking_data.insider_form_uploaded,
        "insider_form_path": None,
        # Business Partner fields (takes precedence over RP)
        "business_partner_id": bp_info.get("id") if is_bp_booking else None,
        "bp_name": bp_info.get("name") if is_bp_booking else None,
        "bp_revenue_share_percent": effective_bp_share if is_bp_booking else None,
        "is_bp_booking": is_bp_booking,
        # BP Revenue Share Override tracking
        "bp_revenue_share_override": bp_revenue_override,
        "bp_original_revenue_share": bp_revenue_share if bp_revenue_override is not None else None,
        "bp_override_approval_status": "pending" if bp_override_requires_approval else "not_required",
        "bp_override_approved_by": None,
        "bp_override_approved_at": None,
        "bp_override_rejection_reason": None,
        # Referral Partner fields (may be zeroed if client=RP or if BP booking)
        "referral_partner_id": None if is_bp_booking else (booking_data.referral_partner_id if not rp_share_auto_zeroed else None),
        "rp_code": None if is_bp_booking else (rp_code if not rp_share_auto_zeroed else None),
        "rp_name": None if is_bp_booking else (rp_name if not rp_share_auto_zeroed else None),
        "rp_revenue_share_percent": 0 if is_bp_booking else (booking_data.rp_revenue_share_percent if not rp_share_auto_zeroed else 0),
        "rp_share_auto_zeroed": rp_share_auto_zeroed,  # Flag for audit trail
        # Employee Revenue Share - calculated based on RP/BP allocation
        "base_employee_share_percent": 100.0,  # Full share before deduction
        "employee_revenue_share_percent": 100.0 - effective_bp_share if is_bp_booking else (100.0 if rp_share_auto_zeroed else 100.0 - (booking_data.rp_revenue_share_percent or 0)),
        "employee_commission_amount": None,  # Calculated when stock transfer confirmed
        "employee_commission_status": "pending",
        # Client confirmation
        "client_confirmation_status": "pending",
        "client_confirmation_token": confirmation_token,
        "client_confirmed_at": None,
        "client_denial_reason": None,
        "is_loss_booking": is_loss_booking,
        "loss_approval_status": loss_approval_status,
        "loss_approved_by": None,
        "loss_approved_at": None,
        "notes": f"[BP Booking: {bp_info.get('name')} @ {bp_revenue_share}%] {booking_data.notes or ''}" if is_bp_booking else (f"[AUTO: RP share zeroed - Client is also an RP] {booking_data.notes or ''}" if rp_share_auto_zeroed else booking_data.notes),
        "user_id": current_user["id"],
        "created_by": current_user["id"],
        "created_by_name": current_user["name"],
        "created_by_role": user_role,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    # Mark as demo data if created by demo user
    booking_doc = mark_as_demo(booking_doc, current_user)
    
    # Insert booking
    await db.bookings.insert_one(booking_doc)
    
    # Create audit log
    await create_audit_log(
        action="BOOKING_CREATE",
        entity_type="booking",
        entity_id=booking_id,
        user_id=current_user["id"],
        user_name=current_user["name"],
        user_role=user_role,
        entity_name=f"{stock['symbol']} - {client['name']} ({booking_number})",
        details={
            "client_id": booking_data.client_id,
            "client_name": client["name"],
            "stock_id": booking_data.stock_id,
            "stock_symbol": stock["symbol"],
            "quantity": booking_data.quantity,
            "buying_price": buying_price,
            "selling_price": booking_data.selling_price,
            "rp_share": booking_data.rp_revenue_share_percent,
            "employee_share": 100.0 - (booking_data.rp_revenue_share_percent or 0)
        }
    )
    
    # Send email notification
    if client.get("email"):
        await send_templated_email(
            "booking_created",
            client["email"],
            {
                "client_name": client["name"],
                "booking_number": booking_number,
                "stock_symbol": stock["symbol"],
                "stock_name": stock["name"],
                "quantity": booking_data.quantity
            },
            cc_email=current_user.get("email")
        )
    
    # Send WhatsApp activity alerts
    try:
        from services.activity_alerts import notify_booking_created
        
        # Get RP and BP info for alerts
        rp_info = None
        if booking_doc.get("referral_partner_id"):
            rp_info = await db.referral_partners.find_one({"id": booking_doc["referral_partner_id"]}, {"_id": 0})
        
        bp_data = None
        if is_bp_booking:
            bp_data = await db.business_partners.find_one({"id": bp_info.get("id")}, {"_id": 0})
        
        await notify_booking_created(
            booking_doc,
            client,
            rp=rp_info,
            bp=bp_data
        )
    except Exception as e:
        print(f"Activity alert error (non-critical): {e}")
    
    # Real-time notification to PE Desk
    await notify_roles(
        [1],
        "booking_pending",
        "New Booking Pending Approval",
        f"Booking {booking_number} for '{client['name']}' - {stock['symbol']} x {booking_data.quantity} awaiting PE Desk approval",
        {"booking_id": booking_id, "booking_number": booking_number, "client_name": client["name"], "stock_symbol": stock["symbol"]}
    )
    
    return Booking(**{k: v for k, v in booking_doc.items() if k not in ["user_id", "created_by_name"]})


@router.put("/bookings/{booking_id}/approve")
async def approve_booking(
    booking_id: str,
    approve: bool = True,
    current_user: dict = Depends(get_current_user),
    _: None = Depends(require_permission("bookings.approve", "approve bookings"))
):
    """
    Approve or reject a booking with atomic inventory update.
    
    When approved, inventory is atomically updated to prevent race conditions.
    """
    user_role = current_user.get("role", 6)
    
    booking = await db.bookings.find_one({"id": booking_id}, {"_id": 0})
    if not booking:
        raise HTTPException(status_code=404, detail="Booking not found")
    
    if booking.get("approval_status") != "pending":
        raise HTTPException(status_code=400, detail="Booking already processed")
    
    if approve:
        # HIGH-CONCURRENCY: Atomic inventory reservation
        success, message, inventory = await check_and_reserve_inventory(
            booking["stock_id"],
            booking["quantity"],
            booking_id
        )
        
        if not success:
            raise HTTPException(status_code=400, detail=message)
    
    # Update booking status
    update_data = {
        "approval_status": "approved" if approve else "rejected",
        "approved_by": current_user["id"],
        "approved_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.bookings.update_one({"id": booking_id}, {"$set": update_data})
    
    # Get related entities for notifications
    client = await db.clients.find_one({"id": booking["client_id"]}, {"_id": 0})
    stock = await db.stocks.find_one({"id": booking["stock_id"]}, {"_id": 0})
    creator = await db.users.find_one({"id": booking["created_by"]}, {"_id": 0})
    booking_number = booking.get("booking_number", booking_id[:8].upper())
    
    if approve:
        # Create audit log
        await create_audit_log(
            action="BOOKING_APPROVE",
            entity_type="booking",
            entity_id=booking_id,
            user_id=current_user["id"],
            user_name=current_user["name"],
            user_role=user_role,
            details={"stock_id": booking["stock_id"], "quantity": booking["quantity"]}
        )
        
        # Send confirmation email to client
        is_loss_pending = booking.get("is_loss_booking") and booking.get("loss_approval_status") == "pending"
        client_emails = get_client_emails(client) if client else []
        primary_email = client_emails[0] if client_emails else None
        additional_emails = client_emails[1:] if len(client_emails) > 1 else None
        
        if client and primary_email:
            confirmation_token = booking.get("client_confirmation_token")
            
            # Get frontend URL - Priority order:
            # 1. custom_domain from Company Master (UI configurable)
            # 2. FRONTEND_URL env variable
            # 3. REACT_APP_BACKEND_URL env variable
            # 4. Default fallback
            company_master = await db.company_master.find_one({"_id": "company_settings"})
            frontend_url = (
                (company_master.get("custom_domain") if company_master else None) or
                os.environ.get('FRONTEND_URL') or 
                os.environ.get('REACT_APP_BACKEND_URL', 'https://quote-refresh.preview.emergentagent.com')
            )
            # Remove trailing slash if present
            frontend_url = frontend_url.rstrip('/') if frontend_url else frontend_url
            
            if is_loss_pending:
                await send_templated_email(
                    "booking_pending_loss_review",
                    primary_email,
                    {
                        "client_name": client["name"],
                        "booking_number": booking_number,
                        "stock_symbol": stock["symbol"] if stock else "N/A"
                    },
                    cc_email=creator.get("email") if creator else None,
                    additional_emails=additional_emails
                )
            else:
                await send_templated_email(
                    "booking_confirmation_request",
                    primary_email,
                    {
                        "client_name": client["name"],
                        "booking_number": booking_number,
                        "otc_ucc": client.get("otc_ucc", "N/A"),
                        "stock_symbol": stock["symbol"] if stock else "N/A",
                        "stock_name": stock["name"] if stock else "",
                        "quantity": booking["quantity"],
                        "selling_price": f"{booking.get('selling_price', 0):,.2f}",
                        "total_value": f"{(booking.get('selling_price', 0) * booking.get('quantity', 0)):,.2f}",
                        "approved_by": current_user["name"],
                        "accept_url": f"{frontend_url}/booking-confirm/{booking_id}/{confirmation_token}/accept",
                        "deny_url": f"{frontend_url}/booking-confirm/{booking_id}/{confirmation_token}/deny"
                    },
                    cc_email=creator.get("email") if creator else None,
                    additional_emails=additional_emails
                )
            
            # Send payment request email with bank details and company documents
            company_master = await db.company_master.find_one({"_id": "company_settings"}, {"_id": 0})
            if company_master:
                await send_payment_request_email(
                    booking_id=booking_id,
                    client=client,
                    stock=stock,
                    booking=booking,
                    company_master=company_master,
                    approved_by=current_user["name"],
                    cc_email=creator.get("email") if creator else None
                )
        
        # Notify booking creator
        if booking.get("created_by"):
            await create_notification(
                booking["created_by"],
                "booking_approved",
                "Booking Approved - Awaiting Client Confirmation",
                f"Your booking {booking_number} for '{stock['symbol'] if stock else 'N/A'}' has been approved.",
                {"booking_id": booking_id, "stock_symbol": stock['symbol'] if stock else None}
            )
        
        # Send WhatsApp activity alert for approval
        try:
            from services.activity_alerts import notify_booking_approved
            if client and client.get("id"):
                await notify_booking_approved(booking, client["id"])
        except Exception as e:
            print(f"Activity alert error (non-critical): {e}")
            
    else:
        # Rejection - create audit log
        await create_audit_log(
            action="BOOKING_REJECT",
            entity_type="booking",
            entity_id=booking_id,
            user_id=current_user["id"],
            user_name=current_user["name"],
            user_role=user_role,
            details={"stock_id": booking["stock_id"], "quantity": booking["quantity"]}
        )
        
        # Notify booking creator
        if booking.get("created_by"):
            await create_notification(
                booking["created_by"],
                "booking_rejected",
                "Booking Rejected",
                f"Your booking for '{stock['symbol'] if stock else 'N/A'}' has been rejected",
                {"booking_id": booking_id, "stock_symbol": stock['symbol'] if stock else None}
            )
        
        # Send WhatsApp activity alert for rejection
        try:
            from services.activity_alerts import notify_booking_rejected
            if client and client.get("id"):
                await notify_booking_rejected(booking, client["id"], "Booking did not meet approval criteria")
        except Exception as e:
            print(f"Activity alert error (non-critical): {e}")
    
    return {"message": f"Booking {'approved' if approve else 'rejected'} successfully"}


@router.put("/bookings/{booking_id}/void")
async def void_booking(
    booking_id: str,
    reason: str = Query(..., description="Reason for voiding the booking"),
    current_user: dict = Depends(get_current_user),
    _: None = Depends(require_permission("bookings.delete", "void bookings"))
):
    """
    Void a booking with atomic inventory release.
    
    This releases the blocked inventory back to available.
    If the booking had payments, creates a refund request.
    """
    user_role = current_user.get("role", 6)
    
    booking = await db.bookings.find_one({"id": booking_id}, {"_id": 0})
    if not booking:
        raise HTTPException(status_code=404, detail="Booking not found")
    
    if booking.get("is_voided"):
        raise HTTPException(status_code=400, detail="Booking is already voided")
    
    if booking.get("stock_transferred"):
        raise HTTPException(status_code=400, detail="Cannot void - stock has already been transferred")
    
    # Release inventory if booking was approved
    if booking.get("approval_status") == "approved":
        success, message = await release_inventory_reservation(
            booking["stock_id"],
            booking["quantity"],
            booking_id
        )
        if not success:
            # Log warning but don't fail - manual reconciliation may be needed
            import logging
            logging.warning(f"Failed to release inventory for voided booking {booking_id}: {message}")
    
    # Update booking
    update_data = {
        "is_voided": True,
        "void_reason": reason,
        "voided_by": current_user["id"],
        "voided_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.bookings.update_one({"id": booking_id}, {"$set": update_data})
    
    # Create refund request if there are payments
    payments = booking.get("payments", [])
    total_paid = sum(p.get("amount", 0) for p in payments)
    
    refund_request_id = None
    if total_paid > 0:
        client = await db.clients.find_one({"id": booking["client_id"]}, {"_id": 0})
        stock = await db.stocks.find_one({"id": booking["stock_id"]}, {"_id": 0})
        
        # Get client's primary bank account for refund
        bank_details = {}
        if client and client.get("bank_accounts") and len(client["bank_accounts"]) > 0:
            primary_bank = client["bank_accounts"][0]
            bank_details = {
                "bank_name": primary_bank.get("bank_name", ""),
                "account_number": primary_bank.get("account_number", ""),
                "ifsc_code": primary_bank.get("ifsc_code", ""),
                "account_holder_name": primary_bank.get("account_holder_name", ""),
                "branch": primary_bank.get("branch", "")
            }
        
        refund_request_id = str(uuid.uuid4())
        refund_request = {
            "id": refund_request_id,
            "booking_id": booking_id,
            "booking_number": booking.get("booking_number", ""),
            "client_id": booking["client_id"],
            "client_name": client["name"] if client else "Unknown",
            "client_email": client.get("email", "") if client else "",
            "stock_id": booking["stock_id"],
            "stock_symbol": stock["symbol"] if stock else "Unknown",
            "quantity": booking["quantity"],
            "refund_amount": total_paid,
            "bank_details": bank_details,
            "void_reason": reason,
            "voided_by": current_user["id"],
            "voided_by_name": current_user["name"],
            "status": "pending",
            "reference_number": None,
            "notes": None,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": None,
            "updated_by": None
        }
        
        await db.refund_requests.insert_one(refund_request)
        
        # Send refund notification email
        if client and client.get("email"):
            await send_templated_email(
                "refund_request_created",
                client["email"],
                {
                    "client_name": client["name"],
                    "booking_number": booking.get("booking_number", ""),
                    "stock_symbol": stock["symbol"] if stock else "Unknown",
                    "refund_amount": f"{total_paid:,.2f}",
                    "void_reason": reason
                }
            )
    
    # Send voided booking notification email to client
    client = await db.clients.find_one({"id": booking["client_id"]}, {"_id": 0})
    stock = await db.stocks.find_one({"id": booking["stock_id"]}, {"_id": 0})
    
    if client and client.get("email"):
        try:
            await send_templated_email(
                "booking_voided",
                client["email"],
                {
                    "client_name": client.get("name", "Valued Customer"),
                    "booking_number": booking.get("booking_number", booking_id),
                    "stock_symbol": stock.get("symbol", "Unknown") if stock else "Unknown",
                    "stock_name": stock.get("name", "") if stock else "",
                    "quantity": str(booking.get("quantity", 0)),
                    "booking_date": booking.get("created_at", "")[:10] if booking.get("created_at") else "",
                    "voided_date": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
                    "voided_by": current_user.get("name", "PE Desk"),
                    "void_reason": reason
                }
            )
        except Exception as e:
            import logging
            logging.error(f"Failed to send voided booking email: {e}")
    
    # Create audit log
    await create_audit_log(
        action="BOOKING_VOID",
        entity_type="booking",
        entity_id=booking_id,
        user_id=current_user["id"],
        user_name=current_user["name"],
        user_role=user_role,
        details={
            "reason": reason,
            "refund_amount": total_paid if total_paid > 0 else None,
            "refund_request_id": refund_request_id
        }
    )
    
    return {
        "message": "Booking voided successfully",
        "refund_request_created": total_paid > 0,
        "refund_amount": total_paid if total_paid > 0 else None,
        "refund_request_id": refund_request_id
    }


@router.get("/bookings", response_model=List[BookingWithDetails])
async def get_bookings(
    status: Optional[str] = None,
    approval_status: Optional[str] = None,
    client_id: Optional[str] = None,
    stock_id: Optional[str] = None,
    search: Optional[str] = None,
    skip: int = Query(0, ge=0),
    limit: int = Query(1000, ge=1, le=10000),
    current_user: dict = Depends(get_current_user),
    _: None = Depends(require_permission("bookings.view", "view bookings"))
):
    """Get all bookings with optional filters based on hierarchy.
    
    Args:
        search: Search query to filter by booking number, client name, PAN, or stock symbol
        skip: Number of records to skip (for pagination)
        limit: Maximum number of records to return
    """
    from services.hierarchy_service import get_team_user_ids
    
    query = {}
    user_role = current_user.get("role", 5)
    user_id = current_user.get("id")
    hierarchy_level = current_user.get("hierarchy_level", 1)
    
    # PE Level, Finance (6), and Viewer (8) see all bookings
    # Employee (7) only sees their own bookings
    if is_pe_level(user_role) or user_role in [6, 8]:
        pass  # No creator filter
    else:
        # Use hierarchy to determine visibility - includes Employee role
        team_ids = await get_team_user_ids(user_id, include_self=True)
        
        # Get client IDs mapped to this user or their team
        mapped_clients = await db.clients.find(
            {"mapped_employee_id": {"$in": team_ids}},
            {"_id": 0, "id": 1}
        ).to_list(10000)
        mapped_client_ids = [c["id"] for c in mapped_clients]
        
        # Show bookings that are either:
        # 1. Created by user/team, OR
        # 2. For clients mapped to user/team
        query["$or"] = [
            {"created_by": {"$in": team_ids}},
            {"client_id": {"$in": mapped_client_ids}} if mapped_client_ids else {"_never_match_": True}
        ]
    
    # Apply filters
    if status:
        query["status"] = status
    if approval_status:
        query["approval_status"] = approval_status
    if client_id:
        query["client_id"] = client_id
    if stock_id:
        query["stock_id"] = stock_id
    
    # Server-side search filter
    if search:
        search_regex = {"$regex": search, "$options": "i"}
        # First, find matching client IDs by name or PAN
        matching_clients = await db.clients.find(
            {"$or": [{"name": search_regex}, {"pan_number": search_regex}]},
            {"_id": 0, "id": 1}
        ).to_list(1000)
        client_ids = [c["id"] for c in matching_clients]
        
        # Find matching stock IDs by symbol
        matching_stocks = await db.stocks.find(
            {"symbol": search_regex},
            {"_id": 0, "id": 1}
        ).to_list(1000)
        stock_ids = [s["id"] for s in matching_stocks]
        
        # Build search conditions
        search_conditions = [
            {"booking_number": search_regex},
            {"created_by_name": search_regex}
        ]
        if client_ids:
            search_conditions.append({"client_id": {"$in": client_ids}})
        if stock_ids:
            search_conditions.append({"stock_id": {"$in": stock_ids}})
        
        # Combine search with existing query using $and
        if "$or" in query:
            # Already have visibility filter, combine with $and
            query = {"$and": [{"$or": query["$or"]}, {"$or": search_conditions}]}
        else:
            query["$or"] = search_conditions
    
    # CRITICAL: Add demo data isolation filter
    # Demo users only see demo data, live users don't see demo data
    query = add_demo_filter(query, current_user)
    
    bookings = await db.bookings.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    
    # Enrich with client and stock details
    result = []
    for booking in bookings:
        client = await db.clients.find_one({"id": booking["client_id"]}, {"_id": 0})
        stock = await db.stocks.find_one({"id": booking["stock_id"]}, {"_id": 0})
        
        # Calculate total amount
        quantity = booking.get("quantity", 0)
        selling_price = booking.get("selling_price", 0)
        buying_price = booking.get("buying_price", 0)
        total_amount = quantity * selling_price
        
        # Calculate payment info
        payments = booking.get("payments", [])
        total_paid = round(sum(p.get("amount", 0) for p in payments), 2)
        
        booking_with_details = {
            **booking,
            "client_name": client["name"] if client else "Unknown",
            "client_email": client.get("email") if client else None,
            "stock_symbol": stock["symbol"] if stock else "Unknown",
            "stock_name": stock["name"] if stock else "Unknown",
            "total_amount": round(total_amount, 2),
            "profit_loss": round((selling_price - buying_price) * quantity, 2),
            "total_paid": total_paid,
            "payment_status": "paid" if total_paid >= total_amount else ("partial" if total_paid > 0 else "pending")
        }
        result.append(BookingWithDetails(**booking_with_details))
    
    return result


# ============== DP Transfer Endpoints (Client Stock Transfers) ==============
# NOTE: These routes MUST be defined before /bookings/{booking_id} to avoid path conflicts

@router.get("/bookings/dp-ready")
async def get_dp_ready_bookings(
    current_user: dict = Depends(get_current_user),
    _: None = Depends(require_permission("dp.view_receivables", "view DP ready bookings"))
):
    """Get all bookings with DP ready status (fully paid, ready to transfer)"""
    user_role = current_user.get("role", 6)
    
    # Build query with demo isolation
    query = {"dp_status": "ready", "approval_status": "approved"}
    query = add_demo_filter(query, current_user)
    
    # Get bookings with dp_status = "ready"
    bookings_list = await db.bookings.find(query, {"_id": 0}).sort("dp_ready_at", -1).to_list(1000)
    
    # Enrich with client and stock details
    for booking in bookings_list:
        client = await db.clients.find_one(
            {"id": booking.get("client_id")},
            {"_id": 0, "name": 1, "email": 1, "otc_ucc": 1, "pan_number": 1, "dp_id": 1, "depository": 1}
        )
        if client:
            booking["client_name"] = client.get("name")
            booking["client_email"] = client.get("email")
            booking["client_otc_ucc"] = client.get("otc_ucc")
            booking["client_pan"] = client.get("pan_number")
            booking["client_dp_id"] = client.get("dp_id")
            booking["client_depository"] = client.get("depository")
        
        stock = await db.stocks.find_one(
            {"id": booking.get("stock_id")},
            {"_id": 0, "name": 1, "symbol": 1}
        )
        if stock:
            booking["stock_name"] = stock.get("name")
            booking["stock_symbol"] = stock.get("symbol")
        
        # Calculate total amount
        booking["total_amount"] = (booking.get("selling_price") or 0) * booking.get("quantity", 0)
    
    return bookings_list


@router.get("/bookings/dp-transferred")
async def get_dp_transferred_bookings(
    current_user: dict = Depends(get_current_user),
    _: None = Depends(require_permission("dp.view_transfers", "view transferred bookings"))
):
    """Get all bookings where stock has been transferred"""
    user_role = current_user.get("role", 6)
    
    # Build query with demo isolation
    query = {"dp_status": "transferred"}
    query = add_demo_filter(query, current_user)
    
    # Get bookings with dp_status = "transferred"
    bookings_list = await db.bookings.find(query, {"_id": 0}).sort("dp_transferred_at", -1).to_list(1000)
    
    # Enrich with client and stock details
    for booking in bookings_list:
        client = await db.clients.find_one(
            {"id": booking.get("client_id")},
            {"_id": 0, "name": 1, "email": 1, "otc_ucc": 1, "pan_number": 1, "dp_id": 1, "depository": 1}
        )
        if client:
            booking["client_name"] = client.get("name")
            booking["client_email"] = client.get("email")
            booking["client_otc_ucc"] = client.get("otc_ucc")
            booking["client_pan"] = client.get("pan_number")
            booking["client_dp_id"] = client.get("dp_id")
            booking["client_depository"] = client.get("depository")
        
        stock = await db.stocks.find_one(
            {"id": booking.get("stock_id")},
            {"_id": 0, "name": 1, "symbol": 1}
        )
        if stock:
            booking["stock_name"] = stock.get("name")
            booking["stock_symbol"] = stock.get("symbol")
        
        booking["total_amount"] = (booking.get("selling_price") or 0) * booking.get("quantity", 0)
    
    return bookings_list


@router.get("/bookings-export")
async def export_bookings(
    format: str = "xlsx",  # "xlsx" or "csv"
    status: Optional[str] = None,
    approval_status: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
    _: None = Depends(require_permission("bookings.export", "export bookings"))
):
    """Export bookings to Excel or CSV"""
    # Build query
    query = {"is_voided": {"$ne": True}}
    if status:
        query["status"] = status
    if approval_status:
        query["approval_status"] = approval_status
    
    # Get bookings
    bookings_data = await db.bookings.find(query, {"_id": 0}).sort("created_at", -1).to_list(10000)
    
    # Get all clients, stocks for lookup
    clients = {c["id"]: c for c in await db.clients.find({}, {"_id": 0, "id": 1, "name": 1, "pan_number": 1}).to_list(10000)}
    stocks = {s["id"]: s for s in await db.stocks.find({}, {"_id": 0, "id": 1, "name": 1, "symbol": 1}).to_list(10000)}
    users = {u["id"]: u for u in await db.users.find({}, {"_id": 0, "id": 1, "name": 1, "email": 1}).to_list(10000)}
    
    if format == "csv":
        # CSV Export
        import csv
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Headers
        headers = [
            "Booking Number", "Booking Date", "Client Name", "Client PAN",
            "Stock Symbol", "Stock Name", "Quantity", "Selling Price",
            "Total Amount", "Landing Price", "Profit/Loss", "Status",
            "Approval Status", "Created By", "Created At", "Notes"
        ]
        writer.writerow(headers)
        
        # Data rows
        for booking in bookings_data:
            client = clients.get(booking.get("client_id"), {})
            stock = stocks.get(booking.get("stock_id"), {})
            created_by_user = users.get(booking.get("created_by"), {})
            
            quantity = booking.get("quantity", 0)
            selling_price = booking.get("selling_price", 0)
            landing_price = booking.get("landing_price", 0)
            total_amount = quantity * selling_price
            profit_loss = (selling_price - landing_price) * quantity if landing_price else 0
            
            row = [
                booking.get("booking_number", ""),
                booking.get("booking_date", ""),
                client.get("name", ""),
                client.get("pan_number", ""),
                stock.get("symbol", ""),
                stock.get("name", ""),
                quantity,
                selling_price,
                total_amount,
                landing_price or "",
                profit_loss if landing_price else "",
                booking.get("status", ""),
                booking.get("approval_status", ""),
                created_by_user.get("name", ""),
                booking.get("created_at", "")[:19] if booking.get("created_at") else "",
                booking.get("notes", "")
            ]
            writer.writerow(row)
        
        output.seek(0)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=bookings_export_{datetime.now().strftime('%Y%m%d')}.csv"}
        )
    
    else:
        # Excel Export
        wb = Workbook()
        ws = wb.active
        ws.title = "Bookings"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="064E3B", end_color="064E3B", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = [
            "Booking Number", "Booking Date", "Client Name", "Client PAN",
            "Stock Symbol", "Stock Name", "Quantity", "Selling Price",
            "Total Amount", "Landing Price", "Profit/Loss", "Status",
            "Approval Status", "Created By", "Created At", "Notes"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Data rows
        for row_num, booking in enumerate(bookings_data, 2):
            client = clients.get(booking.get("client_id"), {})
            stock = stocks.get(booking.get("stock_id"), {})
            created_by_user = users.get(booking.get("created_by"), {})
            
            quantity = booking.get("quantity", 0)
            selling_price = booking.get("selling_price", 0)
            landing_price = booking.get("landing_price", 0)
            total_amount = quantity * selling_price
            profit_loss = (selling_price - landing_price) * quantity if landing_price else 0
            
            row_data = [
                booking.get("booking_number", ""),
                booking.get("booking_date", ""),
                client.get("name", ""),
                client.get("pan_number", ""),
                stock.get("symbol", ""),
                stock.get("name", ""),
                quantity,
                selling_price,
                total_amount,
                landing_price or "",
                profit_loss if landing_price else "",
                booking.get("status", ""),
                booking.get("approval_status", ""),
                created_by_user.get("name", ""),
                booking.get("created_at", "")[:19] if booking.get("created_at") else "",
                booking.get("notes", "")
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.border = thin_border
                if col in [7, 8, 9, 10, 11]:  # Number columns
                    cell.alignment = Alignment(horizontal="right")
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col) if col <= 26 else 'A' + chr(64 + col - 26)].width = 15
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=bookings_export_{datetime.now().strftime('%Y%m%d')}.xlsx"}
        )


@router.get("/bookings/dp-export")
async def export_dp_transfer_excel(
    status: str = "all",  # "ready", "transferred", or "all"
    current_user: dict = Depends(get_current_user),
    _: None = Depends(require_permission("dp.view_transfers", "export DP data"))
):
    """Export DP transfer data to Excel"""
    user_role = current_user.get("role", 6)
    
    # Build query based on status
    query = {"approval_status": "approved"}
    if status == "ready":
        query["dp_status"] = "ready"
    elif status == "transferred":
        query["dp_status"] = "transferred"
    else:
        query["dp_status"] = {"$in": ["ready", "transferred"]}
    
    # Get bookings
    bookings_data = await db.bookings.find(query, {"_id": 0}).to_list(10000)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "DP Transfer"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        "Booking #", "Client Name", "Client DP ID", "Client PAN",
        "Stock Symbol", "Stock Name", "ISIN", "Quantity",
        "Amount", "Status", "DP Type", "Transfer Date"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Data rows
    row = 2
    for booking in bookings_data:
        # Get client details
        client = await db.clients.find_one(
            {"id": booking.get("client_id")},
            {"_id": 0, "name": 1, "dp_id": 1, "pan": 1, "otc_ucc": 1}
        )
        
        # Get stock details
        stock = await db.stocks.find_one(
            {"id": booking.get("stock_id")},
            {"_id": 0, "name": 1, "symbol": 1, "isin": 1}
        )
        
        # Calculate total amount
        total_amount = (booking.get("selling_price") or 0) * booking.get("quantity", 0)
        
        # Get DP ID - use otc_ucc or dp_id
        dp_id = ""
        if client:
            dp_id = client.get("otc_ucc") or client.get("dp_id") or ""
        
        data = [
            booking.get("booking_number", ""),
            client.get("name", "") if client else "",
            dp_id,
            client.get("pan", "") if client else "",
            stock.get("symbol", "") if stock else "",
            stock.get("name", "") if stock else "",
            stock.get("isin", "") if stock else "",
            booking.get("quantity", 0),
            total_amount,
            "READY" if booking.get("dp_status") == "ready" else "TRANSFERRED",
            booking.get("dp_type", ""),
            booking.get("dp_transferred_at", "")[:10] if booking.get("dp_transferred_at") else ""
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin_border
            if col in [8, 9]:  # Quantity and Amount columns
                cell.alignment = Alignment(horizontal="right")
        
        row += 1
    
    # Adjust column widths
    column_widths = [15, 25, 20, 15, 15, 30, 20, 12, 15, 15, 10, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"dp_transfer_{status}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# IMPORTANT: Static routes MUST be defined BEFORE dynamic routes like /bookings/{booking_id}
# to prevent FastAPI from matching 'pending-approval' as a booking_id parameter

@router.get("/bookings/pending-approval", response_model=List[BookingWithDetails])
async def get_pending_bookings(
    current_user: dict = Depends(get_current_user),
    _: None = Depends(require_permission("bookings.approve", "view pending bookings"))
):
    """Get bookings pending approval (PE Level only)."""
    user_role = current_user.get("role", 6)
    
    # Build query with demo isolation
    query = {"approval_status": "pending"}
    query = add_demo_filter(query, current_user)
    
    bookings = await db.bookings.find(
        query,
        {"_id": 0, "user_id": 0}
    ).to_list(1000)
    
    if not bookings:
        return []
    
    # Enrich with client and stock details
    client_ids = list(set(b["client_id"] for b in bookings))
    stock_ids = list(set(b["stock_id"] for b in bookings))
    
    clients = await db.clients.find({"id": {"$in": client_ids}}, {"_id": 0}).to_list(1000)
    stocks = await db.stocks.find({"id": {"$in": stock_ids}}, {"_id": 0}).to_list(1000)
    
    client_map = {c["id"]: c for c in clients}
    stock_map = {s["id"]: s for s in stocks}
    
    enriched = []
    for b in bookings:
        client = client_map.get(b["client_id"], {})
        stock = stock_map.get(b["stock_id"], {})
        enriched.append(BookingWithDetails(
            **b,
            client_name=client.get("name", "Unknown"),
            stock_symbol=stock.get("symbol", "Unknown"),
            stock_name=stock.get("name", "Unknown")
        ))
    
    return enriched


@router.get("/bookings/pending-loss-approval", response_model=List[BookingWithDetails])
async def get_pending_loss_bookings(
    current_user: dict = Depends(get_current_user),
    _: None = Depends(require_permission("bookings.approve", "view pending loss bookings"))
):
    """Get loss bookings pending approval (PE Level only)."""
    
    # Build query with demo isolation
    query = {"is_loss_booking": True, "loss_approval_status": "pending"}
    query = add_demo_filter(query, current_user)
    
    bookings = await db.bookings.find(
        query,
        {"_id": 0, "user_id": 0}
    ).to_list(1000)
    
    if not bookings:
        return []
    
    client_ids = list(set(b["client_id"] for b in bookings))
    stock_ids = list(set(b["stock_id"] for b in bookings))
    
    clients = await db.clients.find({"id": {"$in": client_ids}}, {"_id": 0}).to_list(1000)
    stocks = await db.stocks.find({"id": {"$in": stock_ids}}, {"_id": 0}).to_list(1000)
    
    client_map = {c["id"]: c for c in clients}
    stock_map = {s["id"]: s for s in stocks}
    
    enriched = []
    for b in bookings:
        client = client_map.get(b["client_id"], {})
        stock = stock_map.get(b["stock_id"], {})
        enriched.append(BookingWithDetails(
            **b,
            client_name=client.get("name", "Unknown"),
            stock_symbol=stock.get("symbol", "Unknown"),
            stock_name=stock.get("name", "Unknown")
        ))
    
    return enriched


@router.get("/bookings/pending-bp-overrides")
async def get_pending_bp_overrides(
    current_user: dict = Depends(get_current_user),
    _: None = Depends(require_permission("bookings.approve_revenue_override", "view pending BP overrides"))
):
    """
    Get all bookings with pending BP revenue share overrides.
    Only users with 'bookings.approve_revenue_override' permission can view these.
    """
    bookings_list = await db.bookings.find(
        {"bp_override_approval_status": "pending"},
        {"_id": 0}
    ).sort("created_at", -1).to_list(1000)
    
    # Enrich with client and stock details
    for booking in bookings_list:
        client = await db.clients.find_one(
            {"id": booking.get("client_id")},
            {"_id": 0, "name": 1, "otc_ucc": 1}
        )
        if client:
            booking["client_name"] = client.get("name")
            booking["client_otc_ucc"] = client.get("otc_ucc")
        
        stock = await db.stocks.find_one(
            {"id": booking.get("stock_id")},
            {"_id": 0, "name": 1, "symbol": 1}
        )
        if stock:
            booking["stock_name"] = stock.get("name")
            booking["stock_symbol"] = stock.get("symbol")
        
        # Get creator details
        creator = await db.users.find_one(
            {"id": booking.get("created_by")},
            {"_id": 0, "name": 1}
        )
        if creator:
            booking["created_by_name"] = creator.get("name")
    
    return bookings_list


@router.post("/bookings/{booking_id}/refresh-status")
async def refresh_booking_status(
    booking_id: str,
    current_user: dict = Depends(get_current_user),
    _: None = Depends(require_permission("bookings.edit", "refresh booking status"))
):
    """
    Refresh booking status by checking:
    1. If all payments are done -> update payment_status to 'paid'
    2. If client approval is pending and 1st payment is made -> auto-approve client
    3. Recalculate any derived statuses
    
    Returns updated booking information and any actions taken.
    """
    booking = await db.bookings.find_one({"id": booking_id}, {"_id": 0})
    if not booking:
        raise HTTPException(status_code=404, detail="Booking not found")
    
    actions_taken = []
    updates = {}
    
    # Get client info
    client = await db.clients.find_one({"id": booking["client_id"]}, {"_id": 0})
    
    # Calculate payment status
    quantity = booking.get("quantity", 0)
    selling_price = booking.get("selling_price", 0)
    total_amount = quantity * selling_price
    payments = booking.get("payments", [])
    total_paid = sum(p.get("amount", 0) for p in payments)
    
    # Check payment status
    if total_paid >= total_amount and total_amount > 0:
        if booking.get("payment_status") != "paid":
            updates["payment_status"] = "paid"
            actions_taken.append("Payment status updated to 'paid' (fully paid)")
    elif total_paid > 0:
        if booking.get("payment_status") != "partial":
            updates["payment_status"] = "partial"
            actions_taken.append(f"Payment status updated to 'partial' (₹{total_paid:,.2f} of ₹{total_amount:,.2f})")
    else:
        if booking.get("payment_status") != "pending":
            updates["payment_status"] = "pending"
            actions_taken.append("Payment status updated to 'pending' (no payments)")
    
    # Check if client needs auto-approval (1st payment made)
    if client and client.get("approval_status") == "pending":
        if len(payments) > 0 and total_paid > 0:
            # Auto-approve client when first payment is made
            await db.clients.update_one(
                {"id": booking["client_id"]},
                {"$set": {
                    "approval_status": "approved",
                    "approved_at": datetime.now(timezone.utc).isoformat(),
                    "approved_by": current_user["id"],
                    "approved_by_name": current_user["name"],
                    "auto_approved_reason": f"Auto-approved due to first payment on booking {booking.get('booking_number')}"
                }}
            )
            actions_taken.append(f"Client '{client.get('name')}' auto-approved (first payment received)")
    
    # Check DP status based on payment completion
    if booking.get("approval_status") == "approved" and total_paid >= total_amount and total_amount > 0:
        if booking.get("dp_status") != "ready" and not booking.get("stock_transferred"):
            updates["dp_status"] = "ready"
            actions_taken.append("DP status updated to 'ready' (booking approved and fully paid)")
    
    # Apply updates if any
    if updates:
        updates["updated_at"] = datetime.now(timezone.utc).isoformat()
        await db.bookings.update_one({"id": booking_id}, {"$set": updates})
    
    if not actions_taken:
        actions_taken.append("No status changes needed - booking is up to date")
    
    # Fetch updated booking
    updated_booking = await db.bookings.find_one({"id": booking_id}, {"_id": 0})
    
    return {
        "message": "Booking status refreshed",
        "actions_taken": actions_taken,
        "booking": {
            "id": updated_booking.get("id"),
            "booking_number": updated_booking.get("booking_number"),
            "payment_status": updated_booking.get("payment_status"),
            "approval_status": updated_booking.get("approval_status"),
            "dp_status": updated_booking.get("dp_status"),
            "total_amount": round(total_amount, 2),
            "total_paid": round(total_paid, 2),
            "remaining": round(total_amount - total_paid, 2)
        }
    }


@router.get("/bookings/{booking_id}", response_model=BookingWithDetails)
async def get_booking(
    booking_id: str, 
    current_user: dict = Depends(get_current_user),
    _: None = Depends(require_permission("bookings.view", "view booking details"))
):
    """Get a specific booking by ID."""
    booking = await db.bookings.find_one({"id": booking_id}, {"_id": 0})
    if not booking:
        raise HTTPException(status_code=404, detail="Booking not found")
    
    # Verify demo data access
    require_demo_access(booking, current_user)
    
    # Employee can only view their own bookings
    user_role = current_user.get("role", 5)
    if user_role == 4 and booking.get("created_by") != current_user["id"]:
        raise HTTPException(status_code=403, detail="You can only view your own bookings")
    
    client = await db.clients.find_one({"id": booking["client_id"]}, {"_id": 0})
    stock = await db.stocks.find_one({"id": booking["stock_id"]}, {"_id": 0})
    
    quantity = booking.get("quantity", 0)
    selling_price = booking.get("selling_price", 0)
    buying_price = booking.get("buying_price", 0)
    total_amount = quantity * selling_price
    
    payments = booking.get("payments", [])
    total_paid = round(sum(p.get("amount", 0) for p in payments), 2)
    
    return BookingWithDetails(**{
        **booking,
        "client_name": client["name"] if client else "Unknown",
        "client_email": client.get("email") if client else None,
        "stock_symbol": stock["symbol"] if stock else "Unknown",
        "stock_name": stock["name"] if stock else "Unknown",
        "total_amount": round(total_amount, 2),
        "profit_loss": round((selling_price - buying_price) * quantity, 2),
        "total_paid": total_paid,
        "payment_status": "paid" if total_paid >= total_amount else ("partial" if total_paid > 0 else "pending")
    })


@router.put("/bookings/{booking_id}", response_model=Booking)
async def update_booking(
    booking_id: str, 
    booking_data: BookingCreate, 
    current_user: dict = Depends(get_current_user),
    _: None = Depends(require_permission("bookings.edit", "edit bookings"))
):
    """Update a booking."""
    user_role = current_user.get("role", 7)
    
    await check_permission(current_user, "bookings.edit", "edit bookings")
    
    old_booking = await db.bookings.find_one({"id": booking_id}, {"_id": 0})
    if not old_booking:
        raise HTTPException(status_code=404, detail="Booking not found")
    
    client = await db.clients.find_one({"id": booking_data.client_id}, {"_id": 0})
    stock = await db.stocks.find_one({"id": booking_data.stock_id}, {"_id": 0})
    
    result = await db.bookings.update_one(
        {"id": booking_id},
        {"$set": booking_data.model_dump()}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Booking not found")
    
    await update_inventory(booking_data.stock_id)
    
    await create_audit_log(
        action="BOOKING_UPDATE",
        entity_type="booking",
        entity_id=booking_id,
        user_id=current_user["id"],
        user_name=current_user["name"],
        user_role=user_role,
        entity_name=f"{stock['symbol'] if stock else 'Unknown'} - {client['name'] if client else 'Unknown'}"
    )
    
    updated_booking = await db.bookings.find_one({"id": booking_id}, {"_id": 0, "user_id": 0})
    return updated_booking


@router.put("/bookings/{booking_id}/referral-partner")
async def update_booking_referral_partner(
    booking_id: str,
    referral_partner_id: Optional[str] = None,
    rp_revenue_share_percent: Optional[float] = None,
    current_user: dict = Depends(get_current_user),
    _: None = Depends(require_permission("bookings.edit", "update booking RP"))
):
    """
    Update the referral partner mapping for a booking.
    
    - Can change or remove RP assignment
    - Can update RP revenue share percentage
    - Only allowed before stock transfer is completed
    """
    booking = await db.bookings.find_one({"id": booking_id}, {"_id": 0})
    if not booking:
        raise HTTPException(status_code=404, detail="Booking not found")
    
    if booking.get("stock_transferred"):
        raise HTTPException(
            status_code=400, 
            detail="Cannot change RP mapping after stock has been transferred"
        )
    
    if booking.get("is_bp_booking"):
        raise HTTPException(
            status_code=400,
            detail="Cannot assign RP to a Business Partner booking"
        )
    
    updates = {"updated_at": datetime.now(timezone.utc).isoformat()}
    old_rp_id = booking.get("referral_partner_id")
    old_rp_share = booking.get("rp_revenue_share_percent")
    
    if referral_partner_id:
        # Validate RP exists and is active
        rp = await db.referral_partners.find_one(
            {"id": referral_partner_id, "is_active": True, "approval_status": "approved"},
            {"_id": 0}
        )
        if not rp:
            raise HTTPException(status_code=404, detail="Referral Partner not found or not approved")
        
        updates["referral_partner_id"] = referral_partner_id
        updates["rp_code"] = rp.get("rp_code")
        updates["rp_name"] = rp.get("name")
        updates["rp_revenue_share_percent"] = rp_revenue_share_percent if rp_revenue_share_percent is not None else 30.0
        updates["employee_revenue_share_percent"] = 100.0 - updates["rp_revenue_share_percent"]
    else:
        # Remove RP assignment
        updates["referral_partner_id"] = None
        updates["rp_code"] = None
        updates["rp_name"] = None
        updates["rp_revenue_share_percent"] = 0
        updates["employee_revenue_share_percent"] = 100.0
    
    await db.bookings.update_one({"id": booking_id}, {"$set": updates})
    
    # Create audit log
    await create_audit_log(
        action="BOOKING_RP_UPDATE",
        entity_type="booking",
        entity_id=booking_id,
        user_id=current_user["id"],
        user_name=current_user["name"],
        user_role=current_user.get("role", 6),
        entity_name=booking.get("booking_number"),
        details={
            "old_rp_id": old_rp_id,
            "new_rp_id": referral_partner_id,
            "old_rp_share": old_rp_share,
            "new_rp_share": updates.get("rp_revenue_share_percent")
        }
    )
    
    return {
        "message": "Referral partner mapping updated",
        "referral_partner_id": updates.get("referral_partner_id"),
        "rp_name": updates.get("rp_name"),
        "rp_revenue_share_percent": updates.get("rp_revenue_share_percent")
    }


@router.delete("/bookings/{booking_id}")
async def delete_booking(
    booking_id: str,
    current_user: dict = Depends(get_current_user),
    _: None = Depends(require_permission("bookings.delete", "delete bookings"))
):
    """Delete a booking (requires bookings.delete permission)."""
    user_role = current_user.get("role", 6)
    
    booking = await db.bookings.find_one({"id": booking_id}, {"_id": 0})
    if not booking:
        raise HTTPException(status_code=404, detail="Booking not found")
    
    if booking.get("stock_transferred"):
        raise HTTPException(status_code=400, detail="Cannot delete a booking where stock has already been transferred")
    
    stock_id = booking["stock_id"]
    
    result = await db.bookings.delete_one({"id": booking_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Booking not found")
    
    await update_inventory(stock_id)
    
    await create_audit_log(
        action="BOOKING_DELETE",
        entity_type="booking",
        entity_id=booking_id,
        user_id=current_user["id"],
        user_name=current_user["name"],
        user_role=user_role,
        entity_name=booking.get("booking_number", booking_id[:8])
    )
    
    return {"message": "Booking deleted successfully"}


class PaymentRecordRequest(BaseModel):
    """Request model for recording a payment tranche"""
    amount: float
    payment_date: Optional[str] = None
    payment_mode: str = "bank_transfer"
    reference_number: Optional[str] = None
    notes: Optional[str] = None
    proof_url: Optional[str] = None


@router.post("/bookings/{booking_id}/payments")
async def add_payment_tranche(
    booking_id: str,
    payment_data: PaymentRecordRequest,
    current_user: dict = Depends(get_current_user),
    _: None = Depends(require_permission("bookings.record_payment", "record payments"))
):
    """Add a payment tranche to a booking."""
    user_role = current_user.get("role", 6)
    
    booking = await db.bookings.find_one({"id": booking_id}, {"_id": 0})
    if not booking:
        raise HTTPException(status_code=404, detail="Booking not found")
    
    payments = booking.get("payments", [])
    total_amount = (booking.get("selling_price") or 0) * booking.get("quantity", 0)
    current_paid = sum(p.get("amount", 0) for p in payments)
    remaining = total_amount - current_paid
    
    amount = payment_data.amount
    if amount > remaining + 0.01:
        raise HTTPException(
            status_code=400,
            detail=f"Payment amount ({amount}) exceeds remaining balance ({remaining:.2f})"
        )
    
    tranche_number = len(payments) + 1
    payment = {
        "tranche_number": tranche_number,
        "amount": amount,
        "payment_mode": payment_data.payment_mode,
        "payment_date": payment_data.payment_date or datetime.now(timezone.utc).strftime("%Y-%m-%d"),
        "reference_number": payment_data.reference_number,
        "notes": payment_data.notes,
        "proof_url": payment_data.proof_url,
        "recorded_by": current_user["id"],
        "recorded_by_name": current_user["name"],
        "recorded_at": datetime.now(timezone.utc).isoformat()
    }
    
    new_total_paid = current_paid + amount
    is_complete = abs(new_total_paid - total_amount) < 0.01
    is_first_payment = len(payments) == 0
    
    update_data = {
        "$push": {"payments": payment},
        "$set": {
            "total_paid": new_total_paid,
            "payment_complete": is_complete,
            "payment_status": "paid" if is_complete else ("partial" if new_total_paid > 0 else "pending")
        }
    }
    
    # First payment implies client has accepted - update client_confirmation_status
    if is_first_payment:
        update_data["$set"]["client_confirmation_status"] = "accepted"
        update_data["$set"]["client_confirmed_at"] = datetime.now(timezone.utc).isoformat()
    
    # If payment is complete, mark as DP Ready for transfer
    if is_complete:
        update_data["$set"]["dp_status"] = "ready"
        update_data["$set"]["dp_transfer_ready"] = True
        update_data["$set"]["dp_ready_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.bookings.update_one({"id": booking_id}, update_data)
    
    # Send DP Ready email to client when payment is complete
    if is_complete:
        try:
            from services.email_service import send_dp_ready_email
            client = await db.clients.find_one({"id": booking.get("client_id")}, {"_id": 0})
            stock = await db.stocks.find_one({"id": booking.get("stock_id")}, {"_id": 0})
            company_master = await db.company_master.find_one({"_id": "company_settings"}, {"_id": 0})
            
            if client and stock:
                await send_dp_ready_email(
                    client=client,
                    booking=booking,
                    stock=stock,
                    company_master=company_master or {},
                    cc_email=current_user.get("email")
                )
        except Exception as e:
            logger.error(f"Failed to send DP Ready email: {e}")
    
    # Send role-based notification for payment received
    try:
        from services.role_notification_service import notify_payment_received
        client = await db.clients.find_one({"id": booking.get("client_id")}, {"_id": 0, "name": 1})
        await notify_payment_received(
            booking_number=booking.get("booking_number", booking_id),
            client_name=client.get("name", "Unknown") if client else "Unknown",
            amount=amount,
            payment_mode=payment_data.payment_mode,
            recorded_by=current_user["name"],
            exclude_user_id=current_user["id"]
        )
    except Exception as e:
        logger.error(f"Failed to send payment notification: {e}")
    
    return {
        "message": f"Payment of ₹{amount:,.2f} recorded successfully",
        "tranche_number": tranche_number,
        "total_paid": new_total_paid,
        "remaining": remaining - amount,
        "payment_complete": is_complete,
        "dp_transfer_ready": is_complete,
        "client_confirmation_status": "accepted" if is_first_payment else booking.get("client_confirmation_status", "pending")
    }


@router.get("/bookings/{booking_id}/payments")
async def get_booking_payments(
    booking_id: str,
    current_user: dict = Depends(get_current_user),
    _: None = Depends(require_permission("bookings.view", "view booking payments"))
):
    """Get all payment tranches for a booking."""
    booking = await db.bookings.find_one({"id": booking_id}, {"_id": 0})
    if not booking:
        raise HTTPException(status_code=404, detail="Booking not found")
    
    total_amount = (booking.get("selling_price") or 0) * booking.get("quantity", 0)
    payments = booking.get("payments", [])
    total_paid = sum(p.get("amount", 0) for p in payments)
    
    return {
        "booking_id": booking_id,
        "booking_number": booking.get("booking_number"),
        "total_amount": total_amount,
        "total_paid": total_paid,
        "remaining": total_amount - total_paid,
        "payment_complete": booking.get("payment_complete", False),
        "payments": payments
    }


@router.delete("/bookings/{booking_id}/payments/{tranche_number}")
async def delete_payment_tranche(
    booking_id: str,
    tranche_number: int,
    current_user: dict = Depends(get_current_user),
    _: None = Depends(require_permission("bookings.delete_payment", "delete payments"))
):
    """Delete a payment tranche (requires bookings.delete_payment permission)."""
    
    booking = await db.bookings.find_one({"id": booking_id}, {"_id": 0})
    if not booking:
        raise HTTPException(status_code=404, detail="Booking not found")
    
    payments = booking.get("payments", [])
    payment_to_delete = None
    
    for p in payments:
        if p.get("tranche_number") == tranche_number:
            payment_to_delete = p
            break
    
    if not payment_to_delete:
        raise HTTPException(status_code=404, detail="Payment tranche not found")
    
    new_payments = [p for p in payments if p.get("tranche_number") != tranche_number]
    new_total_paid = sum(p.get("amount", 0) for p in new_payments)
    total_amount = (booking.get("selling_price") or 0) * booking.get("quantity", 0)
    is_complete = abs(new_total_paid - total_amount) < 0.01
    
    update_data = {
        "payments": new_payments,
        "total_paid": new_total_paid,
        "payment_complete": is_complete
    }
    
    # If no payments left, reset client confirmation status
    if len(new_payments) == 0:
        update_data["client_confirmation_status"] = "pending"
        update_data["client_confirmed_at"] = None
    
    # ALWAYS reset DP ready status when any payment is deleted
    # DP Ready should only be true when payment is 100% complete
    update_data["dp_status"] = "ready" if is_complete else None
    update_data["dp_transfer_ready"] = is_complete
    if not is_complete:
        update_data["dp_ready_at"] = None
    
    await db.bookings.update_one({"id": booking_id}, {"$set": update_data})
    
    return {
        "message": f"Payment tranche {tranche_number} deleted successfully",
        "total_paid": new_total_paid,
        "remaining": total_amount - new_total_paid,
        "payment_complete": is_complete,
        "dp_transfer_ready": is_complete,
        "dp_status": "ready" if is_complete else None
    }


# ============== CLIENT BOOKING CONFIRMATION (Public Endpoint) ==============

@router.get("/booking-confirm/{booking_id}/{token}/accept")
async def client_confirm_booking_accept(booking_id: str, token: str):
    """
    Client accepts booking via email link (public endpoint - no auth required).
    This is called when client clicks "I Confirm" in the email.
    """
    booking = await db.bookings.find_one({"id": booking_id}, {"_id": 0})
    
    if not booking:
        raise HTTPException(status_code=404, detail="Booking not found or has been cancelled")
    
    # Verify the confirmation token
    if booking.get("client_confirmation_token") != token:
        raise HTTPException(status_code=400, detail="Invalid or expired confirmation link")
    
    # Check if already confirmed
    if booking.get("client_confirmation_status") == "accepted":
        return {
            "status": "accepted",
            "message": "This booking has already been confirmed",
            "booking_number": booking.get("booking_number")
        }
    
    if booking.get("client_confirmation_status") == "denied":
        raise HTTPException(status_code=400, detail="This booking has already been denied by the client")
    
    # Check booking approval status
    if booking.get("approval_status") != "approved":
        return {
            "status": "pending_approval",
            "message": "This booking is still pending internal approval",
            "booking_number": booking.get("booking_number")
        }
    
    # Check if booking is voided
    if booking.get("is_voided"):
        raise HTTPException(status_code=400, detail="This booking has been cancelled")
    
    # Update booking with client confirmation
    await db.bookings.update_one(
        {"id": booking_id},
        {
            "$set": {
                "client_confirmation_status": "accepted",
                "client_confirmed": True,
                "client_confirmed_at": datetime.now(timezone.utc).isoformat()
            }
        }
    )
    
    # Notify booking creator
    if booking.get("created_by"):
        await create_notification(
            booking["created_by"],
            "client_confirmed_booking",
            f"Client has confirmed booking {booking.get('booking_number')}",
            {"booking_id": booking_id}
        )
    
    # Create audit log
    from services.audit_service import create_audit_log
    await create_audit_log(
        action="CLIENT_CONFIRMED_BOOKING",
        entity_type="booking",
        entity_id=booking_id,
        user_id="client",
        user_name="Client",
        user_role=0,
        entity_name=booking.get("booking_number"),
        details={"confirmation": "accepted"}
    )
    
    return {
        "status": "accepted",
        "message": "Thank you! Your booking has been confirmed successfully.",
        "booking_number": booking.get("booking_number")
    }


@router.post("/booking-confirm/{booking_id}/{token}/deny")
async def client_deny_booking(booking_id: str, token: str, body: Optional[dict] = None):
    """
    Client denies booking via email link (public endpoint - no auth required).
    """
    reason = body.get("reason") if body else None
    
    booking = await db.bookings.find_one({"id": booking_id}, {"_id": 0})
    
    if not booking:
        raise HTTPException(status_code=404, detail="Booking not found or has been cancelled")
    
    # Verify the confirmation token
    if booking.get("client_confirmation_token") != token:
        raise HTTPException(status_code=400, detail="Invalid or expired confirmation link")
    
    # Check if already processed
    if booking.get("client_confirmation_status") == "accepted":
        raise HTTPException(status_code=400, detail="This booking has already been confirmed and cannot be denied")
    
    if booking.get("client_confirmation_status") == "denied":
        return {
            "status": "denied",
            "message": "This booking has already been denied",
            "booking_number": booking.get("booking_number")
        }
    
    # Check if booking is voided
    if booking.get("is_voided"):
        raise HTTPException(status_code=400, detail="This booking has been cancelled")
    
    # Update booking with client denial
    await db.bookings.update_one(
        {"id": booking_id},
        {
            "$set": {
                "client_confirmation_status": "denied",
                "client_confirmed": False,
                "client_denied_at": datetime.now(timezone.utc).isoformat(),
                "client_denial_reason": reason or "No reason provided"
            }
        }
    )
    
    # Notify booking creator and PE Desk
    if booking.get("created_by"):
        await create_notification(
            booking["created_by"],
            "client_denied_booking",
            f"Client has denied booking {booking.get('booking_number')}" + (f": {reason}" if reason else ""),
            {"booking_id": booking_id, "reason": reason}
        )
    
    # Notify PE roles
    await notify_roles(
        [1, 2],  # PE Desk and PE Manager
        "client_denied_booking",
        f"Client denied booking {booking.get('booking_number')}",
        {"booking_id": booking_id, "reason": reason}
    )
    
    # Create audit log
    from services.audit_service import create_audit_log
    await create_audit_log(
        action="CLIENT_DENIED_BOOKING",
        entity_type="booking",
        entity_id=booking_id,
        user_id="client",
        user_name="Client",
        user_role=0,
        entity_name=booking.get("booking_number"),
        details={"confirmation": "denied", "reason": reason}
    )
    
    return {
        "status": "denied",
        "message": "The booking has been cancelled as per your request.",
        "booking_number": booking.get("booking_number")
    }


@router.put("/bookings/{booking_id}/confirm-transfer")
async def confirm_stock_transfer(
    booking_id: str,
    dp_receipt_number: str = Query(None),
    notes: str = Query(None),
    current_user: dict = Depends(get_current_user),
    _: None = Depends(require_permission("dp.transfer", "confirm stock transfers"))
):
    """Confirm DP stock transfer (requires dp.transfer permission)."""
    
    booking = await db.bookings.find_one({"id": booking_id}, {"_id": 0})
    if not booking:
        raise HTTPException(status_code=404, detail="Booking not found")
    
    if booking.get("approval_status") != "approved":
        raise HTTPException(status_code=400, detail="Booking must be approved before transfer")
    
    if booking.get("stock_transferred"):
        raise HTTPException(status_code=400, detail="Stock already transferred")
    
    update_data = {
        "stock_transferred": True,
        "dp_receipt_number": dp_receipt_number,
        "transfer_notes": notes,
        "transfer_date": datetime.now(timezone.utc).isoformat(),
        "transferred_by": current_user["id"],
        "transferred_by_name": current_user["name"],
        "status": "completed"
    }
    
    await db.bookings.update_one({"id": booking_id}, {"$set": update_data})
    
    # Update inventory
    await update_inventory(booking["stock_id"])
    
    # Audit log
    await create_audit_log(
        action="BOOKING_TRANSFER",
        entity_type="booking",
        entity_id=booking_id,
        user_id=current_user["id"],
        user_name=current_user["name"],
        user_role=current_user.get("role", 1),
        entity_name=booking.get("booking_number", booking_id[:8])
    )
    
    return {
        "message": "Stock transfer confirmed successfully",
        "booking_id": booking_id,
        "dp_receipt_number": dp_receipt_number
    }


@router.put("/bookings/{booking_id}/approve-loss")
async def approve_loss_booking(
    booking_id: str,
    approve: bool = True,
    current_user: dict = Depends(get_current_user),
    _: None = Depends(require_permission("bookings.approve", "approve loss bookings"))
):
    """Approve or reject a loss booking (requires bookings.approve permission)."""
    
    booking = await db.bookings.find_one({"id": booking_id}, {"_id": 0})
    if not booking:
        raise HTTPException(status_code=404, detail="Booking not found")
    
    if not booking.get("is_loss_booking", False):
        raise HTTPException(status_code=400, detail="This is not a loss booking")
    
    if booking.get("loss_approval_status") != "pending":
        raise HTTPException(status_code=400, detail="Loss booking already processed")
    
    update_data = {
        "loss_approval_status": "approved" if approve else "rejected",
        "loss_approved_by": current_user["id"],
        "loss_approved_at": datetime.now(timezone.utc).isoformat()
    }
    
    if not approve:
        update_data["status"] = "rejected"
        update_data["rejection_reason"] = "Loss booking rejected by PE"
    
    await db.bookings.update_one({"id": booking_id}, {"$set": update_data})
    
    return {"message": f"Loss booking {'approved' if approve else 'rejected'} successfully"}


@router.put("/bookings/{booking_id}/dp-transfer")
async def mark_dp_transferred(
    booking_id: str,
    dp_type: str,  # "NSDL" or "CDSL"
    current_user: dict = Depends(get_current_user),
    _: None = Depends(require_permission("dp.transfer", "mark DP as transferred"))
):
    """Mark a booking as DP transferred, generate contract note, and send notification to client"""
    from services.email_service import send_stock_transferred_email
    from services.contract_note_service import create_and_save_contract_note
    
    user_role = current_user.get("role", 6)
    
    if dp_type not in ["NSDL", "CDSL"]:
        raise HTTPException(status_code=400, detail="dp_type must be 'NSDL' or 'CDSL'")
    
    booking = await db.bookings.find_one({"id": booking_id}, {"_id": 0})
    if not booking:
        raise HTTPException(status_code=404, detail="Booking not found")
    
    if booking.get("dp_status") != "ready":
        raise HTTPException(status_code=400, detail="Booking is not in DP ready status")
    
    transfer_time = datetime.now(timezone.utc)
    
    # Update booking with transferred status
    await db.bookings.update_one(
        {"id": booking_id},
        {"$set": {
            "dp_status": "transferred",
            "dp_type": dp_type,
            "stock_transferred": True,
            "stock_transfer_date": transfer_time.strftime("%d-%b-%Y"),
            "dp_transferred_at": transfer_time.isoformat(),
            "dp_transferred_by": current_user["id"],
            "dp_transferred_by_name": current_user["name"]
        }}
    )
    
    # Update inventory - deduct from available stock
    inventory = await db.inventory.find_one({"stock_id": booking.get("stock_id")}, {"_id": 0})
    if inventory:
        await db.inventory.update_one(
            {"stock_id": booking.get("stock_id")},
            {"$inc": {"available_quantity": -booking.get("quantity", 0)}}
        )
    
    # Auto-generate and store contract note
    contract_note = None
    try:
        # Check if contract note already exists
        existing_cn = await db.contract_notes.find_one({"booking_id": booking_id}, {"_id": 0})
        if not existing_cn:
            contract_note = await create_and_save_contract_note(
                booking_id=booking_id,
                user_id=current_user["id"],
                user_name=current_user["name"]
            )
            logger.info(f"Contract note {contract_note.get('contract_note_number')} auto-generated for booking {booking_id}")
        else:
            contract_note = existing_cn
            logger.info(f"Contract note already exists for booking {booking_id}")
    except Exception as e:
        logger.error(f"Failed to auto-generate contract note for booking {booking_id}: {str(e)}")
        # Continue even if contract note generation fails
    
    # Get client and stock details for email
    client = await db.clients.find_one({"id": booking.get("client_id")}, {"_id": 0})
    stock = await db.stocks.find_one({"id": booking.get("stock_id")}, {"_id": 0})
    company_master = await db.company_master.find_one({"_id": "company_settings"}, {"_id": 0})
    
    # Send email to client
    if client and stock and company_master:
        await send_stock_transferred_email(
            booking_id=booking_id,
            client=client,
            booking=booking,
            stock=stock,
            dp_type=dp_type,
            transfer_date=transfer_time.isoformat(),
            company_master=company_master,
            cc_email=current_user.get("email")
        )
    
    await create_audit_log(
        action="DP_TRANSFERRED",
        entity_type="booking",
        entity_id=booking_id,
        user_id=current_user["id"],
        user_name=current_user["name"],
        user_role=user_role,
        details={
            "dp_type": dp_type,
            "quantity": booking.get("quantity"),
            "stock_id": booking.get("stock_id"),
            "client_id": booking.get("client_id"),
            "contract_note_generated": contract_note is not None,
            "contract_note_number": contract_note.get("contract_note_number") if contract_note else None
        },
        entity_name=booking.get("booking_number", booking_id)
    )
    
    response = {
        "message": f"Stock transferred via {dp_type}. Client notified about T+2 settlement.",
        "dp_type": dp_type,
        "quantity": booking.get("quantity"),
        "transferred_at": transfer_time.isoformat()
    }
    
    if contract_note:
        response["contract_note"] = {
            "id": contract_note.get("id"),
            "number": contract_note.get("contract_note_number"),
            "status": "generated"
        }
    
    return response



# ============== BP REVENUE OVERRIDE ENDPOINTS ==============

class BPOverrideApprovalRequest(BaseModel):
    approve: bool
    rejection_reason: Optional[str] = None


@router.put("/bookings/{booking_id}/bp-override-approval")
async def approve_bp_revenue_override(
    booking_id: str,
    request: BPOverrideApprovalRequest,
    current_user: dict = Depends(get_current_user),
    _: None = Depends(require_permission("bookings.approve_revenue_override", "approve BP revenue override"))
):
    """
    Approve or reject a BP revenue share override.
    
    - If approved: The override percentage becomes effective
    - If rejected: The original BP revenue share is restored
    
    Only users with 'bookings.approve_revenue_override' permission can perform this action.
    """
    booking = await db.bookings.find_one({"id": booking_id}, {"_id": 0})
    if not booking:
        raise HTTPException(status_code=404, detail="Booking not found")
    
    if booking.get("bp_override_approval_status") != "pending":
        raise HTTPException(
            status_code=400, 
            detail=f"BP override is not pending approval. Current status: {booking.get('bp_override_approval_status')}"
        )
    
    if not booking.get("is_bp_booking"):
        raise HTTPException(status_code=400, detail="This is not a BP booking")
    
    if not request.approve and not request.rejection_reason:
        raise HTTPException(status_code=400, detail="Rejection reason is required when rejecting an override")
    
    update_data = {
        "bp_override_approval_status": "approved" if request.approve else "rejected",
        "bp_override_approved_by": current_user["id"],
        "bp_override_approved_at": datetime.now(timezone.utc).isoformat()
    }
    
    if request.approve:
        # Keep the override percentage - it's already in bp_revenue_share_percent
        pass
    else:
        # Rejected - restore original BP revenue share
        original_share = booking.get("bp_original_revenue_share", 0)
        update_data["bp_revenue_share_percent"] = original_share
        update_data["bp_override_rejection_reason"] = request.rejection_reason
        # Recalculate employee share based on restored BP share
        update_data["employee_revenue_share_percent"] = 100.0 - original_share
    
    await db.bookings.update_one({"id": booking_id}, {"$set": update_data})
    
    # Create audit log
    await create_audit_log(
        action="BP_OVERRIDE_APPROVAL" if request.approve else "BP_OVERRIDE_REJECTION",
        entity_type="booking",
        entity_id=booking_id,
        user_id=current_user["id"],
        user_name=current_user["name"],
        user_role=current_user.get("role", 6),
        entity_name=booking.get("booking_number"),
        details={
            "approved": request.approve,
            "override_percent": booking.get("bp_revenue_share_override"),
            "original_percent": booking.get("bp_original_revenue_share"),
            "rejection_reason": request.rejection_reason if not request.approve else None,
            "bp_name": booking.get("bp_name")
        }
    )
    
    # Notify the booking creator
    if booking.get("created_by"):
        stock = await db.stocks.find_one({"id": booking.get("stock_id")}, {"_id": 0, "symbol": 1})
        await create_notification(
            booking["created_by"],
            "bp_override_approved" if request.approve else "bp_override_rejected",
            f"BP Override {'Approved' if request.approve else 'Rejected'}",
            f"Your BP revenue share override for booking {booking.get('booking_number')} ({stock.get('symbol', 'N/A')}) has been {'approved' if request.approve else 'rejected'}. " +
            (f"Override of {booking.get('bp_revenue_share_override')}% is now effective." if request.approve else f"Reason: {request.rejection_reason}"),
            {"booking_id": booking_id, "booking_number": booking.get("booking_number")}
        )
    
    return {
        "message": f"BP revenue override {'approved' if request.approve else 'rejected'} successfully",
        "booking_id": booking_id,
        "override_status": "approved" if request.approve else "rejected",
        "effective_bp_share": booking.get("bp_revenue_share_percent") if request.approve else booking.get("bp_original_revenue_share", 0)
    }


@router.put("/bookings/{booking_id}/bp-override")
async def update_bp_revenue_override(
    booking_id: str,
    override_percent: float = Query(..., ge=0, le=100, description="New override percentage"),
    current_user: dict = Depends(get_current_user),
    _: None = Depends(require_permission("bookings.edit_revenue_override", "edit BP revenue override"))
):
    """
    Edit the BP revenue share override for a booking.
    
    - Can be used to modify the override before approval
    - Override must be <= the BP's default revenue share
    - Resets approval status to 'pending' if already approved
    
    Only users with 'bookings.edit_revenue_override' permission can perform this action.
    """
    booking = await db.bookings.find_one({"id": booking_id}, {"_id": 0})
    if not booking:
        raise HTTPException(status_code=404, detail="Booking not found")
    
    if not booking.get("is_bp_booking"):
        raise HTTPException(status_code=400, detail="This is not a BP booking")
    
    if booking.get("stock_transferred"):
        raise HTTPException(status_code=400, detail="Cannot modify override after stock has been transferred")
    
    # Get the original BP revenue share to validate
    original_share = booking.get("bp_original_revenue_share")
    if original_share is None:
        # If no original stored, get from BP profile
        bp_info = await db.business_partners.find_one(
            {"id": booking.get("business_partner_id")},
            {"_id": 0, "revenue_share_percent": 1}
        )
        original_share = bp_info.get("revenue_share_percent", 0) if bp_info else 0
    
    if override_percent > original_share:
        raise HTTPException(
            status_code=400,
            detail=f"Override ({override_percent}%) cannot be higher than the BP's default share ({original_share}%)"
        )
    
    # Determine if this needs re-approval
    needs_approval = override_percent != original_share
    
    update_data = {
        "bp_revenue_share_override": override_percent,
        "bp_revenue_share_percent": override_percent,
        "bp_original_revenue_share": original_share,
        "employee_revenue_share_percent": 100.0 - override_percent,
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    if needs_approval:
        update_data["bp_override_approval_status"] = "pending"
        update_data["bp_override_approved_by"] = None
        update_data["bp_override_approved_at"] = None
        update_data["bp_override_rejection_reason"] = None
    else:
        update_data["bp_override_approval_status"] = "not_required"
    
    await db.bookings.update_one({"id": booking_id}, {"$set": update_data})
    
    # Create audit log
    await create_audit_log(
        action="BP_OVERRIDE_EDIT",
        entity_type="booking",
        entity_id=booking_id,
        user_id=current_user["id"],
        user_name=current_user["name"],
        user_role=current_user.get("role", 6),
        entity_name=booking.get("booking_number"),
        details={
            "old_override": booking.get("bp_revenue_share_override"),
            "new_override": override_percent,
            "original_share": original_share,
            "needs_approval": needs_approval
        }
    )
    
    return {
        "message": "BP revenue override updated successfully",
        "override_percent": override_percent,
        "needs_approval": needs_approval,
        "approval_status": "pending" if needs_approval else "not_required"
    }

