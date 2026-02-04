"""
Business Intelligence Report Builder
Generates custom reports with multiple dimensions and filters
"""
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from typing import Optional, List, Any
from datetime import datetime, timezone, timedelta
from pydantic import BaseModel
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from database import db
from utils.auth import get_current_user
from services.permission_service import require_permission

router = APIRouter(prefix="/bi-reports", tags=["Business Intelligence"])


class ReportFilter(BaseModel):
    field: str
    operator: str  # eq, ne, gt, lt, gte, lte, in, contains, between
    value: Any
    value2: Optional[Any] = None  # For 'between' operator


class ReportRequest(BaseModel):
    report_type: str  # bookings, clients, revenue, inventory, payments, pnl
    dimensions: List[str]  # Fields to group by
    metrics: List[str]  # Fields to aggregate
    filters: Optional[List[ReportFilter]] = []
    date_from: Optional[str] = None
    date_to: Optional[str] = None
    sort_by: Optional[str] = None
    sort_order: Optional[str] = "desc"
    limit: Optional[int] = 1000


# Report type configurations
REPORT_CONFIGS = {
    "bookings": {
        "collection": "bookings",
        "dimensions": [
            {"key": "client_name", "label": "Client Name"},
            {"key": "stock_symbol", "label": "Stock Symbol"},
            {"key": "created_by_name", "label": "Created By"},
            {"key": "approval_status", "label": "Approval Status"},
            {"key": "payment_status", "label": "Payment Status"},
            {"key": "dp_status", "label": "DP Status"},
            {"key": "booking_type", "label": "Booking Type"},
            {"key": "is_bp_booking", "label": "BP Booking"},
            {"key": "referral_partner_name", "label": "Referral Partner"},
            {"key": "bp_name", "label": "Business Partner"},
        ],
        "metrics": [
            {"key": "count", "label": "Count", "agg": "count"},
            {"key": "quantity", "label": "Total Quantity", "agg": "sum"},
            {"key": "buying_price", "label": "Avg Buying Price", "agg": "avg"},
            {"key": "selling_price", "label": "Avg Selling Price", "agg": "avg"},
            {"key": "total_value", "label": "Total Value", "agg": "sum", "calc": "quantity * buying_price"},
            {"key": "total_revenue", "label": "Total Revenue", "agg": "sum", "calc": "quantity * (selling_price - buying_price)"},
        ],
        "date_field": "created_at"
    },
    "clients": {
        "collection": "clients",
        "dimensions": [
            {"key": "name", "label": "Client Name"},
            {"key": "pan_number", "label": "PAN Number"},
            {"key": "approval_status", "label": "Approval Status"},
            {"key": "mapped_to_name", "label": "Mapped To"},
            {"key": "client_type", "label": "Client Type"},
            {"key": "is_suspended", "label": "Suspended"},
            {"key": "is_rp", "label": "Is RP"},
        ],
        "metrics": [
            {"key": "count", "label": "Count", "agg": "count"},
            {"key": "total_bookings", "label": "Total Bookings", "agg": "lookup"},
        ],
        "date_field": "created_at"
    },
    "revenue": {
        "collection": "bookings",
        "dimensions": [
            {"key": "referral_partner_name", "label": "Referral Partner"},
            {"key": "bp_name", "label": "Business Partner"},
            {"key": "created_by_name", "label": "Employee"},
            {"key": "stock_symbol", "label": "Stock"},
            {"key": "client_name", "label": "Client"},
        ],
        "metrics": [
            {"key": "count", "label": "Booking Count", "agg": "count"},
            {"key": "rp_revenue", "label": "RP Revenue", "agg": "sum"},
            {"key": "bp_revenue", "label": "BP Revenue", "agg": "sum"},
            {"key": "employee_revenue", "label": "Employee Revenue", "agg": "sum"},
            {"key": "company_revenue", "label": "Company Revenue", "agg": "sum"},
        ],
        "date_field": "created_at"
    },
    "inventory": {
        "collection": "inventory",
        "dimensions": [
            {"key": "stock_name", "label": "Stock Name"},
            {"key": "stock_symbol", "label": "Stock Symbol"},
        ],
        "metrics": [
            {"key": "available_quantity", "label": "Available Qty", "agg": "sum"},
            {"key": "reserved_quantity", "label": "Reserved Qty", "agg": "sum"},
            {"key": "weighted_avg_price", "label": "WAP", "agg": "avg"},
            {"key": "landing_price", "label": "Landing Price", "agg": "avg"},
            {"key": "total_value", "label": "Total Value", "agg": "sum", "calc": "available_quantity * landing_price"},
        ],
        "date_field": "updated_at"
    },
    "payments": {
        "collection": "bookings",
        "dimensions": [
            {"key": "client_name", "label": "Client"},
            {"key": "payment_status", "label": "Payment Status"},
            {"key": "created_by_name", "label": "Created By"},
        ],
        "metrics": [
            {"key": "count", "label": "Booking Count", "agg": "count"},
            {"key": "total_amount", "label": "Total Amount", "agg": "sum"},
            {"key": "paid_amount", "label": "Paid Amount", "agg": "sum"},
            {"key": "pending_amount", "label": "Pending Amount", "agg": "sum"},
        ],
        "date_field": "created_at"
    },
    "pnl": {
        "collection": "bookings",
        "dimensions": [
            {"key": "stock_symbol", "label": "Stock"},
            {"key": "client_name", "label": "Client"},
            {"key": "created_by_name", "label": "Employee"},
            {"key": "booking_type", "label": "Booking Type"},
        ],
        "metrics": [
            {"key": "count", "label": "Booking Count", "agg": "count"},
            {"key": "gross_profit", "label": "Gross Profit", "agg": "sum"},
            {"key": "gross_loss", "label": "Gross Loss", "agg": "sum"},
            {"key": "net_pnl", "label": "Net P&L", "agg": "sum"},
        ],
        "date_field": "created_at"
    }
}


REPORT_TYPE_PERMISSIONS = {
    "bookings": "reports.bi_bookings",
    "clients": "reports.bi_clients",
    "revenue": "reports.bi_revenue",
    "inventory": "reports.bi_inventory",
    "payments": "reports.bi_payments",
    "pnl": "reports.bi_pnl"
}


@router.get("/config")
async def get_report_config(
    current_user: dict = Depends(get_current_user)
):
    """Get available report configurations based on user permissions"""
    from services.permission_service import has_permission
    
    user_permissions = current_user.get("permissions", [])
    
    # Filter report types based on user permissions
    available_types = []
    for report_type, perm_key in REPORT_TYPE_PERMISSIONS.items():
        if has_permission(user_permissions, perm_key):
            label_map = {
                "bookings": "Bookings Report",
                "clients": "Clients Report", 
                "revenue": "Revenue Report",
                "inventory": "Inventory Report",
                "payments": "Payments Report",
                "pnl": "P&L Report"
            }
            desc_map = {
                "bookings": "Analyze booking data by various dimensions",
                "clients": "Client analysis and segmentation",
                "revenue": "Revenue breakdown by partners and employees",
                "inventory": "Stock inventory analysis",
                "payments": "Payment status and collection analysis",
                "pnl": "Profit and loss analysis"
            }
            available_types.append({
                "key": report_type,
                "label": label_map[report_type],
                "description": desc_map[report_type],
                "permission": perm_key
            })
    
    # Filter configs to only include available types
    available_configs = {k: v for k, v in REPORT_CONFIGS.items() if k in [t["key"] for t in available_types]}
    
    return {
        "report_types": available_types,
        "configs": available_configs,
        "can_export": has_permission(user_permissions, "reports.bi_export"),
        "can_save_templates": has_permission(user_permissions, "reports.bi_save_templates")
    }


@router.post("/generate")
async def generate_report(
    request: ReportRequest,
    current_user: dict = Depends(get_current_user)
):
    """Generate a custom report based on specified parameters"""
    from services.permission_service import has_permission
    
    if request.report_type not in REPORT_CONFIGS:
        raise HTTPException(status_code=400, detail=f"Invalid report type: {request.report_type}")
    
    # Check permission for specific report type
    required_perm = REPORT_TYPE_PERMISSIONS.get(request.report_type)
    if required_perm:
        user_permissions = current_user.get("permissions", [])
        if not has_permission(user_permissions, required_perm):
            raise HTTPException(
                status_code=403, 
                detail=f"Permission denied. You need '{required_perm}' permission for {request.report_type} reports"
            )
    
    config = REPORT_CONFIGS[request.report_type]
    collection = getattr(db, config["collection"])
    
    # Build query
    query = {}
    
    # Add date filters
    if request.date_from or request.date_to:
        date_filter = {}
        if request.date_from:
            date_filter["$gte"] = request.date_from
        if request.date_to:
            date_filter["$lte"] = request.date_to + "T23:59:59"
        query[config["date_field"]] = date_filter
    
    # Add custom filters
    for f in request.filters or []:
        if f.operator == "eq":
            query[f.field] = f.value
        elif f.operator == "ne":
            query[f.field] = {"$ne": f.value}
        elif f.operator == "gt":
            query[f.field] = {"$gt": f.value}
        elif f.operator == "lt":
            query[f.field] = {"$lt": f.value}
        elif f.operator == "gte":
            query[f.field] = {"$gte": f.value}
        elif f.operator == "lte":
            query[f.field] = {"$lte": f.value}
        elif f.operator == "in":
            query[f.field] = {"$in": f.value if isinstance(f.value, list) else [f.value]}
        elif f.operator == "contains":
            query[f.field] = {"$regex": f.value, "$options": "i"}
        elif f.operator == "between" and f.value2:
            query[f.field] = {"$gte": f.value, "$lte": f.value2}
    
    # Exclude voided/cancelled for booking-related reports
    if config["collection"] == "bookings":
        query["is_voided"] = {"$ne": True}
        query["status"] = {"$ne": "cancelled"}
    
    # Build aggregation pipeline
    pipeline = [{"$match": query}]
    
    # Group stage
    if request.dimensions:
        group_id = {dim: f"${dim}" for dim in request.dimensions}
        group_stage = {"$group": {"_id": group_id}}
        
        # Add metrics
        for metric in request.metrics:
            metric_config = next((m for m in config["metrics"] if m["key"] == metric), None)
            if metric_config:
                if metric_config["agg"] == "count":
                    group_stage["$group"][metric] = {"$sum": 1}
                elif metric_config["agg"] == "sum":
                    if "calc" in metric_config:
                        # Calculated field - handle in project stage
                        pass
                    else:
                        group_stage["$group"][metric] = {"$sum": f"${metric}"}
                elif metric_config["agg"] == "avg":
                    group_stage["$group"][metric] = {"$avg": f"${metric}"}
        
        # Add count by default
        if "count" not in request.metrics:
            group_stage["$group"]["count"] = {"$sum": 1}
        
        pipeline.append(group_stage)
        
        # Project stage to flatten results
        project_stage = {"$project": {"_id": 0}}
        for dim in request.dimensions:
            project_stage["$project"][dim] = f"$_id.{dim}"
        for metric in request.metrics:
            project_stage["$project"][metric] = 1
        if "count" not in request.metrics:
            project_stage["$project"]["count"] = 1
        
        pipeline.append(project_stage)
    
    # Sort
    if request.sort_by:
        sort_dir = -1 if request.sort_order == "desc" else 1
        pipeline.append({"$sort": {request.sort_by: sort_dir}})
    
    # Limit
    pipeline.append({"$limit": request.limit})
    
    # Execute
    results = await collection.aggregate(pipeline).to_list(request.limit)
    
    # Calculate summary
    summary = {}
    for metric in request.metrics:
        values = [r.get(metric, 0) for r in results if r.get(metric) is not None]
        if values:
            summary[metric] = {
                "total": sum(values),
                "avg": sum(values) / len(values),
                "min": min(values),
                "max": max(values),
                "count": len(values)
            }
    
    return {
        "report_type": request.report_type,
        "dimensions": request.dimensions,
        "metrics": request.metrics,
        "filters": request.filters,
        "date_range": {"from": request.date_from, "to": request.date_to},
        "total_rows": len(results),
        "data": results,
        "summary": summary,
        "generated_at": datetime.now(timezone.utc).isoformat()
    }


@router.post("/export")
async def export_report(
    request: ReportRequest,
    current_user: dict = Depends(get_current_user),
    _: None = Depends(require_permission("reports.bi_builder", "export BI report"))
):
    """Export report to Excel"""
    # Generate report data first
    if request.report_type not in REPORT_CONFIGS:
        raise HTTPException(status_code=400, detail=f"Invalid report type: {request.report_type}")
    
    config = REPORT_CONFIGS[request.report_type]
    collection = getattr(db, config["collection"])
    
    # Build query (same as generate)
    query = {}
    if request.date_from or request.date_to:
        date_filter = {}
        if request.date_from:
            date_filter["$gte"] = request.date_from
        if request.date_to:
            date_filter["$lte"] = request.date_to + "T23:59:59"
        query[config["date_field"]] = date_filter
    
    for f in request.filters or []:
        if f.operator == "eq":
            query[f.field] = f.value
        elif f.operator == "contains":
            query[f.field] = {"$regex": f.value, "$options": "i"}
    
    if config["collection"] == "bookings":
        query["is_voided"] = {"$ne": True}
        query["status"] = {"$ne": "cancelled"}
    
    # Build pipeline
    pipeline = [{"$match": query}]
    
    if request.dimensions:
        group_id = {dim: f"${dim}" for dim in request.dimensions}
        group_stage = {"$group": {"_id": group_id}}
        
        for metric in request.metrics:
            metric_config = next((m for m in config["metrics"] if m["key"] == metric), None)
            if metric_config:
                if metric_config["agg"] == "count":
                    group_stage["$group"][metric] = {"$sum": 1}
                elif metric_config["agg"] == "sum":
                    group_stage["$group"][metric] = {"$sum": f"${metric}"}
                elif metric_config["agg"] == "avg":
                    group_stage["$group"][metric] = {"$avg": f"${metric}"}
        
        if "count" not in request.metrics:
            group_stage["$group"]["count"] = {"$sum": 1}
        
        pipeline.append(group_stage)
        
        project_stage = {"$project": {"_id": 0}}
        for dim in request.dimensions:
            project_stage["$project"][dim] = f"$_id.{dim}"
        for metric in request.metrics:
            project_stage["$project"][metric] = 1
        if "count" not in request.metrics:
            project_stage["$project"]["count"] = 1
        
        pipeline.append(project_stage)
    
    if request.sort_by:
        sort_dir = -1 if request.sort_order == "desc" else 1
        pipeline.append({"$sort": {request.sort_by: sort_dir}})
    
    pipeline.append({"$limit": request.limit})
    
    results = await collection.aggregate(pipeline).to_list(request.limit)
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{request.report_type.title()} Report"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = request.dimensions + request.metrics
    if "count" not in request.metrics:
        headers.append("count")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header.replace("_", " ").title())
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")
    
    # Data rows
    for row_idx, row_data in enumerate(results, 2):
        for col_idx, header in enumerate(headers, 1):
            value = row_data.get(header, "")
            if isinstance(value, float):
                value = round(value, 2)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Summary sheet
    ws_summary = wb.create_sheet(title="Summary")
    ws_summary.cell(row=1, column=1, value="Metric").font = header_font
    ws_summary.cell(row=1, column=1).fill = header_fill
    ws_summary.cell(row=1, column=2, value="Total").font = header_font
    ws_summary.cell(row=1, column=2).fill = header_fill
    ws_summary.cell(row=1, column=3, value="Average").font = header_font
    ws_summary.cell(row=1, column=3).fill = header_fill
    
    row = 2
    for metric in request.metrics:
        values = [r.get(metric, 0) for r in results if r.get(metric) is not None]
        if values:
            ws_summary.cell(row=row, column=1, value=metric.replace("_", " ").title())
            ws_summary.cell(row=row, column=2, value=round(sum(values), 2))
            ws_summary.cell(row=row, column=3, value=round(sum(values) / len(values), 2))
            row += 1
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"bi_report_{request.report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/saved")
async def get_saved_reports(
    current_user: dict = Depends(get_current_user),
    _: None = Depends(require_permission("reports.bi_builder", "view saved reports"))
):
    """Get saved report templates"""
    reports = await db.bi_report_templates.find(
        {"$or": [{"created_by": current_user["id"]}, {"is_public": True}]},
        {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    return reports


@router.post("/save")
async def save_report_template(
    name: str = Query(...),
    description: str = Query(None),
    is_public: bool = Query(False),
    request: ReportRequest = ...,
    current_user: dict = Depends(get_current_user),
    _: None = Depends(require_permission("reports.bi_builder", "save report template"))
):
    """Save a report configuration as a template"""
    import uuid
    
    template = {
        "id": str(uuid.uuid4()),
        "name": name,
        "description": description,
        "is_public": is_public,
        "config": request.dict(),
        "created_by": current_user["id"],
        "created_by_name": current_user["name"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.bi_report_templates.insert_one(template)
    if "_id" in template:
        del template["_id"]
    
    return {"message": "Report template saved successfully", "template": template}
